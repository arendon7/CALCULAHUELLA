from __future__ import annotations

from datetime import UTC, date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from .database import (
    ConsolidationFinding,
    EmissionFactor,
    EmissionFactorVersion,
    FactorDocumentation,
    Gas,
    MethodologySourceDocument,
    Notification,
    AppUser,
    Organization,
    PilotProject,
    PilotSourceRequirement,
    ReferenceCalculationCase,
    ReleaseGate,
)

SECTOR_LIBRARY_VERSION = "0.28.0"

SOURCE_DOCUMENT_SPECS: list[dict[str, Any]] = [
    {
        "code": "IPCC-WASTE-2006-CH4-N2O",
        "title": "2006 IPCC Guidelines · Volume 5, Chapter 4: Biological Treatment of Solid Waste",
        "issuing_body": "IPCC Task Force on National Greenhouse Gas Inventories",
        "document_type": "Factor Tier 1 para tratamiento biológico",
        "publication_date": date(2006, 1, 1),
        "jurisdiction": "Internacional",
        "source_url": "https://www.ipcc-nggip.iges.or.jp/public/2006gl/pdf/5_Volume5/V5_4_Ch4_Bio_Treat.pdf",
        "citation": "IPCC. 2006 Guidelines, Volume 5, Chapter 4, Table 4.1, with official corrigenda.",
        "status": "Vigente junto con el Refinamiento 2019",
        "notes": "Incluye factores por defecto para compostaje y digestión anaerobia; se deben preferir mediciones representativas de planta cuando existan.",
    },
    {
        "code": "IPCC-WASTE-CORRIGENDA",
        "title": "IPCC 2006 Guidelines · official corrigenda for Volume 5, Chapter 4",
        "issuing_body": "IPCC Task Force on National Greenhouse Gas Inventories",
        "document_type": "Corrección oficial",
        "publication_date": date(2006, 1, 1),
        "jurisdiction": "Internacional",
        "source_url": "https://www.ipcc-nggip.iges.or.jp/public/2006gl/corrigenda9.html",
        "citation": "IPCC. Corrigenda to 2006 Guidelines: composting N2O wet basis 0.24 g/kg and anaerobic digestion CH4 wet basis 0.8 g/kg.",
        "status": "Vigente",
        "notes": "Aclara además que el factor de CH4 para digestión anaerobia ya considera la recuperación de metano.",
    },
    {
        "code": "IPCC-WASTEWATER-2019",
        "title": "2019 Refinement · Volume 5, Chapter 6: Wastewater Treatment and Discharge",
        "issuing_body": "IPCC Task Force on National Greenhouse Gas Inventories",
        "document_type": "Metodología de aguas residuales",
        "publication_date": date(2019, 5, 1),
        "jurisdiction": "Internacional",
        "source_url": "https://www.ipcc-nggip.iges.or.jp/public/2019rf/pdf/5_Volume5/19R_V5_6_Ch06_Wastewater.pdf",
        "citation": "IPCC. 2019 Refinement, Volume 5, Chapter 6.",
        "status": "Vigente junto con las Directrices 2006",
        "notes": "Requiere caracterizar sistema, carga orgánica, remoción de lodos y recuperación de CH4; no se reduce a un único factor por volumen.",
    },
    {
        "code": "UPME-FECOC-2016",
        "title": "Calculadora de Factores de Emisión de Combustibles Colombianos · FECOC 2016",
        "issuing_body": "Unidad de Planeación Minero Energética · UPME",
        "document_type": "Herramienta oficial de combustibles",
        "publication_date": date(2016, 1, 1),
        "jurisdiction": "Colombia",
        "source_url": "https://app.upme.gov.co/Calculadora_Emisiones1/new/calculadora.html",
        "citation": "UPME. Calculadora de Emisiones de Combustibles Colombianos FECOC 2016.",
        "status": "Fuente oficial identificada · transcripción controlada pendiente",
        "notes": "La V0.23 registra la fuente, pero no promueve valores a uso formal hasta completar doble revisión de combustible, unidad, PCI, densidad y alcance tecnológico.",
    },
    {
        "code": "UPME-FECOCPLUS-3-2023",
        "title": "Factores de Emisión de los Combustibles Colombianos FECOC+ 3.0",
        "issuing_body": "Unidad de Planeación Minero Energética · UPME y Universidad de Antioquia",
        "document_type": "Estudio nacional de fuentes móviles",
        "publication_date": date(2023, 11, 30),
        "jurisdiction": "Colombia",
        "source_url": "https://docs.upme.gov.co/DemandayEficiencia/Doc_Hemeroteca/FECOC%2B2-3.pdf",
        "citation": "UPME y Universidad de Antioquia. FECOC+ 3.0, informe final.",
        "status": "Fuente oficial identificada · parametrización pendiente",
        "notes": "Útil para transporte por tecnología y ciclo de conducción. No debe sustituirse por un factor genérico sin documentar vehículo, combustible y actividad.",
    },
    {
        "code": "COL-DECRETO-926-2017",
        "title": "Decreto 926 de 2017 · factores de conversión de combustibles",
        "issuing_body": "Ministerio de Hacienda y Crédito Público · República de Colombia",
        "document_type": "Norma reglamentaria",
        "publication_date": date(2017, 6, 1),
        "jurisdiction": "Colombia",
        "source_url": "https://www.suin-juriscol.gov.co/viewDocument.asp?id=30030671",
        "citation": "Colombia. Decreto 926 de 2017, factores de conversión para combustibles fósiles.",
        "status": "Vigente como referencia regulatoria",
        "notes": "Las equivalencias se cargan como uso piloto condicionado. No deben asumirse automáticamente como el mejor factor para un inventario corporativo sin evaluar objetivo, periodo y metodología.",
    },
    {
        "code": "EAAB-FECOC-2016",
        "title": "Anexo de factores de emisión para herramienta MCV · transcripción de FECOC 2016",
        "issuing_body": "Empresa de Acueducto y Alcantarillado de Bogotá · referencia a UPME FECOC",
        "document_type": "Fuente secundaria de transcripción",
        "publication_date": date(2018, 1, 1),
        "jurisdiction": "Colombia",
        "source_url": "https://www.acueducto.com.co/wps/html/resources/2018ag/huella_carbono/feb12/18Anexo_17Factores_emision_herramienta_MCV_V6.pdf",
        "citation": "EAAB. Anexo 17 de factores de emisión; valores atribuidos a UPME FECOC 2016.",
        "status": "En revisión documental",
        "notes": "Se utiliza para pruebas piloto de B10 y E10. Debe contrastarse contra la fuente primaria FECOC antes de promoverse a uso formal.",
    },
    {
        "code": "IPCC-AFOLU-2019",
        "title": "2019 Refinement · Volume 4, Chapter 11: N2O Emissions from Managed Soils",
        "issuing_body": "IPCC Task Force on National Greenhouse Gas Inventories",
        "document_type": "Metodología de suelos y fertilización",
        "publication_date": date(2019, 5, 1),
        "jurisdiction": "Internacional",
        "source_url": "https://www.ipcc-nggip.iges.or.jp/public/2019rf/pdf/4_Volume4/19R_V4_Ch11_Soils_N2O_CO2.pdf",
        "citation": "IPCC. 2019 Refinement, Volume 4, Chapter 11, Tables 11.1 and 11.3.",
        "status": "Vigente junto con las Directrices 2006",
        "notes": "Los factores se aplican sobre kg de N incorporado al suelo y deben separarse de la fabricación del fertilizante y de cambios en carbono del suelo.",
    },
]

FACTOR_SPECS: list[dict[str, Any]] = [
    {
        "name": "Compostaje de residuos orgánicos húmedos · CH4 Tier 1",
        "activity_type": "Compostaje de residuos orgánicos",
        "country": "Internacional",
        "sector": "Residuos",
        "gas_code": "CH4",
        "version": "IPCC-2006-CORR",
        "value": 4.0,
        "input_unit": "t",
        "output_unit": "kg CH4",
        "source_code": "IPCC-WASTE-2006-CH4-N2O",
        "page_reference": "Página 4.6 (PDF p. 5)",
        "table_reference": "Tabla 4.1 · base húmeda",
        "source_value": 4.0,
        "source_unit": "g CH4/kg residuo húmedo",
        "conversion": "4 g/kg equivale numéricamente a 4 kg/t",
        "quality": "B",
        "uncertainty": 100.0,
        "notes": "Factor Tier 1. Depende del tipo de residuo, humedad, aireación, temperatura y material estructurante. Sustituir por mediciones representativas cuando estén disponibles.",
    },
    {
        "name": "Compostaje de residuos orgánicos húmedos · N2O Tier 1",
        "activity_type": "Compostaje de residuos orgánicos",
        "country": "Internacional",
        "sector": "Residuos",
        "gas_code": "N2O",
        "version": "IPCC-2006-CORR",
        "value": 0.24,
        "input_unit": "t",
        "output_unit": "kg N2O",
        "source_code": "IPCC-WASTE-CORRIGENDA",
        "page_reference": "Corrección oficial de la Tabla 4.1",
        "table_reference": "Compostaje · base húmeda",
        "source_value": 0.24,
        "source_unit": "g N2O/kg residuo húmedo",
        "conversion": "0,24 g/kg equivale numéricamente a 0,24 kg/t",
        "quality": "B",
        "uncertainty": 125.0,
        "notes": "Valor corregido oficialmente. Aplicación condicionada a la pertinencia del supuesto Tier 1 y a la revelación de incertidumbre.",
    },
    {
        "name": "Digestión anaerobia en instalación de biogás · CH4 Tier 1",
        "activity_type": "Digestión anaerobia de residuos orgánicos",
        "country": "Internacional",
        "sector": "Residuos",
        "gas_code": "CH4",
        "version": "IPCC-2006-CORR",
        "value": 0.8,
        "input_unit": "t",
        "output_unit": "kg CH4",
        "source_code": "IPCC-WASTE-CORRIGENDA",
        "page_reference": "Corrección oficial de la Tabla 4.1",
        "table_reference": "Digestión anaerobia · base húmeda",
        "source_value": 0.8,
        "source_unit": "g CH4/kg residuo húmedo",
        "conversion": "0,8 g/kg equivale numéricamente a 0,8 kg/t",
        "quality": "B",
        "uncertainty": 250.0,
        "notes": "El factor ya considera recuperación de CH4. Evitar restar nuevamente el metano recuperado y evitar doble conteo con el sector energía.",
    },
    {
        "name": "Liberación directa de HFC-134a · balance de masa",
        "activity_type": "Emisiones fugitivas de refrigerante",
        "country": "Internacional",
        "sector": "Multisectorial",
        "gas_code": "HFC-134a",
        "version": "DIRECT-MASS-AR6",
        "value": 1.0,
        "input_unit": "kg",
        "output_unit": "kg HFC-134a",
        "source_code": "GHGP-GWP-2024",
        "page_reference": "Tabla de GWP100 AR6",
        "table_reference": "HFC-134a",
        "source_value": 1.0,
        "source_unit": "kg gas emitido/kg gas liberado",
        "conversion": "Masa liberada × 1; posteriormente × GWP100 AR6 1530",
        "quality": "A",
        "uncertainty": 0.0,
        "notes": "Requiere identificar el refrigerante y calcular la masa efectivamente liberada mediante balance de inventario, recargas, recuperación y disposición.",
    },
    {
        "name": "Gas natural · equivalencia regulatoria Colombia",
        "activity_type": "Combustión de gas natural",
        "country": "Colombia",
        "sector": "Multisectorial",
        "gas_code": "CO2",
        "version": "COL-D926-2017",
        "value": 1.952,
        "input_unit": "m3",
        "output_unit": "kg CO2",
        "source_code": "COL-DECRETO-926-2017",
        "page_reference": "Artículo y tabla de factores de conversión",
        "table_reference": "Gas natural",
        "source_value": 1.952,
        "source_unit": "kg CO2/m3 estándar",
        "conversion": "Actividad en m3 estándar × 1,952 kg CO2/m3",
        "quality": "B",
        "uncertainty": 0.0,
        "factor_kind": "Factor regulatorio colombiano",
        "reporting_use": "Piloto",
        "review_status": "Aprobado documentalmente",
        "approved_by": "Control metodológico V0.28",
        "notes": "Equivalencia regulatoria. Evaluar periodo, composición del gas y pertinencia frente a FECOC o factor específico antes de uso formal.",
    },
    {
        "name": "GLP · equivalencia regulatoria Colombia",
        "activity_type": "Combustión de GLP",
        "country": "Colombia",
        "sector": "Multisectorial",
        "gas_code": "CO2",
        "version": "COL-D926-2017",
        "value": 6.333,
        "input_unit": "gal",
        "output_unit": "kg CO2",
        "source_code": "COL-DECRETO-926-2017",
        "page_reference": "Tabla de factores de conversión",
        "table_reference": "GLP",
        "source_value": 6.333,
        "source_unit": "kg CO2/gal US",
        "conversion": "Actividad en galones US × 6,333 kg CO2/gal",
        "quality": "B", "uncertainty": 0.0, "factor_kind": "Factor regulatorio colombiano", "reporting_use": "Piloto", "review_status": "Aprobado documentalmente", "approved_by": "Control metodológico V0.28",
        "notes": "Equivalencia regulatoria condicionada; confirmar tipo de GLP, unidad y objetivo del reporte.",
    },
    {
        "name": "Gasolina · equivalencia regulatoria Colombia",
        "activity_type": "Combustión de gasolina",
        "country": "Colombia", "sector": "Transporte y multisectorial", "gas_code": "CO2", "version": "COL-D926-2017", "value": 9.0, "input_unit": "gal", "output_unit": "kg CO2", "source_code": "COL-DECRETO-926-2017",
        "page_reference": "Tabla de factores de conversión", "table_reference": "Gasolina", "source_value": 9.0, "source_unit": "kg CO2/gal US", "conversion": "Actividad en galones US × 9,000 kg CO2/gal", "quality": "B", "uncertainty": 0.0, "factor_kind": "Factor regulatorio colombiano", "reporting_use": "Piloto", "review_status": "Aprobado documentalmente", "approved_by": "Control metodológico V0.28",
        "notes": "Equivalencia regulatoria; no incorpora de forma explícita CH4 y N2O ni caracteriza la mezcla comercial del periodo.",
    },
    {
        "name": "Kerosene / Jet Fuel · equivalencia regulatoria Colombia",
        "activity_type": "Combustión de combustible de aviación",
        "country": "Colombia", "sector": "Transporte", "gas_code": "CO2", "version": "COL-D926-2017", "value": 9.867, "input_unit": "gal", "output_unit": "kg CO2", "source_code": "COL-DECRETO-926-2017",
        "page_reference": "Tabla de factores de conversión", "table_reference": "Kerosene / Jet Fuel", "source_value": 9.867, "source_unit": "kg CO2/gal US", "conversion": "Actividad en galones US × 9,867 kg CO2/gal", "quality": "B", "uncertainty": 0.0, "factor_kind": "Factor regulatorio colombiano", "reporting_use": "Piloto", "review_status": "Aprobado documentalmente", "approved_by": "Control metodológico V0.28",
        "notes": "Referencia regulatoria. Para viajes de negocio debe evaluarse el método por pasajero-km, clase, distancia y forzamiento radiativo cuando proceda.",
    },
    {
        "name": "ACPM · equivalencia regulatoria Colombia",
        "activity_type": "Combustión de ACPM",
        "country": "Colombia", "sector": "Multisectorial", "gas_code": "CO2", "version": "COL-D926-2017", "value": 10.133, "input_unit": "gal", "output_unit": "kg CO2", "source_code": "COL-DECRETO-926-2017",
        "page_reference": "Tabla de factores de conversión", "table_reference": "ACPM", "source_value": 10.133, "source_unit": "kg CO2/gal US", "conversion": "Actividad en galones US × 10,133 kg CO2/gal", "quality": "B", "uncertainty": 0.0, "factor_kind": "Factor regulatorio colombiano", "reporting_use": "Piloto", "review_status": "Aprobado documentalmente", "approved_by": "Control metodológico V0.28",
        "notes": "Referencia regulatoria. Confirmar mezcla comercial, fuente primaria y gases no CO2 antes de uso formal.",
    },
    {
        "name": "Fuel oil · equivalencia regulatoria Colombia",
        "activity_type": "Combustión de fuel oil",
        "country": "Colombia", "sector": "Multisectorial", "gas_code": "CO2", "version": "COL-D926-2017", "value": 11.8, "input_unit": "gal", "output_unit": "kg CO2", "source_code": "COL-DECRETO-926-2017",
        "page_reference": "Tabla de factores de conversión", "table_reference": "Fuel oil", "source_value": 11.8, "source_unit": "kg CO2/gal US", "conversion": "Actividad en galones US × 11,800 kg CO2/gal", "quality": "B", "uncertainty": 0.0, "factor_kind": "Factor regulatorio colombiano", "reporting_use": "Piloto", "review_status": "Aprobado documentalmente", "approved_by": "Control metodológico V0.28",
        "notes": "Referencia regulatoria condicionada a la caracterización del combustible y la fuente de combustión.",
    },
    {
        "name": "Diésel B10 Colombia · CO2 FECOC transcrito",
        "activity_type": "Combustión de diésel B10",
        "country": "Colombia", "sector": "Multisectorial", "gas_code": "CO2", "version": "FECOC-2016-PILOT", "value": 10.2765, "input_unit": "gal", "output_unit": "kg CO2", "source_code": "EAAB-FECOC-2016",
        "page_reference": "Tabla de combustibles", "table_reference": "Diésel B10 · CO2", "source_value": 10.2765, "source_unit": "kg CO2/gal US", "conversion": "Actividad en galones US × 10,2765 kg CO2/gal", "quality": "C", "uncertainty": 0.0, "factor_kind": "FECOC transcrito", "reporting_use": "Piloto", "review_status": "En revisión", "approved_by": "Control piloto V0.28",
        "notes": "Valor transcrito desde fuente secundaria. Debe verificarse contra FECOC primaria antes de promoverlo a uso formal.",
    },
    {
        "name": "Gasolina E10 Colombia · CO2 FECOC transcrito",
        "activity_type": "Combustión de gasolina E10",
        "country": "Colombia", "sector": "Transporte", "gas_code": "CO2", "version": "FECOC-2016-PILOT", "value": 7.6181, "input_unit": "gal", "output_unit": "kg CO2", "source_code": "EAAB-FECOC-2016",
        "page_reference": "Tabla de combustibles", "table_reference": "Gasolina E10 · CO2", "source_value": 7.6181, "source_unit": "kg CO2/gal US", "conversion": "Actividad en galones US × 7,6181 kg CO2/gal", "quality": "C", "uncertainty": 0.0, "factor_kind": "FECOC transcrito", "reporting_use": "Piloto", "review_status": "En revisión", "approved_by": "Control piloto V0.28",
        "notes": "Valor transcrito desde fuente secundaria. Confirmar mezcla y fuente primaria antes de uso formal.",
    },
    {
        "name": "N aplicado al suelo · EF1 agregado",
        "activity_type": "Aplicación de nitrógeno al suelo",
        "country": "Internacional", "sector": "Agricultura", "gas_code": "N2O", "version": "IPCC-2019-EF1-AGG", "value": 0.015714285714285715, "input_unit": "kg", "output_unit": "kg N2O", "source_code": "IPCC-AFOLU-2019",
        "page_reference": "Capítulo 11, página 11.13", "table_reference": "Tabla 11.1 · EF1 agregado", "source_value": 0.010, "source_unit": "kg N2O-N/kg N", "conversion": "EF1 × 44/28 para convertir N2O-N a N2O", "quality": "B", "uncertainty": 80.0, "factor_kind": "IPCC Tier 1", "reporting_use": "Formal", "review_status": "Aprobado documentalmente", "approved_by": "Control metodológico V0.28",
        "notes": "Aplicar a kg de N incorporado al suelo. Revelar clima, tipo de aporte, incertidumbre y cualquier componente indirecto calculado por separado.",
    },
    {
        "name": "Fertilizante sintético en clima húmedo · EF1",
        "activity_type": "Aplicación de fertilizante sintético al suelo",
        "country": "Internacional", "sector": "Agricultura", "gas_code": "N2O", "version": "IPCC-2019-EF1-WET-SYN", "value": 0.025142857142857144, "input_unit": "kg", "output_unit": "kg N2O", "source_code": "IPCC-AFOLU-2019",
        "page_reference": "Capítulo 11, página 11.13", "table_reference": "Tabla 11.1 · sintético, clima húmedo", "source_value": 0.016, "source_unit": "kg N2O-N/kg N", "conversion": "0,016 × 44/28", "quality": "B", "uncertainty": 18.75, "factor_kind": "IPCC Tier 1", "reporting_use": "Formal", "review_status": "Aprobado documentalmente", "approved_by": "Control metodológico V0.28",
        "notes": "Solo emisiones directas por aplicación de N al suelo en clima húmedo; no incluye fabricación ni emisiones indirectas.",
    },
    {
        "name": "Otros aportes de N en clima húmedo · EF1",
        "activity_type": "Aplicación de aportes orgánicos de N al suelo",
        "country": "Internacional", "sector": "Agricultura", "gas_code": "N2O", "version": "IPCC-2019-EF1-WET-OTHER", "value": 0.009428571428571429, "input_unit": "kg", "output_unit": "kg N2O", "source_code": "IPCC-AFOLU-2019",
        "page_reference": "Capítulo 11, página 11.13", "table_reference": "Tabla 11.1 · otros aportes, clima húmedo", "source_value": 0.006, "source_unit": "kg N2O-N/kg N", "conversion": "0,006 × 44/28", "quality": "B", "uncertainty": 83.0, "factor_kind": "IPCC Tier 1", "reporting_use": "Formal", "review_status": "Aprobado documentalmente", "approved_by": "Control metodológico V0.28",
        "notes": "Solo emisiones directas por N aplicado; definir si el producto corresponde a aporte sintético u otro aporte de N.",
    },
    {
        "name": "Aportes de N en clima seco · EF1",
        "activity_type": "Aplicación de nitrógeno al suelo",
        "country": "Internacional", "sector": "Agricultura", "gas_code": "N2O", "version": "IPCC-2019-EF1-DRY", "value": 0.007857142857142858, "input_unit": "kg", "output_unit": "kg N2O", "source_code": "IPCC-AFOLU-2019",
        "page_reference": "Capítulo 11, página 11.13", "table_reference": "Tabla 11.1 · todos los aportes, clima seco", "source_value": 0.005, "source_unit": "kg N2O-N/kg N", "conversion": "0,005 × 44/28", "quality": "B", "uncertainty": 100.0, "factor_kind": "IPCC Tier 1", "reporting_use": "Formal", "review_status": "Aprobado documentalmente", "approved_by": "Control metodológico V0.28",
        "notes": "Aplicable al clima seco según la definición IPCC; documentar criterio climático e irrigación.",
    },
]

COVERAGE_SPECS: list[dict[str, str]] = [
    {"category": "Electricidad adquirida", "status": "Formal", "source": "UPME SIN 2024", "next_action": "Actualizar por año de inventario y confirmar pertenencia al SIN."},
    {"category": "Compostaje", "status": "Formal condicionado", "source": "IPCC 2006 + corrigenda", "next_action": "Levantar humedad, tipo de residuo, aireación y evaluar medición de planta."},
    {"category": "Digestión anaerobia", "status": "Formal condicionado", "source": "IPCC 2006 + corrigenda", "next_action": "Construir balance de CH4 producido, usado, quemado, venteado y fugado."},
    {"category": "Refrigerantes", "status": "Formal por balance de masa", "source": "GHG Protocol GWP v2.0", "next_action": "Inventariar equipos, gases, cargas, recargas y recuperación."},
    {"category": "Combustibles fijos", "status": "Piloto condicionado", "source": "Decreto 926 / FECOC", "next_action": "Validar mezcla, fuente primaria y gases no CO2 antes de uso formal."},
    {"category": "Combustión móvil", "status": "Piloto parcial", "source": "FECOC / FECOC+", "next_action": "Completar tecnología vehicular, CH4, N2O y método por distancia cuando corresponda."},
    {"category": "Transporte contratado", "status": "Pendiente de parametrización", "source": "FECOC+ y datos de proveedor", "next_action": "Priorizar combustible real, toneladas, kilómetros y factor específico del transportador."},
    {"category": "Aguas residuales", "status": "Método paramétrico disponible", "source": "IPCC 2019 Refinement", "next_action": "Caracterizar DBO/DQO, MCF, lodos y CH4 recuperado por sistema."},
    {"category": "Fertilizantes y suelos", "status": "Factores de aplicación disponibles", "source": "IPCC AFOLU 2019", "next_action": "Separar aplicación de N, fabricación, emisiones indirectas y cambios en carbono del suelo."},
    {"category": "Bienes y servicios", "status": "Piloto", "source": "Datos específicos de proveedor", "next_action": "Definir reglas de asignación, límites cradle-to-gate y aseguramiento de datos."},
]

PILOT_REQUIREMENT_SPECS: list[tuple[Any, ...]] = [
    ("YAR-ELEC", "Yarumal", 2, "Energía adquirida", "Electricidad comprada del SIN", "kWh", "Mensual", "Administración de planta", "Facturas y consolidado mensual del medidor", "Formal", "UPME SIN 2024", "Alta"),
    ("YAR-DIESEL", "Yarumal", 1, "Combustión fija", "Diésel en equipos y respaldo", "L", "Mensual", "Operaciones", "Facturas, vales y horómetros", "Piloto condicionado", "Decreto 926 / FECOC V0.28", "Alta"),
    ("YAR-MOVIL", "Yarumal", 1, "Combustión móvil", "Combustible de vehículos bajo control", "L", "Mensual", "Logística", "Vales, placas, kilometraje y tipo de combustible", "Piloto parcial", "FECOC/FECOC+ V0.28", "Media"),
    ("YAR-REFR", "Yarumal", 1, "Emisiones fugitivas", "Refrigerantes en equipos", "kg", "Anual", "Mantenimiento", "Inventario de equipos, gas, carga, recarga y recuperación", "Formal por balance de masa", "GWP AR6 por gas", "Media"),
    ("YAR-REC", "Yarumal", 3, "Residuos tratados", "Residuos orgánicos recibidos", "t", "Mensual", "Operaciones", "Pesajes de ingreso y origen del material", "Dato de actividad", "Sin factor directo", "Alta"),
    ("YAR-COMP", "Yarumal", 1, "Tratamiento biológico", "Material tratado por compostaje", "t", "Mensual", "Operaciones", "Bitácora de lotes, humedad, mezcla y aireación", "Formal condicionado", "IPCC 2006 Table 4.1", "Alta"),
    ("YAR-AD", "Yarumal", 1, "Tratamiento biológico", "Material alimentado a digestión anaerobia", "t", "Mensual", "Operaciones", "Bitácora de carga, sólidos y operación del digestor", "Formal condicionado", "IPCC 2006 corrigenda", "Alta"),
    ("YAR-BIOGAS", "Yarumal", 1, "Biogás", "Biogás producido, usado, quemado y venteado", "m3", "Mensual", "Operaciones", "Medición o estimación separada por destino", "Balance operativo disponible", "Medición específica preferida · V0.28", "Alta"),
    ("YAR-PROD-S", "Yarumal", 0, "Indicador", "Fertilizante sólido producido", "t", "Mensual", "Producción", "Ordenes y registros de producción", "Indicador", "Intensidad de emisiones", "Media"),
    ("YAR-PROD-L", "Yarumal", 0, "Indicador", "Fertilizante líquido producido", "L", "Mensual", "Producción", "Ordenes y registros de producción", "Indicador", "Intensidad de emisiones", "Media"),
    ("TAM-ELEC", "Támesis", 2, "Energía adquirida", "Electricidad comprada del SIN", "kWh", "Mensual", "Administración de planta", "Facturas y consolidado mensual del medidor", "Formal", "UPME SIN 2024", "Alta"),
    ("TAM-DIESEL", "Támesis", 1, "Combustión fija", "Diésel en equipos y respaldo", "L", "Mensual", "Operaciones", "Facturas, vales y horómetros", "Piloto condicionado", "Decreto 926 / FECOC V0.28", "Alta"),
    ("TAM-REFR", "Támesis", 1, "Emisiones fugitivas", "Refrigerantes en equipos", "kg", "Anual", "Mantenimiento", "Inventario de equipos, gas, carga, recarga y recuperación", "Formal por balance de masa", "GWP AR6 por gas", "Media"),
    ("TAM-REC", "Támesis", 3, "Residuos tratados", "Residuos orgánicos recibidos", "t", "Mensual", "Operaciones", "Pesajes de ingreso y origen del material", "Dato de actividad", "Sin factor directo", "Alta"),
    ("TAM-COMP", "Támesis", 1, "Tratamiento biológico", "Material tratado por compostaje", "t", "Mensual", "Operaciones", "Bitácora de lotes, humedad, mezcla y aireación", "Formal condicionado", "IPCC 2006 Table 4.1", "Alta"),
    ("TAM-AD", "Támesis", 1, "Tratamiento biológico", "Material alimentado a digestión anaerobia", "t", "Mensual", "Operaciones", "Bitácora de carga, sólidos y operación del digestor", "Formal condicionado", "IPCC 2006 corrigenda", "Alta"),
    ("TAM-BIOGAS", "Támesis", 1, "Biogás", "Biogás producido, usado, quemado y venteado", "m3", "Mensual", "Operaciones", "Medición o estimación separada por destino", "Balance operativo disponible", "Medición específica preferida · V0.28", "Alta"),
    ("TAM-PROD-S", "Támesis", 0, "Indicador", "Fertilizante sólido producido", "t", "Mensual", "Producción", "Ordenes y registros de producción", "Indicador", "Intensidad de emisiones", "Media"),
    ("TAM-PROD-L", "Támesis", 0, "Indicador", "Fertilizante líquido producido", "L", "Mensual", "Producción", "Ordenes y registros de producción", "Indicador", "Intensidad de emisiones", "Media"),
    ("CORP-TRANSP", "Corporativo", 3, "Transporte y distribución", "Transporte contratado de insumos y productos", "t·km", "Mensual", "Logística", "Origen, destino, toneladas, kilómetros, vehículo y combustible", "Pendiente de parametrización", "Datos de proveedor / FECOC+", "Alta"),
    ("CORP-COMPRA", "Corporativo", 3, "Bienes y servicios adquiridos", "Materias primas y precursores", "kg", "Mensual", "Compras", "Proveedor, producto, cantidad, origen y huella específica", "Piloto", "Factor específico de proveedor", "Alta"),
    ("CORP-VIAJES", "Corporativo", 3, "Viajes de negocio", "Viajes aéreos y terrestres", "pasajero·km", "Mensual", "Administración", "Tiquetes, trayecto, clase y distancia", "Pendiente", "Biblioteca de transporte", "Baja"),
    ("CORP-COMMUTE", "Corporativo", 3, "Desplazamiento de empleados", "Movilidad casa-trabajo", "pasajero·km", "Anual", "Gestión humana", "Encuesta modal, distancia, frecuencia y ocupación", "Pendiente", "Encuesta + factores de transporte", "Baja"),
    ("CORP-AGUA", "Corporativo", 3, "Agua y aguas residuales", "Consumo y tratamiento de agua", "m3", "Mensual", "Administración", "Facturas, sistema de tratamiento y parámetros DBO/DQO", "Método paramétrico disponible", "IPCC 2019 · Bo × MCF", "Media"),
]

REFERENCE_CASE_SPECS = [
    ("REF-009", "Compostaje húmedo · CH4", "Residuos", "100 t tratadas × 4 kg CH4/t × GWP AR6 27.", 100.0, "t", 4.0, "t", "CH4", 27.0, 100.0, 400.0, 10800.0, 1e-9, "Calculado", "IPCC 2006 Table 4.1"),
    ("REF-010", "Compostaje húmedo · N2O", "Residuos", "100 t tratadas × 0,24 kg N2O/t × GWP AR6 273.", 100.0, "t", 0.24, "t", "N2O", 273.0, 100.0, 24.0, 6552.0, 1e-9, "Calculado", "IPCC 2006 corrigenda"),
    ("REF-011", "Digestión anaerobia húmeda · CH4", "Residuos", "100 t tratadas × 0,8 kg CH4/t × GWP AR6 27.", 100.0, "t", 0.8, "t", "CH4", 27.0, 100.0, 80.0, 2160.0, 1e-9, "Calculado", "IPCC 2006 corrigenda"),
    ("REF-012", "Fuga directa de HFC-134a", "Refrigerantes", "2 kg liberados × GWP AR6 1530.", 2.0, "kg", 1.0, "kg", "HFC-134a", 1530.0, 2.0, 2.0, 3060.0, 1e-9, "Calculado", "GHG Protocol GWP v2.0"),
    ("REF-013", "Gas natural regulatorio Colombia", "Combustibles", "100 m3 × 1,952 kg CO2/m3.", 100.0, "m3", 1.952, "m3", "CO2", 1.0, 100.0, 195.2, 195.2, 1e-9, "Calculado", "Decreto 926 de 2017"),
    ("REF-014", "GLP regulatorio Colombia", "Combustibles", "10 gal × 6,333 kg CO2/gal.", 10.0, "gal", 6.333, "gal", "CO2", 1.0, 10.0, 63.33, 63.33, 1e-9, "Calculado", "Decreto 926 de 2017"),
    ("REF-015", "Gasolina regulatoria Colombia", "Combustibles", "10 gal × 9,000 kg CO2/gal.", 10.0, "gal", 9.0, "gal", "CO2", 1.0, 10.0, 90.0, 90.0, 1e-9, "Calculado", "Decreto 926 de 2017"),
    ("REF-016", "ACPM regulatorio Colombia", "Combustibles", "10 gal × 10,133 kg CO2/gal.", 10.0, "gal", 10.133, "gal", "CO2", 1.0, 10.0, 101.33, 101.33, 1e-9, "Calculado", "Decreto 926 de 2017"),
    ("REF-017", "Diésel B10 FECOC transcrito", "Combustibles", "10 gal × 10,2765 kg CO2/gal.", 10.0, "gal", 10.2765, "gal", "CO2", 1.0, 10.0, 102.765, 102.765, 1e-9, "Calculado", "EAAB · FECOC 2016, uso piloto"),
    ("REF-018", "Fertilizante sintético húmedo · directo", "Suelos", "1000 kg N × EF1 0,016 × 44/28 × GWP AR6 273.", 1000.0, "kg", 0.025142857142857144, "kg", "N2O", 273.0, 1000.0, 25.142857142857146, 6864.0, 1e-8, "Calculado", "IPCC 2019 Refinement, Table 11.1"),
    ("REF-019", "Aguas residuales industriales · COD", "Aguas residuales", "1000 kg COD × Bo 0,25 × MCF 0,8 × GWP AR6 27.", 1000.0, "kg", 0.2, "kg", "CH4", 27.0, 1000.0, 200.0, 5400.0, 1e-9, "Calculado", "IPCC 2019 Refinement, Equations 6.4–6.5"),
    ("REF-020", "Urea · volatilización indirecta húmeda", "Suelos", "1000 kg N × FracGASF 0,15 × EF4 0,014 × 44/28 × GWP 273.", 1000.0, "kg", 0.0033, "kg", "N2O", 273.0, 1000.0, 3.3, 900.9, 1e-8, "Calculado", "IPCC 2019 Refinement, Table 11.3"),
]


def _upsert_documents(session: Session) -> dict[str, MethodologySourceDocument]:
    documents = {item.code: item for item in session.scalars(select(MethodologySourceDocument)).all()}
    for spec in SOURCE_DOCUMENT_SPECS:
        item = documents.get(spec["code"])
        if not item:
            item = MethodologySourceDocument(code=spec["code"], title=spec["title"], issuing_body=spec["issuing_body"])
            session.add(item)
            documents[spec["code"]] = item
        for key, value in spec.items():
            setattr(item, key, value)
        item.accessed_at = date.today()
    session.flush()
    return documents


def _upsert_factors(session: Session, documents: dict[str, MethodologySourceDocument]) -> None:
    gases = {item.code: item for item in session.scalars(select(Gas)).all()}
    for spec in FACTOR_SPECS:
        gas = gases.get(spec["gas_code"])
        if not gas:
            continue
        factor = session.scalar(select(EmissionFactor).where(EmissionFactor.name == spec["name"]))
        if not factor:
            factor = EmissionFactor(name=spec["name"], activity_type=spec["activity_type"], country=spec["country"], sector=spec["sector"], status="Activo", is_demo=False)
            session.add(factor)
            session.flush()
        factor.activity_type = spec["activity_type"]
        factor.country = spec["country"]
        factor.sector = spec["sector"]
        factor.status = "Activo"
        factor.is_demo = False
        version = session.scalar(select(EmissionFactorVersion).where(
            EmissionFactorVersion.factor_id == factor.id,
            EmissionFactorVersion.version == spec["version"],
            EmissionFactorVersion.gas_id == gas.id,
        ))
        if not version:
            version = EmissionFactorVersion(factor_id=factor.id, gas_id=gas.id, version=spec["version"], value=spec["value"], input_unit=spec["input_unit"])
            session.add(version)
        version.value = spec["value"]
        version.input_unit = spec["input_unit"]
        version.output_unit = spec["output_unit"]
        version.source_organization = documents[spec["source_code"]].issuing_body
        version.source_document = documents[spec["source_code"]].title
        version.publication_year = documents[spec["source_code"]].publication_date.year if documents[spec["source_code"]].publication_date else 2006
        version.geographic_scope = spec["country"]
        version.technology_scope = spec["activity_type"]
        version.uncertainty_percentage = spec["uncertainty"]
        version.status = spec.get("status", "Aprobado")
        version.notes = spec["notes"]
        version.approved_by = spec.get("approved_by", "Control metodológico V0.28")
        version.approved_at = datetime.now(UTC)
        session.flush()
        doc = session.scalar(select(FactorDocumentation).where(FactorDocumentation.factor_version_id == version.id))
        if not doc:
            doc = FactorDocumentation(factor_version_id=version.id)
            session.add(doc)
        doc.source_document_id = documents[spec["source_code"]].id
        doc.factor_kind = spec.get("factor_kind", "IPCC Tier 1" if "IPCC" in spec["version"] else "Balance de masa directo")
        doc.reporting_use = spec.get("reporting_use", "Formal")
        doc.page_reference = spec["page_reference"]
        doc.table_reference = spec["table_reference"]
        doc.data_year = version.publication_year
        doc.source_value = spec["source_value"]
        doc.source_unit = spec["source_unit"]
        doc.conversion_expression = spec["conversion"]
        doc.aggregated_co2e = spec.get("aggregated_co2e", False)
        doc.gwp_embedded = spec.get("gwp_embedded", "No; aplicar GWP del inventario")
        doc.methane_origin = spec.get("methane_origin", "No fósil / biogénico" if spec["gas_code"] == "CH4" else "No aplica")
        doc.quality_grade = spec["quality"]
        doc.review_status = spec.get("review_status", "Aprobado documentalmente")
        doc.reviewer = spec.get("approved_by", "Control metodológico V0.28")
        doc.reviewed_at = datetime.now(UTC)
        doc.next_review_date = spec.get("next_review_date", date(2027, 8, 1))
        doc.restriction_notes = spec["notes"]
    session.flush()


def _upsert_reference_cases(session: Session) -> None:
    for spec in REFERENCE_CASE_SPECS:
        (code, title, category, description, activity_value, activity_unit, factor_value, factor_input_unit, gas_code, gwp_value, expected_normalized, expected_gas, expected_co2e, tolerance, expected_status, source_reference) = spec
        case = session.scalar(select(ReferenceCalculationCase).where(ReferenceCalculationCase.code == code))
        if not case:
            case = ReferenceCalculationCase(code=code, title=title, activity_value=activity_value, activity_unit=activity_unit, factor_value=factor_value, factor_input_unit=factor_input_unit, gas_code=gas_code, expected_normalized_value=expected_normalized, expected_gas_kg=expected_gas, expected_co2e_kg=expected_co2e)
            session.add(case)
        case.title = title
        case.category = category
        case.description = description
        case.activity_value = activity_value
        case.activity_unit = activity_unit
        case.factor_value = factor_value
        case.factor_input_unit = factor_input_unit
        case.gas_code = gas_code
        case.gwp_value = gwp_value
        case.expected_normalized_value = expected_normalized
        case.expected_gas_kg = expected_gas
        case.expected_co2e_kg = expected_co2e
        case.tolerance = tolerance
        case.expected_status = expected_status
        case.source_reference = source_reference
        case.active = True
    session.flush()


def _upsert_pilot(session: Session) -> None:
    for org in session.scalars(select(Organization)).all():
        pilot = session.scalar(select(PilotProject).where(PilotProject.organization_id == org.id, PilotProject.code == "GREENATICS-2026"))
        if not pilot:
            pilot = PilotProject(
                organization_id=org.id,
                code="GREENATICS-2026",
                name="Piloto real Greenatics 2026",
                reporting_year=2026,
                consolidation_approach="Control operacional",
                organizational_boundary="Operación corporativa y plantas Yarumal y Támesis bajo control operacional.",
                operational_boundary="Alcances 1 y 2 completos; categorías materiales de alcance 3 priorizadas por datos disponibles y relevancia.",
                status="Preparación metodológica",
                lead="Dirección ambiental y operaciones",
                notes="El piloto debe contrastarse con una memoria de cálculo independiente antes de cualquier declaración externa.",
            )
            session.add(pilot)
            session.flush()
        for spec in PILOT_REQUIREMENT_SPECS:
            code, site, scope, category, source_name, unit, frequency, owner, evidence, factor_status, factor_reference, materiality = spec
            item = session.scalar(select(PilotSourceRequirement).where(PilotSourceRequirement.pilot_id == pilot.id, PilotSourceRequirement.code == code))
            if not item:
                item = PilotSourceRequirement(pilot_id=pilot.id, code=code, site=site, scope=scope, category=category, source_name=source_name, activity_unit=unit)
                session.add(item)
            item.site = site
            item.scope = scope
            item.category = category
            item.source_name = source_name
            item.activity_unit = unit
            item.frequency = frequency
            item.data_owner = owner
            item.evidence_expected = evidence
            item.factor_status = factor_status
            item.factor_reference = factor_reference
            item.materiality = materiality
        consultant = session.scalar(select(AppUser).where(AppUser.organization_id == org.id, AppUser.role == "Consultor"))
        existing_notice = session.scalar(select(Notification).where(Notification.organization_id == org.id, Notification.title == "Biblioteca Colombia V0.28 disponible"))
        if consultant and not existing_notice:
            session.add(Notification(organization_id=org.id, user_id=consultant.id, title="Biblioteca Colombia V0.28 disponible", message="La biblioteca incorpora combustibles colombianos condicionados, métodos de aguas residuales, fertilización nitrogenada y balance operativo de biogás.", link="/metodologia/colombia", category="Metodología", priority="Alta", status="Entregada"))
    session.flush()


def _update_release_controls(session: Session) -> None:
    for finding in session.scalars(select(ConsolidationFinding).where(ConsolidationFinding.code == "MET-001")).all():
        finding.status = "En curso"
        finding.target_version = "V0.28"
        finding.evidence = "V0.28 incorpora equivalencias regulatorias colombianas para combustibles, valores FECOC piloto con restricción, método paramétrico de aguas residuales, factores IPCC de N aplicado al suelo y balance operativo de biogás. Transporte por tecnología y factores primarios FECOC completos siguen pendientes."
    for gate in session.scalars(select(ReleaseGate).where(ReleaseGate.code.in_(["GATE-METH", "GATE-CALC"]))).all():
        gate.status = "Parcial"
        gate.evidence = "Biblioteca V0.28 con factores colombianos y métodos paramétricos, además de 20 casos patrón reproducibles."
        gate.notes = "No aprobar V1.0 hasta validar FECOC primaria, transporte por tecnología, parámetros reales de planta y contraste independiente del piloto."
    session.flush()


def ensure_sector_library_defaults(session: Session) -> None:
    documents = _upsert_documents(session)
    _upsert_factors(session, documents)
    _upsert_reference_cases(session)
    _upsert_pilot(session)
    _update_release_controls(session)
    session.flush()


def pilot_summary(session: Session, organization_id: int) -> dict[str, Any]:
    pilot = session.scalar(
        select(PilotProject)
        .where(PilotProject.organization_id == organization_id, PilotProject.code == "GREENATICS-2026")
        .options(selectinload(PilotProject.requirements))
    )
    requirements = sorted(pilot.requirements if pilot else [], key=lambda item: (item.site, item.scope, item.code))
    completed = [item for item in requirements if item.status in {"Disponible", "Validado", "No aplica"}]
    validated = [item for item in requirements if item.status == "Validado"]
    formal = [item for item in requirements if item.factor_status.startswith("Formal") or item.factor_status in {"Indicador", "Dato de actividad"}]
    high = [item for item in requirements if item.materiality == "Alta"]
    high_pending = [item for item in high if item.status not in {"Disponible", "Validado", "No aplica"}]
    data_score = round(100 * len(completed) / max(len(requirements), 1))
    factor_score = round(100 * len(formal) / max(len(requirements), 1))
    readiness_score = round(data_score * 0.6 + factor_score * 0.4)
    by_site: dict[str, list[PilotSourceRequirement]] = {}
    for item in requirements:
        by_site.setdefault(item.site, []).append(item)
    return {
        "pilot": pilot,
        "requirements": requirements,
        "by_site": by_site,
        "coverage": COVERAGE_SPECS,
        "metrics": {
            "requirements": len(requirements),
            "completed": len(completed),
            "validated": len(validated),
            "formal_or_indicator": len(formal),
            "high_pending": len(high_pending),
            "data_score": data_score,
            "factor_score": factor_score,
            "readiness_score": readiness_score,
        },
    }


def build_pilot_workbook(summary: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan piloto"
    pilot = summary["pilot"]
    ws.append(["Campo", "Valor"])
    for label, value in [
        ("Código", pilot.code if pilot else ""), ("Nombre", pilot.name if pilot else ""),
        ("Año", pilot.reporting_year if pilot else ""), ("Enfoque", pilot.consolidation_approach if pilot else ""),
        ("Límite organizacional", pilot.organizational_boundary if pilot else ""),
        ("Límite operacional", pilot.operational_boundary if pilot else ""), ("Estado", pilot.status if pilot else ""),
        ("Preparación", summary["metrics"]["readiness_score"]),
    ]:
        ws.append([label, value])

    ws = wb.create_sheet("Fuentes requeridas")
    ws.append(["Código", "Sede", "Alcance", "Categoría", "Fuente", "Unidad", "Frecuencia", "Responsable", "Evidencia esperada", "Estado factor", "Referencia", "Materialidad", "Estado dato", "Notas"])
    for item in summary["requirements"]:
        ws.append([item.code, item.site, item.scope, item.category, item.source_name, item.activity_unit, item.frequency, item.data_owner, item.evidence_expected, item.factor_status, item.factor_reference, item.materiality, item.status, item.notes])

    ws = wb.create_sheet("Cobertura metodológica")
    ws.append(["Categoría", "Estado", "Fuente", "Siguiente acción"])
    for item in summary["coverage"]:
        ws.append([item["category"], item["status"], item["source"], item["next_action"]])

    ws = wb.create_sheet("Plantilla de datos")
    ws.append(["Código fuente", "Periodo inicio", "Periodo fin", "Valor", "Unidad", "Documento soporte", "Responsable", "Calidad", "Observaciones"])
    for item in summary["requirements"]:
        if item.scope != 0:
            ws.append([item.code, "2026-01-01", "2026-01-31", "", item.activity_unit, "", item.data_owner, "Pendiente", ""])

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 52)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
