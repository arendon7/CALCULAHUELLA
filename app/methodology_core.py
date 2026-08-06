from __future__ import annotations

import json
import secrets
from datetime import UTC, date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from .database import (
    ConsolidationFinding,
    EmissionFactor,
    EmissionFactorVersion,
    EmissionSource,
    FactorDocumentation,
    FactorSelectionRule,
    Gas,
    GWPValue,
    MethodologySourceDocument,
    MethodologyValidationRun,
    ReferenceCalculationCase,
    ReferenceCaseResult,
    ReleaseGate,
    SourceFactorAssignment,
)

ENGINE_VERSION = "1.1.0"
QUALITY_RANK = {"A": 4, "B": 3, "C": 2, "D": 1, "N/A": 0}

OFFICIAL_SOURCE_SPECS: list[dict[str, Any]] = [
    {
        "code": "GHGP-CORP-2004",
        "title": "GHG Protocol Corporate Accounting and Reporting Standard",
        "issuing_body": "GHG Protocol · WRI/WBCSD",
        "document_type": "Estándar corporativo",
        "publication_date": date(2004, 3, 1),
        "jurisdiction": "Internacional",
        "source_url": "https://ghgprotocol.org/corporate-standard",
        "citation": "GHG Protocol. Corporate Accounting and Reporting Standard, revised edition.",
        "status": "Vigente",
        "notes": "Marco de contabilidad y reporte corporativo de siete gases de efecto invernadero.",
    },
    {
        "code": "GHGP-S2-2015",
        "title": "GHG Protocol Scope 2 Guidance",
        "issuing_body": "GHG Protocol · WRI/WBCSD",
        "document_type": "Guía de alcance 2",
        "publication_date": date(2015, 1, 1),
        "jurisdiction": "Internacional",
        "source_url": "https://ghgprotocol.org/scope-2-guidance",
        "citation": "GHG Protocol. Scope 2 Guidance: An amendment to the GHG Protocol Corporate Standard.",
        "status": "Vigente con revisión en curso",
        "notes": "Incluye criterios de calidad para instrumentos contractuales y reporte transparente de energía adquirida.",
    },
    {
        "code": "GHGP-S3-2011",
        "title": "Corporate Value Chain (Scope 3) Accounting and Reporting Standard",
        "issuing_body": "GHG Protocol · WRI/WBCSD",
        "document_type": "Estándar de alcance 3",
        "publication_date": date(2011, 9, 1),
        "jurisdiction": "Internacional",
        "source_url": "https://ghgprotocol.org/corporate-value-chain-scope-3-standard",
        "citation": "GHG Protocol. Corporate Value Chain (Scope 3) Accounting and Reporting Standard.",
        "status": "Vigente",
        "notes": "Metodología corporativa para las quince categorías de emisiones de la cadena de valor.",
    },
    {
        "code": "GHGP-GWP-2024",
        "title": "IPCC Global Warming Potential Values · version 2.0",
        "issuing_body": "GHG Protocol",
        "document_type": "Tabla de GWP",
        "publication_date": date(2024, 8, 7),
        "jurisdiction": "Internacional",
        "source_url": "https://ghgprotocol.org/sites/default/files/2024-08/Global-Warming-Potential-Values%20%28August%202024%29.pdf",
        "citation": "GHG Protocol. IPCC Global Warming Potential Values, version 2.0, 7 August 2024.",
        "status": "Vigente",
        "notes": "Compila GWP100 de AR4, AR5 y AR6 y diferencia metano fósil y no fósil.",
    },
    {
        "code": "GHGP-LSR-2026",
        "title": "GHG Protocol Land Sector and Removals Standard v1.1",
        "issuing_body": "GHG Protocol · WRI/WBCSD",
        "document_type": "Estándar de sector tierra y remociones",
        "publication_date": date(2026, 6, 1),
        "jurisdiction": "Internacional",
        "source_url": "https://ghgprotocol.org/land-sector-and-removals-standard",
        "citation": "GHG Protocol. Land Sector and Removals Standard, version 1.1, 2026.",
        "status": "Publicado · vigente desde 2027",
        "notes": "Requisitos para emisiones del sector tierra, productos biogénicos y remociones de CO2; fecha efectiva 1 de enero de 2027.",
    },
    {
        "code": "IPCC-2006-GL",
        "title": "2006 IPCC Guidelines for National Greenhouse Gas Inventories",
        "issuing_body": "IPCC Task Force on National Greenhouse Gas Inventories",
        "document_type": "Directrices metodológicas",
        "publication_date": date(2006, 1, 1),
        "jurisdiction": "Internacional",
        "source_url": "https://www.ipcc-nggip.iges.or.jp/public/2006gl/",
        "citation": "IPCC. 2006 IPCC Guidelines for National Greenhouse Gas Inventories.",
        "status": "Vigente junto con el Refinamiento 2019",
        "notes": "Base metodológica para inventarios nacionales; sus factores requieren evaluación de aplicabilidad corporativa.",
    },
    {
        "code": "IPCC-2019-RF",
        "title": "2019 Refinement to the 2006 IPCC Guidelines for National Greenhouse Gas Inventories",
        "issuing_body": "IPCC Task Force on National Greenhouse Gas Inventories",
        "document_type": "Refinamiento metodológico",
        "publication_date": date(2019, 5, 1),
        "jurisdiction": "Internacional",
        "source_url": "https://www.ipcc-nggip.iges.or.jp/home/2019refinement.html",
        "citation": "IPCC. 2019 Refinement to the 2006 IPCC Guidelines for National Greenhouse Gas Inventories.",
        "status": "Vigente junto con las Directrices 2006",
        "notes": "Actualiza, complementa y desarrolla las Directrices 2006; no las sustituye.",
    },
    {
        "code": "UPME-R085-2026",
        "title": "Resolución UPME 000085 de 2026 · factor de emisión del SIN 2024",
        "issuing_body": "Unidad de Planeación Minero Energética · UPME",
        "document_type": "Resolución y factor oficial",
        "publication_date": date(2026, 2, 23),
        "jurisdiction": "Colombia",
        "source_url": "https://docs.upme.gov.co/Normatividad/085_2026.pdf",
        "citation": "UPME. Resolución 000085 de 23 de febrero de 2026, artículo 1, literal B.",
        "status": "Vigente",
        "notes": "Establece 0,220 tCO2e/MWh para inventarios de GEI del Sistema Interconectado Nacional, año 2024.",
    },
    {
        "code": "XM-SIN-2025-PRELIM",
        "title": "Resultado preliminar del factor de emisión del SIN 2025",
        "issuing_body": "XM Compañía de Expertos en Mercados S.A. E.S.P.",
        "document_type": "Resultado preliminar y vigilancia metodológica",
        "publication_date": date(2026, 1, 30),
        "jurisdiction": "Colombia",
        "source_url": "https://www.xm.com.co/noticias/8688-resultado-preliminar-del-calculo-de-factor-de-emision-del-sistema-interconectado",
        "citation": "XM. Resultado preliminar del cálculo de Factor de Emisión del Sistema Interconectado Nacional 2025, 30 de enero de 2026.",
        "status": "Preliminar · no incorporado al cálculo",
        "notes": "Fuente en vigilancia regulatoria. No se crea una versión de factor ni se habilita su uso hasta verificar el valor en una fuente oficial reproducible, resolver su condición preliminar y completar revisión documental independiente.",
    },
    {
        "code": "CTH-DEMO",
        "title": "Biblioteca demostrativa Calcula tu Huella",
        "issuing_body": "Calcula tu Huella",
        "document_type": "Datos sintéticos de prueba",
        "publication_date": date(2026, 7, 31),
        "jurisdiction": "No aplica",
        "source_url": "",
        "citation": "Datos sintéticos incluidos exclusivamente para probar el flujo y el motor.",
        "status": "Demostrativo",
        "notes": "No apto para inventarios formales, declaraciones públicas ni verificación.",
    },
]

GWP_SPECS: dict[str, dict[str, Any]] = {
    "CO2": {"name": "Dióxido de carbono", "formula": "CO₂", "values": {"AR4": 1.0, "AR5": 1.0, "AR6": 1.0}},
    "CH4": {"name": "Metano no fósil / combustión", "formula": "CH₄", "values": {"AR4": 25.0, "AR5": 28.0, "AR6": 27.0}},
    "CH4-FOSSIL": {"name": "Metano fósil fugitivo o de proceso", "formula": "CH₄ fósil", "values": {"AR5": 30.0, "AR6": 29.8}},
    "N2O": {"name": "Óxido nitroso", "formula": "N₂O", "values": {"AR4": 298.0, "AR5": 265.0, "AR6": 273.0}},
    "NF3": {"name": "Trifluoruro de nitrógeno", "formula": "NF₃", "values": {"AR4": 17200.0, "AR5": 16100.0, "AR6": 17400.0}},
    "SF6": {"name": "Hexafluoruro de azufre", "formula": "SF₆", "values": {"AR4": 22800.0, "AR5": 23500.0, "AR6": 24300.0}},
    "HFC-134a": {"name": "HFC-134a", "formula": "CH₂FCF₃", "values": {"AR4": 1430.0, "AR5": 1300.0, "AR6": 1530.0}},
    "CO2e": {"name": "Dióxido de carbono equivalente directo", "formula": "CO₂e", "values": {"AR6": 1.0}},
}

RULE_SPECS = [
    ("SEL-001", 10, "Priorizar factor específico verificado del proveedor", "*", "*", "*", "*", "Específico de proveedor", False, 3, "A", "Formal", "La especificidad y verificación primaria prevalecen cuando los límites y la asignación son comparables."),
    ("SEL-002", 20, "Priorizar factor oficial nacional del periodo", "*", "Colombia", "*", "*", "Oficial nacional", True, 1, "A", "Formal", "Favorece representatividad geográfica y temporal para inventarios corporativos en Colombia."),
    ("SEL-003", 30, "Usar factor sectorial reconocido documentado", "*", "*", "*", "*", "Sectorial reconocido", True, 2, "B", "Formal", "Aplica cuando no existe factor específico u oficial nacional pertinente."),
    ("SEL-004", 40, "Usar factor IPCC por gas con conversión explícita", "*", "*", "*", "*", "IPCC por gas", False, 5, "B", "Formal", "Exige documentar poder calorífico, densidad, tecnología y GWP cuando corresponda."),
    ("SEL-005", 90, "Restringir factores agregados con GWP no identificado", "*", "*", "*", "CO2e", "Agregado CO2e", False, 5, "C", "Piloto", "Puede usarse de forma controlada cuando no es posible desagregar gases; debe revelarse el GWP implícito o su desconocimiento."),
    ("SEL-006", 999, "Bloquear factores demostrativos en reportes formales", "*", "*", "*", "*", "Demostrativo", False, 99, "D", "Demostrativo", "Los datos sintéticos solo validan el software y no constituyen evidencia metodológica."),
]

REFERENCE_CASE_SPECS = [
    ("REF-001", "Electricidad SIN · kWh", "Alcance 2", "Aplicación directa del factor oficial para inventarios 2024.", 1000.0, "kWh", 0.220, "kWh", "CO2e", 1.0, 1000.0, 220.0, 220.0, 1e-9, "Calculado", "UPME-R085-2026 · artículo 1.B"),
    ("REF-002", "Electricidad SIN · conversión MWh", "Unidades", "Comprueba MWh → kWh antes de aplicar el factor.", 1.0, "MWh", 0.220, "kWh", "CO2e", 1.0, 1000.0, 220.0, 220.0, 1e-9, "Calculado", "UPME-R085-2026 y conversión SI"),
    ("REF-003", "Metano no fósil AR6", "GWP", "Comprueba el GWP100 AR6 aplicable a combustión y fuentes no fósiles.", 2.0, "kg", 1.0, "kg", "CH4", 27.0, 2.0, 2.0, 54.0, 1e-9, "Calculado", "GHGP-GWP-2024"),
    ("REF-004", "Metano fósil AR6", "GWP", "Comprueba el GWP100 AR6 para metano fósil fugitivo o de proceso.", 2.0, "kg", 1.0, "kg", "CH4-FOSSIL", 29.8, 2.0, 2.0, 59.6, 1e-9, "Calculado", "GHGP-GWP-2024"),
    ("REF-005", "Óxido nitroso AR6", "GWP", "Comprueba el GWP100 AR6 del N2O.", 1.5, "kg", 1.0, "kg", "N2O", 273.0, 1.5, 1.5, 409.5, 1e-9, "Calculado", "GHGP-GWP-2024"),
    ("REF-006", "Conversión tonelada a kilogramo", "Unidades", "Comprueba conversión de masa y aplicación del factor.", 1.0, "t", 0.5, "kg", "CO2e", 1.0, 1000.0, 500.0, 500.0, 1e-9, "Calculado", "Sistema Internacional"),
    ("REF-007", "Conversión galón a litro", "Unidades", "Comprueba la conversión configurada de galón estadounidense a litro.", 10.0, "gal", 2.0, "L", "CO2e", 1.0, 37.85411784, 75.70823568, 75.70823568, 1e-8, "Calculado", "Conversión US gallon"),
    ("REF-008", "Rechazo de dimensiones incompatibles", "Control", "El motor debe rechazar energía convertida a volumen.", 100.0, "kWh", 2.0, "L", "CO2e", 1.0, 0.0, 0.0, 0.0, 1e-9, "Error", "Control dimensional del motor"),
]


def _upsert_source_documents(session: Session) -> dict[str, MethodologySourceDocument]:
    result: dict[str, MethodologySourceDocument] = {}
    today = date.today()
    for spec in OFFICIAL_SOURCE_SPECS:
        item = session.scalar(select(MethodologySourceDocument).where(MethodologySourceDocument.code == spec["code"]))
        if not item:
            item = MethodologySourceDocument(code=spec["code"], title=spec["title"], issuing_body=spec["issuing_body"])
            session.add(item)
        for key, value in spec.items():
            setattr(item, key, value)
        item.accessed_at = today
        result[item.code] = item
    session.flush()
    return result


def _upsert_gwp_values(session: Session) -> dict[str, Gas]:
    gases: dict[str, Gas] = {}
    source = "GHG Protocol · IPCC Global Warming Potential Values v2.0 (2024)"
    for code, spec in GWP_SPECS.items():
        gas = session.scalar(select(Gas).where(Gas.code == code))
        if not gas:
            gas = Gas(code=code, name=spec["name"], formula=spec["formula"])
            session.add(gas)
            session.flush()
        else:
            gas.name = spec["name"]
            gas.formula = spec["formula"]
        gases[code] = gas
        for assessment, value in spec["values"].items():
            gwp = session.scalar(select(GWPValue).where(
                GWPValue.gas_id == gas.id,
                GWPValue.assessment == assessment,
                GWPValue.horizon_years == 100,
            ))
            if not gwp:
                gwp = GWPValue(gas_id=gas.id, assessment=assessment, horizon_years=100, value=value)
                session.add(gwp)
            gwp.value = value
            gwp.source = source
            gwp.status = "Aprobado"
    session.flush()
    return gases


def _upsert_official_electricity_factor(
    session: Session,
    gases: dict[str, Gas],
    documents: dict[str, MethodologySourceDocument],
) -> EmissionFactorVersion:
    factor_name = "Electricidad SIN Colombia · inventarios 2024"
    factor = session.scalar(select(EmissionFactor).where(EmissionFactor.name == factor_name))
    if not factor:
        factor = EmissionFactor(
            name=factor_name,
            activity_type="Electricidad adquirida",
            country="Colombia",
            sector="Multisectorial",
            status="Activo",
            is_demo=False,
        )
        session.add(factor)
        session.flush()
    version = session.scalar(select(EmissionFactorVersion).where(
        EmissionFactorVersion.factor_id == factor.id,
        EmissionFactorVersion.version == "UPME-2024-R085",
        EmissionFactorVersion.gas_id == gases["CO2e"].id,
    ))
    if not version:
        version = EmissionFactorVersion(
            factor_id=factor.id,
            gas_id=gases["CO2e"].id,
            version="UPME-2024-R085",
            value=0.220,
            input_unit="kWh",
            output_unit="kg CO2e",
        )
        session.add(version)
    version.value = 0.220
    version.input_unit = "kWh"
    version.output_unit = "kg CO2e"
    version.source_organization = "Unidad de Planeación Minero Energética · UPME"
    version.source_document = "Resolución 000085 de 2026 · factor de emisión SIN 2024"
    version.publication_year = 2026
    version.geographic_scope = "Colombia · Sistema Interconectado Nacional"
    version.technology_scope = "Electricidad adquirida del SIN"
    version.uncertainty_percentage = 0
    version.status = "Aprobado"
    version.notes = "Factor agregado para inventarios GEI. Validar que el consumo corresponda al SIN y revelar el año del factor."
    version.approved_by = "Fuente oficial UPME · control documental V0.22"
    version.approved_at = datetime.now(UTC)
    session.flush()

    documentation = session.scalar(select(FactorDocumentation).where(FactorDocumentation.factor_version_id == version.id))
    if not documentation:
        documentation = FactorDocumentation(factor_version_id=version.id)
        session.add(documentation)
    documentation.source_document_id = documents["UPME-R085-2026"].id
    documentation.factor_kind = "Oficial nacional"
    documentation.reporting_use = "Formal"
    documentation.page_reference = "Página 3 de 4"
    documentation.table_reference = "Artículo 1 · literal B"
    documentation.data_year = 2024
    documentation.source_value = 0.220
    documentation.source_unit = "tCO2e/MWh"
    documentation.conversion_expression = "0,220 tCO2e/MWh × 1000 kg/t ÷ 1000 kWh/MWh = 0,220 kgCO2e/kWh"
    documentation.aggregated_co2e = True
    documentation.gwp_embedded = "Factor agregado; composición por gas no publicada en la resolución"
    documentation.methane_origin = "No aplica"
    documentation.quality_grade = "A"
    documentation.review_status = "Aprobado documentalmente"
    documentation.reviewer = "Control metodológico V0.22"
    documentation.reviewed_at = datetime.now(UTC)
    documentation.restriction_notes = "No confundir con los factores de proyectos MDL del mismo artículo. Aplicar el literal B para inventarios de GEI."
    session.flush()
    return version


def _document_existing_factors(
    session: Session,
    documents: dict[str, MethodologySourceDocument],
    official_version: EmissionFactorVersion,
) -> None:
    versions = session.scalars(
        select(EmissionFactorVersion)
        .options(selectinload(EmissionFactorVersion.factor), selectinload(EmissionFactorVersion.gas))
    ).all()
    for version in versions:
        existing = session.scalar(select(FactorDocumentation).where(FactorDocumentation.factor_version_id == version.id))
        if existing:
            continue
        is_demo = bool(version.factor.is_demo) or "demo" in version.version.lower() or "demostr" in version.source_organization.lower()
        documentation = FactorDocumentation(
            factor_version_id=version.id,
            source_document_id=documents["CTH-DEMO"].id if is_demo else None,
            factor_kind="Demostrativo" if is_demo else "Interno documentado",
            reporting_use="Demostrativo" if is_demo else "Piloto",
            page_reference="",
            table_reference="",
            data_year=version.publication_year,
            source_value=version.value,
            source_unit=f"{version.output_unit}/{version.input_unit}",
            conversion_expression="Sin conversión documental registrada" if not is_demo else "Dato sintético cargado directamente",
            aggregated_co2e=version.gas.code == "CO2e",
            gwp_embedded="No documentado" if version.gas.code == "CO2e" else "No aplica; factor por gas",
            methane_origin="No documentado" if version.gas.code == "CH4" else "No aplica",
            quality_grade="D" if is_demo else "C",
            review_status="Demostrativo" if is_demo else "Pendiente",
            reviewer="",
            restriction_notes="No apto para uso formal." if is_demo else "Requiere completar fuente primaria, referencia y aprobación.",
        )
        session.add(documentation)
    session.flush()

    # El factor oficial sustituye la asignación demostrativa de electricidad sin alterar el resultado numérico.
    electricity_sources = session.scalars(select(EmissionSource).where(EmissionSource.category == "Energía adquirida")).all()
    for source in electricity_sources:
        assignments = session.scalars(select(SourceFactorAssignment).where(SourceFactorAssignment.source_id == source.id)).all()
        official_assignment = next((item for item in assignments if item.factor_version_id == official_version.id), None)
        if not official_assignment:
            official_assignment = SourceFactorAssignment(
                source_id=source.id,
                factor_version_id=official_version.id,
                active=True,
                assigned_by="migración V0.22",
                notes="Reemplazo controlado del factor demostrativo por UPME 2024 para inventarios.",
            )
            session.add(official_assignment)
        else:
            official_assignment.active = True
        for assignment in assignments:
            if assignment.factor_version_id != official_version.id:
                factor_version = session.get(EmissionFactorVersion, assignment.factor_version_id)
                if factor_version and factor_version.factor.is_demo and factor_version.factor.activity_type == "Electricidad adquirida":
                    assignment.active = False
    session.flush()


def _upsert_selection_rules(session: Session) -> None:
    for code, priority, name, activity_type, country, input_unit, gas_code, kind, year_match, gap, quality, use, rationale in RULE_SPECS:
        rule = session.scalar(select(FactorSelectionRule).where(FactorSelectionRule.code == code))
        if not rule:
            rule = FactorSelectionRule(code=code, name=name)
            session.add(rule)
        rule.priority = priority
        rule.name = name
        rule.activity_type = activity_type
        rule.country = country
        rule.input_unit = input_unit
        rule.gas_code = gas_code
        rule.preferred_factor_kind = kind
        rule.requires_year_match = year_match
        rule.max_year_gap = gap
        rule.minimum_quality_grade = quality
        rule.allowed_reporting_use = use
        rule.status = "Activa"
        rule.rationale = rationale
    session.flush()


def _upsert_reference_cases(session: Session) -> None:
    for spec in REFERENCE_CASE_SPECS:
        (
            code, title, category, description, activity_value, activity_unit, factor_value,
            factor_input_unit, gas_code, gwp_value, expected_normalized, expected_gas,
            expected_co2e, tolerance, expected_status, source_reference,
        ) = spec
        case = session.scalar(select(ReferenceCalculationCase).where(ReferenceCalculationCase.code == code))
        if not case:
            case = ReferenceCalculationCase(code=code, title=title, activity_value=activity_value, activity_unit=activity_unit, factor_value=factor_value, factor_input_unit=factor_input_unit, gas_code=gas_code, expected_normalized_value=expected_normalized, expected_gas_kg=expected_gas, expected_co2e_kg=expected_co2e)
            session.add(case)
        case.title = title
        case.category = category
        case.description = description
        case.activity_value = activity_value
        case.activity_unit = activity_unit
        case.factor_value = factor_value
        case.factor_input_unit = factor_input_unit
        case.gas_code = gas_code
        case.gwp_value = gwp_value
        case.expected_normalized_value = expected_normalized
        case.expected_gas_kg = expected_gas
        case.expected_co2e_kg = expected_co2e
        case.tolerance = tolerance
        case.expected_status = expected_status
        case.source_reference = source_reference
        case.active = True
    session.flush()


def _update_consolidation_controls(session: Session) -> None:
    for finding in session.scalars(select(ConsolidationFinding).where(ConsolidationFinding.code.in_(["MET-001", "MET-002"]))):
        finding.status = "En curso"
        finding.evidence = (
            "V0.33 incorpora cierre metodológico, incertidumbre, partidas separadas y registro documental, UPME SIN 2024, GWP AR4/AR5/AR6, reglas de selección y casos patrón. "
            "La biblioteca multisectorial completa y la revisión profesional externa siguen pendientes."
        )
        finding.target_version = "V0.23" if finding.code == "MET-001" else "V0.22"
        finding.updated_at = datetime.now(UTC)
    for gate in session.scalars(select(ReleaseGate).where(ReleaseGate.code.in_(["GATE-METH", "GATE-CALC"]))):
        gate.status = "Parcial"
        gate.evidence = "Núcleo metodológico V0.33 con trazabilidad documental y suite de referencia reproducible."
        gate.notes = "No aprobar V1.0 hasta completar factores sectoriales, piloto real y revisión independiente."
    session.flush()


def ensure_methodology_core_defaults(session: Session) -> None:
    documents = _upsert_source_documents(session)
    gases = _upsert_gwp_values(session)
    official = _upsert_official_electricity_factor(session, gases, documents)
    _document_existing_factors(session, documents, official)
    _upsert_selection_rules(session)
    _upsert_reference_cases(session)
    _update_consolidation_controls(session)
    session.flush()


def _convert_for_case(session: Session, value: float, from_unit: str, to_unit: str) -> tuple[float | None, str]:
    from .calculations import convert_value

    return convert_value(session, value, from_unit, to_unit)


def run_reference_suite(session: Session, executed_by: str = "sistema") -> MethodologyValidationRun:
    cases = session.scalars(select(ReferenceCalculationCase).where(ReferenceCalculationCase.active.is_(True)).order_by(ReferenceCalculationCase.code)).all()
    run = MethodologyValidationRun(
        run_code=f"VAL-{datetime.now(UTC).strftime('%Y%m%dT%H%M%S')}-{secrets.token_hex(3)}",
        engine_version=ENGINE_VERSION,
        executed_by=executed_by,
        total_cases=len(cases),
    )
    session.add(run)
    session.flush()
    details: list[dict[str, Any]] = []
    passed_count = 0
    for case in cases:
        normalized, conversion_note = _convert_for_case(session, case.activity_value, case.activity_unit, case.factor_input_unit)
        if normalized is None:
            status = "Error"
            actual_normalized = 0.0
            gas_kg = 0.0
            co2e_kg = 0.0
        else:
            status = "Calculado"
            actual_normalized = normalized
            gas_kg = normalized * case.factor_value
            co2e_kg = gas_kg * case.gwp_value
        differences = [
            abs(actual_normalized - case.expected_normalized_value),
            abs(gas_kg - case.expected_gas_kg),
            abs(co2e_kg - case.expected_co2e_kg),
        ]
        difference = max(differences)
        passed = status == case.expected_status and difference <= case.tolerance
        passed_count += int(passed)
        detail = (
            f"{case.activity_value:g} {case.activity_unit} → {actual_normalized:g} {case.factor_input_unit}; "
            f"× {case.factor_value:g} = {gas_kg:g} kg {case.gas_code}; × GWP {case.gwp_value:g} = {co2e_kg:g} kgCO2e. "
            f"{conversion_note}"
        )
        session.add(ReferenceCaseResult(
            run_id=run.id,
            case_id=case.id,
            normalized_value=actual_normalized,
            gas_kg=gas_kg,
            co2e_kg=co2e_kg,
            absolute_difference=difference,
            passed=passed,
            status="Aprobado" if passed else "Fallido",
            detail=detail,
        ))
        details.append({"case": case.code, "passed": passed, "difference": difference, "status": status})
    run.passed_cases = passed_count
    run.failed_cases = len(cases) - passed_count
    run.status = "Aprobado" if run.failed_cases == 0 and run.total_cases > 0 else "Fallido"
    run.details_json = json.dumps(details, ensure_ascii=False)
    session.flush()
    return run


def factor_rows(session: Session) -> list[dict[str, Any]]:
    versions = session.scalars(
        select(EmissionFactorVersion)
        .options(selectinload(EmissionFactorVersion.factor), selectinload(EmissionFactorVersion.gas))
        .order_by(EmissionFactorVersion.id)
    ).all()
    documentation = {
        item.factor_version_id: item
        for item in session.scalars(select(FactorDocumentation).options(selectinload(FactorDocumentation.source_document))).all()
    }
    rows: list[dict[str, Any]] = []
    for version in versions:
        doc = documentation.get(version.id)
        rows.append({
            "version": version,
            "factor": version.factor,
            "gas": version.gas,
            "documentation": doc,
            "formal": bool(doc and doc.reporting_use == "Formal" and doc.review_status.startswith("Aprobado") and version.status == "Aprobado"),
            "documented": bool(doc and doc.source_document_id and doc.page_reference and doc.source_unit),
        })
    return rows


def methodology_summary(session: Session) -> dict[str, Any]:
    sources = session.scalars(select(MethodologySourceDocument).order_by(MethodologySourceDocument.code)).all()
    rules = session.scalars(select(FactorSelectionRule).order_by(FactorSelectionRule.priority, FactorSelectionRule.code)).all()
    cases = session.scalars(select(ReferenceCalculationCase).where(ReferenceCalculationCase.active.is_(True)).order_by(ReferenceCalculationCase.code)).all()
    gwps = session.scalars(select(GWPValue).options(selectinload(GWPValue.gas)).order_by(GWPValue.assessment, GWPValue.gas_id)).all()
    rows = factor_rows(session)
    last_run = session.scalar(
        select(MethodologyValidationRun)
        .options(selectinload(MethodologyValidationRun.results).selectinload(ReferenceCaseResult.case))
        .order_by(MethodologyValidationRun.executed_at.desc(), MethodologyValidationRun.id.desc())
        .limit(1)
    )
    official_sources = [item for item in sources if item.status != "Demostrativo"]
    formal_factors = [row for row in rows if row["formal"]]
    demo_factors = [row for row in rows if row["documentation"] and row["documentation"].reporting_use == "Demostrativo"]
    undocumented = [row for row in rows if not row["documented"]]
    active_rules = [rule for rule in rules if rule.status == "Activa"]
    score = 0
    score += 20 if len(official_sources) >= 6 else round(20 * len(official_sources) / 6)
    score += 20 if formal_factors else 0
    score += 15 if active_rules else 0
    score += 25 if last_run and last_run.status == "Aprobado" else 0
    score += 10 if len(gwps) >= 15 else round(10 * min(len(gwps), 15) / 15)
    score += 10 if not undocumented else max(0, 10 - min(len(undocumented), 10))
    return {
        "sources": sources,
        "rules": rules,
        "cases": cases,
        "gwps": gwps,
        "factor_rows": rows,
        "last_run": last_run,
        "metrics": {
            "score": min(score, 100),
            "official_sources": len(official_sources),
            "formal_factors": len(formal_factors),
            "demo_factors": len(demo_factors),
            "undocumented_factors": len(undocumented),
            "active_rules": len(active_rules),
            "reference_cases": len(cases),
            "passed_cases": last_run.passed_cases if last_run else 0,
            "failed_cases": last_run.failed_cases if last_run else 0,
        },
    }


def select_factor_candidates(
    session: Session,
    *,
    activity_type: str,
    country: str,
    input_unit: str,
    data_year: int,
) -> list[dict[str, Any]]:
    rows = factor_rows(session)
    candidates: list[dict[str, Any]] = []
    for row in rows:
        factor = row["factor"]
        version = row["version"]
        doc: FactorDocumentation | None = row["documentation"]
        score = 0
        reasons: list[str] = []
        if factor.activity_type == activity_type:
            score += 35
            reasons.append("actividad coincidente")
        else:
            continue
        if factor.country == country or factor.country in {"Global", "Internacional"}:
            score += 20
            reasons.append("geografía compatible")
        if version.input_unit == input_unit:
            score += 20
            reasons.append("unidad directa")
        if doc:
            if doc.reporting_use == "Formal":
                score += 20
                reasons.append("apto para reporte formal")
            elif doc.reporting_use == "Demostrativo":
                score -= 100
                reasons.append("solo demostrativo")
            if doc.data_year is not None:
                gap = abs(doc.data_year - data_year)
                score += max(0, 10 - 5 * gap)
                reasons.append(f"diferencia temporal {gap} año(s)")
            score += QUALITY_RANK.get(doc.quality_grade, 0)
        if version.status != "Aprobado":
            score -= 50
            reasons.append("factor no aprobado")
        candidates.append({"version": version, "documentation": doc, "score": score, "reasons": reasons})
    return sorted(candidates, key=lambda item: (-item["score"], item["version"].id))


def build_methodology_workbook(summary: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fuentes"
    ws.append(["Código", "Título", "Emisor", "Tipo", "Fecha", "Jurisdicción", "Estado", "URL", "Citación"])
    for item in summary["sources"]:
        ws.append([item.code, item.title, item.issuing_body, item.document_type, item.publication_date, item.jurisdiction, item.status, item.source_url, item.citation])

    ws = wb.create_sheet("Factores")
    ws.append(["Factor", "Versión", "Actividad", "Gas", "Valor", "Entrada", "Salida", "Tipo", "Uso", "Calidad", "Año dato", "Fuente", "Referencia", "Estado documental", "Formal"])
    for row in summary["factor_rows"]:
        version = row["version"]
        doc = row["documentation"]
        ws.append([
            row["factor"].name, version.version, row["factor"].activity_type, row["gas"].code,
            version.value, version.input_unit, version.output_unit,
            doc.factor_kind if doc else "", doc.reporting_use if doc else "", doc.quality_grade if doc else "",
            doc.data_year if doc else None, doc.source_document.code if doc and doc.source_document else "",
            f"{doc.page_reference} {doc.table_reference}".strip() if doc else "",
            doc.review_status if doc else "Sin documentación", "Sí" if row["formal"] else "No",
        ])

    ws = wb.create_sheet("GWP")
    ws.append(["Gas", "Nombre", "Evaluación", "Horizonte", "Valor", "Fuente", "Estado"])
    for item in summary["gwps"]:
        ws.append([item.gas.code, item.gas.name, item.assessment, item.horizon_years, item.value, item.source, item.status])

    ws = wb.create_sheet("Reglas")
    ws.append(["Código", "Prioridad", "Nombre", "Actividad", "País", "Unidad", "Gas", "Tipo preferido", "Año", "Brecha máxima", "Calidad mínima", "Uso", "Estado", "Justificación"])
    for item in summary["rules"]:
        ws.append([item.code, item.priority, item.name, item.activity_type, item.country, item.input_unit, item.gas_code, item.preferred_factor_kind, "Sí" if item.requires_year_match else "No", item.max_year_gap, item.minimum_quality_grade, item.allowed_reporting_use, item.status, item.rationale])

    ws = wb.create_sheet("Casos patrón")
    ws.append(["Código", "Título", "Categoría", "Actividad", "Unidad", "Factor", "Unidad factor", "Gas", "GWP", "Normalizado esperado", "Gas esperado kg", "CO2e esperado kg", "Estado esperado", "Fuente"])
    for item in summary["cases"]:
        ws.append([item.code, item.title, item.category, item.activity_value, item.activity_unit, item.factor_value, item.factor_input_unit, item.gas_code, item.gwp_value, item.expected_normalized_value, item.expected_gas_kg, item.expected_co2e_kg, item.expected_status, item.source_reference])

    ws = wb.create_sheet("Última validación")
    ws.append(["Métrica", "Valor"])
    last_run = summary["last_run"]
    if last_run:
        for key, value in [
            ("Código", last_run.run_code), ("Motor", last_run.engine_version), ("Ejecutada", last_run.executed_at),
            ("Ejecutor", last_run.executed_by), ("Casos", last_run.total_cases), ("Aprobados", last_run.passed_cases),
            ("Fallidos", last_run.failed_cases), ("Estado", last_run.status),
        ]:
            ws.append([key, value])
        ws.append([])
        ws.append(["Caso", "Estado", "Normalizado", "Gas kg", "CO2e kg", "Diferencia", "Detalle"])
        for result in sorted(last_run.results, key=lambda item: item.case.code):
            ws.append([result.case.code, result.status, result.normalized_value, result.gas_kg, result.co2e_kg, result.absolute_difference, result.detail])
    else:
        ws.append(["Estado", "Sin ejecución"])

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 48)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def summary_json(summary: dict[str, Any]) -> str:
    last_run = summary["last_run"]
    payload = {
        "engine_version": ENGINE_VERSION,
        "metrics": summary["metrics"],
        "last_validation": None if not last_run else {
            "run_code": last_run.run_code,
            "executed_at": last_run.executed_at.isoformat(),
            "status": last_run.status,
            "total": last_run.total_cases,
            "passed": last_run.passed_cases,
            "failed": last_run.failed_cases,
        },
        "sources": [{"code": item.code, "title": item.title, "status": item.status, "url": item.source_url} for item in summary["sources"]],
        "formal_factors": [
            {"name": row["factor"].name, "version": row["version"].version, "value": row["version"].value, "input_unit": row["version"].input_unit}
            for row in summary["factor_rows"] if row["formal"]
        ],
    }
    return json.dumps(payload, ensure_ascii=False, default=str)
