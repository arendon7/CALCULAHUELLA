from __future__ import annotations

import hashlib
import json
import statistics
from collections import defaultdict
from datetime import UTC, date, datetime, timedelta
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload

from .calculations import convert_value, recalculate_source
from .database import (
    ActivityData,
    DataImportBatch,
    DataImportRow,
    DataQualityFinding,
    EmissionSource,
    PilotExecution,
    PilotExecutionSourceLink,
    PilotIssue,
    PilotProject,
    Inventory,
    add_audit,
    refresh_progress,
)

DATA_QUALITY_VERSION = "0.28.0"
ALLOWED_ORIGINS = {
    "Medición directa",
    "Factura",
    "Registro operativo",
    "Registro contable",
    "Información de proveedor",
    "Certificado",
    "Encuesta",
    "Estimación",
}


def _execution_query(organization_id: int):
    return (
        select(PilotExecution)
        .join(PilotProject)
        .where(PilotProject.organization_id == organization_id, PilotProject.code == "GREENATICS-2026")
        .options(
            selectinload(PilotExecution.inventory)
            .selectinload(Inventory.sources),
            selectinload(PilotExecution.source_links)
            .selectinload(PilotExecutionSourceLink.requirement),
            selectinload(PilotExecution.source_links)
            .selectinload(PilotExecutionSourceLink.source)
            .selectinload(EmissionSource.activity_records),
            selectinload(PilotExecution.source_links).selectinload(PilotExecutionSourceLink.request),
        )
    )


def get_execution(session: Session, organization_id: int) -> PilotExecution | None:
    return session.scalar(_execution_query(organization_id).execution_options(populate_existing=True))


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _bool_value(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "si", "sí", "true", "x", "estimado"}


def _quality_level(origin: str, estimated: bool, evidence: str) -> str:
    if origin == "Medición directa" and evidence and not estimated:
        return "A"
    if origin in {"Factura", "Registro operativo", "Registro contable", "Información de proveedor", "Certificado"} and not estimated:
        return "B"
    if estimated or origin in {"Encuesta", "Estimación"}:
        return "C"
    return "D"


def _periods_for_frequency(year: int, frequency: str) -> list[tuple[date, date]]:
    frequency = (frequency or "Mensual").lower()
    if frequency == "anual":
        return [(date(year, 1, 1), date(year, 12, 31))]
    if frequency == "trimestral":
        return [
            (date(year, 1, 1), date(year, 3, 31)),
            (date(year, 4, 1), date(year, 6, 30)),
            (date(year, 7, 1), date(year, 9, 30)),
            (date(year, 10, 1), date(year, 12, 31)),
        ]
    result: list[tuple[date, date]] = []
    for month in range(1, 13):
        if month == 12:
            end = date(year, 12, 31)
        else:
            end = date(year, month + 1, 1) - timedelta(days=1)
        result.append((date(year, month, 1), end))
    return result


def build_data_template(session: Session, organization_id: int) -> bytes:
    execution = get_execution(session, organization_id)
    if not execution or not execution.inventory:
        raise ValueError("Primero inicia la ejecución del piloto Greenatics.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Carga de datos"
    headers = [
        "Código fuente",
        "Fecha inicio",
        "Fecha fin",
        "Valor",
        "Unidad",
        "Origen",
        "Estimado",
        "Referencia de evidencia",
        "Notas",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F6B49")

    year = execution.inventory.start_date.year
    for link in sorted(execution.source_links, key=lambda item: (item.requirement.site, item.requirement.code)):
        if not link.source:
            continue
        for start, end in _periods_for_frequency(year, link.requirement.frequency):
            ws.append([
                link.requirement.code,
                start,
                end,
                None,
                link.source.preferred_unit or link.requirement.activity_unit,
                "Registro operativo",
                "No",
                "",
                f"{link.requirement.site} · {link.requirement.source_name}",
            ])

    widths = [18, 14, 14, 14, 14, 24, 12, 34, 48]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    guide = wb.create_sheet("Instrucciones")
    guide.append(["Regla", "Descripción"])
    guide_rows = [
        ("Un registro por periodo", "No repitas el mismo código y periodo dentro del archivo."),
        ("Unidad", "Usa la unidad sugerida. La plataforma admite conversiones dimensionalmente compatibles."),
        ("Evidencia", "Para fuentes de materialidad alta, registra factura, bitácora, certificado o referencia verificable."),
        ("Estimaciones", "Marca Sí únicamente cuando el valor no provenga de medición o soporte primario."),
        ("Valores", "No se admiten valores negativos. Los ceros deben estar justificados en notas."),
        ("Aplicación", "La carga primero se valida. Nada modifica el inventario hasta pulsar Aplicar lote."),
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 25
    guide.column_dimensions["B"].width = 100

    sources = wb.create_sheet("Catálogo de fuentes")
    sources.append(["Código", "Sede", "Fuente", "Alcance", "Categoría", "Frecuencia", "Unidad", "Materialidad", "Responsable"])
    for link in sorted(execution.source_links, key=lambda item: item.requirement.code):
        req = link.requirement
        sources.append([req.code, req.site, req.source_name, req.scope, req.category, req.frequency, req.activity_unit, req.materiality, req.data_owner])
    for col in range(1, 10):
        sources.column_dimensions[chr(64 + col)].width = 22

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _add_finding(
    session: Session,
    batch: DataImportBatch,
    row: DataImportRow | None,
    code: str,
    severity: str,
    message: str,
) -> None:
    session.add(
        DataQualityFinding(
            batch=batch,
            row=row,
            rule_code=code,
            severity=severity,
            message=message,
            status="Abierto",
        )
    )


def create_import_batch(
    session: Session,
    organization_id: int,
    filename: str,
    content: bytes,
    user_email: str,
) -> DataImportBatch:
    execution = get_execution(session, organization_id)
    if not execution or not execution.inventory:
        raise ValueError("Primero inicia la ejecución del piloto Greenatics.")
    if not content:
        raise ValueError("El archivo está vacío.")

    digest = hashlib.sha256(content).hexdigest()
    duplicate = session.scalar(
        select(DataImportBatch).where(
            DataImportBatch.organization_id == organization_id,
            DataImportBatch.file_hash == digest,
        )
    )
    if duplicate:
        raise ValueError(f"Este archivo ya fue cargado como {duplicate.code}.")

    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError("El archivo no es un Excel XLSX válido.") from exc
    if "Carga de datos" not in wb.sheetnames:
        raise ValueError("No se encontró la hoja 'Carga de datos'. Descarga la plantilla oficial.")

    count = session.scalar(select(func.count(DataImportBatch.id)).where(DataImportBatch.organization_id == organization_id)) or 0
    batch = DataImportBatch(
        organization_id=organization_id,
        execution_id=execution.id,
        inventory_id=execution.inventory.id,
        code=f"DQ-{execution.inventory.start_date.year}-{count + 1:04d}",
        filename=filename[:220],
        file_hash=digest,
        source_format="XLSX",
        source_sheet="Carga de datos",
        mapping_json='{"mode":"pilot_template"}',
        original_headers_json='["Código fuente","Fecha inicio","Fecha fin","Valor","Unidad","Origen","Estimado","Referencia de evidencia","Notas"]',
        status="Validando",
        uploaded_by=user_email,
    )
    session.add(batch)
    session.flush()

    links = {link.requirement.code.strip().upper(): link for link in execution.source_links if link.source}
    ws = wb["Carga de datos"]
    parsed: list[DataImportRow] = []
    seen: set[tuple[str, date | None, date | None]] = set()
    values_by_source: dict[int, list[float]] = defaultdict(list)

    for excel_row in range(2, ws.max_row + 1):
        values = [ws.cell(excel_row, column).value for column in range(1, 10)]
        if not any(value not in (None, "") for value in values):
            continue
        code = str(values[0] or "").strip().upper()
        start = _parse_date(values[1])
        end = _parse_date(values[2])
        try:
            numeric_value = float(values[3]) if values[3] not in (None, "") else None
        except (TypeError, ValueError):
            numeric_value = None
        unit = str(values[4] or "").strip()
        origin = str(values[5] or "Registro operativo").strip()
        if origin not in ALLOWED_ORIGINS:
            origin = "Registro operativo"
        estimated = _bool_value(values[6])
        evidence = str(values[7] or "").strip()
        notes = str(values[8] or "").strip()
        link = links.get(code)
        row = DataImportRow(
            batch=batch,
            row_number=excel_row,
            requirement_code=code,
            source_id=link.source_id if link else None,
            period_start=start,
            period_end=end,
            value=numeric_value,
            unit=unit,
            evidence_reference=evidence,
            data_origin=origin,
            is_estimated=estimated,
            quality_level=_quality_level(origin, estimated, evidence),
            status="Pendiente",
        )
        session.add(row)
        session.flush()
        messages: list[dict[str, str]] = []

        def issue(rule: str, severity: str, message: str) -> None:
            messages.append({"rule": rule, "severity": severity, "message": message})
            _add_finding(session, batch, row, rule, severity, message)

        if not code or not link:
            issue("DQ-001", "Error", "Código de fuente vacío o no reconocido.")
        if not start or not end or start > end:
            issue("DQ-002", "Error", "El periodo es inválido o está incompleto.")
        elif start < execution.inventory.start_date or end > execution.inventory.end_date:
            issue("DQ-003", "Error", "El periodo está fuera del inventario piloto.")
        if numeric_value is None:
            issue("DQ-004", "Error", "El valor no es numérico.")
        elif numeric_value < 0:
            issue("DQ-005", "Error", "No se permiten valores negativos.")
        elif numeric_value == 0 and not notes:
            issue("DQ-006", "Advertencia", "El valor cero requiere justificación en notas.")

        if link and numeric_value is not None:
            preferred = link.source.preferred_unit or link.requirement.activity_unit
            if not unit:
                issue("DQ-007", "Error", "La unidad está vacía.")
            else:
                converted, explanation = convert_value(session, numeric_value, unit, preferred)
                if converted is None:
                    issue("DQ-009", "Error", explanation or f"La unidad {unit} no es compatible con {preferred}.")
                elif unit != preferred:
                    issue("DQ-008", "Advertencia", explanation or f"La unidad {unit} será convertida a {preferred}.")

            key = (code, start, end)
            if key in seen:
                issue("DQ-010", "Error", "Registro duplicado dentro del mismo archivo.")
            seen.add(key)

            if start and end:
                existing = session.scalar(
                    select(ActivityData).where(
                        ActivityData.source_id == link.source_id,
                        ActivityData.period_start == start,
                        ActivityData.period_end == end,
                    )
                )
                if existing:
                    issue("DQ-011", "Advertencia", "Ya existe un registro para el periodo; se actualizará al aplicar el lote.")

            if link.requirement.materiality == "Alta" and not evidence:
                issue("DQ-012", "Advertencia", "La fuente de materialidad alta no tiene referencia de evidencia.")
            if estimated:
                issue("DQ-013", "Advertencia", "El dato está marcado como estimado y requiere método documentado.")
            values_by_source[link.source_id].append(numeric_value)

        row.validation_messages = json.dumps(messages, ensure_ascii=False)
        parsed.append(row)

    if not parsed:
        _add_finding(session, batch, None, "DQ-000", "Error", "El archivo no contiene filas de datos.")

    session.flush()
    # Detecta valores extremos dentro del propio lote sin imponer una regla estadística rígida.
    for row in parsed:
        if row.source_id and row.value is not None:
            peers = values_by_source[row.source_id]
            if len(peers) >= 3:
                median = statistics.median(peers)
                if median > 0 and (row.value > median * 3 or row.value < median * 0.2):
                    message = f"Valor atípico frente a la mediana del lote ({median:,.2f} {row.unit})."
                    _add_finding(session, batch, row, "DQ-014", "Advertencia", message)
                    messages = json.loads(row.validation_messages or "[]")
                    messages.append({"rule": "DQ-014", "severity": "Advertencia", "message": message})
                    row.validation_messages = json.dumps(messages, ensure_ascii=False)

    session.flush()
    for row in parsed:
        severities = {item["severity"] for item in json.loads(row.validation_messages or "[]")}
        row.status = "Error" if "Error" in severities else ("Advertencia" if "Advertencia" in severities else "Válido")

    batch.total_rows = len(parsed)
    batch.error_rows = sum(row.status == "Error" for row in parsed)
    batch.warning_rows = sum(row.status == "Advertencia" for row in parsed)
    batch.valid_rows = sum(row.status == "Válido" for row in parsed)
    batch.quality_score = max(0, round(100 - batch.error_rows * 15 - batch.warning_rows * 4))
    batch.status = "Con errores" if batch.error_rows else "Validado"
    batch.validated_at = datetime.now(UTC)
    batch.notes = "La validación no modifica datos del inventario. Revisa hallazgos antes de aplicar."

    if batch.error_rows:
        issue_sequence = session.scalar(select(func.count(PilotIssue.id)).where(PilotIssue.execution_id == execution.id)) or 0
        session.add(
            PilotIssue(
                execution_id=execution.id,
                code=f"PIL-026-{issue_sequence + 1:03d}",
                category="Calidad de datos",
                title=f"Corregir lote {batch.code}",
                description=f"El lote contiene {batch.error_rows} filas con error y {batch.warning_rows} advertencias.",
                severity="Alta",
                status="Abierto",
                owner="Equipo de datos",
                created_by=user_email,
            )
        )

    add_audit(
        session,
        organization_id,
        user_email,
        "VALIDAR",
        "Lote de datos",
        batch.code,
        detail=f"{batch.total_rows} filas · {batch.error_rows} errores · {batch.warning_rows} advertencias",
    )
    session.flush()
    return batch


def apply_import_batch(session: Session, organization_id: int, batch_id: int, user_email: str) -> DataImportBatch:
    batch = session.scalar(
        select(DataImportBatch)
        .where(DataImportBatch.id == batch_id, DataImportBatch.organization_id == organization_id)
        .options(
            selectinload(DataImportBatch.rows).selectinload(DataImportRow.source),
            selectinload(DataImportBatch.execution).selectinload(PilotExecution.inventory),
        )
    )
    if not batch:
        raise ValueError("Lote no encontrado.")
    if batch.status == "Aplicado":
        return batch
    if batch.error_rows:
        raise ValueError("El lote contiene errores y no puede aplicarse.")
    if not batch.execution or not batch.execution.inventory:
        raise ValueError("El lote no está vinculado con un inventario piloto.")

    from .period_close import assert_periods_editable

    editable_periods = [
        (row.period_start, row.period_end)
        for row in batch.rows
        if row.status in {"Válido", "Advertencia"} and row.period_start and row.period_end
    ]
    assert_periods_editable(session, batch.execution.inventory.id, editable_periods)

    touched: set[int] = set()
    applied = 0
    for row in batch.rows:
        if row.status not in {"Válido", "Advertencia"} or not row.source or row.value is None or not row.period_start or not row.period_end:
            continue
        preferred = row.source.preferred_unit or row.unit
        normalized, explanation = convert_value(session, row.value, row.unit, preferred)
        if normalized is None:
            raise ValueError(explanation or f"No fue posible convertir {row.unit} a {preferred}.")
        record = session.scalar(
            select(ActivityData).where(
                ActivityData.source_id == row.source_id,
                ActivityData.period_start == row.period_start,
                ActivityData.period_end == row.period_end,
            )
        )
        if not record:
            record = ActivityData(
                source_id=row.source_id,
                period_start=row.period_start,
                period_end=row.period_end,
                value=normalized,
                unit=preferred,
                created_by=user_email,
            )
            session.add(record)
            session.flush()
        else:
            record.value = normalized
            record.unit = preferred
            record.updated_at = datetime.now(UTC)
        record.data_origin = row.data_origin
        record.quality_level = row.quality_level
        record.is_estimated = row.is_estimated
        record.status = "Provisional" if row.is_estimated or not row.evidence_reference else "Cargado"
        record.notes = " · ".join(filter(None, [f"Lote {batch.code}", f"Evidencia: {row.evidence_reference}" if row.evidence_reference else "", "Carga controlada V0.27"]))
        row.activity_data_id = record.id
        row.status = "Aplicado"
        touched.add(row.source_id)
        applied += 1

    for source_id in touched:
        source = session.get(EmissionSource, source_id)
        if source:
            recalculate_source(session, source)
            for request in source.requests:
                request.status = "Completado" if source.progress >= 100 else "En curso"
                if request.status == "Completado":
                    request.completed_at = datetime.now(UTC)

    refresh_progress(session, batch.execution.inventory)
    batch.applied_rows = applied
    batch.status = "Aplicado"
    batch.applied_at = datetime.now(UTC)
    add_audit(
        session,
        organization_id,
        user_email,
        "APLICAR",
        "Lote de datos",
        batch.code,
        detail=f"{applied} registros aplicados a {len(touched)} fuentes.",
    )
    session.flush()
    return batch


def resolve_finding(session: Session, organization_id: int, finding_id: int, resolution: str, user_email: str) -> DataQualityFinding:
    finding = session.scalar(
        select(DataQualityFinding)
        .join(DataImportBatch)
        .where(DataQualityFinding.id == finding_id, DataImportBatch.organization_id == organization_id)
    )
    if not finding:
        raise ValueError("Hallazgo no encontrado.")
    finding.status = "Cerrado"
    finding.resolution = resolution.strip()
    finding.resolved_at = datetime.now(UTC)
    add_audit(session, organization_id, user_email, "CERRAR", "Hallazgo de calidad", finding.rule_code, detail=finding.resolution)
    return finding


def data_quality_summary(session: Session, organization_id: int, batch_id: int | None = None) -> dict[str, Any]:
    execution = get_execution(session, organization_id)
    batches = list(
        session.scalars(
            select(DataImportBatch)
            .where(DataImportBatch.organization_id == organization_id, DataImportBatch.execution_id.is_not(None))
            .order_by(DataImportBatch.id.desc())
        )
    )
    selected: DataImportBatch | None = None
    if batch_id:
        selected = session.scalar(
            select(DataImportBatch)
            .where(DataImportBatch.id == batch_id, DataImportBatch.organization_id == organization_id)
            .options(
                selectinload(DataImportBatch.rows).selectinload(DataImportRow.source),
                selectinload(DataImportBatch.findings).selectinload(DataQualityFinding.row),
            )
        )
    elif batches:
        selected = session.scalar(
            select(DataImportBatch)
            .where(DataImportBatch.id == batches[0].id)
            .options(
                selectinload(DataImportBatch.rows).selectinload(DataImportRow.source),
                selectinload(DataImportBatch.findings).selectinload(DataQualityFinding.row),
            )
        )

    if selected:
        for row in selected.rows:
            try:
                messages = json.loads(row.validation_messages or "[]")
                row.validation_display = " · ".join(item.get("message", "") for item in messages if item.get("message"))
            except (TypeError, ValueError, json.JSONDecodeError):
                row.validation_display = row.validation_messages or ""

    total_rows = sum(item.total_rows for item in batches)
    applied_rows = sum(item.applied_rows for item in batches)
    open_findings = session.scalar(
        select(func.count(DataQualityFinding.id))
        .join(DataImportBatch)
        .where(DataImportBatch.organization_id == organization_id, DataQualityFinding.status == "Abierto")
    ) or 0

    coverage: list[dict[str, Any]] = []
    if execution:
        for link in sorted(execution.source_links, key=lambda item: (item.requirement.site, item.requirement.code)):
            source = link.source
            if not source:
                continue
            expected = 1 if link.requirement.frequency.lower() == "anual" else (4 if link.requirement.frequency.lower() == "trimestral" else 12)
            actual = len(source.activity_records)
            coverage.append(
                {
                    "code": link.requirement.code,
                    "site": link.requirement.site,
                    "name": link.requirement.source_name,
                    "expected": expected,
                    "actual": actual,
                    "percent": min(100, round(actual / expected * 100)) if expected else 0,
                    "quality": min((item.quality_level for item in source.activity_records), default="D"),
                    "evidence": sum(bool(item.evidence_id or "Evidencia:" in (item.notes or "")) for item in source.activity_records),
                }
            )

    return {
        "execution": execution,
        "batches": batches,
        "selected": selected,
        "metrics": {
            "batches": len(batches),
            "rows": total_rows,
            "applied": applied_rows,
            "open_findings": open_findings,
            "coverage": round(sum(item["percent"] for item in coverage) / max(len(coverage), 1)),
        },
        "coverage": coverage,
        "rules": [
            ("DQ-001–003", "Identidad y periodo", "Bloquea códigos desconocidos y periodos fuera del inventario."),
            ("DQ-004–006", "Valor", "Bloquea valores inválidos y exige justificar ceros."),
            ("DQ-007–009", "Unidad", "Comprueba compatibilidad dimensional y conversión controlada."),
            ("DQ-010–011", "Duplicidad", "Evita duplicados en el archivo y controla actualizaciones de periodos existentes."),
            ("DQ-012–013", "Trazabilidad", "Advierte ausencia de evidencia y uso de estimaciones."),
            ("DQ-014", "Atípicos", "Señala valores extremos frente a la mediana del mismo lote."),
        ],
    }


def build_quality_report(summary: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Indicador", "Valor"])
    for key, value in summary["metrics"].items():
        ws.append([key, value])

    batches = wb.create_sheet("Lotes")
    batches.append(["Código", "Archivo", "Estado", "Filas", "Válidas", "Advertencias", "Errores", "Aplicadas", "Puntaje", "Cargado por", "Fecha"])
    for item in summary["batches"]:
        batches.append([item.code, item.filename, item.status, item.total_rows, item.valid_rows, item.warning_rows, item.error_rows, item.applied_rows, item.quality_score, item.uploaded_by, item.uploaded_at])

    coverage = wb.create_sheet("Cobertura")
    coverage.append(["Código", "Sede", "Fuente", "Esperados", "Cargados", "Cobertura %", "Calidad", "Referencias de evidencia"])
    for item in summary["coverage"]:
        coverage.append([item["code"], item["site"], item["name"], item["expected"], item["actual"], item["percent"], item["quality"], item["evidence"]])

    selected = summary.get("selected")
    if selected:
        rows = wb.create_sheet("Filas último lote")
        rows.append(["Fila", "Código", "Inicio", "Fin", "Valor", "Unidad", "Calidad", "Estado", "Mensajes"])
        for item in selected.rows:
            rows.append([item.row_number, item.requirement_code, item.period_start, item.period_end, item.value, item.unit, item.quality_level, item.status, item.validation_messages])
        findings = wb.create_sheet("Hallazgos")
        findings.append(["Regla", "Severidad", "Fila", "Mensaje", "Estado", "Resolución"])
        for item in selected.findings:
            findings.append([item.rule_code, item.severity, item.row.row_number if item.row else None, item.message, item.status, item.resolution])

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
