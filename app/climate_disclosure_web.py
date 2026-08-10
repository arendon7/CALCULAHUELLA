from __future__ import annotations

from datetime import UTC, date, datetime
from io import BytesIO

from fastapi import Depends, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from .climate_disclosure import board_summary, build_board_pdf, disclosure_summary, scenario_comparison
from .database import add_audit, get_db
from .db.models import (
    ClimateBoardBriefing,
    ClimateBoardDecision,
    ClimateDisclosureRequirement,
    ClimateDisclosureStatement,
    ClimateScenarioDefinition,
    Inventory,
    Organization,
)


def _require_climate_disclosure_view(user: dict[str, object]) -> None:
    capabilities = user["capabilities"]
    if "view_climate_disclosure" not in capabilities and "manage_climate_disclosure" not in capabilities:
        raise HTTPException(403, "Tu rol no tiene acceso a divulgación climática")


def register_climate_disclosure_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash, parse_date
) -> None:
    @app.get("/divulgacion-climatica", response_class=HTMLResponse)
    def climate_disclosure_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
        _require_climate_disclosure_view(user)
        organization_id = int(user["organization_id"])
        comparison = scenario_comparison(session, organization_id)
        disclosure = disclosure_summary(session, organization_id)
        board = board_summary(session, organization_id)
        inventories = list(session.scalars(
            select(Inventory).where(Inventory.organization_id == organization_id).order_by(Inventory.start_date.desc())
        ))
        return templates.TemplateResponse(
            request=request, name="climate_disclosure.html",
            context=common_context(
                request, session, user, "climate_disclosure", comparison=comparison,
                disclosure=disclosure, board=board, inventories=inventories, today=date.today(),
            ),
        )

    @app.post("/divulgacion-climatica/escenarios/nuevo")
    def climate_scenario_create(
        request: Request, name: str = Form(...), code: str = Form(...), scenario_type: str = Form(...),
        temperature_pathway: str = Form("No especificada"), physical_multiplier: float = Form(1.0),
        transition_multiplier: float = Form(1.0), opportunity_multiplier: float = Form(1.0),
        carbon_price_2030: float = Form(0), energy_cost_change_pct: float = Form(0),
        demand_change_pct: float = Form(0), probability_weight: float = Form(0), narrative: str = Form(""),
        source_reference: str = Form(""), status: str = Form("Activo"),
        session: Session = Depends(get_db), user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_climate_disclosure")
        organization_id = int(user["organization_id"])
        normalized_code = code.strip().upper()[:40]
        duplicate = session.scalar(select(ClimateScenarioDefinition).where(
            ClimateScenarioDefinition.organization_id == organization_id,
            ClimateScenarioDefinition.code == normalized_code,
        ))
        if duplicate:
            raise HTTPException(409, "Ya existe un escenario con ese código")
        comparison = scenario_comparison(session, organization_id)
        assessment = comparison["risk_summary"]["assessment"]
        scenario = ClimateScenarioDefinition(
            organization_id=organization_id, assessment_id=assessment.id if assessment else None,
            name=name.strip(), code=normalized_code, scenario_type=scenario_type.strip(),
            temperature_pathway=temperature_pathway.strip(), physical_multiplier=max(0.1, min(3.0, physical_multiplier)),
            transition_multiplier=max(0.1, min(3.0, transition_multiplier)),
            opportunity_multiplier=max(0.1, min(3.0, opportunity_multiplier)),
            carbon_price_2030=max(0, carbon_price_2030), energy_cost_change_pct=max(-100, min(500, energy_cost_change_pct)),
            demand_change_pct=max(-100, min(500, demand_change_pct)), probability_weight=max(0, min(100, probability_weight)),
            narrative=narrative.strip(), source_reference=source_reference.strip(), status=status, created_by=str(user["email"]),
        )
        session.add(scenario)
        add_audit(session, organization_id, str(user["email"]), "CREAR", "Escenario climático", scenario.name, new_value=scenario.code)
        session.commit(); set_flash(request, "Escenario climático registrado.")
        return RedirectResponse("/divulgacion-climatica", status_code=303)

    @app.post("/divulgacion-climatica/escenarios/{scenario_id}/actualizar")
    def climate_scenario_update(
        scenario_id: int, request: Request, physical_multiplier: float = Form(...), transition_multiplier: float = Form(...),
        opportunity_multiplier: float = Form(...), carbon_price_2030: float = Form(0),
        energy_cost_change_pct: float = Form(0), demand_change_pct: float = Form(0),
        probability_weight: float = Form(0), narrative: str = Form(""), source_reference: str = Form(""),
        status: str = Form("Activo"), session: Session = Depends(get_db), user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_climate_disclosure")
        scenario = session.get(ClimateScenarioDefinition, scenario_id)
        if not scenario or scenario.organization_id != int(user["organization_id"]):
            raise HTTPException(404, "Escenario no encontrado")
        scenario.physical_multiplier = max(0.1, min(3.0, physical_multiplier))
        scenario.transition_multiplier = max(0.1, min(3.0, transition_multiplier))
        scenario.opportunity_multiplier = max(0.1, min(3.0, opportunity_multiplier))
        scenario.carbon_price_2030 = max(0, carbon_price_2030)
        scenario.energy_cost_change_pct = max(-100, min(500, energy_cost_change_pct))
        scenario.demand_change_pct = max(-100, min(500, demand_change_pct))
        scenario.probability_weight = max(0, min(100, probability_weight))
        scenario.narrative = narrative.strip(); scenario.source_reference = source_reference.strip(); scenario.status = status
        add_audit(session, scenario.organization_id, str(user["email"]), "ACTUALIZAR", "Escenario climático", scenario.name, new_value=f"Peso {scenario.probability_weight}%")
        session.commit(); set_flash(request, "Supuestos del escenario actualizados.")
        return RedirectResponse("/divulgacion-climatica", status_code=303)

    @app.post("/divulgacion-climatica/declaracion")
    def climate_disclosure_save(
        request: Request, title: str = Form(...), inventory_id: str = Form(""), framework: str = Form(...),
        reporting_period: str = Form(...), scope_description: str = Form(""), materiality_basis: str = Form(""),
        owner: str = Form(...), status: str = Form("Borrador"), notes: str = Form(""),
        session: Session = Depends(get_db), user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_climate_disclosure")
        organization_id = int(user["organization_id"])
        summary = disclosure_summary(session, organization_id)
        statement = summary["statement"] or ClimateDisclosureStatement(organization_id=organization_id, title=title.strip(), created_by=str(user["email"]))
        if not summary["statement"]:
            session.add(statement)
        statement.title = title.strip(); statement.inventory_id = int(inventory_id) if inventory_id else None
        statement.framework = framework.strip(); statement.reporting_period = reporting_period.strip()
        statement.scope_description = scope_description.strip(); statement.materiality_basis = materiality_basis.strip()
        statement.owner = owner.strip(); statement.status = status; statement.notes = notes.strip()
        if status == "Aprobada":
            ensure_capability(user, "approve")
            statement.approved_by = str(user["email"]); statement.approved_at = datetime.now(UTC)
        add_audit(session, organization_id, str(user["email"]), "ACTUALIZAR", "Divulgación climática", statement.title, new_value=status)
        session.commit(); set_flash(request, "Ficha de divulgación guardada.")
        return RedirectResponse("/divulgacion-climatica", status_code=303)

    @app.post("/divulgacion-climatica/requisitos/nuevo")
    def climate_requirement_create(
        request: Request, pillar: str = Form(...), code: str = Form(...), requirement: str = Form(...),
        response: str = Form(""), status: str = Form("Pendiente"), evidence_reference: str = Form(""),
        owner: str = Form(...), due_date: str = Form(""), session: Session = Depends(get_db),
        user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_climate_disclosure")
        organization_id = int(user["organization_id"]); summary = disclosure_summary(session, organization_id)
        if status not in {"Completo", "Parcial", "Pendiente", "No aplica"}:
            raise HTTPException(400, "Estado de requisito inválido")
        statement = summary["statement"]
        if not statement:
            raise HTTPException(409, "Primero crea la ficha de divulgación")
        normalized_code = code.strip().upper()[:40]
        duplicate = session.scalar(select(ClimateDisclosureRequirement).where(
            ClimateDisclosureRequirement.statement_id == statement.id,
            ClimateDisclosureRequirement.code == normalized_code,
        ))
        if duplicate:
            raise HTTPException(409, "Ya existe un requisito con ese código")
        item = ClimateDisclosureRequirement(
            statement_id=statement.id, organization_id=organization_id, pillar=pillar.strip(), code=normalized_code,
            requirement=requirement.strip(), response=response.strip(), status=status,
            evidence_reference=evidence_reference.strip(), owner=owner.strip(),
            due_date=parse_date(due_date) if due_date else None, updated_by=str(user["email"]),
        )
        session.add(item); add_audit(session, organization_id, str(user["email"]), "CREAR", "Requisito de divulgación", item.code, new_value=item.pillar)
        session.commit(); set_flash(request, "Requisito de divulgación registrado.")
        return RedirectResponse("/divulgacion-climatica", status_code=303)

    @app.post("/divulgacion-climatica/requisitos/{requirement_id}/actualizar")
    def climate_requirement_update(
        requirement_id: int, request: Request, response: str = Form(""), status: str = Form(...),
        evidence_reference: str = Form(""), owner: str = Form(...), due_date: str = Form(""),
        session: Session = Depends(get_db), user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_climate_disclosure")
        requirement = session.get(ClimateDisclosureRequirement, requirement_id)
        if not requirement or requirement.organization_id != int(user["organization_id"]):
            raise HTTPException(404, "Requisito no encontrado")
        if status not in {"Completo", "Parcial", "Pendiente", "No aplica"}:
            raise HTTPException(400, "Estado de requisito inválido")
        requirement.response = response.strip(); requirement.status = status
        requirement.evidence_reference = evidence_reference.strip(); requirement.owner = owner.strip()
        requirement.due_date = parse_date(due_date) if due_date else None; requirement.updated_by = str(user["email"])
        add_audit(session, requirement.organization_id, str(user["email"]), "ACTUALIZAR", "Requisito de divulgación", requirement.code, new_value=status)
        session.commit(); set_flash(request, "Requisito de divulgación actualizado.")
        return RedirectResponse("/divulgacion-climatica", status_code=303)

    @app.post("/divulgacion-climatica/comite")
    def climate_board_save(
        request: Request, title: str = Form(...), meeting_date: str = Form(""), audience: str = Form("Comité directivo"),
        status: str = Form("Borrador"), executive_summary: str = Form(""), decisions_required: str = Form(""),
        key_message: str = Form(""), prepared_by: str = Form(...),
        session: Session = Depends(get_db), user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_climate_disclosure")
        organization_id = int(user["organization_id"])
        summary = board_summary(session, organization_id)
        briefing = summary["briefing"] or ClimateBoardBriefing(organization_id=organization_id, title=title.strip(), created_by=str(user["email"]))
        if not summary["briefing"]:
            session.add(briefing)
        risk_assessment = summary["risk_summary"]["assessment"]
        statement = summary["disclosure"]["statement"]
        briefing.assessment_id = risk_assessment.id if risk_assessment else None
        briefing.disclosure_id = statement.id if statement else None
        briefing.title = title.strip(); briefing.meeting_date = parse_date(meeting_date) if meeting_date else None
        briefing.audience = audience.strip(); briefing.status = status; briefing.executive_summary = executive_summary.strip()
        briefing.decisions_required = decisions_required.strip(); briefing.key_message = key_message.strip(); briefing.prepared_by = prepared_by.strip()
        if status == "Aprobado":
            ensure_capability(user, "approve")
            briefing.approved_by = str(user["email"]); briefing.approved_at = datetime.now(UTC)
        add_audit(session, organization_id, str(user["email"]), "ACTUALIZAR", "Informe de comité", briefing.title, new_value=status)
        session.commit(); set_flash(request, "Informe para comité actualizado.")
        return RedirectResponse("/divulgacion-climatica", status_code=303)

    @app.post("/divulgacion-climatica/decisiones/nueva")
    def climate_board_decision_create(
        request: Request, topic: str = Form(...), decision: str = Form(""), owner: str = Form(...),
        due_date: str = Form(""), status: str = Form("Pendiente"), rationale: str = Form(""),
        evidence_reference: str = Form(""), session: Session = Depends(get_db), user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_climate_disclosure")
        organization_id = int(user["organization_id"]); summary = board_summary(session, organization_id)
        briefing = summary["briefing"]
        if not briefing:
            raise HTTPException(409, "Primero crea el informe para comité")
        item = ClimateBoardDecision(
            briefing_id=briefing.id, organization_id=organization_id, topic=topic.strip(), decision=decision.strip(),
            owner=owner.strip(), due_date=parse_date(due_date) if due_date else None, status=status,
            rationale=rationale.strip(), evidence_reference=evidence_reference.strip(), created_by=str(user["email"]),
        )
        session.add(item); add_audit(session, organization_id, str(user["email"]), "CREAR", "Decisión de comité", item.topic, new_value=status)
        session.commit(); set_flash(request, "Decisión registrada.")
        return RedirectResponse("/divulgacion-climatica", status_code=303)

    @app.post("/divulgacion-climatica/decisiones/{decision_id}/estado")
    def climate_board_decision_update(
        decision_id: int, request: Request, decision: str = Form(""), owner: str = Form(...),
        due_date: str = Form(""), status: str = Form(...), rationale: str = Form(""),
        evidence_reference: str = Form(""), session: Session = Depends(get_db), user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_climate_disclosure")
        item = session.get(ClimateBoardDecision, decision_id)
        if not item or item.organization_id != int(user["organization_id"]):
            raise HTTPException(404, "Decisión no encontrada")
        item.decision = decision.strip(); item.owner = owner.strip(); item.due_date = parse_date(due_date) if due_date else None
        item.status = status; item.rationale = rationale.strip(); item.evidence_reference = evidence_reference.strip()
        add_audit(session, item.organization_id, str(user["email"]), "ACTUALIZAR", "Decisión de comité", item.topic, new_value=status)
        session.commit(); set_flash(request, "Estado de la decisión actualizado.")
        return RedirectResponse("/divulgacion-climatica", status_code=303)

    @app.get("/divulgacion-climatica/exportar.xlsx")
    def climate_disclosure_export(session: Session = Depends(get_db), user: dict = Depends(require_user)):
        _require_climate_disclosure_view(user)
        organization_id = int(user["organization_id"])
        comparison = scenario_comparison(session, organization_id); disclosure = disclosure_summary(session, organization_id); board = board_summary(session, organization_id)
        wb = Workbook(); ws = wb.active; ws.title = "Comparación"
        ws.append(["Escenario", "Código", "Trayectoria", "Probabilidad %", "Exposición", "Costo carbono", "Oportunidad", "Presión neta", "Presión ponderada", "Resiliencia", "Riesgos críticos"])
        for result in comparison["results"]:
            scenario = result["scenario"]
            ws.append([scenario.name, scenario.code, scenario.temperature_pathway, scenario.probability_weight, result["downside_exposure"], result["carbon_cost"], result["opportunity_value"], result["net_financial_pressure"], result["weighted_pressure"], result["resilience_score"], result["critical_risks"]])
        ws2 = wb.create_sheet("Supuestos"); ws2.append(["Escenario", "Tipo", "Multiplicador físico", "Multiplicador transición", "Multiplicador oportunidad", "Precio carbono 2030", "Cambio energía %", "Cambio demanda %", "Narrativa", "Fuente", "Estado"])
        for scenario in comparison["scenarios"]:
            ws2.append([scenario.name, scenario.scenario_type, scenario.physical_multiplier, scenario.transition_multiplier, scenario.opportunity_multiplier, scenario.carbon_price_2030, scenario.energy_cost_change_pct, scenario.demand_change_pct, scenario.narrative, scenario.source_reference, scenario.status])
        ws3 = wb.create_sheet("Divulgación"); ws3.append(["Pilar", "Código", "Requisito", "Respuesta", "Estado", "Evidencia", "Responsable", "Vencimiento"])
        for requirement in disclosure["requirements"]:
            ws3.append([requirement.pillar, requirement.code, requirement.requirement, requirement.response, requirement.status, requirement.evidence_reference, requirement.owner, requirement.due_date])
        ws4 = wb.create_sheet("Decisiones"); ws4.append(["Tema", "Decisión", "Responsable", "Vencimiento", "Estado", "Justificación", "Evidencia"])
        for item in board["decisions"]:
            ws4.append([item.topic, item.decision, item.owner, item.due_date, item.status, item.rationale, item.evidence_reference])
        buffer = BytesIO(); wb.save(buffer)
        return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="divulgacion_climatica_{organization_id}.xlsx"'})

    @app.get("/divulgacion-climatica/comite.pdf")
    def climate_board_pdf(session: Session = Depends(get_db), user: dict = Depends(require_user)):
        _require_climate_disclosure_view(user)
        organization_id = int(user["organization_id"]); org = session.get(Organization, organization_id)
        summary = board_summary(session, organization_id)
        content, digest = build_board_pdf(summary, org.name)
        briefing = summary["briefing"]
        if briefing:
            briefing.document_hash = digest
            session.commit()
        return Response(content=content, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="informe_comite_climatico_{organization_id}.pdf"', "X-Document-SHA256": digest})
