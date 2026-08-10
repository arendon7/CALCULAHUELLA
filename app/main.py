from __future__ import annotations

import hashlib
import hmac
import json
import math
import os
import re
import secrets
from datetime import UTC, date, datetime
from io import BytesIO
from pathlib import Path

from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload
from starlette.middleware.sessions import SessionMiddleware
from starlette.middleware.trustedhost import TrustedHostMiddleware

from .database import (
    ActivityData,
    ActivityIndicator,
    AppUser,
    AuditEvent,
    DataRequest,
    EmissionCalculation,
    EmissionFactor,
    EmissionFactorVersion,
    EmissionSource,
    EmissionTarget,
    ReductionScenario,
    ReductionScenarioAction,
    VerificationFinding,
    EvidenceDocument,
    Facility,
    Gas,
    GWPValue,
    Inventory,
    InventoryFacility,
    InventoryDecision,
    Notification,
    NotificationPreference,
    OrganizationMembership,
    ScheduledAutomation,
    AutomationRun,
    PlatformSetting,
    MethodologyRelease,
    InventoryMethodologySnapshot,
    ComplianceRequirement,
    ComplianceAssessment,
    DocumentControlRecord,
    CommercialReadinessItem,
    ServicePlan,
    OrganizationSubscription,
    UsageCounter,
    CustomerOnboardingItem,
    SupportTicket,
    SupportMessage,
    BillingInvoice,
    CommercialLead,
    CommercialProposal,
    PaymentTransaction,
    ServiceContract,
    ServiceOrder,
    CollectionAction,
    BillingDocumentRecord,
    CustomerSuccessProfile,
    AccountHealthSnapshot,
    ValueMilestone,
    SuccessCommitment,
    RenewalOpportunity,
    BenchmarkReference,
    ImpactSnapshot,
    ClimateRiskAssessment,
    ClimateRisk,
    ClimateRiskControl,
    ClimateTransitionRoadmap,
    ClimateTransitionAction,
    ClimateScenarioDefinition,
    ClimateDisclosureStatement,
    ClimateDisclosureRequirement,
    ClimateBoardBriefing,
    ClimateBoardDecision,
    ConsolidationFinding,
    ReleaseGate,
    JourneyValidation,
    Organization,
    ReviewObservation,
    ReductionAction,
    ReportArtifact,
    SectorTemplate,
    SectorTemplateSource,
    SourceFactorAssignment,
    Supplier,
    SupplierCampaign,
    Scope3CategoryAssessment,
    SupplierDataRequest,
    SupplierResponse,
    UnitConversion,
    UnitDefinition,
    INSTANCE_DIR,
    UPLOAD_DIR,
    SessionLocal,
    add_audit,
    get_db,
    refresh_progress,
    hash_password,
)
from .config import settings
from .accounting import is_gross_source
from .inventory_context import (
    ensure_inventory_editable,
    get_inventory,
    get_source_for_user,
    inventory_metrics,
)
from .inventory_lifecycle import clone_inventory_version, next_inventory_version
from .access_control import ROLE_CAPABILITIES, can_open_route
from .product_registry import PRODUCT_MODULES
from .product_experience import demo_story_for, journey_detail, navigation_for, normalize_view_mode, role_profile
from .onboarding_experience import onboarding_summary
from .guided_onboarding import load_profile as load_guided_profile, decision_plan as guided_decision_plan
from .consolidation import consolidation_summary, build_consolidation_workbook, summary_json
from .architecture import domain_architecture_summary
from .methodology_web import register_methodology_core_routes
from .factor_library_web import register_factor_library_routes
from .methodology_closure_web import register_methodology_closure_routes
from .land_removals_web import register_land_removals_routes
from .product_project_assurance_web import register_product_project_assurance_routes
from .colombia_library_web import register_colombia_library_routes
from .pilot_web import register_pilot_routes
from .pilot_execution_web import register_pilot_execution_routes
from .data_quality_web import register_data_quality_routes
from .period_close_web import register_period_close_routes
from .operational_imports_web import register_operational_import_routes
from .operations_web import register_operations_routes
from .integrations_web import register_integration_routes
from .users_web import register_user_routes
from .inventories_web import register_inventory_routes
from .reports_web import register_report_routes
from .reduction_web import register_reduction_routes
from .delivery_web import register_delivery_routes
from .delivery_readiness import professional_delivery_summary
from .organizations_web import register_organization_routes
from .information_web import register_information_routes
from .capture_web import register_capture_routes
from .review_web import register_review_routes
from .demo_web import register_demo_routes
from .product_intelligence_web import register_product_intelligence_routes
from .guided_onboarding_web import register_guided_onboarding_routes
from .service_operations_web import register_service_operations_routes
from .experience_web import register_experience_routes
from .workflow_web import register_workflow_routes
from .legal_web import register_legal_routes
from .public_web import register_public_routes
from .auth_web import register_auth_routes
from .user_context import resolve_current_user
from .period_close import assert_periods_editable
from .pilot_execution import guided_workspace
from .storage import storage, StorageError
from .notifications import create_notification, notify_roles, get_or_create_preference, process_pending_notifications
from .automations import AUTOMATION_TYPES, CADENCES, ROLE_OPTIONS, calculate_next_run, execute_automation, process_due_automations
from .security import (CSRFMiddleware, RequestBodyLimitMiddleware, RequestContextMiddleware, SecurityHeadersMiddleware, login_throttle, password_needs_upgrade, validate_upload_bytes, verify_password)
from .operations import diagnostic_snapshot
from .observability import OperationalMetricsMiddleware
from .calculations import convert_value, normalize_factor_output, recalculate_inventory, recalculate_source, source_calculation_summary
from .analytics import full_analysis, indicator_metrics, reduction_summary
from .reporting import create_report_artifact
from .scenarios import get_scenario, portfolio_macc, scenario_summary
from .reduction_portfolio import build_portfolio_workbook, portfolio_json, portfolio_summary
from .verification import create_verification_package
from .customer_success import account_metrics, refresh_account_health, sync_renewal_opportunity
from .support_workflow import (
    OPEN_STATUSES, CLOSED_STATUSES, add_support_message, ensure_reference, response_deadline,
    route_assignment, status_class, support_summary, ticket_context, ticket_overdue, ticket_waiting_days,
)
from .impact_intelligence import impact_metrics, refresh_impact_snapshot, compare_benchmarks, portfolio_comparison
from .climate_risk import assessment_summary, calculate_risk_scores, risk_level, synchronize_control_effectiveness, refresh_assessment_status
from .climate_disclosure import scenario_comparison, disclosure_summary, board_summary, build_board_pdf
from .supply_chain import (
    approved_duplicate_responses,
    calculate_supplier_response,
    campaign_summary,
    ensure_scope3_assessments,
    inventory_supply_chain_summary,
    quality_level as supplier_quality_level,
    sync_supplier_source,
    validate_supplier_response,
)
from .scope3_catalog import canonical_category_label, category_from_value

BASE_DIR = Path(__file__).resolve().parent
app = FastAPI(title=settings.app_name, version=settings.version, docs_url=None if settings.is_production else "/docs", redoc_url=None)
app.add_middleware(OperationalMetricsMiddleware)
app.add_middleware(SecurityHeadersMiddleware)
if settings.trusted_hosts:
    app.add_middleware(TrustedHostMiddleware, allowed_hosts=list(settings.trusted_hosts))
app.add_middleware(
    SessionMiddleware,
    secret_key=settings.session_secret,
    max_age=60 * 60 * 8,
    same_site="lax",
    https_only=settings.session_https_only,
)
app.add_middleware(CSRFMiddleware)
app.add_middleware(RequestContextMiddleware)
app.add_middleware(RequestBodyLimitMiddleware)
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
templates = Jinja2Templates(directory=BASE_DIR / "templates")

def format_number(value: float, decimals: int = 1) -> str:
    formatted = f"{value:,.{decimals}f}"
    return formatted.replace(",", "X").replace(".", ",").replace("X", ".")

templates.env.filters["number_es"] = format_number

def parse_date(value: str):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(400, "Fecha inválida") from exc

ALLOWED_UPLOAD_EXTENSIONS = {".pdf", ".xlsx", ".xls", ".csv", ".jpg", ".jpeg", ".png"}
MAX_UPLOAD_SIZE = settings.max_upload_mb * 1024 * 1024
ALLOWED_UNITS = ["kWh", "MWh", "L", "gal", "m³", "kg", "t", "km", "t·km", "pasajero·km", "COP", "unidad", "servicio anual", "tCO₂e"]
DATA_ORIGINS = ["Medición directa", "Factura", "Registro operativo", "Registro contable", "Información de proveedor", "Certificado", "Encuesta", "Estimación"]

def safe_filename(name: str) -> str:
    base = Path(name).name.strip().replace(" ", "_")
    base = re.sub(r"[^A-Za-z0-9._-]", "_", base)
    return base[:140] or "archivo"

def quality_from(origin: str, is_estimated: bool, has_evidence: bool) -> str:
    if origin == "Medición directa" and has_evidence and not is_estimated:
        return "A"
    if origin in {"Factura", "Registro operativo", "Registro contable", "Información de proveedor", "Certificado"} and not is_estimated:
        return "B"
    if origin in {"Encuesta", "Estimación"} or is_estimated:
        return "C"
    return "D"

def format_bytes(value: int) -> str:
    if value < 1024:
        return f"{value} B"
    if value < 1024 * 1024:
        return f"{value / 1024:.1f} KB"
    return f"{value / (1024 * 1024):.1f} MB"

templates.env.filters["bytes"] = format_bytes

def current_user(request: Request) -> dict[str, object] | None:
    return resolve_current_user(request)

def require_user(request: Request) -> dict[str, object]:
    user = current_user(request)
    if not user:
        raise HTTPException(status_code=401)
    return user

def ensure_capability(user: dict[str, object], capability: str) -> None:
    if capability not in user["capabilities"]:
        raise HTTPException(403, "Tu rol no tiene permiso para esta acción")

def set_flash(request: Request, message: str, level: str = "success") -> None:
    request.session["flash"] = {"message": message, "level": level}

def common_context(request: Request, session: Session, user: dict[str, object], active: str, **extra):
    org = session.get(Organization, int(user["organization_id"]))
    flash = request.session.pop("flash", None)
    unread_notifications = session.scalar(select(func.count(Notification.id)).where(
        Notification.organization_id == int(user["organization_id"]),
        Notification.user_id == int(user["id"]),
        Notification.read_at.is_(None),
    )) or 0
    view_mode = normalize_view_mode(user.get("view_mode"))
    return {
        "user": user,
        "org": org,
        "active": active,
        "flash": flash,
        "unread_notifications": unread_notifications,
        "navigation": navigation_for(user, view_mode),
        "role_profile": role_profile(str(user.get("role", "Cliente"))),
        "can_open_route": can_open_route,
        **extra,
    }

def review_gate_summary(session: Session, inventory: Inventory) -> dict[str, object]:
    included_sources = [source for source in inventory.sources if source.included]
    incomplete = [source for source in included_sources if source.progress < 100]
    missing_factors = [source for source in included_sources if source.category != "Datos específicos de proveedores" and not any(item.active and item.factor_version.status == "Aprobado" for item in source.factor_assignments)]
    error_count = session.scalar(
        select(func.count()).select_from(EmissionCalculation).join(ActivityData).join(EmissionSource).where(
            EmissionSource.inventory_id == inventory.id, EmissionCalculation.status == "Error"
        )
    ) or 0
    open_observations = [item for item in inventory.observations if item.status != "Cerrada"]
    blocking_observations = [item for item in open_observations if item.severity in {"Mayor", "Crítica"}]
    config_ok = bool(inventory.methodology and inventory.gwp_version and inventory.facility_links)
    sources_ok = not incomplete
    factors_ok = not missing_factors and error_count == 0
    review_ok = not blocking_observations
    gates = [
        {"name": "Configuración metodológica", "status": "Aprobado" if config_ok else "Pendiente", "detail": "Metodología, GWP, periodo y límites definidos." if config_ok else "Completa metodología, GWP y sedes incluidas."},
        {"name": "Cobertura de fuentes", "status": "Aprobado" if sources_ok else "En progreso", "detail": "Todas las fuentes incluidas tienen cobertura completa." if sources_ok else f"{len(incomplete)} fuente(s) todavía no tienen cobertura completa."},
        {"name": "Factores y cálculos", "status": "Aprobado" if factors_ok else "En progreso", "detail": "Todas las fuentes tienen factores aprobados y cálculos sin error." if factors_ok else f"{len(missing_factors)} fuente(s) sin factor y {error_count} error(es) de cálculo."},
        {"name": "Revisión profesional", "status": "Aprobado" if review_ok else "En progreso", "detail": "No existen observaciones mayores o críticas abiertas." if review_ok else f"{len(blocking_observations)} observación(es) bloqueante(s) abierta(s)."},
    ]
    blockers = []
    if not config_ok:
        blockers.append("Configuración metodológica incompleta")
    if incomplete:
        blockers.append(f"{len(incomplete)} fuentes incompletas")
    if missing_factors:
        blockers.append(f"{len(missing_factors)} fuentes sin factor aprobado")
    if error_count:
        blockers.append(f"{error_count} errores de cálculo")
    if blocking_observations:
        blockers.append(f"{len(blocking_observations)} observaciones mayores o críticas")
    return {
        "gates": gates,
        "blockers": blockers,
        "can_approve": not blockers,
        "open_observations": open_observations,
        "blocking_observations": blocking_observations,
        "incomplete_sources": incomplete,
        "error_count": error_count,
    }

@app.exception_handler(401)
async def unauthorized_handler(request: Request, exc: HTTPException):
    if request.url.path.startswith("/api/"):
        return JSONResponse({"detail": str(exc.detail)}, status_code=401)
    return RedirectResponse("/login", status_code=303)

@app.exception_handler(403)
async def forbidden_handler(request: Request, exc: HTTPException):
    user = current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    with SessionLocal() as session:
        return templates.TemplateResponse(
            request=request,
            name="error.html",
            context=common_context(request, session, user, "", title="Acceso restringido", message=str(exc.detail)),
            status_code=403,
        )

@app.post("/preferencias/vista")
def update_view_mode(
    request: Request,
    mode: str = Form(...),
    return_url: str = Form("/dashboard"),
    user: dict = Depends(require_user),
):
    request.session["view_mode"] = normalize_view_mode(mode)
    destination = return_url if return_url.startswith("/") and not return_url.startswith("//") else "/dashboard"
    set_flash(
        request,
        "Vista esencial activada: se prioriza el flujo del inventario."
        if request.session["view_mode"] == "essential"
        else "Vista completa activada: se muestran capacidades avanzadas e internas.",
    )
    return RedirectResponse(destination, status_code=303)

@app.get("/recorrido-inventario", response_class=HTMLResponse)
def inventory_journey_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    inventory = get_inventory(session, user)
    workspace = guided_workspace(session, user, inventory)
    journey = journey_detail(workspace, str(user["role"]))
    session.commit()
    return templates.TemplateResponse(
        request=request,
        name="inventory_journey.html",
        context=common_context(
            request,
            session,
            user,
            "journey",
            inventory=inventory,
            journey=journey,
        ),
    )

@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    inventory = get_inventory(session, user)
    metrics = inventory_metrics(inventory)
    inventories = list(session.scalars(select(Inventory).where(Inventory.organization_id == int(user["organization_id"])).order_by(Inventory.start_date.desc())))
    tasks = list(session.scalars(
        select(DataRequest)
        .where(
            DataRequest.inventory_id == inventory.id,
            DataRequest.status.notin_(["Completada", "Cerrada"]),
        )
        .order_by(DataRequest.due_date)
    ))
    workspace = guided_workspace(session, user, inventory)
    delivery = professional_delivery_summary(session, inventory)
    dashboard_action = delivery["next_action"]
    if str(user["role"]) == "Cliente":
        if tasks:
            dashboard_action = {
                "name": "Atender solicitudes de información",
                "detail": f"Tienes {len(tasks)} requerimiento(s) activo(s). Completa los datos o soportes solicitados antes de la revisión técnica.",
                "owner": "Responsable de información",
                "acceptance": "Solicitudes respondidas y evidencias vinculadas al periodo correcto.",
                "href": "/informacion#solicitudes",
                "action": "Abrir pendientes",
            }
        else:
            dashboard_action = {
                "name": "Completar datos y evidencias",
                "detail": "Revisa los periodos pendientes y conserva un soporte verificable para cada valor relevante.",
                "owner": "Responsable de información",
                "acceptance": "Fuentes del periodo completas y soportes vinculados.",
                "href": "/captura-guiada",
                "action": "Continuar captura",
            }
    onboarding_rows = list(session.scalars(select(CustomerOnboardingItem).where(
        CustomerOnboardingItem.organization_id == int(user["organization_id"])
    ).order_by(CustomerOnboardingItem.display_order)))
    onboarding_state = onboarding_summary(onboarding_rows, inventory_id=inventory.id)
    guided_profile = load_guided_profile(session, inventory.organization)
    guided_setup = guided_decision_plan(guided_profile, inventory.organization, inventory=inventory)
    session.commit()
    return templates.TemplateResponse(
        request=request,
        name="dashboard.html",
        context=common_context(
            request, session, user, "dashboard", inventory=inventory, inventories=inventories,
            tasks=tasks, sources=inventory.sources, workspace=workspace, delivery=delivery,
            dashboard_action=dashboard_action,
            journey=journey_detail(workspace, str(user["role"])), onboarding=onboarding_state,
            guided_setup=guided_setup, demo_story=demo_story_for(inventory.organization.trade_name), **metrics,
        ),
    )

@app.get("/sectorizacion", response_class=HTMLResponse)
def sectorization_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    inventory = get_inventory(session, user)
    templates_list = list(session.scalars(select(SectorTemplate).where(SectorTemplate.active.is_(True)).options(selectinload(SectorTemplate.source_items)).order_by(SectorTemplate.sector)))
    selected_template = next((item for item in templates_list if item.sector == inventory.organization.sector), None)
    return templates.TemplateResponse(request=request, name="sectorization.html", context=common_context(request, session, user, "sectorization", inventory=inventory, templates_list=templates_list, selected_template=selected_template, facilities=[link.facility for link in inventory.facility_links if link.included]))

@app.post("/sectorizacion/aplicar")
def apply_sector_template(request: Request, inventory_id: int = Form(...), template_id: int = Form(...), facility_id: int | None = Form(None), include_optional: bool = Form(False), session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_sources")
    inventory = get_inventory(session, user, inventory_id)
    ensure_inventory_editable(inventory)
    template = session.scalar(select(SectorTemplate).where(SectorTemplate.id == template_id, SectorTemplate.active.is_(True)).options(selectinload(SectorTemplate.source_items)))
    if not template:
        raise HTTPException(404, "Plantilla sectorial no encontrada")
    allowed_facilities = {link.facility_id for link in inventory.facility_links if link.included}
    selected_facility_id = facility_id if facility_id in allowed_facilities else (next(iter(allowed_facilities)) if allowed_facilities else None)
    existing_keys = {(source.name.strip().lower(), source.category.strip().lower()) for source in inventory.sources}
    created = 0
    assigned = 0
    for item in template.source_items:
        if not item.recommended and not include_optional:
            continue
        key = (item.name.strip().lower(), item.category.strip().lower())
        if key in existing_keys:
            continue
        source = EmissionSource(inventory_id=inventory.id, facility_id=selected_facility_id, name=item.name, scope=item.scope, category=item.category, responsible=item.responsible, materiality=item.materiality, data_frequency=item.data_frequency, preferred_unit=item.preferred_unit, icon=item.icon, included=True, status="Pendiente", progress=0)
        session.add(source)
        session.flush()
        existing_keys.add(key)
        created += 1
        if item.factor_activity_type:
            factor_versions = list(session.scalars(select(EmissionFactorVersion).join(EmissionFactor).where(EmissionFactor.activity_type == item.factor_activity_type, EmissionFactorVersion.status == "Aprobado")))
            for version in factor_versions:
                session.add(SourceFactorAssignment(source_id=source.id, factor_version_id=version.id, active=True, assigned_by=str(user["email"]), notes=f"Asignación automática desde {template.name}"))
                assigned += 1
    refresh_progress(session, inventory)
    add_audit(session, int(user["organization_id"]), str(user["email"]), "APLICAR", "Plantilla sectorial", template.name, f"{created} fuentes creadas; {assigned} factores asignados")
    session.commit()
    set_flash(request, f"Plantilla aplicada: {created} fuentes nuevas y {assigned} asignaciones de factor.")
    return RedirectResponse("/sectorizacion", status_code=303)

def _parse_excel_period(value: object, inventory: Inventory) -> tuple[date, date]:
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    else:
        text = str(value or "").strip()
        for pattern in ("%Y-%m", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                parsed = datetime.strptime(text, pattern).date()
                break
            except ValueError:
                parsed = None
        if parsed is None:
            raise ValueError("periodo inválido")
    start = date(parsed.year, parsed.month, 1)
    if parsed.month == 12:
        next_month = date(parsed.year + 1, 1, 1)
    else:
        next_month = date(parsed.year, parsed.month + 1, 1)
    end = date.fromordinal(next_month.toordinal() - 1)
    if start < inventory.start_date or end > inventory.end_date:
        raise ValueError("periodo fuera del inventario")
    return start, end

@app.get("/calculos", response_class=HTMLResponse)
def calculations_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    inventory = get_inventory(session, user)
    source_rows = []
    total_calculations = 0
    total_alerts = 0
    total_errors = 0
    for source in inventory.sources:
        summary = source_calculation_summary(session, source.id)
        source_rows.append({"source": source, "summary": summary, "assignments": len([item for item in source.factor_assignments if item.active])})
        total_calculations += len(summary["calculations"])
        total_alerts += int(summary["alerts"])
        total_errors += int(summary["errors"])
    return templates.TemplateResponse(
        request=request,
        name="calculations.html",
        context=common_context(
            request,
            session,
            user,
            "calculations",
            inventory=inventory,
            source_rows=source_rows,
            total_calculations=total_calculations,
            total_alerts=total_alerts,
            total_errors=total_errors,
        ),
    )

@app.post("/inventarios/{inventory_id}/recalcular")
def inventory_recalculate(inventory_id: int, request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "view_methodology")
    inventory = get_inventory(session, user, inventory_id)
    ensure_inventory_editable(inventory)
    result = recalculate_inventory(session, inventory)
    add_audit(session, int(user["organization_id"]), str(user["email"]), "RECALCULAR", "Inventario", inventory.name, f"{result['sources']} fuentes · {result['calculations']} cálculos · {len(result['warnings'])} alertas")
    session.commit()
    set_flash(request, f"Inventario recalculado: {result['calculations']} componentes y {len(result['warnings'])} alertas.", "error" if result["warnings"] else "success")
    return RedirectResponse("/calculos", status_code=303)

@app.get("/metodologia", response_class=HTMLResponse)
def methodology_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "view_methodology")
    factors = list(
        session.scalars(
            select(EmissionFactor)
            .options(selectinload(EmissionFactor.versions).selectinload(EmissionFactorVersion.gas))
            .order_by(EmissionFactor.activity_type, EmissionFactor.name)
        )
    )
    units = list(session.scalars(select(UnitDefinition).order_by(UnitDefinition.dimension, UnitDefinition.code)))
    conversions = list(session.scalars(select(UnitConversion).where(UnitConversion.active.is_(True)).order_by(UnitConversion.from_unit, UnitConversion.to_unit)))
    gases = list(session.scalars(select(Gas).options(selectinload(Gas.gwp_values)).order_by(Gas.code)))
    return templates.TemplateResponse(
        request=request,
        name="methodology.html",
        context=common_context(
            request,
            session,
            user,
            "methodology",
            factors=factors,
            units=units,
            conversions=conversions,
            gases=gases,
            allowed_units=ALLOWED_UNITS,
        ),
    )

@app.post("/metodologia/factores/nuevo")
def factor_create(
    request: Request,
    name: str = Form(...),
    activity_type: str = Form(...),
    gas_id: int = Form(...),
    value: float = Form(...),
    input_unit: str = Form(...),
    output_unit: str = Form("kg gas"),
    version: str = Form("1.0"),
    source_organization: str = Form(...),
    source_document: str = Form(""),
    publication_year: int = Form(...),
    geographic_scope: str = Form("Colombia"),
    technology_scope: str = Form("Genérico"),
    uncertainty_percentage: float = Form(0),
    notes: str = Form(""),
    session: Session = Depends(get_db),
    user: dict = Depends(require_user),
):
    ensure_capability(user, "view_methodology")
    if input_unit not in ALLOWED_UNITS:
        raise HTTPException(400, "Unidad no autorizada")
    gas = session.get(Gas, gas_id)
    if not gas or not math.isfinite(value) or value < 0:
        raise HTTPException(400, "Gas o valor inválido")
    if not math.isfinite(uncertainty_percentage) or uncertainty_percentage < 0:
        raise HTTPException(400, "La incertidumbre debe ser un número finito mayor o igual a cero")
    normalized_output, output_error = normalize_factor_output(1.0, output_unit, gas.code)
    if normalized_output is None:
        raise HTTPException(400, output_error)
    factor = session.scalar(select(EmissionFactor).where(EmissionFactor.name == name.strip()))
    if not factor:
        factor = EmissionFactor(name=name.strip(), activity_type=activity_type.strip(), country="Colombia", sector="Multisectorial", status="Activo", is_demo=False)
        session.add(factor)
        session.flush()
    duplicate = session.scalar(select(EmissionFactorVersion).where(EmissionFactorVersion.factor_id == factor.id, EmissionFactorVersion.version == version.strip(), EmissionFactorVersion.gas_id == gas.id))
    if duplicate:
        set_flash(request, "Ya existe esa versión para el gas seleccionado.", "error")
        return RedirectResponse("/metodologia", status_code=303)
    factor_version = EmissionFactorVersion(
        factor_id=factor.id,
        gas_id=gas.id,
        version=version.strip(),
        value=value,
        input_unit=input_unit,
        output_unit=output_unit.strip(),
        source_organization=source_organization.strip(),
        source_document=source_document.strip(),
        publication_year=publication_year,
        geographic_scope=geographic_scope.strip(),
        technology_scope=technology_scope.strip(),
        uncertainty_percentage=max(0, uncertainty_percentage),
        status="Pendiente de revisión",
        notes=notes.strip(),
    )
    session.add(factor_version)
    add_audit(session, int(user["organization_id"]), str(user["email"]), "CREAR", "Factor", factor.name, f"Versión {factor_version.version} · {gas.code} · pendiente")
    session.commit()
    set_flash(request, "El factor fue creado como pendiente de revisión.")
    return RedirectResponse("/metodologia", status_code=303)

@app.post("/metodologia/factores/{version_id}/estado")
def factor_status_update(
    version_id: int,
    request: Request,
    status: str = Form(...),
    session: Session = Depends(get_db),
    user: dict = Depends(require_user),
):
    if not user["can_review"]:
        raise HTTPException(403, "Solo un revisor puede aprobar factores")
    factor_version = session.scalar(select(EmissionFactorVersion).where(EmissionFactorVersion.id == version_id).options(selectinload(EmissionFactorVersion.factor)))
    if not factor_version:
        raise HTTPException(404, "Factor no encontrado")
    if status not in {"Pendiente de revisión", "Aprobado", "Retirado"}:
        raise HTTPException(400, "Estado inválido")
    if status == "Aprobado":
        normalized_output, output_error = normalize_factor_output(
            1.0,
            factor_version.output_unit,
            factor_version.gas.code,
        )
        if normalized_output is None:
            raise HTTPException(409, f"El factor no puede aprobarse: {output_error}")
        if not math.isfinite(factor_version.value) or factor_version.value < 0:
            raise HTTPException(409, "El factor no puede aprobarse porque su valor no es válido")
    factor_version.status = status
    factor_version.approved_by = str(user["name"]) if status == "Aprobado" else factor_version.approved_by
    factor_version.approved_at = datetime.now(UTC) if status == "Aprobado" else factor_version.approved_at
    add_audit(session, int(user["organization_id"]), str(user["email"]), "REVISAR", "Factor", factor_version.factor.name, f"Estado {status}")
    session.commit()
    set_flash(request, "El estado del factor fue actualizado.")
    return RedirectResponse("/metodologia", status_code=303)

@app.post("/metodologia/conversiones/nueva")
def conversion_create(
    request: Request,
    from_unit: str = Form(...),
    to_unit: str = Form(...),
    multiplier: float = Form(...),
    source: str = Form("Conversión interna aprobada"),
    session: Session = Depends(get_db),
    user: dict = Depends(require_user),
):
    ensure_capability(user, "view_methodology")
    if from_unit == to_unit or not math.isfinite(multiplier) or multiplier <= 0:
        raise HTTPException(400, "Conversión inválida")
    source_definition = session.scalar(select(UnitDefinition).where(UnitDefinition.code == from_unit))
    target_definition = session.scalar(select(UnitDefinition).where(UnitDefinition.code == to_unit))
    if not source_definition or not target_definition or source_definition.dimension != target_definition.dimension:
        raise HTTPException(400, "Las unidades no existen o tienen dimensiones incompatibles")
    conversion = session.scalar(select(UnitConversion).where(UnitConversion.from_unit == from_unit, UnitConversion.to_unit == to_unit))
    if conversion:
        conversion.multiplier = multiplier
        conversion.source = source.strip()
        conversion.active = True
    else:
        session.add(UnitConversion(from_unit=from_unit, to_unit=to_unit, multiplier=multiplier, source=source.strip(), active=True))
    add_audit(session, int(user["organization_id"]), str(user["email"]), "CONFIGURAR", "Conversión", f"{from_unit} → {to_unit}", f"Multiplicador {multiplier:g}")
    session.commit()
    set_flash(request, "La conversión fue guardada.")
    return RedirectResponse("/metodologia", status_code=303)

@app.get("/analisis", response_class=HTMLResponse)
def analysis_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    inventory = get_inventory(session, user)
    analysis = full_analysis(session, inventory)
    return templates.TemplateResponse(
        request=request,
        name="analysis.html",
        context=common_context(
            request, session, user, "analysis", inventory=inventory,
            indicator_types=[("Producción", "t"), ("Empleados", "personas"), ("Ingresos", "COP"), ("Servicios", "servicios"), ("Área", "m²")],
            **analysis,
        ),
    )

@app.post("/analisis/indicadores/nuevo")
def create_indicator(
    request: Request,
    inventory_id: int = Form(...),
    indicator_type: str = Form(...),
    value: float = Form(...),
    unit: str = Form(...),
    period_start: str = Form(...),
    period_end: str = Form(...),
    source_name: str = Form("Registro operativo"),
    facility_id: int | None = Form(None),
    notes: str = Form(""),
    session: Session = Depends(get_db),
    user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_inventory")
    inventory = get_inventory(session, user, inventory_id)
    ensure_inventory_editable(inventory)
    if value < 0:
        raise HTTPException(400, "El valor del indicador no puede ser negativo")
    indicator = ActivityIndicator(
        inventory_id=inventory.id, facility_id=facility_id or None,
        period_start=parse_date(period_start), period_end=parse_date(period_end),
        indicator_type=indicator_type.strip(), value=value, unit=unit.strip(),
        source_name=source_name.strip() or "Registro operativo", notes=notes.strip(),
        status="Cargado", created_by=str(user["email"]),
    )
    session.add(indicator)
    add_audit(session, int(user["organization_id"]), str(user["email"]), "CREAR", "Indicador", indicator.indicator_type, f"{value} {unit}")
    session.commit()
    set_flash(request, "Indicador operativo registrado.")
    return RedirectResponse("/analisis", status_code=303)

@app.post("/analisis/indicadores/{indicator_id}/editar")
def update_indicator(
    indicator_id: int, request: Request, value: float = Form(...), unit: str = Form(...),
    source_name: str = Form("Registro operativo"), notes: str = Form(""),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_inventory")
    indicator = session.scalar(
        select(ActivityIndicator).join(Inventory).where(
            ActivityIndicator.id == indicator_id, Inventory.organization_id == int(user["organization_id"])
        )
    )
    if not indicator:
        raise HTTPException(404, "Indicador no encontrado")
    inventory = get_inventory(session, user, indicator.inventory_id)
    ensure_inventory_editable(inventory)
    previous = f"{indicator.value} {indicator.unit}"
    indicator.value = value
    indicator.unit = unit.strip()
    indicator.source_name = source_name.strip() or indicator.source_name
    indicator.notes = notes.strip()
    add_audit(session, int(user["organization_id"]), str(user["email"]), "EDITAR", "Indicador", indicator.indicator_type, previous_value=previous, new_value=f"{value} {unit}")
    session.commit()
    set_flash(request, "Indicador actualizado.")
    return RedirectResponse("/analisis", status_code=303)

@app.get("/cadena-valor", response_class=HTMLResponse)
def supply_chain_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    if not (user["can_manage_supply_chain"] or user["can_review"] or user["can_approve"]):
        raise HTTPException(403, "Tu rol no tiene acceso a la cadena de valor")
    inventory = get_inventory(session, user)
    suppliers = list(session.scalars(
        select(Supplier).where(Supplier.organization_id == int(user["organization_id"])).order_by(Supplier.strategic.desc(), Supplier.name)
    ))
    campaigns = list(session.scalars(
        select(SupplierCampaign)
        .where(SupplierCampaign.inventory_id == inventory.id)
        .options(
            selectinload(SupplierCampaign.requests).selectinload(SupplierDataRequest.supplier),
            selectinload(SupplierCampaign.requests).selectinload(SupplierDataRequest.response),
        )
        .order_by(SupplierCampaign.created_at.desc())
    ))
    selected_id = request.query_params.get("campaign_id")
    selected = next((item for item in campaigns if str(item.id) == selected_id), campaigns[0] if campaigns else None)
    selected_summary = campaign_summary(session, selected) if selected else None
    summary = inventory_supply_chain_summary(session, inventory)
    supplier_source = sync_supplier_source(session, inventory.id)
    session.commit()
    return templates.TemplateResponse(
        request=request,
        name="supply_chain.html",
        context=common_context(
            request, session, user, "supply_chain", inventory=inventory, suppliers=suppliers,
            campaigns=campaigns, selected=selected, selected_summary=selected_summary,
            supply_summary=summary, supplier_source=supplier_source,
            portal_base=str(request.base_url).rstrip("/"),
        ),
    )

@app.get("/api/cadena-valor/resumen")
def supply_chain_summary_api(session: Session = Depends(get_db), user: dict = Depends(require_user)):
    if not (user["can_manage_supply_chain"] or user["can_review"] or user["can_approve"]):
        raise HTTPException(403, "Tu rol no tiene acceso a la cadena de valor")
    inventory = get_inventory(session, user)
    summary = inventory_supply_chain_summary(session, inventory)
    session.commit()
    return {
        "inventory_id": inventory.id,
        "campaign_count": summary["campaign_count"],
        "request_count": summary["request_count"],
        "response_count": summary["response_count"],
        "approved_count": summary["approved_count"],
        "emissions_tco2e": summary["emissions"],
        "screening_coverage": summary["screening_coverage"],
        "assessed_category_count": summary["assessed_category_count"],
        "material_category_count": summary["material_category_count"],
        "active_category_count": summary["active_category_count"],
        "approved_category_count": summary["approved_category_count"],
        "quality_score": summary["quality_score"],
        "duplicate_count": summary["duplicate_count"],
        "direction_emissions": summary["direction_emissions"],
        "categories": summary["categories"],
        "warnings": summary["warnings"],
    }

@app.post("/cadena-valor/categorias/{category_code}/evaluar")
def assess_scope3_category(
    category_code: str,
    request: Request,
    status: str = Form(...),
    relevance_score: float = Form(0),
    rationale: str = Form(""),
    owner: str = Form("Responsable ambiental"),
    data_strategy: str = Form("Por definir"),
    session: Session = Depends(get_db),
    user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_supply_chain")
    inventory = get_inventory(session, user)
    ensure_inventory_editable(inventory)
    category = category_from_value(category_code)
    if not category:
        raise HTTPException(404, "Categoría de Alcance 3 no encontrada")
    allowed_statuses = {"Pendiente", "Material", "No material", "No aplica"}
    if status not in allowed_statuses:
        raise HTTPException(400, "Estado de evaluación no permitido")
    clean_rationale = rationale.strip()
    if status in {"Material", "No material", "No aplica"} and not clean_rationale:
        raise HTTPException(400, "La conclusión de materialidad debe incluir una justificación.")
    ensure_scope3_assessments(session, inventory.id)
    assessment = session.scalar(
        select(Scope3CategoryAssessment).where(
            Scope3CategoryAssessment.inventory_id == inventory.id,
            Scope3CategoryAssessment.category_code == category.code,
        )
    )
    if not assessment:
        raise HTTPException(404, "Evaluación de categoría no disponible")
    assessment.status = status
    assessment.relevance_score = min(5.0, max(0.0, relevance_score))
    assessment.rationale = clean_rationale
    assessment.owner = owner.strip() or "Responsable ambiental"
    assessment.data_strategy = data_strategy.strip() or "Por definir"
    assessment.updated_by = str(user["email"])
    assessment.updated_at = datetime.now(UTC)
    add_audit(
        session,
        int(user["organization_id"]),
        str(user["email"]),
        "EVALUAR",
        "Categoría Scope 3",
        f"{category.code} · {category.name}",
        f"{status}; relevancia {assessment.relevance_score:.1f}/5",
    )
    session.commit()
    set_flash(request, f"{category.code} actualizada como {status.lower()}.")
    return RedirectResponse("/cadena-valor", status_code=303)


@app.post("/cadena-valor/proveedores/nuevo")
def create_supplier(
    request: Request, name: str = Form(...), tax_id: str = Form(""), sector: str = Form(""), country: str = Form("Colombia"),
    contact_name: str = Form(""), contact_email: str = Form(""), annual_spend_cop: float = Form(0),
    strategic: str | None = Form(None), risk_level: str = Form("Medio"),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_supply_chain")
    supplier = Supplier(
        organization_id=int(user["organization_id"]), name=name.strip(), tax_id=tax_id.strip(), sector=sector.strip(),
        country=country.strip(), contact_name=contact_name.strip(), contact_email=contact_email.strip().lower(),
        annual_spend_cop=max(0, annual_spend_cop), strategic=strategic is not None, risk_level=risk_level,
    )
    session.add(supplier)
    add_audit(session, int(user["organization_id"]), str(user["email"]), "CREAR", "Proveedor", supplier.name, f"Sector {supplier.sector}")
    session.commit()
    set_flash(request, "Proveedor registrado en la cadena de valor.")
    return RedirectResponse("/cadena-valor", status_code=303)

@app.post("/cadena-valor/campanas/nueva")
def create_supplier_campaign(
    request: Request, inventory_id: int = Form(...), name: str = Form(...), category: str = Form("Bienes y servicios adquiridos"),
    due_date: str = Form(...), methodology: str = Form("GHG Protocol Scope 3"), description: str = Form(""),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_supply_chain")
    inventory = get_inventory(session, user, inventory_id)
    ensure_inventory_editable(inventory)
    canonical_category = canonical_category_label(category)
    campaign = SupplierCampaign(
        inventory_id=inventory.id, name=name.strip(), category=canonical_category, due_date=parse_date(due_date),
        status="Borrador", methodology=methodology.strip(), description=description.strip(), created_by=str(user["email"]),
    )
    session.add(campaign)
    ensure_scope3_assessments(session, inventory.id)
    category_profile = category_from_value(campaign.category)
    if category_profile:
        assessment = session.scalar(select(Scope3CategoryAssessment).where(
            Scope3CategoryAssessment.inventory_id == inventory.id,
            Scope3CategoryAssessment.category_code == category_profile.code,
        ))
        if assessment:
            assessment.status = "Material"
            assessment.relevance_score = max(assessment.relevance_score, 4)
            assessment.rationale = assessment.rationale or "Categoría priorizada mediante campaña de levantamiento."
            assessment.data_strategy = "Campaña de proveedores"
            assessment.updated_by = str(user["email"])
            assessment.updated_at = datetime.now(UTC)
    add_audit(session, int(user["organization_id"]), str(user["email"]), "CREAR", "Campaña de proveedores", campaign.name, campaign.category)
    session.commit()
    set_flash(request, "Campaña creada. Ahora agrega solicitudes a proveedores.")
    return RedirectResponse(f"/cadena-valor?campaign_id={campaign.id}", status_code=303)

@app.post("/cadena-valor/solicitudes/nueva")
def create_supplier_request(
    request: Request, campaign_id: int = Form(...), supplier_id: int = Form(...), product_service: str = Form(...),
    quantity: float = Form(0), unit: str = Form("unidad"), spend_cop: float = Form(0),
    requested_method: str = Form("Factor específico del proveedor"), due_date: str = Form(...), notes: str = Form(""),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_supply_chain")
    campaign = session.scalar(
        select(SupplierCampaign).join(Inventory).where(
            SupplierCampaign.id == campaign_id, Inventory.organization_id == int(user["organization_id"])
        )
    )
    supplier = session.scalar(select(Supplier).where(Supplier.id == supplier_id, Supplier.organization_id == int(user["organization_id"])))
    if not campaign or not supplier:
        raise HTTPException(404, "Campaña o proveedor no encontrado")
    ensure_inventory_editable(campaign.inventory)
    token = secrets.token_urlsafe(24)
    data_request = SupplierDataRequest(
        campaign_id=campaign.id, supplier_id=supplier.id, product_service=product_service.strip(),
        quantity=max(0, quantity), unit=unit.strip(), spend_cop=max(0, spend_cop), requested_method=requested_method,
        status="Enviada", due_date=parse_date(due_date), access_token=token,
        token_expires_at=datetime.combine(parse_date(due_date), datetime.max.time(), tzinfo=UTC), sent_at=datetime.now(UTC), notes=notes.strip(),
    )
    campaign.status = "En curso"
    session.add(data_request)
    add_audit(session, int(user["organization_id"]), str(user["email"]), "SOLICITAR", "Dato de proveedor", supplier.name, product_service)
    session.commit()
    set_flash(request, "Solicitud creada. El enlace seguro está disponible en la tabla de la campaña.")
    return RedirectResponse(f"/cadena-valor?campaign_id={campaign.id}", status_code=303)

@app.post("/cadena-valor/solicitudes/{request_id}/renovar")
def renew_supplier_token(
    request_id: int, request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_supply_chain")
    data_request = session.scalar(
        select(SupplierDataRequest).join(SupplierCampaign).join(Inventory).where(
            SupplierDataRequest.id == request_id, Inventory.organization_id == int(user["organization_id"])
        )
    )
    if not data_request:
        raise HTTPException(404, "Solicitud no encontrada")
    data_request.access_token = secrets.token_urlsafe(24)
    data_request.token_expires_at = datetime.combine(data_request.due_date, datetime.max.time(), tzinfo=UTC)
    data_request.sent_at = datetime.now(UTC)
    data_request.status = "Enviada"
    session.commit()
    set_flash(request, "Enlace seguro renovado.")
    return RedirectResponse(f"/cadena-valor?campaign_id={data_request.campaign_id}", status_code=303)

@app.get("/proveedor/responder/{token}", response_class=HTMLResponse)
def supplier_public_portal(token: str, request: Request, session: Session = Depends(get_db)):
    data_request = session.scalar(
        select(SupplierDataRequest)
        .where(SupplierDataRequest.access_token == token)
        .options(
            selectinload(SupplierDataRequest.supplier),
            selectinload(SupplierDataRequest.campaign).selectinload(SupplierCampaign.inventory).selectinload(Inventory.organization),
            selectinload(SupplierDataRequest.response),
        )
    )
    if not data_request:
        raise HTTPException(404, "Enlace de proveedor no encontrado")
    expires = data_request.token_expires_at
    if expires and expires.tzinfo is None:
        expires = expires.replace(tzinfo=UTC)
    expired = bool(expires and expires < datetime.now(UTC))
    return templates.TemplateResponse(
        request=request, name="supplier_portal.html",
        context={"data_request": data_request, "expired": expired, "flash": request.session.pop("supplier_flash", None)},
    )

@app.post("/proveedor/responder/{token}")
async def submit_supplier_response(
    token: str, request: Request, method: str = Form(...), activity_value: float = Form(0), activity_unit: str = Form(""),
    emission_factor: float = Form(0), factor_unit: str = Form("kg CO2e/unidad"), reported_emissions_tco2e: float = Form(0),
    methodology: str = Form(""), boundary: str = Form(""), verified: str | None = Form(None), notes: str = Form(""),
    evidence: UploadFile | None = File(None), session: Session = Depends(get_db),
):
    data_request = session.scalar(
        select(SupplierDataRequest)
        .where(SupplierDataRequest.access_token == token)
        .options(selectinload(SupplierDataRequest.supplier), selectinload(SupplierDataRequest.campaign), selectinload(SupplierDataRequest.response))
    )
    if not data_request:
        raise HTTPException(404, "Enlace de proveedor no encontrado")
    expires = data_request.token_expires_at
    if expires and expires.tzinfo is None:
        expires = expires.replace(tzinfo=UTC)
    if expires and expires < datetime.now(UTC):
        raise HTTPException(410, "El enlace de respuesta venció")
    existing_evidence = bool(data_request.response and data_request.response.evidence_stored_name)
    validation = validate_supplier_response(
        data_request,
        method=method,
        activity_value=max(0, activity_value),
        activity_unit=activity_unit.strip() or data_request.unit,
        emission_factor=max(0, emission_factor),
        factor_unit=factor_unit.strip(),
        reported_emissions_tco2e=max(0, reported_emissions_tco2e),
        methodology=methodology.strip(),
        boundary=boundary.strip(),
        has_evidence=existing_evidence or bool(evidence and evidence.filename),
    )
    if validation["errors"]:
        raise HTTPException(400, " ".join(validation["errors"]))
    file_name = ""
    stored_name = ""
    sha256 = ""
    file_size = 0
    if evidence and evidence.filename:
        extension = Path(evidence.filename).suffix.lower()
        if extension not in ALLOWED_UPLOAD_EXTENSIONS:
            raise HTTPException(400, "Formato de evidencia no permitido")
        content = await evidence.read(MAX_UPLOAD_SIZE + 1)
        if len(content) > MAX_UPLOAD_SIZE:
            raise HTTPException(400, f"El archivo supera el límite de {settings.max_upload_mb} MB")
        file_name = safe_filename(evidence.filename)
        valid_file, validation_message, detected_mime = validate_upload_bytes(
            file_name, content, evidence.content_type, ALLOWED_UPLOAD_EXTENSIONS
        )
        if not valid_file:
            raise HTTPException(400, validation_message)
        stored_name = f"uploads/supplier_portal/request_{data_request.id}/{datetime.now(UTC):%Y%m%d%H%M%S}_{file_name}"
        storage.put_bytes(stored_name, content, detected_mime)
        sha256 = hashlib.sha256(content).hexdigest()
        file_size = len(content)
    calculated = calculate_supplier_response(
        data_request, method=method, activity_value=max(0, activity_value), emission_factor=max(0, emission_factor),
        reported_emissions_tco2e=max(0, reported_emissions_tco2e),
    )
    has_evidence = bool(stored_name or (data_request.response and data_request.response.evidence_stored_name))
    level = supplier_quality_level(method, verified is not None, has_evidence)
    response = data_request.response or SupplierResponse(request_id=data_request.id)
    response.method = method
    response.activity_value = max(0, activity_value)
    response.activity_unit = activity_unit.strip() or data_request.unit
    response.emission_factor = max(0, emission_factor)
    response.factor_unit = factor_unit.strip()
    response.reported_emissions_tco2e = max(0, reported_emissions_tco2e)
    response.calculated_emissions_tco2e = calculated
    response.methodology = methodology.strip()
    response.boundary = boundary.strip()
    response.verified = verified is not None
    response.quality_level = level
    response.notes = notes.strip()
    response.review_status = "Pendiente"
    response.submitted_at = datetime.now(UTC)
    if stored_name:
        response.evidence_name = file_name
        response.evidence_stored_name = stored_name
        response.evidence_sha256 = sha256
        response.evidence_size = file_size
    session.add(response)
    data_request.status = "Respondida"
    data_request.responded_at = datetime.now(UTC)
    session.commit()
    warning_suffix = "" if not validation["warnings"] else f" Observaciones automáticas: {len(validation['warnings'])}."
    request.session["supplier_flash"] = {"message": "Respuesta recibida. La empresa revisará la metodología y el soporte." + warning_suffix, "level": "success"}
    return RedirectResponse(f"/proveedor/responder/{token}", status_code=303)

@app.post("/cadena-valor/respuestas/{response_id}/revisar")
def review_supplier_response(
    response_id: int, request: Request, decision: str = Form(...), reviewer_comments: str = Form(""),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    if not (user["can_review"] or user["can_manage_supply_chain"]):
        raise HTTPException(403, "Tu rol no puede revisar respuestas de proveedores")
    response = session.scalar(
        select(SupplierResponse)
        .join(SupplierDataRequest).join(SupplierCampaign).join(Inventory)
        .where(SupplierResponse.id == response_id, Inventory.organization_id == int(user["organization_id"]))
        .options(selectinload(SupplierResponse.request).selectinload(SupplierDataRequest.campaign))
    )
    if not response:
        raise HTTPException(404, "Respuesta no encontrada")
    if decision == "Aprobado":
        validation = validate_supplier_response(
            response.request,
            method=response.method,
            activity_value=response.activity_value,
            activity_unit=response.activity_unit,
            emission_factor=response.emission_factor,
            factor_unit=response.factor_unit,
            reported_emissions_tco2e=response.reported_emissions_tco2e,
            methodology=response.methodology,
            boundary=response.boundary,
            has_evidence=bool(response.evidence_stored_name),
        )
        blockers = list(validation["errors"])
        if not response.methodology.strip():
            blockers.append("La metodología debe documentarse antes de aprobar.")
        if not response.boundary.strip():
            blockers.append("Los límites del cálculo deben documentarse antes de aprobar.")
        duplicates = approved_duplicate_responses(session, response)
        if duplicates:
            blockers.append("Existe otra respuesta aprobada para la misma categoría, proveedor y producto o servicio.")
        if blockers:
            raise HTTPException(409, " ".join(blockers))
    response.review_status = decision
    response.reviewer_comments = reviewer_comments.strip()
    response.reviewed_by = str(user["email"])
    response.reviewed_at = datetime.now(UTC)
    response.request.status = "Aprobada" if decision == "Aprobado" else "Devuelta"
    session.flush()
    source = sync_supplier_source(session, response.request.campaign.inventory_id)
    add_audit(session, int(user["organization_id"]), str(user["email"]), decision.upper(), "Respuesta de proveedor", response.request.supplier.name, f"{response.calculated_emissions_tco2e:.3f} tCO2e")
    session.commit()
    set_flash(request, f"Respuesta {decision.lower()}. La fuente de alcance 3 quedó en {source.emissions:.3f} tCO₂e.")
    return RedirectResponse(f"/cadena-valor?campaign_id={response.request.campaign_id}", status_code=303)

@app.get("/cadena-valor/respuestas/{response_id}/evidencia")
def download_supplier_evidence(
    response_id: int, session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    response = session.scalar(
        select(SupplierResponse).join(SupplierDataRequest).join(SupplierCampaign).join(Inventory).where(
            SupplierResponse.id == response_id, Inventory.organization_id == int(user["organization_id"])
        )
    )
    if not response or not response.evidence_stored_name:
        raise HTTPException(404, "Evidencia no disponible")
    if not storage.exists(response.evidence_stored_name):
        raise HTTPException(404, "Archivo no encontrado")
    local_path = storage.local_path(response.evidence_stored_name)
    if local_path:
        return FileResponse(local_path, filename=response.evidence_name or local_path.name)
    return RedirectResponse(storage.presigned_url(response.evidence_stored_name), status_code=302)

@app.get("/cadena-valor/plantilla.xlsx")
def supplier_campaign_template(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    if not (user["can_manage_supply_chain"] or user["can_review"]):
        raise HTTPException(403)
    inventory = get_inventory(session, user)
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes proveedores"
    ws.append([
        "Código categoría", "Categoría Scope 3", "Dirección", "Proveedor", "NIT", "Producto o servicio",
        "Cantidad", "Unidad", "Gasto COP", "Método solicitado", "Estado", "Fecha límite", "Método respondido",
        "Calidad A-D", "Revisión", "Emisiones tCO2e", "Metodología", "Límite", "Enlace seguro",
    ])
    requests = session.scalars(
        select(SupplierDataRequest)
        .join(SupplierCampaign).where(SupplierCampaign.inventory_id == inventory.id)
        .options(
            selectinload(SupplierDataRequest.supplier),
            selectinload(SupplierDataRequest.campaign),
            selectinload(SupplierDataRequest.response),
        )
        .order_by(SupplierDataRequest.id)
    )
    base = str(request.base_url).rstrip("/") + "/proveedor/responder/"
    for item in requests:
        category = category_from_value(item.campaign.category)
        response = item.response
        ws.append([
            category.code if category else "", category.name if category else item.campaign.category,
            category.direction if category else "", item.supplier.name, item.supplier.tax_id, item.product_service,
            item.quantity, item.unit, item.spend_cop, item.requested_method, item.status, item.due_date,
            response.method if response else "", response.quality_level if response else "",
            response.review_status if response else "", response.calculated_emissions_tco2e if response else 0,
            response.methodology if response else "", response.boundary if response else "", base + item.access_token,
        ])
    summary = inventory_supply_chain_summary(session, inventory)
    categories_ws = wb.create_sheet("Screening 15 categorías")
    categories_ws.append(["Código", "Categoría", "Dirección", "Evaluación", "Relevancia 0-5", "Justificación", "Responsable", "Estrategia de datos", "Estado de datos", "Campañas", "Solicitudes", "Respuestas", "Aprobadas", "Emisiones tCO2e", "Límite mínimo", "Métodos recomendados"])
    for category in summary["categories"]:
        categories_ws.append([
            category["code"], category["name"], category["direction"], category["assessment_status"],
            category["relevance_score"], category["rationale"], category["owner"], category["data_strategy"],
            category["data_status"], category["campaign_count"], category["request_count"], category["response_count"],
            category["approved_count"], category["emissions"], category["minimum_boundary"], ", ".join(category["recommended_methods"]),
        ])
    stream = BytesIO()
    wb.save(stream)
    return Response(stream.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=cadena_valor_proveedores.xlsx"})

@app.get("/escenarios", response_class=HTMLResponse)
def scenarios_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    inventory = get_inventory(session, user)
    scenarios = list(session.scalars(
        select(ReductionScenario)
        .where(ReductionScenario.inventory_id == inventory.id)
        .options(selectinload(ReductionScenario.action_links).selectinload(ReductionScenarioAction.action))
        .order_by(ReductionScenario.created_at.desc())
    ))
    selected_id = request.query_params.get("scenario_id")
    selected = None
    if selected_id and selected_id.isdigit():
        selected = get_scenario(session, int(selected_id), int(user["organization_id"]))
    if not selected and scenarios:
        selected = get_scenario(session, scenarios[0].id, int(user["organization_id"]))
    selected_summary = scenario_summary(selected) if selected else None
    macc = portfolio_macc(inventory.reduction_actions, selected.discount_rate if selected else 10.0)
    return templates.TemplateResponse(
        request=request,
        name="scenarios.html",
        context=common_context(
            request, session, user, "scenarios", inventory=inventory, scenarios=scenarios,
            selected=selected, selected_summary=selected_summary, actions=inventory.reduction_actions,
            portfolio_macc=macc,
        ),
    )

@app.post("/escenarios/nuevo")
def create_scenario(
    request: Request, inventory_id: int = Form(...), name: str = Form(...), description: str = Form(""),
    start_year: int = Form(...), target_year: int = Form(...), discount_rate: float = Form(10.0),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_inventory")
    inventory = get_inventory(session, user, inventory_id)
    ensure_inventory_editable(inventory)
    if target_year < start_year:
        raise HTTPException(400, "El año objetivo no puede ser anterior al año inicial")
    scenario = ReductionScenario(
        inventory_id=inventory.id, name=name.strip(), description=description.strip(), start_year=start_year,
        target_year=target_year, discount_rate=max(0.0, discount_rate), status="Borrador", created_by=str(user["email"]),
    )
    session.add(scenario)
    session.flush()
    for action in inventory.reduction_actions:
        session.add(ReductionScenarioAction(
            scenario_id=scenario.id, action_id=action.id, included=False,
            implementation_year=action.implementation_year or start_year, adoption_percent=100.0,
        ))
    add_audit(session, int(user["organization_id"]), str(user["email"]), "CREAR", "Escenario", scenario.name, f"Periodo {start_year}-{target_year}")
    session.commit()
    set_flash(request, "Escenario creado. Selecciona las medidas que harán parte del portafolio.")
    return RedirectResponse(f"/escenarios?scenario_id={scenario.id}", status_code=303)

@app.post("/escenarios/{scenario_id}/configurar")
async def configure_scenario(
    scenario_id: int, request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_inventory")
    scenario = get_scenario(session, scenario_id, int(user["organization_id"]))
    if not scenario:
        raise HTTPException(404, "Escenario no encontrado")
    ensure_inventory_editable(scenario.inventory)
    form = await request.form()
    scenario.status = str(form.get("status") or scenario.status)
    discount_rate = float(form.get("discount_rate") or scenario.discount_rate)
    scenario.discount_rate = max(0.0, discount_rate)
    for link in scenario.action_links:
        link.included = f"include_{link.action_id}" in form
        try:
            link.adoption_percent = min(100.0, max(0.0, float(form.get(f"adoption_{link.action_id}") or 100)))
        except (TypeError, ValueError):
            link.adoption_percent = 100.0
        try:
            link.implementation_year = int(form.get(f"year_{link.action_id}") or scenario.start_year)
        except (TypeError, ValueError):
            link.implementation_year = scenario.start_year
    add_audit(session, int(user["organization_id"]), str(user["email"]), "CONFIGURAR", "Escenario", scenario.name, "Portafolio, adopción y cronograma actualizados")
    session.commit()
    set_flash(request, "Escenario recalculado correctamente.")
    return RedirectResponse(f"/escenarios?scenario_id={scenario.id}", status_code=303)

@app.get("/verificacion", response_class=HTMLResponse)
def verification_portal(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    if not (user["can_external_audit"] or user["can_review"] or user["can_approve"]):
        raise HTTPException(403, "Tu rol no tiene acceso al portal de verificación")
    inventory = get_inventory(session, user)
    gate = review_gate_summary(session, inventory)
    findings = list(session.scalars(
        select(VerificationFinding)
        .where(VerificationFinding.inventory_id == inventory.id)
        .options(selectinload(VerificationFinding.source))
        .order_by(VerificationFinding.status == "Cerrado", VerificationFinding.created_at.desc())
    ))
    reports = list(session.scalars(select(ReportArtifact).where(ReportArtifact.inventory_id == inventory.id).order_by(ReportArtifact.generated_at.desc())))
    decisions = list(session.scalars(select(InventoryDecision).where(InventoryDecision.inventory_id == inventory.id).order_by(InventoryDecision.decided_at.desc())))
    calculations_count = session.scalar(
        select(func.count()).select_from(EmissionCalculation).join(ActivityData).join(EmissionSource).where(EmissionSource.inventory_id == inventory.id)
    ) or 0
    evidence_with_files = sum(1 for item in inventory.documents if item.stored_name and storage.exists(item.stored_name))
    finding_counts = {
        "open": sum(1 for item in findings if item.status != "Cerrado"),
        "major": sum(1 for item in findings if item.status != "Cerrado" and item.severity in {"Mayor", "Crítica"}),
        "closed": sum(1 for item in findings if item.status == "Cerrado"),
    }
    return templates.TemplateResponse(
        request=request,
        name="verification.html",
        context=common_context(
            request, session, user, "verification", inventory=inventory, findings=findings, reports=reports,
            decisions=decisions, calculations_count=calculations_count, evidence_with_files=evidence_with_files,
            finding_counts=finding_counts, **gate,
        ),
    )

@app.post("/verificacion/hallazgos/nuevo")
def create_verification_finding(
    request: Request, inventory_id: int = Form(...), title: str = Form(...), description: str = Form(...),
    finding_type: str = Form("Observación"), severity: str = Form("Menor"), source_id: str = Form(""),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "external_audit")
    inventory = get_inventory(session, user, inventory_id)
    selected_source_id = int(source_id) if source_id.isdigit() else None
    if selected_source_id and not any(item.id == selected_source_id for item in inventory.sources):
        raise HTTPException(400, "La fuente no pertenece al inventario")
    finding = VerificationFinding(
        inventory_id=inventory.id, source_id=selected_source_id, title=title.strip(), description=description.strip(),
        finding_type=finding_type, severity=severity, status="Abierto", verifier_email=str(user["email"]),
    )
    session.add(finding)
    add_audit(session, int(user["organization_id"]), str(user["email"]), "CREAR", "Hallazgo de verificación", finding.title, f"{finding_type} · {severity}")
    session.commit()
    set_flash(request, "Hallazgo registrado en el expediente de verificación.")
    return RedirectResponse("/verificacion", status_code=303)

@app.post("/verificacion/hallazgos/{finding_id}/responder")
def respond_verification_finding(
    finding_id: int, request: Request, management_response: str = Form(...),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    if not (user["can_manage_inventory"] or user["can_provide_data"] or user["can_review"]):
        raise HTTPException(403, "Tu rol no puede responder hallazgos")
    finding = session.scalar(
        select(VerificationFinding).join(Inventory).where(
            VerificationFinding.id == finding_id, Inventory.organization_id == int(user["organization_id"])
        )
    )
    if not finding:
        raise HTTPException(404, "Hallazgo no encontrado")
    finding.management_response = management_response.strip()
    finding.response_by = str(user["email"])
    finding.response_at = datetime.now(UTC)
    finding.status = "Respondido"
    add_audit(session, int(user["organization_id"]), str(user["email"]), "RESPONDER", "Hallazgo de verificación", finding.title, management_response[:180])
    session.commit()
    set_flash(request, "Respuesta enviada al verificador.")
    return RedirectResponse("/verificacion", status_code=303)

@app.post("/verificacion/hallazgos/{finding_id}/cerrar")
def close_verification_finding(
    finding_id: int, request: Request, conclusion: str = Form(...), decision: str = Form("Cerrar"),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "external_audit")
    finding = session.scalar(
        select(VerificationFinding).join(Inventory).where(
            VerificationFinding.id == finding_id, Inventory.organization_id == int(user["organization_id"])
        )
    )
    if not finding:
        raise HTTPException(404, "Hallazgo no encontrado")
    finding.conclusion = conclusion.strip()
    finding.closed_by = str(user["email"])
    if decision == "Cerrar":
        finding.status = "Cerrado"
        finding.closed_at = datetime.now(UTC)
    else:
        finding.status = "Abierto"
        finding.closed_at = None
    add_audit(session, int(user["organization_id"]), str(user["email"]), decision.upper(), "Hallazgo de verificación", finding.title, conclusion[:180])
    session.commit()
    set_flash(request, "Decisión del verificador registrada.")
    return RedirectResponse("/verificacion", status_code=303)

@app.post("/verificacion/paquete")
def generate_verification_package(
    request: Request, inventory_id: int = Form(...), session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    if not (user["can_external_audit"] or user["can_review"] or user["can_approve"]):
        raise HTTPException(403, "Tu rol no puede generar el paquete de verificación")
    inventory = get_inventory(session, user, inventory_id)
    artifact = create_verification_package(session, inventory, str(user["email"]))
    add_audit(session, int(user["organization_id"]), str(user["email"]), "GENERAR", "Paquete de verificación", artifact.file_name, artifact.sha256)
    session.commit()
    set_flash(request, "Paquete de verificación generado con manifiesto, índices y archivos disponibles.")
    return RedirectResponse("/verificacion", status_code=303)

@app.get("/notificaciones", response_class=HTMLResponse)
def notifications_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    notifications = list(session.scalars(select(Notification).where(
        Notification.organization_id == int(user["organization_id"]),
        Notification.user_id == int(user["id"]),
    ).order_by(Notification.created_at.desc()).limit(100)))
    preference = get_or_create_preference(session, int(user["id"]))
    session.commit()
    return templates.TemplateResponse(
        request=request,
        name="notifications.html",
        context=common_context(request, session, user, "notifications", notifications=notifications, preference=preference),
    )

@app.post("/notificaciones/{notification_id}/leer")
def notification_read(notification_id: int, request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    notification = session.scalar(select(Notification).where(
        Notification.id == notification_id,
        Notification.organization_id == int(user["organization_id"]),
        Notification.user_id == int(user["id"]),
    ))
    if not notification:
        raise HTTPException(404, "Notificación no encontrada")
    notification.read_at = datetime.now(UTC)
    session.commit()
    return RedirectResponse(notification.link or "/notificaciones", status_code=303)

@app.post("/notificaciones/leer-todas")
def notifications_read_all(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    notifications = list(session.scalars(select(Notification).where(
        Notification.organization_id == int(user["organization_id"]),
        Notification.user_id == int(user["id"]),
        Notification.read_at.is_(None),
    )))
    now = datetime.now(UTC)
    for notification in notifications:
        notification.read_at = now
    session.commit()
    set_flash(request, "Todas las notificaciones quedaron marcadas como leídas.")
    return RedirectResponse("/notificaciones", status_code=303)

@app.post("/notificaciones/preferencias")
def notification_preferences_update(
    request: Request,
    email_enabled: str | None = Form(None),
    in_app_enabled: str | None = Form(None),
    digest_frequency: str = Form("Inmediato"),
    session: Session = Depends(get_db),
    user: dict = Depends(require_user),
):
    preference = get_or_create_preference(session, int(user["id"]))
    preference.email_enabled = email_enabled == "on"
    preference.in_app_enabled = in_app_enabled == "on"
    preference.digest_frequency = digest_frequency if digest_frequency in {"Inmediato", "Diario", "Semanal"} else "Inmediato"
    session.commit()
    set_flash(request, "Preferencias de notificación actualizadas.")
    return RedirectResponse("/notificaciones", status_code=303)

@app.get("/administracion-plataforma", response_class=HTMLResponse)
def platform_admin_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_operations")
    users = list(session.scalars(
        select(AppUser).join(OrganizationMembership).where(
            OrganizationMembership.organization_id == int(user["organization_id"]),
            OrganizationMembership.active.is_(True),
        ).order_by(AppUser.name)
    ))
    settings_rows = list(session.scalars(select(PlatformSetting).where(PlatformSetting.organization_id == int(user["organization_id"])).order_by(PlatformSetting.key)))
    notification_stats = {
        "total": session.scalar(select(func.count(Notification.id)).where(Notification.organization_id == int(user["organization_id"]))) or 0,
        "pending": session.scalar(select(func.count(Notification.id)).where(Notification.organization_id == int(user["organization_id"]), Notification.status.in_(["Pendiente", "Error"]))) or 0,
        "unread": session.scalar(select(func.count(Notification.id)).where(Notification.organization_id == int(user["organization_id"]), Notification.read_at.is_(None))) or 0,
    }
    storage_status = storage.diagnostics()
    return templates.TemplateResponse(
        request=request,
        name="platform_admin.html",
        context=common_context(request, session, user, "platform_admin", users=users, settings_rows=settings_rows, notification_stats=notification_stats, storage_status=storage_status, app_settings=settings),
    )

@app.post("/administracion-plataforma/configuracion")
def platform_setting_update(
    request: Request,
    key: str = Form(...),
    value: str = Form(...),
    description: str = Form(""),
    session: Session = Depends(get_db),
    user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_operations")
    clean_key = re.sub(r"[^a-z0-9_]+", "_", key.strip().lower()).strip("_")
    if not clean_key:
        raise HTTPException(400, "Clave inválida")
    row = session.scalar(select(PlatformSetting).where(PlatformSetting.organization_id == int(user["organization_id"]), PlatformSetting.key == clean_key))
    if not row:
        row = PlatformSetting(organization_id=int(user["organization_id"]), key=clean_key)
        session.add(row)
    row.value = value.strip()
    row.description = description.strip()
    row.updated_by = str(user["email"])
    add_audit(session, int(user["organization_id"]), str(user["email"]), "CONFIGURAR", "Plataforma", clean_key, value.strip())
    session.commit()
    set_flash(request, "Configuración guardada.")
    return RedirectResponse("/administracion-plataforma", status_code=303)

@app.post("/administracion-plataforma/notificaciones/prueba")
def platform_test_notification(
    request: Request,
    role: str = Form("Administrador"),
    title: str = Form("Prueba de notificación"),
    message: str = Form("El centro de notificaciones está operativo."),
    session: Session = Depends(get_db),
    user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_operations")
    created = notify_roles(session, int(user["organization_id"]), {role}, title, message, link="/notificaciones", category="Prueba", email_requested=True)
    add_audit(session, int(user["organization_id"]), str(user["email"]), "NOTIFICAR", "Plataforma", role, f"{len(created)} destinatarios")
    session.commit()
    set_flash(request, f"Se generaron {len(created)} notificaciones de prueba.")
    return RedirectResponse("/administracion-plataforma", status_code=303)

@app.post("/administracion-plataforma/notificaciones/procesar")
def platform_process_notifications(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_operations")
    result = process_pending_notifications(session, settings.notification_batch_size)
    session.commit()
    set_flash(request, f"Cola procesada: {result['sent']} enviadas y {result['failed']} con error.")
    return RedirectResponse("/administracion-plataforma", status_code=303)

@app.get("/portafolio", response_class=HTMLResponse)
def portfolio_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_portfolio")
    memberships = list(session.scalars(
        select(OrganizationMembership)
        .where(OrganizationMembership.user_id == int(user["id"]), OrganizationMembership.active.is_(True))
        .options(selectinload(OrganizationMembership.organization).selectinload(Organization.inventories))
        .order_by(OrganizationMembership.id)
    ))
    portfolio = []
    for membership in memberships:
        org_item = membership.organization
        inventories = org_item.inventories if org_item else []
        portfolio.append({
            "membership": membership,
            "organization": org_item,
            "inventories": inventories,
            "latest_inventory": sorted(inventories, key=lambda item: (item.start_date, item.id), reverse=True)[0] if inventories else None,
            "demo_story": demo_story_for(org_item.trade_name) if org_item else None,
        })
    return templates.TemplateResponse(
        request=request,
        name="portfolio.html",
        context=common_context(request, session, user, "portfolio", portfolio=portfolio),
    )

@app.post("/portafolio/cambiar/{organization_id}")
def portfolio_switch(organization_id: int, request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    membership = session.scalar(select(OrganizationMembership).where(
        OrganizationMembership.user_id == int(user["id"]),
        OrganizationMembership.organization_id == organization_id,
        OrganizationMembership.active.is_(True),
    ))
    if not membership:
        raise HTTPException(403, "No tienes acceso a esta organización")
    request.session["active_org_id"] = organization_id
    add_audit(session, organization_id, str(user["email"]), "CAMBIAR", "Organización activa", str(organization_id), "Cambio de contexto multiempresa")
    session.commit()
    set_flash(request, "Organización activa actualizada.")
    return RedirectResponse("/dashboard", status_code=303)

@app.post("/portafolio/nueva")
def portfolio_create(
    request: Request,
    name: str = Form(...),
    trade_name: str = Form(""),
    tax_id: str = Form(...),
    sector: str = Form(...),
    city: str = Form("Medellín"),
    session: Session = Depends(get_db),
    user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_org")
    if session.scalar(select(Organization).where(func.lower(Organization.name) == name.strip().lower())):
        raise HTTPException(409, "Ya existe una organización con ese nombre")
    organization = Organization(
        name=name.strip(), trade_name=trade_name.strip() or name.strip(), tax_id=tax_id.strip(),
        sector=sector.strip(), country="Colombia", department="Antioquia", city=city.strip(),
        contact_name=str(user["name"]), contact_email=str(user["email"]), status="Activa",
    )
    session.add(organization)
    session.flush()
    session.add(OrganizationMembership(
        user_id=int(user["id"]), organization_id=organization.id, role="Administrador", active=True,
    ))
    add_audit(session, organization.id, str(user["email"]), "CREAR", "Organización", organization.name, "Alta desde portafolio multiempresa")
    session.commit()
    request.session["active_org_id"] = organization.id
    set_flash(request, "Organización creada. Ahora puedes configurar sedes e inventarios.")
    return RedirectResponse("/organizacion", status_code=303)

@app.get("/automatizaciones", response_class=HTMLResponse)
def automations_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_automations")
    automations = list(session.scalars(
        select(ScheduledAutomation)
        .where(ScheduledAutomation.organization_id == int(user["organization_id"]))
        .options(selectinload(ScheduledAutomation.inventory), selectinload(ScheduledAutomation.runs))
        .order_by(ScheduledAutomation.active.desc(), ScheduledAutomation.name)
    ))
    inventories = list(session.scalars(select(Inventory).where(Inventory.organization_id == int(user["organization_id"])).order_by(Inventory.start_date.desc())))
    recent_runs = list(session.scalars(
        select(AutomationRun)
        .join(ScheduledAutomation)
        .where(ScheduledAutomation.organization_id == int(user["organization_id"]))
        .options(selectinload(AutomationRun.automation))
        .order_by(AutomationRun.started_at.desc()).limit(30)
    ))
    return templates.TemplateResponse(
        request=request,
        name="automations.html",
        context=common_context(
            request, session, user, "automations", automations=automations, inventories=inventories,
            recent_runs=recent_runs, automation_types=AUTOMATION_TYPES, cadences=CADENCES,
            role_options=ROLE_OPTIONS, scheduler_enabled=settings.scheduler_enabled,
        ),
    )

@app.post("/automatizaciones/nueva")
def automation_create(
    request: Request,
    name: str = Form(...),
    automation_type: str = Form(...),
    cadence: str = Form("Semanal"),
    schedule_time: str = Form("08:00"),
    inventory_id: int | None = Form(None),
    weekday: int | None = Form(None),
    month_day: int | None = Form(None),
    days_before: int = Form(3),
    recipient_roles: list[str] = Form(default=[]),
    session: Session = Depends(get_db),
    user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_automations")
    if automation_type not in AUTOMATION_TYPES or cadence not in CADENCES:
        raise HTTPException(400, "Tipo o frecuencia inválida")
    if inventory_id:
        get_inventory(session, user, inventory_id)
    automation = ScheduledAutomation(
        organization_id=int(user["organization_id"]), inventory_id=inventory_id or None,
        name=name.strip(), automation_type=automation_type, cadence=cadence,
        schedule_time=schedule_time, weekday=weekday, month_day=month_day,
        timezone="America/Bogota", recipient_roles=json.dumps(recipient_roles or ["Administrador", "Consultor"]),
        days_before=max(0, min(days_before, 60)), active=True, created_by=str(user["email"]),
    )
    session.add(automation)
    session.flush()
    automation.next_run_at = calculate_next_run(automation)
    add_audit(session, int(user["organization_id"]), str(user["email"]), "CREAR", "Automatización", automation.name, automation.automation_type)
    session.commit()
    set_flash(request, "Automatización creada y programada.")
    return RedirectResponse("/automatizaciones", status_code=303)

@app.post("/automatizaciones/{automation_id}/estado")
def automation_toggle(automation_id: int, request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_automations")
    automation = session.scalar(select(ScheduledAutomation).where(
        ScheduledAutomation.id == automation_id,
        ScheduledAutomation.organization_id == int(user["organization_id"]),
    ))
    if not automation:
        raise HTTPException(404, "Automatización no encontrada")
    automation.active = not automation.active
    automation.next_run_at = calculate_next_run(automation) if automation.active else None
    add_audit(session, int(user["organization_id"]), str(user["email"]), "ACTIVAR" if automation.active else "DESACTIVAR", "Automatización", automation.name)
    session.commit()
    set_flash(request, f"Automatización {'activada' if automation.active else 'desactivada'}.")
    return RedirectResponse("/automatizaciones", status_code=303)

@app.post("/automatizaciones/{automation_id}/ejecutar")
def automation_run_now(automation_id: int, request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_automations")
    automation = session.scalar(select(ScheduledAutomation).where(
        ScheduledAutomation.id == automation_id,
        ScheduledAutomation.organization_id == int(user["organization_id"]),
    ))
    if not automation:
        raise HTTPException(404, "Automatización no encontrada")
    run = execute_automation(session, automation, triggered_by=str(user["email"]))
    add_audit(session, int(user["organization_id"]), str(user["email"]), "EJECUTAR", "Automatización", automation.name, run.summary)
    session.commit()
    set_flash(request, f"Ejecución {run.status.lower()}: {run.summary}", "success" if run.status == "Ejecutado" else "warning")
    return RedirectResponse("/automatizaciones", status_code=303)

@app.post("/automatizaciones/procesar-vencidas")
def automations_process_due(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_automations")
    result = process_due_automations(session)
    set_flash(request, f"Programación revisada: {result['executed']} ejecuciones y {result['errors']} errores.")
    return RedirectResponse("/automatizaciones", status_code=303)

def _compliance_score(rows: list[ComplianceAssessment]) -> int:
    applicable = [row for row in rows if row.status != "No aplica"]
    if not applicable:
        return 0
    weights = {"Cumple": 100, "Parcial": 50, "Pendiente": 0, "No cumple": 0}
    return round(sum(weights.get(row.status, 0) for row in applicable) / len(applicable))

@app.get("/direccion-ejecutiva", response_class=HTMLResponse)
def executive_portfolio_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_portfolio")
    memberships = list(session.scalars(
        select(OrganizationMembership)
        .where(OrganizationMembership.user_id == int(user["id"]), OrganizationMembership.active.is_(True))
        .options(selectinload(OrganizationMembership.organization))
        .order_by(OrganizationMembership.id)
    ))
    cards = []
    portfolio_total = 0.0
    total_reduction = 0.0
    for membership in memberships:
        organization = membership.organization
        inventory = session.scalar(select(Inventory).where(Inventory.organization_id == organization.id).order_by(Inventory.start_date.desc(), Inventory.id.desc()).limit(1))
        if inventory:
            emissions = session.scalar(select(func.coalesce(func.sum(EmissionSource.emissions), 0.0)).where(EmissionSource.inventory_id == inventory.id, EmissionSource.included.is_(True))) or 0.0
            assessments = list(session.scalars(select(ComplianceAssessment).where(ComplianceAssessment.inventory_id == inventory.id).options(selectinload(ComplianceAssessment.requirement))))
            open_observations = session.scalar(select(func.count(ReviewObservation.id)).where(ReviewObservation.inventory_id == inventory.id, ReviewObservation.status != "Cerrada")) or 0
            reduction = session.scalar(select(func.coalesce(func.sum(ReductionAction.expected_reduction), 0.0)).where(ReductionAction.inventory_id == inventory.id)) or 0.0
            documents = session.scalar(select(func.count(DocumentControlRecord.id)).where(DocumentControlRecord.organization_id == organization.id)) or 0
            portfolio_total += float(emissions)
            total_reduction += float(reduction)
            cards.append({"organization": organization, "membership": membership, "inventory": inventory, "emissions": float(emissions), "compliance": _compliance_score(assessments), "open_observations": open_observations, "reduction": float(reduction), "documents": documents})
        else:
            cards.append({"organization": organization, "membership": membership, "inventory": None, "emissions": 0.0, "compliance": 0, "open_observations": 0, "reduction": 0.0, "documents": 0})
    average_compliance = round(sum(item["compliance"] for item in cards) / max(len(cards), 1))
    return templates.TemplateResponse(request=request, name="executive_portfolio.html", context=common_context(request, session, user, "executive", cards=cards, portfolio_total=portfolio_total, total_reduction=total_reduction, average_compliance=average_compliance))

@app.get("/cumplimiento", response_class=HTMLResponse)
def compliance_page(request: Request, inventory_id: int | None = None, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "view_compliance")
    inventory = get_inventory(session, user, inventory_id)
    rows = list(session.scalars(
        select(ComplianceAssessment)
        .where(ComplianceAssessment.inventory_id == inventory.id)
        .options(selectinload(ComplianceAssessment.requirement), selectinload(ComplianceAssessment.evidence))
        .join(ComplianceRequirement)
        .order_by(ComplianceRequirement.display_order)
    ))
    inventories = list(session.scalars(select(Inventory).where(Inventory.organization_id == int(user["organization_id"])).order_by(Inventory.start_date.desc())))
    score = _compliance_score(rows)
    by_framework = {}
    for row in rows:
        by_framework.setdefault(row.requirement.framework, []).append(row)
    return templates.TemplateResponse(request=request, name="compliance.html", context=common_context(request, session, user, "compliance", inventory=inventory, inventories=inventories, rows=rows, by_framework=by_framework, compliance_score=score, documents=inventory.documents))

@app.post("/cumplimiento/{assessment_id}/actualizar")
def compliance_update(assessment_id: int, request: Request, status: str = Form(...), owner: str = Form("Responsable ambiental"), evidence_id: int | None = Form(None), notes: str = Form(""), session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_compliance")
    assessment = session.scalar(select(ComplianceAssessment).join(Inventory).where(ComplianceAssessment.id == assessment_id, Inventory.organization_id == int(user["organization_id"])).options(selectinload(ComplianceAssessment.requirement)))
    if not assessment:
        raise HTTPException(404, "Evaluación no encontrada")
    if status not in {"Cumple", "Parcial", "Pendiente", "No cumple", "No aplica"}:
        raise HTTPException(400, "Estado inválido")
    if evidence_id:
        evidence = session.scalar(select(EvidenceDocument).where(EvidenceDocument.id == evidence_id, EvidenceDocument.inventory_id == assessment.inventory_id))
        if not evidence:
            raise HTTPException(400, "La evidencia no pertenece al inventario")
    assessment.status = status
    assessment.owner = owner.strip()
    assessment.evidence_id = evidence_id or None
    assessment.notes = notes.strip()
    assessment.updated_by = str(user["email"])
    add_audit(session, int(user["organization_id"]), str(user["email"]), "EVALUAR", "Cumplimiento", assessment.requirement.code, f"Estado {status}")
    session.commit()
    set_flash(request, "Evaluación de cumplimiento actualizada.")
    return RedirectResponse(f"/cumplimiento?inventory_id={assessment.inventory_id}", status_code=303)

@app.get("/gobierno-metodologico", response_class=HTMLResponse)
def methodology_governance_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_methodology_governance")
    releases = list(session.scalars(select(MethodologyRelease).where(MethodologyRelease.organization_id == int(user["organization_id"])).order_by(MethodologyRelease.created_at.desc())))
    inventories = list(session.scalars(select(Inventory).where(Inventory.organization_id == int(user["organization_id"])).order_by(Inventory.start_date.desc())))
    snapshots = list(session.scalars(select(InventoryMethodologySnapshot).join(Inventory).where(Inventory.organization_id == int(user["organization_id"])).options(selectinload(InventoryMethodologySnapshot.inventory), selectinload(InventoryMethodologySnapshot.release)).order_by(InventoryMethodologySnapshot.created_at.desc())))
    return templates.TemplateResponse(request=request, name="methodology_governance.html", context=common_context(request, session, user, "methodology_governance", releases=releases, inventories=inventories, snapshots=snapshots))

@app.post("/gobierno-metodologico/versiones/nueva")
def methodology_release_create(request: Request, name: str = Form(...), version: str = Form(...), issuing_body: str = Form("Calcula tu Huella"), publication_date: str = Form(""), effective_from: str = Form(""), source_reference: str = Form(""), notes: str = Form(""), session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_methodology_governance")
    if session.scalar(select(MethodologyRelease).where(MethodologyRelease.organization_id == int(user["organization_id"]), MethodologyRelease.name == name.strip(), MethodologyRelease.version == version.strip())):
        raise HTTPException(409, "La versión metodológica ya existe")
    fingerprint = hashlib.sha256(f"{name}|{version}|{source_reference}|{notes}".encode()).hexdigest()
    release = MethodologyRelease(organization_id=int(user["organization_id"]), name=name.strip(), version=version.strip(), issuing_body=issuing_body.strip(), publication_date=parse_date(publication_date) if publication_date else None, effective_from=parse_date(effective_from) if effective_from else None, status="Borrador", source_reference=source_reference.strip(), content_hash=fingerprint, notes=notes.strip())
    session.add(release)
    add_audit(session, int(user["organization_id"]), str(user["email"]), "CREAR", "Versión metodológica", f"{release.name} {release.version}", fingerprint)
    session.commit()
    set_flash(request, "Versión metodológica creada en borrador.")
    return RedirectResponse("/gobierno-metodologico", status_code=303)

@app.post("/gobierno-metodologico/versiones/{release_id}/aprobar")
def methodology_release_approve(release_id: int, request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_methodology_governance")
    release = session.scalar(select(MethodologyRelease).where(MethodologyRelease.id == release_id, MethodologyRelease.organization_id == int(user["organization_id"])))
    if not release:
        raise HTTPException(404, "Versión no encontrada")
    release.status = "Aprobado"
    release.approved_by = str(user["email"])
    release.approved_at = datetime.now(UTC)
    add_audit(session, int(user["organization_id"]), str(user["email"]), "APROBAR", "Versión metodológica", f"{release.name} {release.version}")
    session.commit()
    set_flash(request, "Versión metodológica aprobada.")
    return RedirectResponse("/gobierno-metodologico", status_code=303)

@app.post("/gobierno-metodologico/snapshots/nuevo")
def methodology_snapshot_create(request: Request, inventory_id: int = Form(...), methodology_release_id: int | None = Form(None), snapshot_name: str = Form(...), policy_notes: str = Form(""), session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_methodology_governance")
    inventory = get_inventory(session, user, inventory_id)
    release = None
    if methodology_release_id:
        release = session.scalar(select(MethodologyRelease).where(MethodologyRelease.id == methodology_release_id, MethodologyRelease.organization_id == int(user["organization_id"])))
        if not release:
            raise HTTPException(400, "Versión metodológica inválida")
    snapshot = InventoryMethodologySnapshot(inventory_id=inventory.id, methodology_release_id=release.id if release else None, snapshot_name=snapshot_name.strip(), status="Aprobado", methodology_name=inventory.methodology, methodology_version=inventory.methodology_version, gwp_version=inventory.gwp_version, consolidation_approach=inventory.consolidation_approach, materiality_threshold=inventory.materiality_threshold, policy_json=json.dumps({"notes": policy_notes.strip(), "inventory_version": inventory.version}, ensure_ascii=False), approved_by=str(user["email"]), approved_at=datetime.now(UTC))
    session.add(snapshot)
    add_audit(session, int(user["organization_id"]), str(user["email"]), "CONGELAR", "Snapshot metodológico", snapshot.snapshot_name, f"Inventario #{inventory.id}")
    session.commit()
    set_flash(request, "Configuración metodológica congelada para el inventario.")
    return RedirectResponse("/gobierno-metodologico", status_code=303)

@app.get("/centro-documental", response_class=HTMLResponse)
def document_center_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_documents")
    records = list(session.scalars(select(DocumentControlRecord).where(DocumentControlRecord.organization_id == int(user["organization_id"])).options(selectinload(DocumentControlRecord.inventory), selectinload(DocumentControlRecord.evidence), selectinload(DocumentControlRecord.report)).order_by(DocumentControlRecord.category, DocumentControlRecord.document_code)))
    inventories = list(session.scalars(select(Inventory).where(Inventory.organization_id == int(user["organization_id"])).order_by(Inventory.start_date.desc())))
    evidence = list(session.scalars(select(EvidenceDocument).join(Inventory).where(Inventory.organization_id == int(user["organization_id"])).order_by(EvidenceDocument.uploaded_at.desc()).limit(100)))
    reports = list(session.scalars(select(ReportArtifact).join(Inventory).where(Inventory.organization_id == int(user["organization_id"])).order_by(ReportArtifact.generated_at.desc()).limit(100)))
    due = [row for row in records if row.review_due and row.review_due <= date.today()]
    return templates.TemplateResponse(request=request, name="document_center.html", context=common_context(request, session, user, "documents", records=records, inventories=inventories, evidence=evidence, reports=reports, due=due))

@app.post("/centro-documental/registros/nuevo")
def document_record_create(request: Request, document_code: str = Form(...), title: str = Form(...), category: str = Form("Soporte"), version: str = Form("1.0"), owner: str = Form("Gestión ambiental"), confidentiality: str = Form("Interno"), retention_years: int = Form(7), review_due: str = Form(""), inventory_id: int | None = Form(None), evidence_document_id: int | None = Form(None), report_artifact_id: int | None = Form(None), notes: str = Form(""), session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_documents")
    organization_id = int(user["organization_id"])
    if session.scalar(select(DocumentControlRecord).where(DocumentControlRecord.organization_id == organization_id, DocumentControlRecord.document_code == document_code.strip())):
        raise HTTPException(409, "El código documental ya existe")
    inventory = get_inventory(session, user, inventory_id) if inventory_id else None
    evidence = None
    report = None
    if evidence_document_id:
        evidence = session.scalar(select(EvidenceDocument).join(Inventory).where(EvidenceDocument.id == evidence_document_id, Inventory.organization_id == organization_id))
        if not evidence:
            raise HTTPException(400, "Evidencia inválida")
    if report_artifact_id:
        report = session.scalar(select(ReportArtifact).join(Inventory).where(ReportArtifact.id == report_artifact_id, Inventory.organization_id == organization_id))
        if not report:
            raise HTTPException(400, "Informe inválido")
    row = DocumentControlRecord(organization_id=organization_id, inventory_id=inventory.id if inventory else None, evidence_document_id=evidence.id if evidence else None, report_artifact_id=report.id if report else None, document_code=document_code.strip(), title=title.strip(), category=category.strip(), version=version.strip(), owner=owner.strip(), confidentiality=confidentiality, retention_years=max(1, retention_years), review_due=parse_date(review_due) if review_due else None, status="Vigente", sha256=(evidence.sha256 if evidence else (report.sha256 if report else "")), notes=notes.strip(), created_by=str(user["email"]))
    session.add(row)
    add_audit(session, organization_id, str(user["email"]), "REGISTRAR", "Documento controlado", row.document_code, row.title)
    session.commit()
    set_flash(request, "Documento incorporado al registro maestro.")
    return RedirectResponse("/centro-documental", status_code=303)

@app.post("/centro-documental/registros/{record_id}/actualizar")
def document_record_update(record_id: int, request: Request, status: str = Form(...), version: str = Form(...), owner: str = Form(...), confidentiality: str = Form(...), review_due: str = Form(""), notes: str = Form(""), session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_documents")
    row = session.scalar(select(DocumentControlRecord).where(DocumentControlRecord.id == record_id, DocumentControlRecord.organization_id == int(user["organization_id"])))
    if not row:
        raise HTTPException(404, "Documento no encontrado")
    row.status = status
    row.version = version.strip()
    row.owner = owner.strip()
    row.confidentiality = confidentiality
    row.review_due = parse_date(review_due) if review_due else None
    row.notes = notes.strip()
    add_audit(session, int(user["organization_id"]), str(user["email"]), "ACTUALIZAR", "Documento controlado", row.document_code, f"Versión {row.version} · {row.status}")
    session.commit()
    set_flash(request, "Control documental actualizado.")
    return RedirectResponse("/centro-documental", status_code=303)

@app.get("/alistamiento", response_class=HTMLResponse)
def readiness_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_readiness")
    rows = list(session.scalars(select(CommercialReadinessItem).where(CommercialReadinessItem.organization_id == int(user["organization_id"])).order_by(CommercialReadinessItem.display_order)))
    weights = {"Completado": 100, "En progreso": 50, "Pendiente": 0, "Bloqueado": 0}
    score = round(sum(weights.get(row.status, 0) for row in rows) / max(len(rows), 1))
    categories = {}
    for row in rows:
        categories.setdefault(row.category, []).append(row)
    return templates.TemplateResponse(request=request, name="readiness.html", context=common_context(request, session, user, "readiness", rows=rows, categories=categories, readiness_score=score))

@app.post("/alistamiento/{item_id}/actualizar")
def readiness_update(item_id: int, request: Request, status: str = Form(...), owner: str = Form(...), due_date: str = Form(""), notes: str = Form(""), session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_readiness")
    row = session.scalar(select(CommercialReadinessItem).where(CommercialReadinessItem.id == item_id, CommercialReadinessItem.organization_id == int(user["organization_id"])))
    if not row:
        raise HTTPException(404, "Elemento no encontrado")
    if status not in {"Completado", "En progreso", "Pendiente", "Bloqueado"}:
        raise HTTPException(400, "Estado inválido")
    row.status = status
    row.owner = owner.strip()
    row.due_date = parse_date(due_date) if due_date else None
    row.notes = notes.strip()
    row.updated_by = str(user["email"])
    add_audit(session, int(user["organization_id"]), str(user["email"]), "ACTUALIZAR", "Alistamiento comercial", row.title, status)
    session.commit()
    set_flash(request, "Elemento de alistamiento actualizado.")
    return RedirectResponse("/alistamiento", status_code=303)

def _service_usage(session: Session, organization_id: int, plan: ServicePlan | None) -> dict[str, object]:
    users = session.scalar(select(func.count(OrganizationMembership.id)).where(
        OrganizationMembership.organization_id == organization_id,
        OrganizationMembership.active.is_(True),
    )) or 0
    facilities = session.scalar(select(func.count(Facility.id)).where(Facility.organization_id == organization_id)) or 0
    inventories = session.scalar(select(func.count(Inventory.id)).where(Inventory.organization_id == organization_id)) or 0
    storage_bytes = session.scalar(
        select(func.coalesce(func.sum(EvidenceDocument.file_size), 0))
        .join(Inventory, EvidenceDocument.inventory_id == Inventory.id)
        .where(Inventory.organization_id == organization_id)
    ) or 0
    storage_mb = round(float(storage_bytes) / (1024 * 1024), 2)
    values = {
        "users": {"value": int(users), "limit": plan.max_users if plan else 0, "label": "Usuarios"},
        "facilities": {"value": int(facilities), "limit": plan.max_facilities if plan else 0, "label": "Sedes"},
        "inventories": {"value": int(inventories), "limit": plan.max_inventories if plan else 0, "label": "Inventarios"},
        "storage": {"value": storage_mb, "limit": plan.max_storage_mb if plan else 0, "label": "Almacenamiento MB"},
    }
    period = date.today().replace(day=1)
    metric_map = {"users": users, "facilities": facilities, "inventories": inventories, "storage_mb": storage_mb}
    for metric, value in metric_map.items():
        counter = session.scalar(select(UsageCounter).where(
            UsageCounter.organization_id == organization_id,
            UsageCounter.metric == metric,
            UsageCounter.period_start == period,
        ))
        if counter:
            counter.value = float(value)
        else:
            session.add(UsageCounter(organization_id=organization_id, metric=metric, period_start=period, value=float(value)))
    for item in values.values():
        limit = float(item["limit"] or 0)
        item["percentage"] = min(100, round(float(item["value"]) / limit * 100)) if limit else 0
        item["exceeded"] = bool(limit and float(item["value"]) > limit)
    return values

@app.get("/cuenta-servicio", response_class=HTMLResponse)
def service_account(request: Request, session: Session = Depends(get_db)):
    user = require_user(request)
    subscription = session.scalar(
        select(OrganizationSubscription)
        .where(OrganizationSubscription.organization_id == int(user["organization_id"]))
        .options(selectinload(OrganizationSubscription.plan))
    )
    plans = list(session.scalars(select(ServicePlan).where(ServicePlan.active.is_(True)).order_by(ServicePlan.monthly_fee)))
    usage = _service_usage(session, int(user["organization_id"]), subscription.plan if subscription else None)
    invoices = list(session.scalars(select(BillingInvoice).where(BillingInvoice.organization_id == int(user["organization_id"])).order_by(BillingInvoice.issued_at.desc())))
    session.commit()
    return templates.TemplateResponse(request, "service_account.html", common_context(
        request, session, user, "service_account", subscription=subscription, plans=plans, usage=usage, invoices=invoices,
    ))

@app.post("/cuenta-servicio/suscripcion")
def update_subscription(
    request: Request,
    plan_id: int = Form(...),
    billing_cycle: str = Form("Anual"),
    session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_subscription")
    plan = session.get(ServicePlan, plan_id)
    if not plan or not plan.active:
        raise HTTPException(404, "Plan no disponible")
    subscription = session.scalar(select(OrganizationSubscription).where(OrganizationSubscription.organization_id == int(user["organization_id"])))
    previous = subscription.plan_id if subscription else None
    if subscription:
        subscription.plan_id = plan.id
        subscription.billing_cycle = billing_cycle if billing_cycle in {"Mensual", "Anual"} else "Anual"
        subscription.status = "Activa"
        subscription.start_date = date.today()
    else:
        subscription = OrganizationSubscription(
            organization_id=int(user["organization_id"]), plan_id=plan.id,
            billing_cycle=billing_cycle if billing_cycle in {"Mensual", "Anual"} else "Anual",
            status="Activa", start_date=date.today(),
        )
        session.add(subscription)
    add_audit(session, int(user["organization_id"]), str(user["email"]), "ACTUALIZAR", "Suscripción", plan.name, previous_value=str(previous or ""), new_value=str(plan.id))
    session.commit()
    set_flash(request, f"Plan actualizado a {plan.name}. Esta operación es administrativa y no procesa pagos.")
    return RedirectResponse("/cuenta-servicio", status_code=303)

@app.get("/onboarding", response_class=HTMLResponse)
def onboarding(request: Request, session: Session = Depends(get_db)):
    user = require_user(request)
    inventory = get_inventory(session, user)
    rows = list(session.scalars(select(CustomerOnboardingItem).where(
        CustomerOnboardingItem.organization_id == int(user["organization_id"])
    ).order_by(CustomerOnboardingItem.display_order)))
    onboarding_state = onboarding_summary(rows, inventory_id=inventory.id)
    return templates.TemplateResponse(request, "onboarding.html", common_context(
        request, session, user, "onboarding", inventory=inventory, rows=rows,
        onboarding=onboarding_state, onboarding_score=onboarding_state["score"],
    ))

@app.post("/onboarding/{item_id}/actualizar")
def update_onboarding_item(
    item_id: int,
    request: Request,
    status: str = Form(...),
    owner: str = Form(""),
    due_date: str = Form(""),
    session: Session = Depends(get_db),
):
    user = require_user(request)
    if not (user["can_manage_org"] or user["can_manage_inventory"]):
        raise HTTPException(403, "Tu rol no puede modificar el onboarding")
    row = session.scalar(select(CustomerOnboardingItem).where(
        CustomerOnboardingItem.id == item_id,
        CustomerOnboardingItem.organization_id == int(user["organization_id"]),
    ))
    if not row:
        raise HTTPException(404, "Actividad de onboarding no encontrada")
    row.status = status if status in {"Pendiente", "En progreso", "Completado", "Bloqueado"} else "Pendiente"
    row.owner = owner.strip() or row.owner
    row.due_date = parse_date(due_date) if due_date else None
    row.completed_at = datetime.now(UTC) if row.status == "Completado" else None
    row.updated_by = str(user["email"])
    add_audit(session, int(user["organization_id"]), str(user["email"]), "ACTUALIZAR", "Onboarding", row.title, new_value=row.status)
    session.commit()
    set_flash(request, "Actividad de onboarding actualizada.")
    return RedirectResponse("/onboarding", status_code=303)

@app.get("/soporte", response_class=HTMLResponse)
def support_center(
    request: Request,
    status: str = "",
    category: str = "",
    q: str = "",
    inventory_id: int | None = None,
    source_id: int | None = None,
    activity_data_id: int | None = None,
    session: Session = Depends(get_db),
):
    user = require_user(request)
    organization_id = int(user["organization_id"])
    query = (
        select(SupportTicket)
        .where(SupportTicket.organization_id == organization_id)
        .options(
            selectinload(SupportTicket.messages),
            selectinload(SupportTicket.inventory),
            selectinload(SupportTicket.source),
            selectinload(SupportTicket.activity_data),
        )
        .order_by(SupportTicket.updated_at.desc(), SupportTicket.created_at.desc())
    )
    all_tickets = list(session.scalars(query))
    tickets = all_tickets
    if status:
        tickets = [item for item in tickets if item.status == status]
    if category:
        tickets = [item for item in tickets if item.category == category]
    normalized_q = q.strip().casefold()
    if normalized_q:
        tickets = [
            item for item in tickets
            if normalized_q in " ".join((item.public_reference, item.subject, item.description, item.assigned_to)).casefold()
        ]
    inventories = list(session.scalars(select(Inventory).where(Inventory.organization_id == organization_id).order_by(Inventory.created_at.desc())))
    sources = list(session.scalars(
        select(EmissionSource).join(Inventory).where(Inventory.organization_id == organization_id).order_by(EmissionSource.name)
    ))
    prefill_record = None
    if activity_data_id:
        prefill_record = session.scalar(
            select(ActivityData).join(EmissionSource).join(Inventory).where(
                ActivityData.id == activity_data_id, Inventory.organization_id == organization_id,
            )
        )
    return templates.TemplateResponse(request, "support.html", common_context(
        request, session, user, "support", tickets=tickets, stats=support_summary(all_tickets),
        status_class=status_class, ticket_overdue=ticket_overdue, ticket_waiting_days=ticket_waiting_days,
        ticket_context=ticket_context, filters={"status": status, "category": category, "q": q},
        inventories=inventories, sources=sources, prefill={
            "inventory_id": inventory_id or (prefill_record.source.inventory_id if prefill_record and prefill_record.source else None),
            "source_id": source_id or (prefill_record.source_id if prefill_record else None),
            "activity_data_id": activity_data_id or None,
            "category": "Revisión de factor" if activity_data_id else "",
        },
    ))


@app.get("/soporte/{ticket_id}", response_class=HTMLResponse)
def support_ticket_detail(ticket_id: int, request: Request, session: Session = Depends(get_db)):
    user = require_user(request)
    ticket = session.scalar(
        select(SupportTicket)
        .where(SupportTicket.id == ticket_id, SupportTicket.organization_id == int(user["organization_id"]))
        .options(
            selectinload(SupportTicket.messages),
            selectinload(SupportTicket.inventory),
            selectinload(SupportTicket.source),
            selectinload(SupportTicket.activity_data),
        )
    )
    if not ticket:
        raise HTTPException(404, "Caso no encontrado")
    visible_messages = [
        message for message in ticket.messages
        if message.visible_to_client or user["role"] != "Cliente"
    ]
    return templates.TemplateResponse(request, "support_detail.html", common_context(
        request, session, user, "support", ticket=ticket, messages=visible_messages,
        context_items=ticket_context(ticket), overdue=ticket_overdue(ticket),
        waiting_days=ticket_waiting_days(ticket), status_class=status_class(ticket.status),
    ))


@app.post("/soporte/nuevo")
def create_support_ticket(
    request: Request,
    subject: str = Form(...),
    description: str = Form(...),
    category: str = Form("Soporte funcional"),
    request_type: str = Form("Consulta"),
    priority: str = Form("Normal"),
    desired_outcome: str = Form(""),
    due_date: str = Form(""),
    inventory_id: int | None = Form(None),
    source_id: int | None = Form(None),
    activity_data_id: int | None = Form(None),
    session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_support")
    organization_id = int(user["organization_id"])
    normalized_subject = subject.strip()
    normalized_description = description.strip()
    if len(normalized_subject) < 6 or len(normalized_description) < 12:
        set_flash(request, "Describe el asunto y la necesidad con suficiente detalle.", "error")
        return RedirectResponse("/soporte#nuevo-caso", status_code=303)
    inventory = session.scalar(select(Inventory).where(Inventory.id == inventory_id, Inventory.organization_id == organization_id)) if inventory_id else None
    source = session.scalar(select(EmissionSource).join(Inventory).where(EmissionSource.id == source_id, Inventory.organization_id == organization_id)) if source_id else None
    record = session.scalar(select(ActivityData).join(EmissionSource).join(Inventory).where(ActivityData.id == activity_data_id, Inventory.organization_id == organization_id)) if activity_data_id else None
    if inventory_id and not inventory:
        raise HTTPException(400, "Inventario no válido")
    if source_id and not source:
        raise HTTPException(400, "Fuente no válida")
    if activity_data_id and not record:
        raise HTTPException(400, "Dato no válido")
    if record and source and record.source_id != source.id:
        raise HTTPException(400, "El dato no pertenece a la fuente seleccionada")
    if source and inventory and source.inventory_id != inventory.id:
        raise HTTPException(400, "La fuente no pertenece al inventario seleccionado")
    normalized_priority = priority if priority in {"Baja", "Normal", "Alta", "Crítica"} else "Normal"
    ticket = SupportTicket(
        organization_id=organization_id,
        inventory_id=inventory.id if inventory else (source.inventory_id if source else None),
        source_id=source.id if source else (record.source_id if record else None),
        activity_data_id=record.id if record else None,
        created_by=str(user["email"]),
        request_type=request_type if request_type in {"Consulta", "Requerimiento", "Incidencia", "Decisión metodológica"} else "Consulta",
        category=category,
        priority=normalized_priority,
        subject=normalized_subject,
        description=normalized_description,
        desired_outcome=desired_outcome.strip(),
        status="Abierto",
        assigned_to=route_assignment(category),
        due_date=parse_date(due_date) if due_date else None,
        response_due_at=response_deadline(normalized_priority),
        last_message_at=datetime.now(UTC),
    )
    session.add(ticket)
    session.flush()
    ensure_reference(ticket)
    add_support_message(
        session, ticket, author_email=str(user["email"]), author_role=str(user["role"]),
        body=normalized_description, message_type="Solicitud inicial", visible_to_client=True,
    )
    notify_roles(
        session, organization_id, {"Administrador", "Consultor", "Revisor"},
        f"Nuevo requerimiento {ticket.public_reference}",
        f"{ticket.subject} · prioridad {ticket.priority}", link=f"/soporte/{ticket.id}",
        category="Soporte", priority=ticket.priority,
    )
    add_audit(session, organization_id, str(user["email"]), "CREAR", "Requerimiento", ticket.public_reference, detail=ticket.description)
    session.commit()
    set_flash(request, f"Requerimiento {ticket.public_reference} creado y asignado a {ticket.assigned_to}.")
    return RedirectResponse(f"/soporte/{ticket.id}", status_code=303)


@app.post("/soporte/{ticket_id}/mensajes")
def add_ticket_message(
    ticket_id: int,
    request: Request,
    body: str = Form(...),
    message_type: str = Form("Mensaje"),
    visible_to_client: str | None = Form(None),
    next_status: str = Form(""),
    session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_support")
    ticket = session.scalar(select(SupportTicket).where(
        SupportTicket.id == ticket_id, SupportTicket.organization_id == int(user["organization_id"]),
    ))
    if not ticket:
        raise HTTPException(404, "Caso no encontrado")
    internal = visible_to_client is None and user["role"] != "Cliente"
    if user["role"] == "Cliente":
        internal = False
        message_type = "Mensaje del cliente"
    allowed_types = {"Mensaje", "Mensaje del cliente", "Respuesta técnica", "Solicitud de información", "Nota interna", "Decisión metodológica"}
    normalized_type = message_type if message_type in allowed_types else "Mensaje"
    if internal:
        normalized_type = "Nota interna"
    try:
        add_support_message(
            session, ticket, author_email=str(user["email"]), author_role=str(user["role"]), body=body,
            message_type=normalized_type, visible_to_client=not internal,
        )
    except ValueError as exc:
        set_flash(request, str(exc), "error")
        return RedirectResponse(f"/soporte/{ticket.id}#conversacion", status_code=303)
    allowed_statuses = OPEN_STATUSES | CLOSED_STATUSES
    if next_status in allowed_statuses and user["role"] != "Cliente":
        ticket.status = next_status
    elif user["role"] == "Cliente" and ticket.status == "Esperando cliente":
        ticket.status = "En gestión"
        ticket.response_due_at = response_deadline(ticket.priority)
    ticket.closed_at = datetime.now(UTC) if ticket.status == "Cerrado" else None
    add_audit(session, int(user["organization_id"]), str(user["email"]), "MENSAJE", "Requerimiento", ticket.public_reference, detail=normalized_type)
    session.commit()
    set_flash(request, "Mensaje registrado en la conversación.")
    return RedirectResponse(f"/soporte/{ticket.id}#conversacion", status_code=303)


@app.post("/soporte/{ticket_id}/actualizar")
def update_support_ticket(
    ticket_id: int,
    request: Request,
    status: str = Form(...),
    assigned_to: str = Form(""),
    priority: str = Form("Normal"),
    due_date: str = Form(""),
    resolution: str = Form(""),
    session: Session = Depends(get_db),
):
    user = require_user(request)
    if user["role"] == "Cliente":
        raise HTTPException(403, "El cliente puede crear y responder casos, pero el equipo gestiona su estado")
    ensure_capability(user, "manage_support")
    ticket = session.scalar(select(SupportTicket).where(
        SupportTicket.id == ticket_id,
        SupportTicket.organization_id == int(user["organization_id"]),
    ))
    if not ticket:
        raise HTTPException(404, "Caso no encontrado")
    previous_status = ticket.status
    previous_priority = ticket.priority
    ticket.status = status if status in OPEN_STATUSES | CLOSED_STATUSES else ticket.status
    ticket.priority = priority if priority in {"Baja", "Normal", "Alta", "Crítica"} else ticket.priority
    ticket.assigned_to = assigned_to.strip() or ticket.assigned_to
    ticket.due_date = parse_date(due_date) if due_date else None
    normalized_resolution = resolution.strip()
    if normalized_resolution and normalized_resolution != ticket.resolution:
        ticket.resolution = normalized_resolution
        add_support_message(
            session, ticket, author_email=str(user["email"]), author_role=str(user["role"]),
            body=normalized_resolution, message_type="Respuesta técnica", visible_to_client=True,
        )
    if previous_priority != ticket.priority and ticket.status in OPEN_STATUSES:
        ticket.response_due_at = response_deadline(ticket.priority)
    ticket.closed_at = datetime.now(UTC) if ticket.status == "Cerrado" else None
    if previous_status != ticket.status:
        add_support_message(
            session, ticket, author_email=str(user["email"]), author_role=str(user["role"]),
            body=f"Estado actualizado de {previous_status} a {ticket.status}.",
            message_type="Cambio de estado", visible_to_client=True,
        )
    add_audit(
        session, int(user["organization_id"]), str(user["email"]), "ACTUALIZAR", "Requerimiento",
        ticket.public_reference, previous_value=previous_status, new_value=ticket.status,
    )
    session.commit()
    set_flash(request, f"Requerimiento {ticket.public_reference} actualizado.")
    return RedirectResponse(f"/soporte/{ticket.id}", status_code=303)


@app.get("/api/soporte/resumen")
def support_api_summary(request: Request, session: Session = Depends(get_db)):
    user = require_user(request)
    tickets = list(session.scalars(
        select(SupportTicket).where(SupportTicket.organization_id == int(user["organization_id"])).order_by(SupportTicket.updated_at.desc())
    ))
    return {
        "version": settings.version,
        "summary": support_summary(tickets),
        "recent": [
            {
                "id": ticket.id,
                "reference": ticket.public_reference,
                "subject": ticket.subject,
                "status": ticket.status,
                "priority": ticket.priority,
                "assigned_to": ticket.assigned_to,
                "overdue": ticket_overdue(ticket),
            }
            for ticket in tickets[:10]
        ],
    }


@app.get("/administracion-saas", response_class=HTMLResponse)
def saas_admin(request: Request, session: Session = Depends(get_db)):
    user = require_user(request)
    ensure_capability(user, "manage_saas")
    plans = list(session.scalars(select(ServicePlan).order_by(ServicePlan.monthly_fee)))
    subscriptions = list(session.scalars(
        select(OrganizationSubscription)
        .options(selectinload(OrganizationSubscription.organization), selectinload(OrganizationSubscription.plan))
        .order_by(OrganizationSubscription.id)
    ))
    invoices = list(session.scalars(select(BillingInvoice).order_by(BillingInvoice.issued_at.desc()).limit(50)))
    organizations = list(session.scalars(select(Organization).order_by(Organization.name)))
    summary = {
        "active": sum(1 for item in subscriptions if item.status == "Activa"),
        "trial": sum(1 for item in subscriptions if item.status == "Prueba"),
        "mrr": round(sum((item.custom_monthly_fee or (item.plan.monthly_fee if item.plan else 0)) for item in subscriptions if item.status in {"Activa", "Prueba"})),
        "organizations": len(organizations),
    }
    return templates.TemplateResponse(request, "saas_admin.html", common_context(
        request, session, user, "saas_admin", plans=plans, subscriptions=subscriptions,
        invoices=invoices, organizations=organizations, summary=summary,
    ))

@app.post("/administracion-saas/planes/nuevo")
def create_service_plan(
    request: Request,
    code: str = Form(...), name: str = Form(...), description: str = Form(""),
    monthly_fee: float = Form(0), annual_fee: float = Form(0),
    max_users: int = Form(5), max_facilities: int = Form(3), max_inventories: int = Form(3), max_storage_mb: int = Form(1024),
    includes_scope3: str | None = Form(None), includes_verification_portal: str | None = Form(None),
    session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_saas")
    normalized_code = re.sub(r"[^A-Z0-9_-]", "", code.upper())
    if not normalized_code or session.scalar(select(ServicePlan).where(ServicePlan.code == normalized_code)):
        raise HTTPException(400, "Código de plan inválido o duplicado")
    plan = ServicePlan(
        code=normalized_code, name=name.strip(), description=description.strip(), monthly_fee=max(0, monthly_fee), annual_fee=max(0, annual_fee),
        max_users=max(1, max_users), max_facilities=max(1, max_facilities), max_inventories=max(1, max_inventories), max_storage_mb=max(100, max_storage_mb),
        includes_scope3=bool(includes_scope3), includes_verification_portal=bool(includes_verification_portal), active=True,
    )
    session.add(plan)
    add_audit(session, int(user["organization_id"]), str(user["email"]), "CREAR", "Plan SaaS", plan.name, detail=plan.code)
    session.commit()
    set_flash(request, f"Plan {plan.name} creado.")
    return RedirectResponse("/administracion-saas", status_code=303)

@app.post("/administracion-saas/suscripciones/{subscription_id}/actualizar")
def admin_update_subscription(
    subscription_id: int,
    request: Request,
    plan_id: int = Form(...), status: str = Form(...), billing_cycle: str = Form("Anual"),
    custom_monthly_fee: str = Form(""), renewal_date: str = Form(""), notes: str = Form(""),
    session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_saas")
    subscription = session.get(OrganizationSubscription, subscription_id)
    plan = session.get(ServicePlan, plan_id)
    if not subscription or not plan:
        raise HTTPException(404, "Suscripción o plan no encontrado")
    subscription.plan_id = plan.id
    subscription.status = status if status in {"Prueba", "Activa", "Suspendida", "Cancelada"} else subscription.status
    subscription.billing_cycle = billing_cycle if billing_cycle in {"Mensual", "Anual"} else subscription.billing_cycle
    subscription.custom_monthly_fee = float(custom_monthly_fee) if custom_monthly_fee.strip() else None
    subscription.renewal_date = parse_date(renewal_date) if renewal_date else None
    subscription.notes = notes.strip()
    session.commit()
    set_flash(request, f"Suscripción de {subscription.organization.name} actualizada.")
    return RedirectResponse("/administracion-saas", status_code=303)

@app.post("/administracion-saas/facturas/{invoice_id}/estado")
def update_invoice_status(
    invoice_id: int, request: Request, status: str = Form(...), session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_saas")
    invoice = session.get(BillingInvoice, invoice_id)
    if not invoice:
        raise HTTPException(404, "Registro de cobro no encontrado")
    invoice.status = status if status in {"Pendiente", "Pagada", "Vencida", "Anulada", "Demostrativa"} else invoice.status
    invoice.paid_at = datetime.now(UTC) if invoice.status == "Pagada" else None
    session.commit()
    set_flash(request, f"Estado de {invoice.reference} actualizado.")
    return RedirectResponse("/administracion-saas", status_code=303)

def _lead_complexity(employees_band: str, facilities_count: int, desired_scopes: str, has_previous_inventory: bool, objective: str, urgency: str) -> tuple[int, str]:
    employee_points = {"1 a 20": 1, "21 a 50": 2, "51 a 200": 4, "Más de 200": 6}.get(employees_band, 2)
    score = employee_points + min(max(facilities_count, 1), 8)
    if "3" in desired_scopes:
        score += 3
    if has_previous_inventory:
        score += 1
    if any(keyword in objective.lower() for keyword in ("verificación", "regulator", "licitación")):
        score += 2
    if urgency == "Alta":
        score += 1
    plan = "ESENCIAL" if score <= 5 else "EMPRESARIAL" if score <= 12 else "CORPORATIVO"
    return score, plan

def _proposal_items(raw: str) -> list[str]:
    try:
        parsed = json.loads(raw or "[]")
        return [str(item) for item in parsed] if isinstance(parsed, list) else []
    except (json.JSONDecodeError, TypeError):
        return []

def _proposal_total(implementation_fee: float, recurring_fee: float, discount_amount: float, tax_rate: float) -> float:
    subtotal = max(0.0, implementation_fee) + max(0.0, recurring_fee) - max(0.0, discount_amount)
    return round(max(0.0, subtotal) * (1 + max(0.0, tax_rate) / 100), 2)

@app.get("/comercial", response_class=HTMLResponse)
def commercial_center(request: Request, session: Session = Depends(get_db)):
    user = require_user(request)
    ensure_capability(user, "manage_commercial")
    leads = list(session.scalars(select(CommercialLead).order_by(CommercialLead.created_at.desc())))
    proposals = list(session.scalars(
        select(CommercialProposal).options(selectinload(CommercialProposal.lead), selectinload(CommercialProposal.plan)).order_by(CommercialProposal.created_at.desc())
    ))
    payments = list(session.scalars(
        select(PaymentTransaction).options(selectinload(PaymentTransaction.proposal)).order_by(PaymentTransaction.created_at.desc()).limit(50)
    ))
    plans = list(session.scalars(select(ServicePlan).where(ServicePlan.active.is_(True)).order_by(ServicePlan.monthly_fee)))
    summary = {
        "leads": len(leads), "qualified": sum(1 for item in leads if item.status in {"Calificado", "Propuesta"}),
        "proposals": len(proposals), "accepted": sum(1 for item in proposals if item.status == "Aceptada"),
        "pipeline": round(sum(item.first_year_total for item in proposals if item.status in {"Borrador", "Enviada", "Vista", "Aceptada"})),
        "paid": round(sum(item.amount for item in payments if item.status == "Pagada")),
    }
    return templates.TemplateResponse(request, "commercial.html", common_context(
        request, session, user, "commercial", leads=leads, proposals=proposals, payments=payments, plans=plans, summary=summary,
    ))

@app.post("/comercial/leads/{lead_id}/estado")
def update_commercial_lead(
    lead_id: int, request: Request, status: str = Form(...), assigned_to: str = Form("Equipo comercial"),
    session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_commercial")
    lead = session.get(CommercialLead, lead_id)
    if not lead:
        raise HTTPException(404, "Prospecto no encontrado")
    valid = {"Nuevo", "Contactado", "Calificado", "Propuesta", "Ganado", "Descartado"}
    lead.status = status if status in valid else lead.status
    lead.assigned_to = assigned_to.strip() or lead.assigned_to
    session.commit()
    set_flash(request, f"Prospecto {lead.company_name} actualizado.")
    return RedirectResponse("/comercial", status_code=303)

@app.post("/comercial/propuestas/nueva")
def create_commercial_proposal(
    request: Request, lead_id: int = Form(...), plan_id: int = Form(...), title: str = Form(...),
    implementation_fee: float = Form(0), recurring_fee: float = Form(0), discount_amount: float = Form(0),
    tax_rate: float = Form(19), billing_cycle: str = Form("Anual"), valid_until: str = Form(""),
    scope: str = Form(""), deliverables: str = Form(""), terms: str = Form(""), session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_commercial")
    lead = session.get(CommercialLead, lead_id)
    plan = session.get(ServicePlan, plan_id)
    if not lead or not plan:
        raise HTTPException(404, "Prospecto o plan no encontrado")
    today = date.today()
    sequence = (session.scalar(select(func.count(CommercialProposal.id))) or 0) + 1
    reference = f"PROP-{today.year}-{sequence:04d}"
    while session.scalar(select(CommercialProposal).where(CommercialProposal.reference == reference)):
        sequence += 1
        reference = f"PROP-{today.year}-{sequence:04d}"
    proposal = CommercialProposal(
        lead_id=lead.id, plan_id=plan.id, reference=reference, public_token=secrets.token_urlsafe(24),
        title=title.strip(), company_name=lead.company_name, contact_name=lead.contact_name, contact_email=lead.email,
        status="Borrador", valid_until=parse_date(valid_until) if valid_until else None,
        billing_cycle=billing_cycle if billing_cycle in {"Mensual", "Anual"} else "Anual",
        implementation_fee=max(0, implementation_fee), recurring_fee=max(0, recurring_fee),
        discount_amount=max(0, discount_amount), tax_rate=max(0, tax_rate),
        first_year_total=_proposal_total(implementation_fee, recurring_fee, discount_amount, tax_rate),
        scope_json=json.dumps([item.strip() for item in scope.splitlines() if item.strip()], ensure_ascii=False),
        deliverables_json=json.dumps([item.strip() for item in deliverables.splitlines() if item.strip()], ensure_ascii=False),
        terms=terms.strip(), contract_version="1.0", created_by=str(user["email"]),
    )
    session.add(proposal)
    lead.status = "Propuesta"
    session.commit()
    set_flash(request, f"Propuesta {proposal.reference} creada.")
    return RedirectResponse("/comercial", status_code=303)

@app.post("/comercial/propuestas/{proposal_id}/enviar")
def send_commercial_proposal(proposal_id: int, request: Request, session: Session = Depends(get_db)):
    user = require_user(request)
    ensure_capability(user, "manage_commercial")
    proposal = session.get(CommercialProposal, proposal_id)
    if not proposal:
        raise HTTPException(404, "Propuesta no encontrada")
    proposal.status = "Enviada"
    proposal.sent_at = datetime.now(UTC)
    session.commit()
    set_flash(request, f"Propuesta {proposal.reference} marcada como enviada. Enlace público disponible.")
    return RedirectResponse("/comercial", status_code=303)

@app.get("/propuesta/{token}", response_class=HTMLResponse)
def public_proposal(token: str, request: Request, session: Session = Depends(get_db)):
    proposal = session.scalar(select(CommercialProposal).where(CommercialProposal.public_token == token).options(selectinload(CommercialProposal.plan)))
    if not proposal:
        raise HTTPException(404, "Propuesta no encontrada")
    if proposal.status == "Enviada":
        proposal.status = "Vista"
    if not proposal.viewed_at:
        proposal.viewed_at = datetime.now(UTC)
    session.commit()
    payment = session.scalar(select(PaymentTransaction).where(PaymentTransaction.proposal_id == proposal.id).order_by(PaymentTransaction.id.desc()).limit(1))
    return templates.TemplateResponse(request=request, name="public_proposal.html", context={
        "proposal": proposal, "scope_items": _proposal_items(proposal.scope_json),
        "deliverables": _proposal_items(proposal.deliverables_json), "payment": payment, "app_settings": settings,
    })

@app.post("/propuesta/{token}/aceptar")
def accept_public_proposal(
    token: str, request: Request, accepted_by: str = Form(...), accepted_email: str = Form(...),
    accept_terms: str | None = Form(None), session: Session = Depends(get_db),
):
    proposal = session.scalar(select(CommercialProposal).where(CommercialProposal.public_token == token))
    if not proposal:
        raise HTTPException(404, "Propuesta no encontrada")
    if not accept_terms:
        raise HTTPException(400, "Debes aceptar las condiciones de la propuesta")
    if proposal.valid_until and proposal.valid_until < date.today():
        proposal.status = "Vencida"
        session.commit()
        raise HTTPException(409, "La propuesta está vencida")
    timestamp = datetime.now(UTC)
    client_ip = request.client.host if request.client else "unknown"
    acceptance_source = f"{proposal.reference}|{accepted_by.strip()}|{accepted_email.strip().lower()}|{timestamp.isoformat()}|{proposal.first_year_total}|{proposal.contract_version}"
    proposal.status = "Aceptada"
    proposal.accepted_by = accepted_by.strip()
    proposal.accepted_email = accepted_email.strip().lower()
    proposal.accepted_ip = client_ip
    proposal.accepted_at = timestamp
    proposal.acceptance_hash = hashlib.sha256(acceptance_source.encode("utf-8")).hexdigest()
    payment = session.scalar(select(PaymentTransaction).where(PaymentTransaction.proposal_id == proposal.id).order_by(PaymentTransaction.id.desc()).limit(1))
    if not payment:
        payment = PaymentTransaction(
            proposal_id=proposal.id, public_token=secrets.token_urlsafe(24), gateway="Demo",
            status="Pendiente", amount=proposal.first_year_total, currency="COP",
            external_reference=f"PAY-{proposal.reference}", payer_name=proposal.accepted_by,
            payer_email=proposal.accepted_email, provider_payload='{"mode": "demo"}',
        )
        session.add(payment)
    session.commit()
    return RedirectResponse(f"/pago/{payment.public_token}", status_code=303)

@app.post("/propuesta/{token}/rechazar")
def reject_public_proposal(token: str, request: Request, reason: str = Form(""), session: Session = Depends(get_db)):
    proposal = session.scalar(select(CommercialProposal).where(CommercialProposal.public_token == token))
    if not proposal:
        raise HTTPException(404, "Propuesta no encontrada")
    proposal.status = "Rechazada"
    proposal.rejection_reason = reason.strip()
    session.commit()
    return RedirectResponse(f"/propuesta/{token}", status_code=303)

@app.get("/pago/{token}", response_class=HTMLResponse)
def public_payment(token: str, request: Request, session: Session = Depends(get_db)):
    payment = session.scalar(select(PaymentTransaction).where(PaymentTransaction.public_token == token).options(selectinload(PaymentTransaction.proposal)))
    if not payment:
        raise HTTPException(404, "Pago no encontrado")
    return templates.TemplateResponse(request=request, name="public_payment.html", context={"payment": payment, "proposal": payment.proposal, "app_settings": settings})

@app.post("/pago/{token}/confirmar")
def confirm_demo_payment(
    token: str, request: Request, payer_name: str = Form(...), payer_email: str = Form(...),
    method: str = Form("Transferencia demostrativa"), session: Session = Depends(get_db),
):
    payment = session.scalar(select(PaymentTransaction).where(PaymentTransaction.public_token == token).options(selectinload(PaymentTransaction.proposal)))
    if not payment or not payment.proposal:
        raise HTTPException(404, "Pago no encontrado")
    proposal = payment.proposal
    if proposal.status != "Aceptada":
        raise HTTPException(409, "La propuesta debe aceptarse antes del pago")
    payment.status = "Pagada"
    payment.gateway = "Demo"
    payment.payer_name = payer_name.strip()
    payment.payer_email = payer_email.strip().lower()
    payment.paid_at = datetime.now(UTC)
    payment.provider_payload = json.dumps({"mode": "demo", "method": method, "confirmed_at": payment.paid_at.isoformat()}, ensure_ascii=False)
    if not proposal.organization_id:
        base_name = proposal.company_name.strip()
        organization = session.scalar(select(Organization).where(Organization.name == base_name))
        if not organization:
            lead = session.get(CommercialLead, proposal.lead_id) if proposal.lead_id else None
            organization = Organization(
                name=base_name, trade_name=base_name, tax_id="PENDIENTE", sector=lead.sector if lead else "Por configurar",
                country="Colombia", city=lead.city if lead else "Por configurar", employees=0,
                contact_name=proposal.contact_name, contact_email=proposal.contact_email, status="Activa",
            )
            session.add(organization)
            session.flush()
        proposal.organization_id = organization.id
        subscription = session.scalar(select(OrganizationSubscription).where(OrganizationSubscription.organization_id == organization.id))
        if not subscription and proposal.plan_id:
            renewal = date(date.today().year + 1, date.today().month, min(date.today().day, 28))
            subscription = OrganizationSubscription(
                organization_id=organization.id, plan_id=proposal.plan_id, billing_cycle=proposal.billing_cycle,
                status="Activa", start_date=date.today(), renewal_date=renewal,
                notes=f"Activada desde propuesta {proposal.reference} y pago demostrativo.",
            )
            session.add(subscription)
            session.flush()
        invoice = session.scalar(select(BillingInvoice).where(BillingInvoice.reference == f"COBRO-{proposal.reference}"))
        if not invoice:
            invoice = BillingInvoice(
                organization_id=organization.id, subscription_id=subscription.id if subscription else None,
                reference=f"COBRO-{proposal.reference}", period_start=date.today(),
                period_end=date(date.today().year + 1, date.today().month, min(date.today().day, 28)),
                amount=payment.amount, status="Pagada", issued_at=date.today(), due_date=date.today(),
                paid_at=payment.paid_at, notes="Registro administrativo generado desde el pago demostrativo. No constituye factura electrónica.",
            )
            session.add(invoice)
            session.flush()
        payment.subscription_id = subscription.id if subscription else None
        payment.invoice_id = invoice.id
        onboarding_specs = [
            ("ORG-01", "Organización", "Completar información legal y operativa", 10),
            ("USR-01", "Accesos", "Invitar responsables y definir roles", 20),
            ("MET-01", "Metodología", "Aprobar metodología y límites", 30),
            ("DAT-01", "Información", "Cargar el primer conjunto de datos", 40),
            ("CAL-01", "Cálculo", "Validar el primer cálculo trazable", 50),
            ("REP-01", "Entrega", "Generar el primer informe", 60),
        ]
        for code, category, title, order in onboarding_specs:
            if not session.scalar(select(CustomerOnboardingItem).where(CustomerOnboardingItem.organization_id == organization.id, CustomerOnboardingItem.code == code)):
                session.add(CustomerOnboardingItem(
                    organization_id=organization.id, code=code, category=category, title=title,
                    description="Actividad creada automáticamente después de la contratación.", status="Pendiente",
                    owner="Cliente", display_order=order, updated_by="sistema",
                ))
    if proposal.lead_id:
        lead = session.get(CommercialLead, proposal.lead_id)
        if lead:
            lead.status = "Ganado"
    session.commit()
    return RedirectResponse(f"/pago/{token}", status_code=303)

class PaymentWebhookPayload(BaseModel):
    external_reference: str = Field(min_length=3, max_length=120)
    status: str = Field(min_length=3, max_length=30)
    amount: float = Field(ge=0)
    payer_email: str = ""

@app.post("/api/pagos/webhook")
def payment_webhook(payload: PaymentWebhookPayload, x_payment_secret: str | None = Header(None), session: Session = Depends(get_db)):
    if not settings.payment_webhook_secret or not hmac.compare_digest(x_payment_secret or "", settings.payment_webhook_secret):
        raise HTTPException(401, "Firma de pago inválida")
    payment = session.scalar(select(PaymentTransaction).where(PaymentTransaction.external_reference == payload.external_reference))
    if not payment:
        raise HTTPException(404, "Transacción no encontrada")
    if abs(payment.amount - payload.amount) > 0.01:
        raise HTTPException(409, "El valor informado no coincide")
    normalized_status = payload.status.strip().lower()
    mapping = {"paid": "Pagada", "approved": "Pagada", "pending": "Pendiente", "failed": "Fallida", "declined": "Fallida", "refunded": "Reembolsada"}
    payment.status = mapping.get(normalized_status, payload.status[:30])
    payment.payer_email = payload.payer_email.strip().lower() or payment.payer_email
    payment.paid_at = datetime.now(UTC) if payment.status == "Pagada" else payment.paid_at
    payment.provider_payload = json.dumps(payload.model_dump(), ensure_ascii=False)
    session.commit()
    return {"ok": True, "transaction_id": payment.id, "status": payment.status}

def _contract_signature_hash(contract: ServiceContract, signed_by: str, signed_email: str, signed_at: datetime) -> str:
    payload = "|".join([
        contract.reference, str(contract.organization_id), contract.version,
        contract.start_date.isoformat(), contract.end_date.isoformat() if contract.end_date else "",
        f"{contract.contract_value:.2f}", contract.billing_cycle, contract.terms_snapshot,
        signed_by.strip(), signed_email.strip().lower(), signed_at.isoformat(),
    ])
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()

def _contract_reference(session: Session) -> str:
    year = date.today().year
    current = session.scalar(select(func.count()).select_from(ServiceContract)) or 0
    return f"CTR-{year}-{int(current) + 1:04d}"

def _order_reference(session: Session) -> str:
    year = date.today().year
    current = session.scalar(select(func.count()).select_from(ServiceOrder)) or 0
    return f"OS-{year}-{int(current) + 1:04d}"

@app.get("/operacion-comercial", response_class=HTMLResponse)
def commercial_operations(request: Request, session: Session = Depends(get_db)):
    user = require_user(request)
    ensure_capability(user, "manage_commercial")
    contracts = list(session.scalars(
        select(ServiceContract)
        .options(selectinload(ServiceContract.organization), selectinload(ServiceContract.proposal))
        .order_by(ServiceContract.created_at.desc())
    ))
    orders = list(session.scalars(
        select(ServiceOrder)
        .options(selectinload(ServiceOrder.organization), selectinload(ServiceOrder.contract))
        .order_by(ServiceOrder.created_at.desc())
    ))
    invoices = list(session.scalars(select(BillingInvoice).order_by(BillingInvoice.issued_at.desc(), BillingInvoice.id.desc())))
    actions = list(session.scalars(
        select(CollectionAction)
        .options(selectinload(CollectionAction.organization), selectinload(CollectionAction.invoice))
        .order_by(CollectionAction.created_at.desc())
    ))
    documents = list(session.scalars(
        select(BillingDocumentRecord)
        .options(selectinload(BillingDocumentRecord.organization), selectinload(BillingDocumentRecord.invoice))
        .order_by(BillingDocumentRecord.created_at.desc())
    ))
    subscriptions = list(session.scalars(
        select(OrganizationSubscription)
        .options(selectinload(OrganizationSubscription.organization), selectinload(OrganizationSubscription.plan))
        .order_by(OrganizationSubscription.id)
    ))
    organizations = list(session.scalars(select(Organization).order_by(Organization.name)))
    org_map = {item.id: item for item in organizations}
    proposals = list(session.scalars(
        select(CommercialProposal)
        .where(CommercialProposal.status == "Aceptada", CommercialProposal.organization_id.is_not(None))
        .options(selectinload(CommercialProposal.organization), selectinload(CommercialProposal.plan))
        .order_by(CommercialProposal.accepted_at.desc())
    ))
    today = date.today()
    overdue = [item for item in invoices if item.status in {"Pendiente", "Vencida"} and item.due_date and item.due_date < today]
    outstanding = [item for item in invoices if item.status in {"Pendiente", "Vencida"}]
    renewals = [item for item in contracts if item.status == "Vigente" and item.end_date and 0 <= (item.end_date - today).days <= 120]
    summary = {
        "active_contracts": sum(1 for item in contracts if item.status == "Vigente"),
        "open_orders": sum(1 for item in orders if item.status not in {"Aceptada", "Cancelada"}),
        "outstanding_amount": round(sum(item.amount for item in outstanding), 2),
        "overdue_amount": round(sum(item.amount for item in overdue), 2),
        "renewals": len(renewals),
    }
    return templates.TemplateResponse(request, "commercial_operations.html", common_context(
        request, session, user, "commercial_operations",
        contracts=contracts, orders=orders, invoices=invoices, actions=actions, documents=documents,
        subscriptions=subscriptions, organizations=organizations, org_map=org_map, proposals=proposals,
        overdue=overdue, renewals=renewals, summary=summary, today=today,
    ))

@app.post("/operacion-comercial/contratos/nuevo")
def create_service_contract(
    request: Request,
    organization_id: int = Form(...), proposal_id: str = Form(""), reference: str = Form(""),
    title: str = Form(...), start_date: str = Form(...), end_date: str = Form(""),
    renewal_type: str = Form("Anual"), auto_renew: str | None = Form(None), notice_days: int = Form(30),
    contract_value: float = Form(0), billing_cycle: str = Form("Anual"), owner: str = Form("Equipo comercial"),
    terms_snapshot: str = Form(""), session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_commercial")
    organization = session.get(Organization, organization_id)
    if not organization:
        raise HTTPException(404, "Organización no encontrada")
    proposal = session.get(CommercialProposal, int(proposal_id)) if proposal_id.strip().isdigit() else None
    normalized_reference = reference.strip().upper() or _contract_reference(session)
    if session.scalar(select(ServiceContract).where(ServiceContract.reference == normalized_reference)):
        raise HTTPException(409, "La referencia contractual ya existe")
    contract = ServiceContract(
        organization_id=organization.id, proposal_id=proposal.id if proposal else None,
        reference=normalized_reference, title=title.strip(), version="1.0", status="Borrador",
        start_date=parse_date(start_date), end_date=parse_date(end_date) if end_date else None,
        renewal_type=renewal_type.strip() or "Anual", auto_renew=bool(auto_renew), notice_days=max(0, notice_days),
        contract_value=max(0, contract_value), billing_cycle=billing_cycle if billing_cycle in {"Mensual", "Anual", "Único"} else "Anual",
        owner=owner.strip() or "Equipo comercial", terms_snapshot=terms_snapshot.strip(), created_by=str(user["email"]),
    )
    session.add(contract)
    add_audit(session, organization.id, str(user["email"]), "CREAR", "Contrato de servicio", contract.reference, detail=contract.title)
    session.commit()
    set_flash(request, f"Contrato {contract.reference} creado en borrador.")
    return RedirectResponse("/operacion-comercial", status_code=303)

@app.post("/operacion-comercial/contratos/{contract_id}/firmar")
def sign_service_contract(
    contract_id: int, request: Request, signed_by: str = Form(...), signed_email: str = Form(...),
    session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_commercial")
    contract = session.get(ServiceContract, contract_id)
    if not contract:
        raise HTTPException(404, "Contrato no encontrado")
    if contract.signature_hash:
        raise HTTPException(409, "El contrato ya tiene una firma registrada")
    signed_at = datetime.now(UTC)
    contract.signed_by = signed_by.strip()
    contract.signed_email = signed_email.strip().lower()
    contract.signed_at = signed_at
    contract.signature_hash = _contract_signature_hash(contract, contract.signed_by, contract.signed_email, signed_at)
    contract.status = "Vigente"
    add_audit(session, contract.organization_id, str(user["email"]), "FIRMAR", "Contrato de servicio", contract.reference, new_value=contract.signature_hash)
    session.commit()
    set_flash(request, f"Firma contractual registrada para {contract.reference}.")
    return RedirectResponse("/operacion-comercial", status_code=303)

@app.post("/operacion-comercial/contratos/{contract_id}/estado")
def update_contract_status(
    contract_id: int, request: Request, status: str = Form(...), session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_commercial")
    contract = session.get(ServiceContract, contract_id)
    if not contract:
        raise HTTPException(404, "Contrato no encontrado")
    allowed = {"Borrador", "Vigente", "Suspendido", "Terminado", "Renovado"}
    if status not in allowed:
        raise HTTPException(400, "Estado contractual inválido")
    previous = contract.status
    contract.status = status
    add_audit(session, contract.organization_id, str(user["email"]), "ACTUALIZAR", "Contrato de servicio", contract.reference, previous_value=previous, new_value=status)
    session.commit()
    set_flash(request, f"Estado de {contract.reference} actualizado a {status}.")
    return RedirectResponse("/operacion-comercial", status_code=303)

@app.post("/operacion-comercial/contratos/{contract_id}/renovar")
def renew_service_contract(
    contract_id: int, request: Request, start_date: str = Form(...), end_date: str = Form(...),
    contract_value: float = Form(...), session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_commercial")
    contract = session.get(ServiceContract, contract_id)
    if not contract:
        raise HTTPException(404, "Contrato no encontrado")
    renewal_number = int(session.scalar(select(func.count()).select_from(ServiceContract).where(ServiceContract.parent_contract_id == contract.id)) or 0) + 1
    reference = f"{contract.reference}-R{renewal_number}"
    while session.scalar(select(ServiceContract).where(ServiceContract.reference == reference)):
        renewal_number += 1
        reference = f"{contract.reference}-R{renewal_number}"
    renewed = ServiceContract(
        organization_id=contract.organization_id, proposal_id=contract.proposal_id, parent_contract_id=contract.id,
        reference=reference, title=contract.title, version=f"{renewal_number + 1}.0", status="Borrador",
        start_date=parse_date(start_date), end_date=parse_date(end_date), renewal_type=contract.renewal_type,
        auto_renew=contract.auto_renew, notice_days=contract.notice_days, contract_value=max(0, contract_value),
        billing_cycle=contract.billing_cycle, owner=contract.owner, terms_snapshot=contract.terms_snapshot,
        created_by=str(user["email"]),
    )
    contract.status = "Renovado"
    session.add(renewed)
    add_audit(session, contract.organization_id, str(user["email"]), "RENOVAR", "Contrato de servicio", contract.reference, new_value=reference)
    session.commit()
    set_flash(request, f"Renovación {reference} creada como nueva versión contractual.")
    return RedirectResponse("/operacion-comercial", status_code=303)

@app.post("/operacion-comercial/ordenes/nueva")
def create_service_order(
    request: Request, organization_id: int = Form(...), contract_id: str = Form(""), reference: str = Form(""),
    title: str = Form(...), service_type: str = Form("Implementación"), description: str = Form(""),
    planned_start: str = Form(""), planned_end: str = Form(""), owner: str = Form("Equipo de implementación"),
    acceptance_criteria: str = Form(""), notes: str = Form(""), session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_commercial")
    organization = session.get(Organization, organization_id)
    if not organization:
        raise HTTPException(404, "Organización no encontrada")
    contract = session.get(ServiceContract, int(contract_id)) if contract_id.strip().isdigit() else None
    if contract and contract.organization_id != organization.id:
        raise HTTPException(409, "El contrato no corresponde a la organización seleccionada")
    normalized_reference = reference.strip().upper() or _order_reference(session)
    if session.scalar(select(ServiceOrder).where(ServiceOrder.reference == normalized_reference)):
        raise HTTPException(409, "La referencia de orden ya existe")
    order = ServiceOrder(
        organization_id=organization.id, contract_id=contract.id if contract else None,
        reference=normalized_reference, title=title.strip(), service_type=service_type.strip() or "Implementación",
        description=description.strip(), status="Planeada", planned_start=parse_date(planned_start) if planned_start else None,
        planned_end=parse_date(planned_end) if planned_end else None, owner=owner.strip() or "Equipo de implementación",
        acceptance_criteria=acceptance_criteria.strip(), notes=notes.strip(), created_by=str(user["email"]),
    )
    session.add(order)
    add_audit(session, organization.id, str(user["email"]), "CREAR", "Orden de servicio", order.reference, detail=order.title)
    session.commit()
    set_flash(request, f"Orden {order.reference} creada.")
    return RedirectResponse("/operacion-comercial", status_code=303)

@app.post("/operacion-comercial/ordenes/{order_id}/estado")
def update_service_order(
    order_id: int, request: Request, status: str = Form(...), notes: str = Form(""), session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_commercial")
    order = session.get(ServiceOrder, order_id)
    if not order:
        raise HTTPException(404, "Orden de servicio no encontrada")
    allowed = {"Planeada", "En ejecución", "Bloqueada", "Entregada", "Aceptada", "Cancelada"}
    if status not in allowed:
        raise HTTPException(400, "Estado de orden inválido")
    previous = order.status
    order.status = status
    order.notes = notes.strip() or order.notes
    if status in {"Entregada", "Aceptada"} and not order.delivered_at:
        order.delivered_at = datetime.now(UTC)
    if status == "Aceptada":
        order.accepted_at = datetime.now(UTC)
    add_audit(session, order.organization_id, str(user["email"]), "ACTUALIZAR", "Orden de servicio", order.reference, previous_value=previous, new_value=status)
    session.commit()
    set_flash(request, f"Orden {order.reference} actualizada a {status}.")
    return RedirectResponse("/operacion-comercial", status_code=303)

@app.post("/operacion-comercial/cobros/recurrente")
def generate_recurring_invoice(
    request: Request, subscription_id: int = Form(...), period_start: str = Form(...), period_end: str = Form(...),
    due_date: str = Form(...), reference: str = Form(""), notes: str = Form(""), session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_commercial")
    subscription = session.scalar(
        select(OrganizationSubscription)
        .where(OrganizationSubscription.id == subscription_id)
        .options(selectinload(OrganizationSubscription.plan), selectinload(OrganizationSubscription.organization))
    )
    if not subscription or not subscription.plan:
        raise HTTPException(404, "Suscripción no encontrada")
    start = parse_date(period_start)
    end = parse_date(period_end)
    due = parse_date(due_date)
    normalized_reference = reference.strip().upper() or f"REC-{subscription.organization_id}-{start.strftime('%Y%m')}-{subscription.id}"
    if session.scalar(select(BillingInvoice).where(BillingInvoice.reference == normalized_reference)):
        raise HTTPException(409, "Ya existe un cobro para esa referencia")
    if subscription.billing_cycle == "Mensual":
        amount = subscription.custom_monthly_fee or subscription.plan.monthly_fee
    else:
        amount = (subscription.custom_monthly_fee * 12) if subscription.custom_monthly_fee else subscription.plan.annual_fee
    invoice = BillingInvoice(
        organization_id=subscription.organization_id, subscription_id=subscription.id, reference=normalized_reference,
        period_start=start, period_end=end, amount=max(0, amount), status="Pendiente", issued_at=date.today(),
        due_date=due, notes=notes.strip() or "Cobro recurrente generado desde la suscripción. No constituye factura electrónica.",
    )
    session.add(invoice)
    session.flush()
    session.add(BillingDocumentRecord(
        organization_id=subscription.organization_id, invoice_id=invoice.id,
        document_type="Documento de cobro interno", internal_reference=f"DOC-{invoice.reference}",
        provider="Sin integración", status="Pendiente de integración", issued_at=invoice.issued_at,
        notes="Registro pendiente de emisión mediante proveedor tributario autorizado.", created_by=str(user["email"]),
    ))
    add_audit(session, subscription.organization_id, str(user["email"]), "GENERAR", "Cobro recurrente", invoice.reference, new_value=f"{invoice.amount:.2f}")
    session.commit()
    set_flash(request, f"Cobro {invoice.reference} generado por ${format_number(invoice.amount, 0)} COP.")
    return RedirectResponse("/operacion-comercial", status_code=303)

@app.post("/operacion-comercial/cartera/nueva")
def create_collection_action(
    request: Request, invoice_id: int = Form(...), action_type: str = Form("Recordatorio"),
    channel: str = Form("Correo"), recipient: str = Form(""), due_at: str = Form(""), notes: str = Form(""),
    session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_commercial")
    invoice = session.get(BillingInvoice, invoice_id)
    if not invoice:
        raise HTTPException(404, "Cobro no encontrado")
    action = CollectionAction(
        organization_id=invoice.organization_id, invoice_id=invoice.id, action_type=action_type.strip() or "Recordatorio",
        channel=channel.strip() or "Correo", recipient=recipient.strip(), due_at=parse_date(due_at) if due_at else None,
        status="Pendiente", notes=notes.strip(), created_by=str(user["email"]),
    )
    session.add(action)
    add_audit(session, invoice.organization_id, str(user["email"]), "CREAR", "Gestión de cartera", invoice.reference, detail=action.action_type)
    session.commit()
    set_flash(request, "Gestión de cartera programada.")
    return RedirectResponse("/operacion-comercial", status_code=303)

@app.post("/operacion-comercial/cartera/{action_id}/completar")
def complete_collection_action(
    action_id: int, request: Request, result: str = Form(...), invoice_status: str = Form(""),
    session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_commercial")
    action = session.get(CollectionAction, action_id)
    if not action:
        raise HTTPException(404, "Gestión de cartera no encontrada")
    action.status = "Completada"
    action.result = result.strip()
    action.completed_at = datetime.now(UTC)
    if invoice_status in {"Pendiente", "Pagada", "Vencida", "Anulada"}:
        action.invoice.status = invoice_status
        if invoice_status == "Pagada":
            action.invoice.paid_at = datetime.now(UTC)
    add_audit(session, action.organization_id, str(user["email"]), "COMPLETAR", "Gestión de cartera", action.invoice.reference, detail=action.result)
    session.commit()
    set_flash(request, "Gestión de cartera completada.")
    return RedirectResponse("/operacion-comercial", status_code=303)

@app.post("/operacion-comercial/documentos/{document_id}/actualizar")
def update_billing_document(
    document_id: int, request: Request, status: str = Form(...), provider: str = Form(""),
    external_number: str = Form(""), issued_at: str = Form(""), cufe: str = Form(""),
    document_url: str = Form(""), notes: str = Form(""), session: Session = Depends(get_db),
):
    user = require_user(request)
    ensure_capability(user, "manage_commercial")
    document = session.get(BillingDocumentRecord, document_id)
    if not document:
        raise HTTPException(404, "Documento de cobro no encontrado")
    allowed = {"Borrador", "Pendiente de integración", "Emitido externamente", "Rechazado", "Anulado"}
    if status not in allowed:
        raise HTTPException(400, "Estado documental inválido")
    document.status = status
    document.provider = provider.strip() or document.provider
    document.external_number = external_number.strip()
    document.issued_at = parse_date(issued_at) if issued_at else document.issued_at
    document.cufe = cufe.strip()
    document.document_url = document_url.strip()
    document.notes = notes.strip()
    add_audit(session, document.organization_id, str(user["email"]), "ACTUALIZAR", "Documento de cobro", document.internal_reference, new_value=document.status)
    session.commit()
    set_flash(request, f"Documento {document.internal_reference} actualizado.")
    return RedirectResponse("/operacion-comercial", status_code=303)

def _require_customer_success_view(user: dict[str, object]) -> None:
    capabilities = user.get("capabilities", set())
    if "view_customer_success" not in capabilities and "manage_customer_success" not in capabilities:
        raise HTTPException(403, "Tu rol no tiene acceso a éxito del cliente")

@app.get("/exito-cliente", response_class=HTMLResponse)
def customer_success_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    _require_customer_success_view(user)
    organization_id = int(user["organization_id"])
    profile = session.scalar(select(CustomerSuccessProfile).where(CustomerSuccessProfile.organization_id == organization_id))
    snapshot = session.scalar(
        select(AccountHealthSnapshot)
        .where(AccountHealthSnapshot.organization_id == organization_id)
        .order_by(AccountHealthSnapshot.calculated_at.desc())
    )
    if not snapshot:
        snapshot = refresh_account_health(session, organization_id, str(user["email"]))
        sync_renewal_opportunity(session, organization_id, snapshot, str(user["email"]))
        session.commit()
    metrics = account_metrics(session, organization_id)
    milestones = list(session.scalars(
        select(ValueMilestone)
        .where(ValueMilestone.organization_id == organization_id)
        .options(selectinload(ValueMilestone.inventory))
        .order_by(ValueMilestone.target_date, ValueMilestone.id)
    ))
    commitments = list(session.scalars(
        select(SuccessCommitment)
        .where(SuccessCommitment.organization_id == organization_id)
        .order_by(SuccessCommitment.status, SuccessCommitment.due_date, SuccessCommitment.id)
    ))
    renewal = session.scalar(
        select(RenewalOpportunity)
        .where(RenewalOpportunity.organization_id == organization_id)
        .options(selectinload(RenewalOpportunity.contract))
        .order_by(RenewalOpportunity.renewal_date)
    )
    history = list(session.scalars(
        select(AccountHealthSnapshot)
        .where(AccountHealthSnapshot.organization_id == organization_id)
        .order_by(AccountHealthSnapshot.calculated_at.desc())
        .limit(8)
    ))
    inventories = list(session.scalars(
        select(Inventory).where(Inventory.organization_id == organization_id).order_by(Inventory.start_date.desc())
    ))
    portfolio = []
    if user.get("can_manage_customer_success"):
        for item in user.get("organizations", []):
            org_id = int(item["id"])
            latest = session.scalar(
                select(AccountHealthSnapshot)
                .where(AccountHealthSnapshot.organization_id == org_id)
                .order_by(AccountHealthSnapshot.calculated_at.desc())
            )
            if latest:
                portfolio.append({"organization": item, "snapshot": latest})
    return templates.TemplateResponse(
        request=request,
        name="customer_success.html",
        context=common_context(
            request, session, user, "customer_success",
            profile=profile, snapshot=snapshot, metrics=metrics, milestones=milestones,
            commitments=commitments, renewal=renewal, history=history, inventories=inventories,
            portfolio=portfolio,
        ),
    )

@app.post("/exito-cliente/perfil")
def customer_success_profile_update(
    request: Request,
    lifecycle_stage: str = Form("Adopción"), owner: str = Form("Equipo de éxito del cliente"),
    executive_sponsor: str = Form(""), sponsor_email: str = Form(""),
    primary_objective: str = Form(""), success_plan: str = Form(""),
    risk_override: str = Form(""), risk_reason: str = Form(""),
    last_business_review: str = Form(""), next_business_review: str = Form(""),
    satisfaction_score: str = Form(""), nps_score: str = Form(""),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_customer_success")
    organization_id = int(user["organization_id"])
    profile = session.scalar(select(CustomerSuccessProfile).where(CustomerSuccessProfile.organization_id == organization_id))
    if not profile:
        profile = CustomerSuccessProfile(organization_id=organization_id)
        session.add(profile)
    allowed_stages = {"Implementación", "Adopción", "Valor", "Renovación", "Expansión", "En riesgo"}
    allowed_risks = {"", "Sano", "Atención", "Riesgo", "Crítico"}
    if lifecycle_stage not in allowed_stages or risk_override not in allowed_risks:
        raise HTTPException(400, "Etapa o nivel de riesgo inválido")
    profile.lifecycle_stage = lifecycle_stage
    profile.owner = owner.strip() or "Equipo de éxito del cliente"
    profile.executive_sponsor = executive_sponsor.strip()
    profile.sponsor_email = sponsor_email.strip().lower()
    profile.primary_objective = primary_objective.strip()
    profile.success_plan = success_plan.strip()
    profile.risk_override = risk_override
    profile.risk_reason = risk_reason.strip()
    profile.last_business_review = parse_date(last_business_review) if last_business_review else None
    profile.next_business_review = parse_date(next_business_review) if next_business_review else None
    profile.satisfaction_score = float(satisfaction_score) if satisfaction_score else None
    profile.nps_score = int(nps_score) if nps_score else None
    if profile.satisfaction_score is not None and not 1 <= profile.satisfaction_score <= 5:
        raise HTTPException(400, "La satisfacción debe estar entre 1 y 5")
    if profile.nps_score is not None and not 0 <= profile.nps_score <= 10:
        raise HTTPException(400, "El NPS relacional debe estar entre 0 y 10")
    add_audit(session, organization_id, str(user["email"]), "ACTUALIZAR", "Éxito del cliente", "Perfil de cuenta", new_value=lifecycle_stage)
    session.commit()
    set_flash(request, "Perfil de éxito del cliente actualizado.")
    return RedirectResponse("/exito-cliente", status_code=303)

@app.post("/exito-cliente/salud/recalcular")
def customer_success_recalculate(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_customer_success")
    organization_id = int(user["organization_id"])
    snapshot = refresh_account_health(session, organization_id, str(user["email"]))
    renewal = sync_renewal_opportunity(session, organization_id, snapshot, str(user["email"]))
    add_audit(
        session, organization_id, str(user["email"]), "RECALCULAR", "Salud de cuenta", str(snapshot.id),
        new_value=f"{snapshot.overall_score} · {snapshot.risk_level}",
        detail=f"Renovación: {renewal.probability}%" if renewal else "Sin contrato vigente",
    )
    session.commit()
    set_flash(request, f"Salud recalculada: {snapshot.overall_score}/100 · {snapshot.risk_level}.")
    return RedirectResponse("/exito-cliente", status_code=303)

@app.post("/exito-cliente/hitos/nuevo")
def customer_success_milestone_create(
    request: Request, title: str = Form(...), category: str = Form("Resultado climático"),
    inventory_id: str = Form(""), owner: str = Form("Equipo de éxito del cliente"),
    target_date: str = Form(""), expected_value: float = Form(0), realized_value: float = Form(0),
    unit: str = Form(""), status: str = Form("Planeado"), evidence_note: str = Form(""),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_customer_success")
    organization_id = int(user["organization_id"])
    inventory = session.get(Inventory, int(inventory_id)) if inventory_id.strip().isdigit() else None
    if inventory and inventory.organization_id != organization_id:
        raise HTTPException(409, "El inventario no corresponde a la organización activa")
    allowed = {"Planeado", "En progreso", "Completado", "Cancelado"}
    if status not in allowed:
        raise HTTPException(400, "Estado de hito inválido")
    milestone = ValueMilestone(
        organization_id=organization_id, inventory_id=inventory.id if inventory else None,
        title=title.strip(), category=category.strip() or "Resultado climático", owner=owner.strip(),
        target_date=parse_date(target_date) if target_date else None,
        expected_value=max(0, expected_value), realized_value=max(0, realized_value), unit=unit.strip(),
        status=status, evidence_note=evidence_note.strip(), created_by=str(user["email"]),
        completed_at=datetime.now(UTC) if status == "Completado" else None,
    )
    session.add(milestone)
    add_audit(session, organization_id, str(user["email"]), "CREAR", "Hito de valor", milestone.title, new_value=status)
    session.commit()
    set_flash(request, "Hito de valor creado.")
    return RedirectResponse("/exito-cliente", status_code=303)

@app.post("/exito-cliente/hitos/{milestone_id}/estado")
def customer_success_milestone_update(
    milestone_id: int, request: Request, status: str = Form(...), realized_value: float = Form(0),
    evidence_note: str = Form(""), session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_customer_success")
    milestone = session.get(ValueMilestone, milestone_id)
    if not milestone or milestone.organization_id != int(user["organization_id"]):
        raise HTTPException(404, "Hito no encontrado")
    allowed = {"Planeado", "En progreso", "Completado", "Cancelado"}
    if status not in allowed:
        raise HTTPException(400, "Estado de hito inválido")
    previous = milestone.status
    milestone.status = status
    milestone.realized_value = max(0, realized_value)
    milestone.evidence_note = evidence_note.strip() or milestone.evidence_note
    milestone.completed_at = datetime.now(UTC) if status == "Completado" else None
    add_audit(session, milestone.organization_id, str(user["email"]), "ACTUALIZAR", "Hito de valor", milestone.title, previous_value=previous, new_value=status)
    session.commit()
    set_flash(request, "Hito actualizado.")
    return RedirectResponse("/exito-cliente", status_code=303)

@app.post("/exito-cliente/compromisos/nuevo")
def customer_success_commitment_create(
    request: Request, title: str = Form(...), description: str = Form(""), owner: str = Form(""),
    due_date: str = Form(""), priority: str = Form("Media"), source: str = Form("Plan de éxito"),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_customer_success")
    organization_id = int(user["organization_id"])
    commitment = SuccessCommitment(
        organization_id=organization_id, title=title.strip(), description=description.strip(),
        owner=owner.strip() or "Equipo de éxito del cliente", due_date=parse_date(due_date) if due_date else None,
        priority=priority if priority in {"Baja", "Media", "Alta", "Crítica"} else "Media",
        status="Pendiente", source=source.strip() or "Plan de éxito", created_by=str(user["email"]),
    )
    session.add(commitment)
    add_audit(session, organization_id, str(user["email"]), "CREAR", "Compromiso de éxito", commitment.title, new_value=commitment.priority)
    session.commit()
    set_flash(request, "Compromiso creado.")
    return RedirectResponse("/exito-cliente", status_code=303)

@app.post("/exito-cliente/compromisos/{commitment_id}/estado")
def customer_success_commitment_update(
    commitment_id: int, request: Request, status: str = Form(...),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_customer_success")
    commitment = session.get(SuccessCommitment, commitment_id)
    if not commitment or commitment.organization_id != int(user["organization_id"]):
        raise HTTPException(404, "Compromiso no encontrado")
    allowed = {"Pendiente", "En progreso", "Completado", "Bloqueado", "Cancelado"}
    if status not in allowed:
        raise HTTPException(400, "Estado de compromiso inválido")
    previous = commitment.status
    commitment.status = status
    commitment.completed_at = datetime.now(UTC) if status == "Completado" else None
    add_audit(session, commitment.organization_id, str(user["email"]), "ACTUALIZAR", "Compromiso de éxito", commitment.title, previous_value=previous, new_value=status)
    session.commit()
    set_flash(request, "Compromiso actualizado.")
    return RedirectResponse("/exito-cliente", status_code=303)

@app.post("/exito-cliente/renovacion/{renewal_id}/actualizar")
def customer_success_renewal_update(
    renewal_id: int, request: Request, status: str = Form(...), probability: int = Form(...),
    strategy: str = Form(""), blockers: str = Form(""), next_action: str = Form(""),
    next_action_date: str = Form(""), decision_notes: str = Form(""),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_customer_success")
    renewal = session.get(RenewalOpportunity, renewal_id)
    if not renewal or renewal.organization_id != int(user["organization_id"]):
        raise HTTPException(404, "Oportunidad de renovación no encontrada")
    allowed = {"Por preparar", "Bien encaminada", "En riesgo", "Propuesta enviada", "Renovada", "No renovada"}
    if status not in allowed or not 0 <= probability <= 100:
        raise HTTPException(400, "Estado o probabilidad inválidos")
    previous = f"{renewal.status} · {renewal.probability}%"
    renewal.status = status
    renewal.probability = probability
    renewal.strategy = strategy.strip()
    renewal.blockers = blockers.strip()
    renewal.next_action = next_action.strip()
    renewal.next_action_date = parse_date(next_action_date) if next_action_date else None
    renewal.decision_notes = decision_notes.strip()
    renewal.updated_by = str(user["email"])
    add_audit(session, renewal.organization_id, str(user["email"]), "ACTUALIZAR", "Renovación", str(renewal.id), previous_value=previous, new_value=f"{status} · {probability}%")
    session.commit()
    set_flash(request, "Estrategia de renovación actualizada.")
    return RedirectResponse("/exito-cliente", status_code=303)

def _require_impact_view(user: dict[str, object]) -> None:
    capabilities = user.get("capabilities", set())
    if "view_impact" not in capabilities and "manage_impact" not in capabilities:
        raise HTTPException(403, "Tu rol no tiene acceso a inteligencia de impacto")

@app.get("/inteligencia-impacto", response_class=HTMLResponse)
def impact_intelligence_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    _require_impact_view(user)
    organization_id = int(user["organization_id"])
    metrics = impact_metrics(session, organization_id)
    snapshot = session.scalar(
        select(ImpactSnapshot).where(ImpactSnapshot.organization_id == organization_id).order_by(ImpactSnapshot.calculated_at.desc())
    )
    if not snapshot:
        snapshot = refresh_impact_snapshot(session, organization_id, created_by=str(user["email"]))
        session.commit()
    references = list(session.scalars(
        select(BenchmarkReference).where(BenchmarkReference.organization_id == organization_id, BenchmarkReference.status == "Activo").order_by(BenchmarkReference.metric_name)
    ))
    comparisons = compare_benchmarks(metrics, references)
    history = list(session.scalars(
        select(ImpactSnapshot).where(ImpactSnapshot.organization_id == organization_id).order_by(ImpactSnapshot.calculated_at.desc()).limit(12)
    ))
    organization_ids = [int(item["id"]) for item in user.get("organizations", [])]
    portfolio = portfolio_comparison(session, organization_ids or [organization_id], organization_id)
    return templates.TemplateResponse(
        request=request, name="impact_intelligence.html",
        context=common_context(request, session, user, "impact", metrics=metrics, snapshot=snapshot, references=references, comparisons=comparisons, history=history, portfolio=portfolio),
    )

@app.post("/inteligencia-impacto/recalcular")
def recalculate_impact(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_impact")
    snapshot = refresh_impact_snapshot(session, int(user["organization_id"]), created_by=str(user["email"]))
    add_audit(session, int(user["organization_id"]), str(user["email"]), "RECALCULAR", "Analítica de impacto", str(snapshot.id), new_value=f"Puntaje {snapshot.impact_score}/100")
    session.commit()
    set_flash(request, "Analítica de impacto actualizada.")
    return RedirectResponse("/inteligencia-impacto", status_code=303)

@app.post("/inteligencia-impacto/benchmarks/nuevo")
def create_benchmark(
    request: Request, name: str = Form(...), metric_code: str = Form(...), metric_name: str = Form(...),
    period_label: str = Form("Referencia"), unit: str = Form(...), median_value: float = Form(...),
    top_quartile_value: float = Form(...), lower_is_better: str = Form("true"), source_type: str = Form("Referencia interna"),
    source_reference: str = Form(""), confidence_level: str = Form("Media"), notes: str = Form(""),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_impact")
    allowed_metrics = {"intensity_employee", "intensity_revenue_billion", "intensity_production", "quality_score", "evidence_coverage"}
    if metric_code not in allowed_metrics or median_value < 0 or top_quartile_value < 0:
        raise HTTPException(400, "Métrica o valores de referencia inválidos")
    org = session.get(Organization, int(user["organization_id"]))
    reference = BenchmarkReference(
        organization_id=org.id, name=name.strip(), sector=org.sector, metric_code=metric_code, metric_name=metric_name.strip(),
        period_label=period_label.strip(), unit=unit.strip(), median_value=median_value, top_quartile_value=top_quartile_value,
        lower_is_better=lower_is_better.lower() in {"1", "true", "si", "sí", "on"}, source_type=source_type.strip(),
        source_reference=source_reference.strip(), confidence_level=confidence_level, notes=notes.strip(), created_by=str(user["email"]),
    )
    session.add(reference)
    add_audit(session, org.id, str(user["email"]), "CREAR", "Benchmark", reference.name, f"{metric_name}: mediana {median_value}; cuartil {top_quartile_value}")
    session.commit()
    set_flash(request, "Referencia de benchmark registrada.")
    return RedirectResponse("/inteligencia-impacto", status_code=303)

@app.post("/inteligencia-impacto/benchmarks/{reference_id}/estado")
def update_benchmark_status(reference_id: int, request: Request, status: str = Form(...), session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "manage_impact")
    reference = session.scalar(select(BenchmarkReference).where(BenchmarkReference.id == reference_id, BenchmarkReference.organization_id == int(user["organization_id"])))
    if not reference:
        raise HTTPException(404, "Benchmark no encontrado")
    if status not in {"Activo", "Archivado"}:
        raise HTTPException(400, "Estado inválido")
    reference.status = status
    add_audit(session, reference.organization_id, str(user["email"]), "ACTUALIZAR", "Benchmark", reference.name, new_value=status)
    session.commit()
    set_flash(request, "Estado del benchmark actualizado.")
    return RedirectResponse("/inteligencia-impacto", status_code=303)

@app.get("/inteligencia-impacto/exportar.xlsx")
def export_impact_intelligence(session: Session = Depends(get_db), user: dict = Depends(require_user)):
    _require_impact_view(user)
    organization_id = int(user["organization_id"])
    metrics = impact_metrics(session, organization_id)
    references = list(session.scalars(select(BenchmarkReference).where(BenchmarkReference.organization_id == organization_id)))
    comparisons = compare_benchmarks(metrics, references)
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen de impacto"
    ws.append(["Métrica", "Valor", "Unidad"])
    units = {"total_emissions": "tCO2e", "intensity_employee": "tCO2e/empleado", "intensity_revenue_billion": "tCO2e/mil millones COP", "quality_score": "%", "evidence_coverage": "%", "expected_reduction": "tCO2e", "annual_savings": "COP/año", "value_per_tonne": "COP/tCO2e"}
    for key in ["total_emissions", "intensity_employee", "intensity_revenue_billion", "quality_score", "evidence_coverage", "expected_reduction", "annual_savings", "value_per_tonne"]:
        ws.append([key, metrics.get(key, 0), units[key]])
    ws2 = wb.create_sheet("Benchmark")
    ws2.append(["Referencia", "Métrica", "Actual", "Mediana", "Cuartil superior", "Unidad", "Estado", "Fuente", "Confianza"])
    for row in comparisons:
        ref = row["reference"]
        ws2.append([ref.name, ref.metric_name, row["current"], ref.median_value, ref.top_quartile_value, ref.unit, row["status"], ref.source_reference, ref.confidence_level])
    buffer = BytesIO()
    wb.save(buffer)
    filename = f"inteligencia_impacto_{organization_id}.xlsx"
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

def _require_climate_risk_view(user: dict[str, object]) -> None:
    capabilities = user["capabilities"]
    if "view_climate_risk" not in capabilities and "manage_climate_risk" not in capabilities:
        raise HTTPException(403, "Tu rol no tiene acceso a riesgos climáticos")

@app.get("/riesgos-climaticos", response_class=HTMLResponse)
def climate_risk_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    _require_climate_risk_view(user)
    organization_id = int(user["organization_id"])
    summary = assessment_summary(session, organization_id)
    inventories = list(session.scalars(select(Inventory).where(Inventory.organization_id == organization_id).order_by(Inventory.start_date.desc())))
    return templates.TemplateResponse(
        request=request, name="climate_risk.html",
        context=common_context(request, session, user, "climate_risk", summary=summary, inventories=inventories, risk_level=risk_level),
    )

@app.post("/riesgos-climaticos/evaluacion")
def climate_assessment_save(
    request: Request, name: str = Form(...), inventory_id: str = Form(""), methodology: str = Form("Análisis corporativo de escenarios"),
    scenario: str = Form("Escenario central"), base_year: int = Form(...), short_horizon: int = Form(...), medium_horizon: int = Form(...),
    long_horizon: int = Form(...), currency: str = Form("COP"), owner: str = Form(...), status: str = Form("En evaluación"), notes: str = Form(""),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_climate_risk")
    organization_id = int(user["organization_id"])
    if not (base_year <= short_horizon <= medium_horizon <= long_horizon):
        raise HTTPException(400, "Los horizontes deben estar ordenados desde el año base hasta el largo plazo")
    summary = assessment_summary(session, organization_id)
    assessment = summary["assessment"] or ClimateRiskAssessment(organization_id=organization_id, name=name.strip(), created_by=str(user["email"]))
    if not summary["assessment"]:
        session.add(assessment)
    assessment.name = name.strip(); assessment.inventory_id = int(inventory_id) if inventory_id else None
    assessment.methodology = methodology.strip(); assessment.scenario = scenario.strip(); assessment.base_year = base_year
    assessment.short_horizon = short_horizon; assessment.medium_horizon = medium_horizon; assessment.long_horizon = long_horizon
    assessment.currency = currency.strip().upper()[:20]; assessment.owner = owner.strip(); assessment.status = status; assessment.notes = notes.strip()
    refresh_assessment_status(session, assessment, str(user["email"])) if assessment.id else None
    add_audit(session, organization_id, str(user["email"]), "ACTUALIZAR", "Evaluación climática", assessment.name, new_value=assessment.status)
    session.commit(); set_flash(request, "Evaluación climática guardada.")
    return RedirectResponse("/riesgos-climaticos", status_code=303)

@app.post("/riesgos-climaticos/riesgos/nuevo")
def climate_risk_create(
    request: Request, risk_type: str = Form(...), category: str = Form(...), hazard: str = Form(...), description: str = Form(""),
    location: str = Form("Corporativo"), value_chain_stage: str = Form("Operación propia"), time_horizon: str = Form("Mediano plazo"),
    scenario: str = Form("Escenario central"), likelihood: int = Form(...), financial_impact: int = Form(...), operational_impact: int = Form(...),
    reputational_impact: int = Form(...), control_effectiveness: int = Form(0), financial_exposure: float = Form(0), owner: str = Form(...),
    response_strategy: str = Form("Mitigar"), response_detail: str = Form(""), status: str = Form("Abierto"), source_reference: str = Form(""),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_climate_risk")
    organization_id = int(user["organization_id"]); summary = assessment_summary(session, organization_id)
    assessment = summary["assessment"]
    if not assessment: raise HTTPException(409, "Primero crea la evaluación climática")
    if risk_type not in {"Físico", "Transición", "Oportunidad"}: raise HTTPException(400, "Tipo de riesgo inválido")
    likelihood = max(1, min(5, likelihood)); financial_impact = max(1, min(5, financial_impact))
    operational_impact = max(1, min(5, operational_impact)); reputational_impact = max(1, min(5, reputational_impact))
    inherent, residual = calculate_risk_scores(likelihood, financial_impact, operational_impact, reputational_impact, control_effectiveness)
    risk = ClimateRisk(
        assessment_id=assessment.id, organization_id=organization_id, risk_type=risk_type, category=category.strip(), hazard=hazard.strip(),
        description=description.strip(), location=location.strip(), value_chain_stage=value_chain_stage.strip(), time_horizon=time_horizon,
        scenario=scenario.strip(), likelihood=likelihood, financial_impact=financial_impact, operational_impact=operational_impact,
        reputational_impact=reputational_impact, inherent_score=inherent, control_effectiveness=max(0, min(100, control_effectiveness)),
        residual_score=residual, financial_exposure=max(0, financial_exposure), owner=owner.strip(), response_strategy=response_strategy,
        response_detail=response_detail.strip(), status=status, source_reference=source_reference.strip(), created_by=str(user["email"]),
    )
    session.add(risk); add_audit(session, organization_id, str(user["email"]), "CREAR", "Riesgo climático", risk.hazard, new_value=f"{risk_type} · {risk_level(residual)}")
    session.commit(); set_flash(request, "Riesgo climático registrado.")
    return RedirectResponse("/riesgos-climaticos", status_code=303)

@app.post("/riesgos-climaticos/riesgos/{risk_id}/actualizar")
def climate_risk_update(
    risk_id: int, request: Request, likelihood: int = Form(...), financial_impact: int = Form(...), operational_impact: int = Form(...),
    reputational_impact: int = Form(...), financial_exposure: float = Form(0), owner: str = Form(...), response_strategy: str = Form(...),
    response_detail: str = Form(""), status: str = Form(...), source_reference: str = Form(""),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_climate_risk")
    risk = session.get(ClimateRisk, risk_id)
    if not risk or risk.organization_id != int(user["organization_id"]): raise HTTPException(404, "Riesgo no encontrado")
    risk.likelihood=max(1,min(5,likelihood)); risk.financial_impact=max(1,min(5,financial_impact)); risk.operational_impact=max(1,min(5,operational_impact)); risk.reputational_impact=max(1,min(5,reputational_impact))
    risk.financial_exposure=max(0, financial_exposure); risk.owner=owner.strip(); risk.response_strategy=response_strategy; risk.response_detail=response_detail.strip(); risk.status=status; risk.source_reference=source_reference.strip()
    synchronize_control_effectiveness(session, risk)
    add_audit(session, risk.organization_id, str(user["email"]), "ACTUALIZAR", "Riesgo climático", risk.hazard, new_value=f"Residual {risk.residual_score}")
    session.commit(); set_flash(request, "Riesgo actualizado.")
    return RedirectResponse("/riesgos-climaticos", status_code=303)

@app.post("/riesgos-climaticos/controles/nuevo")
def climate_control_create(
    request: Request, risk_id: int = Form(...), name: str = Form(...), control_type: str = Form(...), owner: str = Form(...),
    status: str = Form(...), effectiveness: int = Form(...), implementation_date: str = Form(""), next_review: str = Form(""),
    annual_cost: float = Form(0), evidence: str = Form(""), session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_climate_risk")
    organization_id = int(user["organization_id"]); risk = session.get(ClimateRisk, risk_id)
    if not risk or risk.organization_id != organization_id: raise HTTPException(404, "Riesgo no encontrado")
    control = ClimateRiskControl(risk_id=risk.id, organization_id=organization_id, name=name.strip(), control_type=control_type,
        owner=owner.strip(), status=status, effectiveness=max(0, min(100, effectiveness)),
        implementation_date=parse_date(implementation_date) if implementation_date else None, next_review=parse_date(next_review) if next_review else None,
        annual_cost=max(0, annual_cost), evidence=evidence.strip(), created_by=str(user["email"]))
    session.add(control); session.flush(); synchronize_control_effectiveness(session, risk)
    add_audit(session, organization_id, str(user["email"]), "CREAR", "Control climático", control.name, new_value=f"Efectividad {control.effectiveness}%")
    session.commit(); set_flash(request, "Control registrado y riesgo residual recalculado.")
    return RedirectResponse("/riesgos-climaticos", status_code=303)

@app.post("/riesgos-climaticos/hoja-ruta")
def climate_roadmap_save(
    request: Request, name: str = Form(...), baseline_year: int = Form(...), target_year: int = Form(...), owner: str = Form(...),
    governance: str = Form(""), approved_budget: float = Form(0), status: str = Form(...), notes: str = Form(""),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_climate_risk")
    organization_id = int(user["organization_id"]); summary=assessment_summary(session, organization_id); assessment=summary["assessment"]
    if target_year < baseline_year: raise HTTPException(400, "El año objetivo no puede ser anterior al año base")
    if not assessment: raise HTTPException(409, "Primero crea la evaluación climática")
    roadmap = summary["roadmap"] or ClimateTransitionRoadmap(organization_id=organization_id, assessment_id=assessment.id, name=name.strip(), created_by=str(user["email"]))
    if not summary["roadmap"]: session.add(roadmap)
    roadmap.name=name.strip(); roadmap.baseline_year=baseline_year; roadmap.target_year=target_year; roadmap.owner=owner.strip(); roadmap.governance=governance.strip(); roadmap.approved_budget=max(0, approved_budget); roadmap.status=status; roadmap.notes=notes.strip()
    add_audit(session, organization_id, str(user["email"]), "ACTUALIZAR", "Hoja de ruta climática", roadmap.name, new_value=roadmap.status)
    session.commit(); set_flash(request, "Hoja de ruta guardada.")
    return RedirectResponse("/riesgos-climaticos", status_code=303)

@app.post("/riesgos-climaticos/acciones/nueva")
def climate_action_create(
    request: Request, risk_id: str = Form(""), category: str = Form(...), title: str = Form(...), description: str = Form(""), owner: str = Form(...),
    start_date: str = Form(""), end_date: str = Form(""), priority: str = Form("Media"), status: str = Form("Planeada"), progress: int = Form(0),
    expected_reduction_tco2e: float = Form(0), capex: float = Form(0), annual_opex: float = Form(0), annual_savings: float = Form(0),
    avoided_loss: float = Form(0), indicator: str = Form(""), target_value: float = Form(0), current_value: float = Form(0), unit: str = Form(""),
    dependencies: str = Form(""), evidence_note: str = Form(""), session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_climate_risk")
    organization_id=int(user["organization_id"]); summary=assessment_summary(session, organization_id); roadmap=summary["roadmap"]
    parsed_start = parse_date(start_date) if start_date else None; parsed_end = parse_date(end_date) if end_date else None
    if parsed_start and parsed_end and parsed_end < parsed_start: raise HTTPException(400, "La fecha final no puede ser anterior a la fecha inicial")
    if not roadmap: raise HTTPException(409, "Primero crea la hoja de ruta")
    linked_risk = session.get(ClimateRisk, int(risk_id)) if risk_id else None
    if linked_risk and linked_risk.organization_id != organization_id: raise HTTPException(404, "Riesgo no encontrado")
    action=ClimateTransitionAction(roadmap_id=roadmap.id, organization_id=organization_id, risk_id=linked_risk.id if linked_risk else None,
        category=category.strip(), title=title.strip(), description=description.strip(), owner=owner.strip(), start_date=parsed_start,
        end_date=parsed_end, priority=priority, status=status, progress=max(0,min(100,progress)), expected_reduction_tco2e=max(0,expected_reduction_tco2e),
        capex=max(0,capex), annual_opex=max(0,annual_opex), annual_savings=max(0,annual_savings), avoided_loss=max(0,avoided_loss), indicator=indicator.strip(),
        target_value=target_value, current_value=current_value, unit=unit.strip(), dependencies=dependencies.strip(), evidence_note=evidence_note.strip(), created_by=str(user["email"]))
    session.add(action); add_audit(session, organization_id, str(user["email"]), "CREAR", "Acción climática", action.title, new_value=action.status)
    session.commit(); set_flash(request, "Acción añadida a la hoja de ruta.")
    return RedirectResponse("/riesgos-climaticos", status_code=303)

@app.post("/riesgos-climaticos/acciones/{action_id}/estado")
def climate_action_update(
    action_id: int, request: Request, status: str = Form(...), progress: int = Form(...), current_value: float = Form(0), evidence_note: str = Form(""),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_climate_risk")
    action=session.get(ClimateTransitionAction, action_id)
    if not action or action.organization_id != int(user["organization_id"]): raise HTTPException(404, "Acción no encontrada")
    action.status=status; action.progress=max(0,min(100,progress)); action.current_value=current_value; action.evidence_note=evidence_note.strip()
    add_audit(session, action.organization_id, str(user["email"]), "ACTUALIZAR", "Acción climática", action.title, new_value=f"{status} · {action.progress}%")
    session.commit(); set_flash(request, "Avance de la acción actualizado.")
    return RedirectResponse("/riesgos-climaticos", status_code=303)

@app.get("/riesgos-climaticos/exportar.xlsx")
def climate_risk_export(session: Session = Depends(get_db), user: dict = Depends(require_user)):
    _require_climate_risk_view(user); organization_id=int(user["organization_id"]); summary=assessment_summary(session, organization_id)
    wb=Workbook(); ws=wb.active; ws.title="Resumen"
    ws.append(["Evaluación", summary["assessment"].name if summary["assessment"] else ""]); ws.append(["Riesgos", summary["counts"]["total"]]); ws.append(["Exposición bruta", summary["financial"]["gross_exposure"]]); ws.append(["Exposición residual", summary["financial"]["residual_exposure"]]); ws.append(["Exposición evitada", summary["financial"]["avoided_exposure"]]); ws.append(["Valor de oportunidades", summary["financial"]["opportunity_value"]]); ws.append(["Costo anual de controles", summary["financial"]["control_cost"]]); ws.append(["Preparación", summary["readiness_score"]])
    ws2=wb.create_sheet("Riesgos"); ws2.append(["Tipo","Categoría","Riesgo u oportunidad","Ubicación","Horizonte","Probabilidad","Impacto máximo","Inherente","Controles %","Residual","Nivel","Exposición","Responsable","Estrategia","Estado","Fuente"])
    for risk in summary["risks"]: ws2.append([risk.risk_type,risk.category,risk.hazard,risk.location,risk.time_horizon,risk.likelihood,max(risk.financial_impact,risk.operational_impact,risk.reputational_impact),risk.inherent_score,risk.control_effectiveness,risk.residual_score,risk_level(risk.residual_score),risk.financial_exposure,risk.owner,risk.response_strategy,risk.status,risk.source_reference])
    ws3=wb.create_sheet("Controles"); ws3.append(["Riesgo","Control","Tipo","Responsable","Estado","Efectividad %","Costo anual","Próxima revisión","Evidencia"])
    risk_map={risk.id:risk.hazard for risk in summary["risks"]}
    for control in summary["controls"]: ws3.append([risk_map.get(control.risk_id,""),control.name,control.control_type,control.owner,control.status,control.effectiveness,control.annual_cost,control.next_review,control.evidence])
    ws4=wb.create_sheet("Hoja de ruta"); ws4.append(["Categoría","Acción","Riesgo vinculado","Responsable","Inicio","Fin","Prioridad","Estado","Avance %","Reducción tCO2e","CAPEX","OPEX anual","Ahorro anual","Pérdida evitada","Indicador","Meta","Actual","Unidad","Dependencias"])
    for action in summary["actions"]: ws4.append([action.category,action.title,risk_map.get(action.risk_id,""),action.owner,action.start_date,action.end_date,action.priority,action.status,action.progress,action.expected_reduction_tco2e,action.capex,action.annual_opex,action.annual_savings,action.avoided_loss,action.indicator,action.target_value,action.current_value,action.unit,action.dependencies])
    buffer=BytesIO(); wb.save(buffer)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="riesgos_climaticos_{organization_id}.xlsx"'})

def _require_climate_disclosure_view(user: dict[str, object]) -> None:
    capabilities = user["capabilities"]
    if "view_climate_disclosure" not in capabilities and "manage_climate_disclosure" not in capabilities:
        raise HTTPException(403, "Tu rol no tiene acceso a divulgación climática")

@app.get("/divulgacion-climatica", response_class=HTMLResponse)
def climate_disclosure_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    _require_climate_disclosure_view(user)
    organization_id = int(user["organization_id"])
    comparison = scenario_comparison(session, organization_id)
    disclosure = disclosure_summary(session, organization_id)
    board = board_summary(session, organization_id)
    inventories = list(session.scalars(
        select(Inventory).where(Inventory.organization_id == organization_id).order_by(Inventory.start_date.desc())
    ))
    return templates.TemplateResponse(
        request=request, name="climate_disclosure.html",
        context=common_context(
            request, session, user, "climate_disclosure", comparison=comparison,
            disclosure=disclosure, board=board, inventories=inventories, today=date.today(),
        ),
    )

@app.post("/divulgacion-climatica/escenarios/nuevo")
def climate_scenario_create(
    request: Request, name: str = Form(...), code: str = Form(...), scenario_type: str = Form(...),
    temperature_pathway: str = Form("No especificada"), physical_multiplier: float = Form(1.0),
    transition_multiplier: float = Form(1.0), opportunity_multiplier: float = Form(1.0),
    carbon_price_2030: float = Form(0), energy_cost_change_pct: float = Form(0),
    demand_change_pct: float = Form(0), probability_weight: float = Form(0), narrative: str = Form(""),
    source_reference: str = Form(""), status: str = Form("Activo"),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_climate_disclosure")
    organization_id = int(user["organization_id"])
    normalized_code = code.strip().upper()[:40]
    duplicate = session.scalar(select(ClimateScenarioDefinition).where(
        ClimateScenarioDefinition.organization_id == organization_id,
        ClimateScenarioDefinition.code == normalized_code,
    ))
    if duplicate:
        raise HTTPException(409, "Ya existe un escenario con ese código")
    comparison = scenario_comparison(session, organization_id)
    assessment = comparison["risk_summary"]["assessment"]
    scenario = ClimateScenarioDefinition(
        organization_id=organization_id, assessment_id=assessment.id if assessment else None,
        name=name.strip(), code=normalized_code, scenario_type=scenario_type.strip(),
        temperature_pathway=temperature_pathway.strip(), physical_multiplier=max(0.1, min(3.0, physical_multiplier)),
        transition_multiplier=max(0.1, min(3.0, transition_multiplier)),
        opportunity_multiplier=max(0.1, min(3.0, opportunity_multiplier)),
        carbon_price_2030=max(0, carbon_price_2030), energy_cost_change_pct=max(-100, min(500, energy_cost_change_pct)),
        demand_change_pct=max(-100, min(500, demand_change_pct)), probability_weight=max(0, min(100, probability_weight)),
        narrative=narrative.strip(), source_reference=source_reference.strip(), status=status, created_by=str(user["email"]),
    )
    session.add(scenario)
    add_audit(session, organization_id, str(user["email"]), "CREAR", "Escenario climático", scenario.name, new_value=scenario.code)
    session.commit(); set_flash(request, "Escenario climático registrado.")
    return RedirectResponse("/divulgacion-climatica", status_code=303)

@app.post("/divulgacion-climatica/escenarios/{scenario_id}/actualizar")
def climate_scenario_update(
    scenario_id: int, request: Request, physical_multiplier: float = Form(...), transition_multiplier: float = Form(...),
    opportunity_multiplier: float = Form(...), carbon_price_2030: float = Form(0),
    energy_cost_change_pct: float = Form(0), demand_change_pct: float = Form(0),
    probability_weight: float = Form(0), narrative: str = Form(""), source_reference: str = Form(""),
    status: str = Form("Activo"), session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_climate_disclosure")
    scenario = session.get(ClimateScenarioDefinition, scenario_id)
    if not scenario or scenario.organization_id != int(user["organization_id"]):
        raise HTTPException(404, "Escenario no encontrado")
    scenario.physical_multiplier = max(0.1, min(3.0, physical_multiplier))
    scenario.transition_multiplier = max(0.1, min(3.0, transition_multiplier))
    scenario.opportunity_multiplier = max(0.1, min(3.0, opportunity_multiplier))
    scenario.carbon_price_2030 = max(0, carbon_price_2030)
    scenario.energy_cost_change_pct = max(-100, min(500, energy_cost_change_pct))
    scenario.demand_change_pct = max(-100, min(500, demand_change_pct))
    scenario.probability_weight = max(0, min(100, probability_weight))
    scenario.narrative = narrative.strip(); scenario.source_reference = source_reference.strip(); scenario.status = status
    add_audit(session, scenario.organization_id, str(user["email"]), "ACTUALIZAR", "Escenario climático", scenario.name, new_value=f"Peso {scenario.probability_weight}%")
    session.commit(); set_flash(request, "Supuestos del escenario actualizados.")
    return RedirectResponse("/divulgacion-climatica", status_code=303)

@app.post("/divulgacion-climatica/declaracion")
def climate_disclosure_save(
    request: Request, title: str = Form(...), inventory_id: str = Form(""), framework: str = Form(...),
    reporting_period: str = Form(...), scope_description: str = Form(""), materiality_basis: str = Form(""),
    owner: str = Form(...), status: str = Form("Borrador"), notes: str = Form(""),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_climate_disclosure")
    organization_id = int(user["organization_id"])
    summary = disclosure_summary(session, organization_id)
    statement = summary["statement"] or ClimateDisclosureStatement(organization_id=organization_id, title=title.strip(), created_by=str(user["email"]))
    if not summary["statement"]:
        session.add(statement)
    statement.title = title.strip(); statement.inventory_id = int(inventory_id) if inventory_id else None
    statement.framework = framework.strip(); statement.reporting_period = reporting_period.strip()
    statement.scope_description = scope_description.strip(); statement.materiality_basis = materiality_basis.strip()
    statement.owner = owner.strip(); statement.status = status; statement.notes = notes.strip()
    if status == "Aprobada":
        ensure_capability(user, "approve")
        statement.approved_by = str(user["email"]); statement.approved_at = datetime.now(UTC)
    add_audit(session, organization_id, str(user["email"]), "ACTUALIZAR", "Divulgación climática", statement.title, new_value=status)
    session.commit(); set_flash(request, "Ficha de divulgación guardada.")
    return RedirectResponse("/divulgacion-climatica", status_code=303)

@app.post("/divulgacion-climatica/requisitos/nuevo")
def climate_requirement_create(
    request: Request, pillar: str = Form(...), code: str = Form(...), requirement: str = Form(...),
    response: str = Form(""), status: str = Form("Pendiente"), evidence_reference: str = Form(""),
    owner: str = Form(...), due_date: str = Form(""), session: Session = Depends(get_db),
    user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_climate_disclosure")
    organization_id = int(user["organization_id"]); summary = disclosure_summary(session, organization_id)
    if status not in {"Completo", "Parcial", "Pendiente", "No aplica"}:
        raise HTTPException(400, "Estado de requisito inválido")
    statement = summary["statement"]
    if not statement:
        raise HTTPException(409, "Primero crea la ficha de divulgación")
    normalized_code = code.strip().upper()[:40]
    duplicate = session.scalar(select(ClimateDisclosureRequirement).where(
        ClimateDisclosureRequirement.statement_id == statement.id,
        ClimateDisclosureRequirement.code == normalized_code,
    ))
    if duplicate:
        raise HTTPException(409, "Ya existe un requisito con ese código")
    item = ClimateDisclosureRequirement(
        statement_id=statement.id, organization_id=organization_id, pillar=pillar.strip(), code=normalized_code,
        requirement=requirement.strip(), response=response.strip(), status=status,
        evidence_reference=evidence_reference.strip(), owner=owner.strip(),
        due_date=parse_date(due_date) if due_date else None, updated_by=str(user["email"]),
    )
    session.add(item); add_audit(session, organization_id, str(user["email"]), "CREAR", "Requisito de divulgación", item.code, new_value=item.pillar)
    session.commit(); set_flash(request, "Requisito de divulgación registrado.")
    return RedirectResponse("/divulgacion-climatica", status_code=303)

@app.post("/divulgacion-climatica/requisitos/{requirement_id}/actualizar")
def climate_requirement_update(
    requirement_id: int, request: Request, response: str = Form(""), status: str = Form(...),
    evidence_reference: str = Form(""), owner: str = Form(...), due_date: str = Form(""),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_climate_disclosure")
    requirement = session.get(ClimateDisclosureRequirement, requirement_id)
    if not requirement or requirement.organization_id != int(user["organization_id"]):
        raise HTTPException(404, "Requisito no encontrado")
    if status not in {"Completo", "Parcial", "Pendiente", "No aplica"}:
        raise HTTPException(400, "Estado de requisito inválido")
    requirement.response = response.strip(); requirement.status = status
    requirement.evidence_reference = evidence_reference.strip(); requirement.owner = owner.strip()
    requirement.due_date = parse_date(due_date) if due_date else None; requirement.updated_by = str(user["email"])
    add_audit(session, requirement.organization_id, str(user["email"]), "ACTUALIZAR", "Requisito de divulgación", requirement.code, new_value=status)
    session.commit(); set_flash(request, "Requisito de divulgación actualizado.")
    return RedirectResponse("/divulgacion-climatica", status_code=303)

@app.post("/divulgacion-climatica/comite")
def climate_board_save(
    request: Request, title: str = Form(...), meeting_date: str = Form(""), audience: str = Form("Comité directivo"),
    status: str = Form("Borrador"), executive_summary: str = Form(""), decisions_required: str = Form(""),
    key_message: str = Form(""), prepared_by: str = Form(...),
    session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_climate_disclosure")
    organization_id = int(user["organization_id"])
    summary = board_summary(session, organization_id)
    briefing = summary["briefing"] or ClimateBoardBriefing(organization_id=organization_id, title=title.strip(), created_by=str(user["email"]))
    if not summary["briefing"]:
        session.add(briefing)
    risk_assessment = summary["risk_summary"]["assessment"]
    statement = summary["disclosure"]["statement"]
    briefing.assessment_id = risk_assessment.id if risk_assessment else None
    briefing.disclosure_id = statement.id if statement else None
    briefing.title = title.strip(); briefing.meeting_date = parse_date(meeting_date) if meeting_date else None
    briefing.audience = audience.strip(); briefing.status = status; briefing.executive_summary = executive_summary.strip()
    briefing.decisions_required = decisions_required.strip(); briefing.key_message = key_message.strip(); briefing.prepared_by = prepared_by.strip()
    if status == "Aprobado":
        ensure_capability(user, "approve")
        briefing.approved_by = str(user["email"]); briefing.approved_at = datetime.now(UTC)
    add_audit(session, organization_id, str(user["email"]), "ACTUALIZAR", "Informe de comité", briefing.title, new_value=status)
    session.commit(); set_flash(request, "Informe para comité actualizado.")
    return RedirectResponse("/divulgacion-climatica", status_code=303)

@app.post("/divulgacion-climatica/decisiones/nueva")
def climate_board_decision_create(
    request: Request, topic: str = Form(...), decision: str = Form(""), owner: str = Form(...),
    due_date: str = Form(""), status: str = Form("Pendiente"), rationale: str = Form(""),
    evidence_reference: str = Form(""), session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_climate_disclosure")
    organization_id = int(user["organization_id"]); summary = board_summary(session, organization_id)
    briefing = summary["briefing"]
    if not briefing:
        raise HTTPException(409, "Primero crea el informe para comité")
    item = ClimateBoardDecision(
        briefing_id=briefing.id, organization_id=organization_id, topic=topic.strip(), decision=decision.strip(),
        owner=owner.strip(), due_date=parse_date(due_date) if due_date else None, status=status,
        rationale=rationale.strip(), evidence_reference=evidence_reference.strip(), created_by=str(user["email"]),
    )
    session.add(item); add_audit(session, organization_id, str(user["email"]), "CREAR", "Decisión de comité", item.topic, new_value=status)
    session.commit(); set_flash(request, "Decisión registrada.")
    return RedirectResponse("/divulgacion-climatica", status_code=303)

@app.post("/divulgacion-climatica/decisiones/{decision_id}/estado")
def climate_board_decision_update(
    decision_id: int, request: Request, decision: str = Form(""), owner: str = Form(...),
    due_date: str = Form(""), status: str = Form(...), rationale: str = Form(""),
    evidence_reference: str = Form(""), session: Session = Depends(get_db), user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_climate_disclosure")
    item = session.get(ClimateBoardDecision, decision_id)
    if not item or item.organization_id != int(user["organization_id"]):
        raise HTTPException(404, "Decisión no encontrada")
    item.decision = decision.strip(); item.owner = owner.strip(); item.due_date = parse_date(due_date) if due_date else None
    item.status = status; item.rationale = rationale.strip(); item.evidence_reference = evidence_reference.strip()
    add_audit(session, item.organization_id, str(user["email"]), "ACTUALIZAR", "Decisión de comité", item.topic, new_value=status)
    session.commit(); set_flash(request, "Estado de la decisión actualizado.")
    return RedirectResponse("/divulgacion-climatica", status_code=303)

@app.get("/divulgacion-climatica/exportar.xlsx")
def climate_disclosure_export(session: Session = Depends(get_db), user: dict = Depends(require_user)):
    _require_climate_disclosure_view(user)
    organization_id = int(user["organization_id"])
    comparison = scenario_comparison(session, organization_id); disclosure = disclosure_summary(session, organization_id); board = board_summary(session, organization_id)
    wb = Workbook(); ws = wb.active; ws.title = "Comparación"
    ws.append(["Escenario", "Código", "Trayectoria", "Probabilidad %", "Exposición", "Costo carbono", "Oportunidad", "Presión neta", "Presión ponderada", "Resiliencia", "Riesgos críticos"])
    for result in comparison["results"]:
        scenario = result["scenario"]
        ws.append([scenario.name, scenario.code, scenario.temperature_pathway, scenario.probability_weight, result["downside_exposure"], result["carbon_cost"], result["opportunity_value"], result["net_financial_pressure"], result["weighted_pressure"], result["resilience_score"], result["critical_risks"]])
    ws2 = wb.create_sheet("Supuestos"); ws2.append(["Escenario", "Tipo", "Multiplicador físico", "Multiplicador transición", "Multiplicador oportunidad", "Precio carbono 2030", "Cambio energía %", "Cambio demanda %", "Narrativa", "Fuente", "Estado"])
    for scenario in comparison["scenarios"]:
        ws2.append([scenario.name, scenario.scenario_type, scenario.physical_multiplier, scenario.transition_multiplier, scenario.opportunity_multiplier, scenario.carbon_price_2030, scenario.energy_cost_change_pct, scenario.demand_change_pct, scenario.narrative, scenario.source_reference, scenario.status])
    ws3 = wb.create_sheet("Divulgación"); ws3.append(["Pilar", "Código", "Requisito", "Respuesta", "Estado", "Evidencia", "Responsable", "Vencimiento"])
    for requirement in disclosure["requirements"]:
        ws3.append([requirement.pillar, requirement.code, requirement.requirement, requirement.response, requirement.status, requirement.evidence_reference, requirement.owner, requirement.due_date])
    ws4 = wb.create_sheet("Decisiones"); ws4.append(["Tema", "Decisión", "Responsable", "Vencimiento", "Estado", "Justificación", "Evidencia"])
    for item in board["decisions"]:
        ws4.append([item.topic, item.decision, item.owner, item.due_date, item.status, item.rationale, item.evidence_reference])
    buffer = BytesIO(); wb.save(buffer)
    return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="divulgacion_climatica_{organization_id}.xlsx"'})

@app.get("/divulgacion-climatica/comite.pdf")
def climate_board_pdf(session: Session = Depends(get_db), user: dict = Depends(require_user)):
    _require_climate_disclosure_view(user)
    organization_id = int(user["organization_id"]); org = session.get(Organization, organization_id)
    summary = board_summary(session, organization_id)
    content, digest = build_board_pdf(summary, org.name)
    briefing = summary["briefing"]
    if briefing:
        briefing.document_hash = digest
        session.commit()
    return Response(content=content, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="informe_comite_climatico_{organization_id}.pdf"', "X-Document-SHA256": digest})

@app.get("/consolidacion", response_class=HTMLResponse)
def consolidation_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "view_consolidation")
    summary = consolidation_summary(session, int(user["organization_id"]), BASE_DIR.parent)
    architecture = domain_architecture_summary(app, BASE_DIR.parent)
    return templates.TemplateResponse(
        request=request,
        name="consolidation.html",
        context=common_context(
            request,
            session,
            user,
            "consolidation",
            summary=summary,
            domain_architecture=architecture,
        ),
    )

@app.post("/consolidacion/hallazgos/{finding_id}")
def update_consolidation_finding(
    finding_id: int,
    request: Request,
    status: str = Form(...),
    owner: str = Form(""),
    target_version: str = Form("V1.0"),
    evidence: str = Form(""),
    session: Session = Depends(get_db),
    user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_consolidation")
    finding = session.scalar(select(ConsolidationFinding).where(
        ConsolidationFinding.id == finding_id,
        ConsolidationFinding.organization_id == int(user["organization_id"]),
    ))
    if not finding:
        raise HTTPException(404, "Hallazgo no encontrado")
    if status not in {"Abierto", "En curso", "Bloqueado", "Resuelto", "Aceptado"}:
        raise HTTPException(400, "Estado inválido")
    previous = finding.status
    finding.status = status
    finding.owner = owner.strip() or finding.owner
    finding.target_version = target_version.strip() or finding.target_version
    finding.evidence = evidence.strip()
    add_audit(session, int(user["organization_id"]), str(user["email"]), "ACTUALIZAR", "Hallazgo de consolidación", finding.code, previous_value=previous, new_value=status, detail=finding.title)
    session.commit()
    set_flash(request, f"Hallazgo {finding.code} actualizado.")
    return RedirectResponse("/consolidacion#hallazgos", status_code=303)

@app.post("/consolidacion/puertas/{gate_id}")
def update_release_gate(
    gate_id: int,
    request: Request,
    status: str = Form(...),
    responsible: str = Form(""),
    evidence: str = Form(""),
    notes: str = Form(""),
    session: Session = Depends(get_db),
    user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_consolidation")
    gate = session.scalar(select(ReleaseGate).where(
        ReleaseGate.id == gate_id,
        ReleaseGate.organization_id == int(user["organization_id"]),
    ))
    if not gate:
        raise HTTPException(404, "Puerta de salida no encontrada")
    if status not in {"Pendiente", "Parcial", "En revisión", "Aprobado", "Bloqueado"}:
        raise HTTPException(400, "Estado inválido")
    previous = gate.status
    gate.status = status
    gate.responsible = responsible.strip() or gate.responsible
    gate.evidence = evidence.strip()
    gate.notes = notes.strip()
    add_audit(session, int(user["organization_id"]), str(user["email"]), "ACTUALIZAR", "Puerta V1.0", gate.code, previous_value=previous, new_value=status, detail=gate.name)
    session.commit()
    set_flash(request, f"Puerta {gate.code} actualizada.")
    return RedirectResponse("/consolidacion#puertas", status_code=303)

@app.post("/consolidacion/recorridos/{validation_id}")
def update_journey_validation(
    validation_id: int,
    request: Request,
    status: str = Form(...),
    notes: str = Form(""),
    session: Session = Depends(get_db),
    user: dict = Depends(require_user),
):
    ensure_capability(user, "manage_consolidation")
    validation = session.scalar(select(JourneyValidation).where(
        JourneyValidation.id == validation_id,
        JourneyValidation.organization_id == int(user["organization_id"]),
    ))
    if not validation:
        raise HTTPException(404, "Recorrido no encontrado")
    if status not in {"No probado", "En prueba", "Con bloqueos", "Aprobado"}:
        raise HTTPException(400, "Estado inválido")
    previous = validation.status
    validation.status = status
    validation.notes = notes.strip()
    validation.tested_by = str(user["email"]) if status != "No probado" else ""
    validation.tested_at = datetime.now(UTC) if status != "No probado" else None
    add_audit(session, int(user["organization_id"]), str(user["email"]), "VALIDAR", "Recorrido por rol", validation.journey_code, previous_value=previous, new_value=status, detail=validation.notes)
    session.commit()
    set_flash(request, f"Recorrido {validation.journey_code} actualizado.")
    return RedirectResponse("/consolidacion#recorridos", status_code=303)

@app.get("/consolidacion/exportar.xlsx")
def export_consolidation(session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "view_consolidation")
    summary = consolidation_summary(session, int(user["organization_id"]), BASE_DIR.parent)
    content = build_consolidation_workbook(summary)
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="consolidacion_v1_0.xlsx"'},
    )

@app.get("/api/arquitectura/resumen")
def architecture_api(user: dict = Depends(require_user)):
    ensure_capability(user, "view_consolidation")
    return domain_architecture_summary(app, BASE_DIR.parent)

@app.get("/api/consolidacion/resumen")
def consolidation_api(session: Session = Depends(get_db), user: dict = Depends(require_user)):
    ensure_capability(user, "view_consolidation")
    summary = consolidation_summary(session, int(user["organization_id"]), BASE_DIR.parent)
    return Response(content=summary_json(summary), media_type="application/json")

register_public_routes(app, templates, current_user)
register_auth_routes(app, templates, current_user)

register_organization_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash
)
register_information_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash,
    parse_date, get_inventory, ensure_inventory_editable, quality_from, safe_filename,
    format_bytes, ALLOWED_UPLOAD_EXTENSIONS, MAX_UPLOAD_SIZE, ALLOWED_UNITS,
    DATA_ORIGINS, _parse_excel_period,
)
register_capture_routes(
    app, templates, common_context, require_user, set_flash,
    parse_date, get_inventory, ensure_inventory_editable, quality_from, safe_filename,
    ALLOWED_UPLOAD_EXTENSIONS, MAX_UPLOAD_SIZE, ALLOWED_UNITS, DATA_ORIGINS,
)
register_review_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash,
    parse_date, get_inventory, get_source_for_user, ensure_inventory_editable,
    review_gate_summary, clone_inventory_version,
)

register_user_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash
)
register_service_operations_routes(
    app, templates, common_context, require_user, ensure_capability
)
register_inventory_routes(
    app,
    templates,
    common_context,
    require_user,
    ensure_capability,
    set_flash,
    parse_date,
    get_inventory,
    get_source_for_user,
    ensure_inventory_editable,
    inventory_metrics,
    ALLOWED_UNITS,
    DATA_ORIGINS,
)
register_report_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash, get_inventory
)
register_reduction_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash,
    parse_date, get_inventory, get_source_for_user, ensure_inventory_editable, inventory_metrics,
)
register_delivery_routes(
    app, templates, common_context, require_user, get_inventory
)
register_workflow_routes(
    app, templates, common_context, require_user
)
register_experience_routes(
    app, templates, common_context, require_user, get_inventory
)
register_legal_routes(
    app, templates, current_user
)

register_methodology_core_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash
)
register_factor_library_routes(
    app, templates, common_context, require_user, ensure_capability
)
register_methodology_closure_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash, get_inventory, ensure_inventory_editable
)
register_land_removals_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash, get_inventory, ensure_inventory_editable
)

register_product_project_assurance_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash, get_inventory, ensure_inventory_editable
)

register_colombia_library_routes(
    app, templates, common_context, require_user, ensure_capability
)

register_pilot_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash
)

register_pilot_execution_routes(
    app, templates, common_context, require_user, set_flash
)

register_data_quality_routes(
    app, templates, common_context, require_user, set_flash
)
register_period_close_routes(
    app, templates, common_context, require_user, set_flash
)
register_operational_import_routes(
    app, templates, common_context, require_user, set_flash, INSTANCE_DIR, MAX_UPLOAD_SIZE
)
register_integration_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash,
    ensure_inventory_editable, ALLOWED_UNITS,
)
register_operations_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash
)
register_demo_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash
)
register_product_intelligence_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash, settings
)
register_guided_onboarding_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash
)

@app.get("/modulos", response_class=HTMLResponse)
def modules_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
    return templates.TemplateResponse(
        request=request,
        name="modules.html",
        context=common_context(request, session, user, "modules", modules=PRODUCT_MODULES),
    )

@app.get("/api/health")
def health():
    return {"status": "ok", "app": settings.app_name, "version": settings.version, "environment": settings.environment}

@app.get("/api/ready")
def ready():
    snapshot = diagnostic_snapshot()
    status_code = 200 if snapshot["status"] == "ready" else 503
    return Response(
        content=json.dumps(snapshot, ensure_ascii=False, default=str),
        media_type="application/json",
        status_code=status_code,
    )
