from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, timedelta
from io import BytesIO
from statistics import median

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from .database import EmissionTarget, Inventory, ReductionAction
from .scenarios import action_economics


STATUS_ORDER = [
    "Identificada",
    "En evaluación",
    "Diseño",
    "Aprobada",
    "En ejecución",
    "Implementada",
    "Verificada",
]


def _clamp(value: float, minimum: float = 0.0, maximum: float = 100.0) -> float:
    return max(minimum, min(maximum, value))


def _baseline(inventory: Inventory) -> float:
    return round(sum(float(source.emissions or 0) for source in inventory.sources if source.included), 6)


def _target_requirement(targets: list[EmissionTarget]) -> tuple[float, EmissionTarget | None]:
    active = [item for item in targets if item.status not in {"Suspendida", "Cancelada"}]
    if not active:
        return 0.0, None
    target = sorted(active, key=lambda item: (item.target_year, item.id))[0]
    required = max(float(target.baseline_value or 0) - float(target.target_value or 0), 0.0)
    return required, target


def action_readiness(action: ReductionAction, today: date | None = None) -> dict[str, object]:
    today = today or date.today()
    checks = [
        ("Impacto cuantificado", float(action.expected_reduction or 0) > 0, 20),
        ("Responsable asignado", bool((action.responsible or "").strip()), 15),
        ("Fecha objetivo", action.target_date is not None, 10),
        ("Fuente vinculada", action.source_id is not None, 10),
        ("Inversión estimada", float(action.investment_cost or 0) > 0, 10),
        ("Ahorro estimado", float(action.annual_savings or 0) > 0, 10),
        ("Año de implementación", action.implementation_year is not None, 10),
        ("Viabilidad evaluada", action.feasibility in {"Alta", "Media", "Baja"}, 5),
        ("Riesgo evaluado", action.risk_level in {"Bajo", "Medio", "Alto"}, 5),
        ("Avance coherente", 0 <= int(action.progress_percent or 0) <= 100, 5),
    ]
    score = sum(weight for _, ok, weight in checks if ok)
    missing = [label for label, ok, _ in checks if not ok]
    overdue = bool(action.target_date and action.target_date < today and action.status not in {"Implementada", "Verificada"})
    due_soon = bool(
        action.target_date
        and today <= action.target_date <= today + timedelta(days=90)
        and action.status not in {"Implementada", "Verificada"}
    )
    if score >= 85:
        level = "Lista para decisión"
    elif score >= 65:
        level = "En estructuración"
    else:
        level = "Información insuficiente"
    return {
        "score": score,
        "level": level,
        "missing": missing,
        "overdue": overdue,
        "due_soon": due_soon,
    }


def _decision_class(action: ReductionAction, readiness: int, median_reduction: float, marginal_cost: float, payback_years: float | None) -> str:
    if readiness < 65:
        return "Por estructurar"
    high_impact = float(action.expected_reduction or 0) >= median_reduction
    if action.status in {"Aprobada", "En ejecución", "Implementada", "Verificada"}:
        return "En ejecución"
    if marginal_cost <= 0 and payback_years is not None and payback_years <= 3.5 and action.feasibility in {"Alta", "Media"}:
        return "Ganancia rápida"
    if (high_impact or marginal_cost <= 0) and action.feasibility in {"Alta", "Media"}:
        return "Apuesta estratégica"
    return "Habilitador"


def portfolio_summary(session: Session, inventory: Inventory, today: date | None = None) -> dict[str, object]:
    del session  # La firma conserva uniformidad con otros servicios y facilita futuras consultas.
    today = today or date.today()
    actions = list(inventory.reduction_actions)
    targets = list(inventory.targets)
    baseline = _baseline(inventory)
    required_reduction, primary_target = _target_requirement(targets)
    reductions = [float(item.expected_reduction or 0) for item in actions if float(item.expected_reduction or 0) > 0]
    median_reduction = median(reductions) if reductions else 0.0

    rows: list[dict[str, object]] = []
    for action in actions:
        readiness = action_readiness(action, today)
        economics = action_economics(action)
        impact = float(action.expected_reduction or 0)
        impact_share = impact / baseline * 100 if baseline else 0.0
        classification = _decision_class(action, int(readiness["score"]), median_reduction, economics.marginal_cost, economics.payback_years)
        rows.append({
            "id": action.id,
            "action": action,
            "title": action.title,
            "source": action.source.name if action.source else "Corporativo",
            "owner": action.responsible or "Sin asignar",
            "status": action.status,
            "priority": action.priority,
            "expected_reduction": impact,
            "actual_reduction": float(action.actual_reduction or 0),
            "investment": float(action.investment_cost or 0),
            "annual_savings": float(action.annual_savings or 0),
            "payback_years": economics.payback_years,
            "marginal_cost": economics.marginal_cost,
            "impact_share": round(impact_share, 1),
            "readiness_score": int(readiness["score"]),
            "readiness_level": readiness["level"],
            "missing": readiness["missing"],
            "overdue": readiness["overdue"],
            "due_soon": readiness["due_soon"],
            "classification": classification,
            "target_date": action.target_date,
            "implementation_year": action.implementation_year,
            "progress_percent": int(action.progress_percent or 0),
            "feasibility": action.feasibility,
            "risk_level": action.risk_level,
        })

    rows.sort(key=lambda item: (
        0 if item["classification"] == "Ganancia rápida" else 1,
        -float(item["expected_reduction"]),
        -int(item["readiness_score"]),
    ))

    expected = sum(float(item.expected_reduction or 0) for item in actions)
    actual = sum(float(item.actual_reduction or 0) for item in actions)
    investment = sum(float(item.investment_cost or 0) for item in actions)
    annual_savings = sum(float(item.annual_savings or 0) for item in actions)
    weighted_progress = (
        sum(float(item.expected_reduction or 0) * int(item.progress_percent or 0) for item in actions) / expected
        if expected else 0.0
    )
    coverage = expected / required_reduction * 100 if required_reduction else 0.0
    gap = max(required_reduction - expected, 0.0)
    actual_coverage = actual / required_reduction * 100 if required_reduction else 0.0
    readiness_score = round(sum(int(item["readiness_score"]) for item in rows) / len(rows)) if rows else 0

    pipeline_counter = Counter(item.status for item in actions)
    pipeline = [{"status": status, "count": pipeline_counter.get(status, 0)} for status in STATUS_ORDER]
    other_statuses = sorted(set(pipeline_counter) - set(STATUS_ORDER))
    pipeline.extend({"status": status, "count": pipeline_counter[status]} for status in other_statuses)

    owner_map: dict[str, dict[str, float | int | str]] = defaultdict(lambda: {"count": 0, "reduction": 0.0})
    for item in rows:
        owner = str(item["owner"])
        owner_map[owner]["count"] = int(owner_map[owner]["count"]) + 1
        owner_map[owner]["reduction"] = float(owner_map[owner]["reduction"]) + float(item["expected_reduction"])
    owners = [
        {"owner": owner, "count": values["count"], "reduction": round(float(values["reduction"]), 2)}
        for owner, values in owner_map.items()
    ]
    owners.sort(key=lambda item: (-float(item["reduction"]), str(item["owner"])))

    classifications = Counter(str(item["classification"]) for item in rows)
    overdue = [item for item in rows if item["overdue"]]
    due_soon = [item for item in rows if item["due_soon"]]
    incomplete = [item for item in rows if int(item["readiness_score"]) < 65]
    quick_wins = [item for item in rows if item["classification"] == "Ganancia rápida"]

    if not actions:
        portfolio_status = "Sin portafolio"
        primary_decision = "Crear las primeras medidas vinculadas con las fuentes de mayor contribución."
    elif required_reduction and coverage < 80:
        portfolio_status = "Brecha material"
        formatted_gap = f"{gap:,.1f}".replace(",", "X").replace(".", ",").replace("X", ".")
        primary_decision = f"Estructurar medidas adicionales por {formatted_gap} tCO₂e/año para acercarse a la meta."
    elif readiness_score < 65:
        portfolio_status = "Requiere estructuración"
        primary_decision = "Completar responsables, fechas y evaluación económica antes de solicitar aprobación."
    elif overdue:
        portfolio_status = "Ejecución en riesgo"
        primary_decision = f"Resolver {len(overdue)} acción(es) vencida(s) y redefinir compromisos de ejecución."
    elif coverage >= 100 and readiness_score >= 85:
        portfolio_status = "Listo para decisión"
        primary_decision = "Aprobar la primera ola de implementación y formalizar el seguimiento trimestral."
    else:
        portfolio_status = "En estructuración"
        primary_decision = "Priorizar ganancias rápidas y cerrar la estructuración de las apuestas estratégicas."

    target_year = primary_target.target_year if primary_target else max([item.implementation_year or today.year for item in actions] + [today.year])
    start_year = min(today.year, inventory.start_date.year)
    timeline = []
    for year in range(start_year, target_year + 1):
        annual_reduction = sum(
            float(item.expected_reduction or 0)
            for item in actions
            if (item.implementation_year or target_year) <= year
        )
        timeline.append({
            "year": year,
            "reduction": round(annual_reduction, 2),
            "projected_emissions": round(max(baseline - annual_reduction, 0.0), 2),
            "target_value": round(float(primary_target.target_value or 0), 2) if primary_target else None,
        })

    return {
        "baseline": round(baseline, 3),
        "primary_target": primary_target,
        "required_reduction": round(required_reduction, 3),
        "expected_reduction": round(expected, 3),
        "actual_reduction": round(actual, 3),
        "coverage_percent": round(_clamp(coverage, 0, 999), 1),
        "actual_coverage_percent": round(_clamp(actual_coverage, 0, 999), 1),
        "gap": round(gap, 3),
        "investment": round(investment, 2),
        "annual_savings": round(annual_savings, 2),
        "payback_years": round(investment / annual_savings, 2) if annual_savings else None,
        "weighted_progress": round(weighted_progress),
        "readiness_score": readiness_score,
        "portfolio_status": portfolio_status,
        "primary_decision": primary_decision,
        "actions": rows,
        "action_count": len(rows),
        "quick_wins": quick_wins,
        "overdue": overdue,
        "due_soon": due_soon,
        "incomplete": incomplete,
        "pipeline": pipeline,
        "owners": owners,
        "classifications": dict(classifications),
        "timeline": timeline,
        "today": today,
        "decision_ready": bool(actions and readiness_score >= 85 and (not required_reduction or coverage >= 100) and not overdue),
    }


def portfolio_json(summary: dict[str, object]) -> dict[str, object]:
    target = summary.get("primary_target")
    return {
        "baseline": summary["baseline"],
        "target": {
            "name": target.name,
            "target_year": target.target_year,
            "target_value": target.target_value,
        } if target else None,
        "required_reduction": summary["required_reduction"],
        "expected_reduction": summary["expected_reduction"],
        "actual_reduction": summary["actual_reduction"],
        "coverage_percent": summary["coverage_percent"],
        "gap": summary["gap"],
        "readiness_score": summary["readiness_score"],
        "portfolio_status": summary["portfolio_status"],
        "primary_decision": summary["primary_decision"],
        "decision_ready": summary["decision_ready"],
        "actions": [
            {
                key: item[key]
                for key in [
                    "id", "title", "source", "owner", "status", "priority", "expected_reduction",
                    "actual_reduction", "investment", "annual_savings", "payback_years", "marginal_cost",
                    "impact_share", "readiness_score", "readiness_level", "missing", "overdue", "due_soon",
                    "classification", "implementation_year", "progress_percent", "feasibility", "risk_level",
                ]
            } | {"target_date": item["target_date"].isoformat() if item["target_date"] else None}
            for item in summary["actions"]
        ],
        "pipeline": summary["pipeline"],
        "owners": summary["owners"],
        "timeline": summary["timeline"],
    }


def build_portfolio_workbook(inventory: Inventory, summary: dict[str, object]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dirección"
    navy = "173B57"
    green = "2E7D5B"
    light = "EAF2EF"
    amber = "F5E7C2"

    ws.append(["CALCULA TU HUELLA · PORTAFOLIO DE REDUCCIÓN"])
    ws.append([inventory.name])
    ws.append([])
    metrics = [
        ("Emisiones base", summary["baseline"], "tCO2e"),
        ("Reducción requerida", summary["required_reduction"], "tCO2e/año"),
        ("Reducción esperada", summary["expected_reduction"], "tCO2e/año"),
        ("Cobertura de la meta", summary["coverage_percent"], "%"),
        ("Brecha", summary["gap"], "tCO2e/año"),
        ("Preparación del portafolio", summary["readiness_score"], "%"),
        ("Inversión", summary["investment"], "COP"),
        ("Ahorro anual", summary["annual_savings"], "COP/año"),
        ("Retorno simple", summary["payback_years"] or 0, "años"),
    ]
    ws.append(["Indicador", "Valor", "Unidad"])
    for row in metrics:
        ws.append(row)
    ws.append([])
    ws.append(["Estado", summary["portfolio_status"]])
    ws.append(["Decisión principal", summary["primary_decision"]])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    ws.merge_cells(start_row=15, start_column=2, end_row=15, end_column=3)
    ws.merge_cells(start_row=16, start_column=2, end_row=16, end_column=3)

    actions_ws = wb.create_sheet("Acciones")
    actions_ws.append([
        "Clasificación", "Acción", "Fuente", "Responsable", "Estado", "Prioridad", "Reducción esperada",
        "Reducción real", "Impacto inventario %", "Inversión COP", "Ahorro anual COP", "Retorno años",
        "Costo marginal COP/tCO2e", "Preparación %", "Nivel", "Pendientes", "Fecha objetivo",
        "Vencida", "Próximos 90 días", "Año implementación", "Avance %", "Viabilidad", "Riesgo",
    ])
    for item in summary["actions"]:
        actions_ws.append([
            item["classification"], item["title"], item["source"], item["owner"], item["status"], item["priority"],
            item["expected_reduction"], item["actual_reduction"], item["impact_share"], item["investment"],
            item["annual_savings"], item["payback_years"], item["marginal_cost"], item["readiness_score"],
            item["readiness_level"], "; ".join(item["missing"]), item["target_date"],
            "Sí" if item["overdue"] else "No", "Sí" if item["due_soon"] else "No",
            item["implementation_year"], item["progress_percent"], item["feasibility"], item["risk_level"],
        ])

    targets_ws = wb.create_sheet("Metas")
    targets_ws.append(["Meta", "Tipo", "Año base", "Año meta", "Línea base", "Valor actual", "Valor objetivo", "Unidad", "Reducción %", "Avance %", "Estado", "Notas"])
    for target in inventory.targets:
        targets_ws.append([
            target.name, target.metric_type, target.baseline_year, target.target_year, target.baseline_value,
            target.current_value, target.target_value, target.unit, target.reduction_percent,
            target.progress_percent, target.status, target.notes,
        ])

    timeline_ws = wb.create_sheet("Trayectoria")
    timeline_ws.append(["Año", "Reducción acumulada", "Emisiones proyectadas", "Valor meta"])
    for item in summary["timeline"]:
        timeline_ws.append([item["year"], item["reduction"], item["projected_emissions"], item["target_value"]])

    owners_ws = wb.create_sheet("Responsables")
    owners_ws.append(["Responsable", "Acciones", "Reducción esperada"])
    for item in summary["owners"]:
        owners_ws.append([item["owner"], item["count"], item["reduction"]])

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2" if sheet.title != "Dirección" else "A5"
        if sheet.title != "Dirección" and sheet.max_row > 1:
            sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if sheet.title == "Dirección":
            sheet[1][0].font = Font(bold=True, color="FFFFFF", size=15)
            sheet[2][0].font = Font(bold=True, color=green, size=12)
            for cell in sheet[4]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=green)
            for row in range(5, 14):
                sheet.cell(row, 1).fill = PatternFill("solid", fgColor=light)
            sheet.cell(15, 1).fill = PatternFill("solid", fgColor=amber)
            sheet.cell(16, 1).fill = PatternFill("solid", fgColor=amber)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in range(1, sheet.max_column + 1):
            letter = get_column_letter(column)
            max_length = max(len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1))
            sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 44)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
