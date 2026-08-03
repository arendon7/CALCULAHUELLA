from __future__ import annotations

import json
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from .database import (
    EmissionFactor,
    EmissionFactorVersion,
    FactorDocumentation,
    MethodologySourceDocument,
    ReferenceCalculationCase,
)

LIBRARY_VERSION = "0.28.0"
US_GALLON_LITRES = 3.785411784
GWP_AR6_CH4 = 27.0
GWP_AR6_N2O = 273.0
N2O_N_TO_N2O = 44.0 / 28.0
METHANE_DENSITY_KG_M3 = 0.7168

COLOMBIA_SOURCE_CODES = {
    "UPME-R085-2026",
    "UPME-FECOC-2016",
    "UPME-FECOCPLUS-3-2023",
    "COL-DECRETO-926-2017",
    "EAAB-FECOC-2016",
    "IPCC-WASTEWATER-2019",
    "IPCC-AFOLU-2019",
}

FUEL_CALCULATOR_FACTORS: dict[str, dict[str, Any]] = {
    "REG-GN": {"label": "Gas natural · factor regulatorio", "value": 1.952, "unit": "m3", "gas": "CO2", "source": "Decreto 926 de 2017", "use": "Piloto condicionado"},
    "REG-GLP": {"label": "GLP · factor regulatorio", "value": 6.333, "unit": "gal", "gas": "CO2", "source": "Decreto 926 de 2017", "use": "Piloto condicionado"},
    "REG-GASOLINA": {"label": "Gasolina · factor regulatorio", "value": 9.000, "unit": "gal", "gas": "CO2", "source": "Decreto 926 de 2017", "use": "Piloto condicionado"},
    "REG-JET": {"label": "Kerosene / Jet Fuel · factor regulatorio", "value": 9.867, "unit": "gal", "gas": "CO2", "source": "Decreto 926 de 2017", "use": "Piloto condicionado"},
    "REG-ACPM": {"label": "ACPM · factor regulatorio", "value": 10.133, "unit": "gal", "gas": "CO2", "source": "Decreto 926 de 2017", "use": "Piloto condicionado"},
    "REG-FUELOIL": {"label": "Fuel oil · factor regulatorio", "value": 11.800, "unit": "gal", "gas": "CO2", "source": "Decreto 926 de 2017", "use": "Piloto condicionado"},
    "FECOC-DIESEL-CO2": {"label": "Diésel B10 · CO2 FECOC transcrito", "value": 10.2765, "unit": "gal", "gas": "CO2", "source": "FECOC 2016 · transcripción secundaria", "use": "Piloto en revisión"},
    "FECOC-GASOLINA-CO2": {"label": "Gasolina E10 · CO2 FECOC transcrito", "value": 7.6181, "unit": "gal", "gas": "CO2", "source": "FECOC 2016 · transcripción secundaria", "use": "Piloto en revisión"},
}

FERTILIZER_DIRECT_EF = {
    "aggregated": {"label": "EF1 agregado", "wet": 0.010, "dry": 0.010},
    "synthetic": {"label": "Fertilizante sintético", "wet": 0.016, "dry": 0.005},
    "other": {"label": "Otros aportes de N", "wet": 0.006, "dry": 0.005},
}
FERTILIZER_VOLATILIZATION = {
    "generic": 0.11,
    "urea": 0.15,
    "ammonium": 0.08,
    "nitrate": 0.01,
    "ammonium_nitrate": 0.05,
}


def _non_negative(value: float, name: str) -> float:
    value = float(value)
    if value < 0:
        raise ValueError(f"{name} no puede ser negativo")
    return value


def calculate_combustion(factor_code: str, amount: float, amount_unit: str) -> dict[str, Any]:
    spec = FUEL_CALCULATOR_FACTORS.get(factor_code)
    if not spec:
        raise ValueError("Factor de combustible no reconocido")
    amount = _non_negative(amount, "La actividad")
    amount_unit = amount_unit.strip()
    target_unit = spec["unit"]
    if target_unit == "m3":
        if amount_unit != "m3":
            raise ValueError("El gas natural debe informarse en m3 estándar")
        normalized = amount
    elif target_unit == "gal":
        if amount_unit == "gal":
            normalized = amount
        elif amount_unit == "L":
            normalized = amount / US_GALLON_LITRES
        else:
            raise ValueError("El combustible líquido debe informarse en L o gal US")
    else:
        raise ValueError("Unidad del factor no soportada")
    gas_kg = normalized * float(spec["value"])
    return {
        "factor_code": factor_code,
        "label": spec["label"],
        "original_value": amount,
        "original_unit": amount_unit,
        "normalized_value": normalized,
        "normalized_unit": target_unit,
        "factor_value": float(spec["value"]),
        "gas": spec["gas"],
        "gas_kg": gas_kg,
        "co2e_kg": gas_kg,
        "co2e_t": gas_kg / 1000.0,
        "source": spec["source"],
        "reporting_use": spec["use"],
        "formula": f"{normalized:.6f} {target_unit} × {spec['value']} kg {spec['gas']}/{target_unit}",
    }


def calculate_wastewater(
    organic_load_kg: float,
    basis: str,
    mcf: float,
    recovered_ch4_kg: float = 0.0,
) -> dict[str, Any]:
    organic_load_kg = _non_negative(organic_load_kg, "La carga orgánica")
    recovered_ch4_kg = _non_negative(recovered_ch4_kg, "El CH4 recuperado o quemado")
    mcf = float(mcf)
    if not 0 <= mcf <= 1:
        raise ValueError("MCF debe estar entre 0 y 1")
    basis = basis.upper().strip()
    if basis not in {"COD", "BOD"}:
        raise ValueError("La base debe ser COD o BOD")
    bo = 0.25 if basis == "COD" else 0.60
    gross_ch4 = organic_load_kg * bo * mcf
    emitted_ch4 = max(0.0, gross_ch4 - recovered_ch4_kg)
    co2e_kg = emitted_ch4 * GWP_AR6_CH4
    return {
        "organic_load_kg": organic_load_kg,
        "basis": basis,
        "bo": bo,
        "mcf": mcf,
        "gross_ch4_kg": gross_ch4,
        "recovered_ch4_kg": recovered_ch4_kg,
        "emitted_ch4_kg": emitted_ch4,
        "gwp": GWP_AR6_CH4,
        "co2e_kg": co2e_kg,
        "co2e_t": co2e_kg / 1000.0,
        "formula": f"({organic_load_kg:.3f} kg {basis} × {bo} × {mcf}) − {recovered_ch4_kg:.3f} kg CH4",
        "warning": "El MCF debe corresponder al sistema real. Deben descontarse únicamente cantidades efectivamente recuperadas o quemadas y documentadas.",
    }


def calculate_fertilizer(
    nitrogen_kg: float,
    climate: str,
    input_type: str,
    include_volatilization: bool = True,
    include_leaching: bool = True,
) -> dict[str, Any]:
    nitrogen_kg = _non_negative(nitrogen_kg, "El nitrógeno aplicado")
    climate = climate.lower().strip()
    if climate not in {"wet", "dry"}:
        raise ValueError("El clima debe ser wet o dry")
    input_type = input_type.lower().strip()
    direct_group = "synthetic" if input_type in {"urea", "ammonium", "nitrate", "ammonium_nitrate", "synthetic"} else "other"
    direct_ef_n2on_per_n = FERTILIZER_DIRECT_EF[direct_group][climate]
    direct_n2o_kg = nitrogen_kg * direct_ef_n2on_per_n * N2O_N_TO_N2O

    volatilization_fraction = FERTILIZER_VOLATILIZATION.get(input_type, FERTILIZER_VOLATILIZATION["generic"])
    ef4 = 0.014 if climate == "wet" else 0.005
    volatilization_n2o_kg = nitrogen_kg * volatilization_fraction * ef4 * N2O_N_TO_N2O if include_volatilization else 0.0

    frac_leach = 0.24 if climate == "wet" else 0.0
    ef5 = 0.011
    leaching_n2o_kg = nitrogen_kg * frac_leach * ef5 * N2O_N_TO_N2O if include_leaching else 0.0

    total_n2o_kg = direct_n2o_kg + volatilization_n2o_kg + leaching_n2o_kg
    co2e_kg = total_n2o_kg * GWP_AR6_N2O
    return {
        "nitrogen_kg": nitrogen_kg,
        "climate": climate,
        "input_type": input_type,
        "direct_ef_n2on_per_n": direct_ef_n2on_per_n,
        "direct_n2o_kg": direct_n2o_kg,
        "volatilization_fraction": volatilization_fraction,
        "ef4": ef4,
        "volatilization_n2o_kg": volatilization_n2o_kg,
        "frac_leach": frac_leach,
        "ef5": ef5,
        "leaching_n2o_kg": leaching_n2o_kg,
        "total_n2o_kg": total_n2o_kg,
        "gwp": GWP_AR6_N2O,
        "co2e_kg": co2e_kg,
        "co2e_t": co2e_kg / 1000.0,
        "warning": "La herramienta estima emisiones por aplicación de N al suelo. No representa la huella de fabricación del fertilizante ni cambios de carbono del suelo.",
    }


def calculate_biogas_balance(
    produced_m3: float,
    used_m3: float,
    flared_m3: float,
    vented_m3: float,
    methane_fraction: float,
    leakage_percent: float,
    methane_density_kg_m3: float = METHANE_DENSITY_KG_M3,
) -> dict[str, Any]:
    produced_m3 = _non_negative(produced_m3, "El biogás producido")
    used_m3 = _non_negative(used_m3, "El biogás utilizado")
    flared_m3 = _non_negative(flared_m3, "El biogás quemado")
    vented_m3 = _non_negative(vented_m3, "El biogás venteado")
    methane_fraction = float(methane_fraction)
    leakage_percent = float(leakage_percent)
    methane_density_kg_m3 = _non_negative(methane_density_kg_m3, "La densidad de CH4")
    if not 0 <= methane_fraction <= 1:
        raise ValueError("La fracción de metano debe estar entre 0 y 1")
    if not 0 <= leakage_percent <= 100:
        raise ValueError("La fuga debe estar entre 0 y 100 %")
    assigned = used_m3 + flared_m3 + vented_m3
    unassigned = produced_m3 - assigned
    leaked_biogas_m3 = produced_m3 * leakage_percent / 100.0
    emitted_biogas_m3 = vented_m3 + leaked_biogas_m3
    emitted_ch4_kg = emitted_biogas_m3 * methane_fraction * methane_density_kg_m3
    co2e_kg = emitted_ch4_kg * GWP_AR6_CH4
    return {
        "produced_m3": produced_m3,
        "used_m3": used_m3,
        "flared_m3": flared_m3,
        "vented_m3": vented_m3,
        "assigned_m3": assigned,
        "unassigned_m3": unassigned,
        "methane_fraction": methane_fraction,
        "leakage_percent": leakage_percent,
        "leaked_biogas_m3": leaked_biogas_m3,
        "emitted_biogas_m3": emitted_biogas_m3,
        "methane_density_kg_m3": methane_density_kg_m3,
        "emitted_ch4_kg": emitted_ch4_kg,
        "gwp": GWP_AR6_CH4,
        "co2e_kg": co2e_kg,
        "co2e_t": co2e_kg / 1000.0,
        "balanced": abs(unassigned) <= max(0.01, produced_m3 * 0.01),
        "warning": "La estimación no sustituye medición de caudal, composición, eficiencia de antorcha ni pruebas de fugas. El volumen usado o quemado no se cuenta como CH4 emitido en este balance simplificado.",
    }


def colombia_library_summary(session: Session) -> dict[str, Any]:
    documents = list(session.scalars(
        select(MethodologySourceDocument)
        .where(MethodologySourceDocument.code.in_(COLOMBIA_SOURCE_CODES))
        .order_by(MethodologySourceDocument.code)
    ))
    versions = list(session.scalars(
        select(EmissionFactorVersion)
        .options(selectinload(EmissionFactorVersion.factor), selectinload(EmissionFactorVersion.gas))
        .join(EmissionFactor)
        .where(EmissionFactor.country.in_(["Colombia", "Internacional"]))
        .order_by(EmissionFactor.activity_type, EmissionFactor.name, EmissionFactorVersion.version)
    ))
    documented: list[dict[str, Any]] = []
    for version in versions:
        doc = session.scalar(select(FactorDocumentation).where(FactorDocumentation.factor_version_id == version.id).options(selectinload(FactorDocumentation.source_document)))
        if not doc or not doc.source_document or doc.source_document.code not in COLOMBIA_SOURCE_CODES:
            continue
        documented.append({"factor": version.factor, "version": version, "gas": version.gas, "documentation": doc, "source": doc.source_document})
    cases = list(session.scalars(select(ReferenceCalculationCase).where(ReferenceCalculationCase.code >= "REF-013").order_by(ReferenceCalculationCase.code)))
    counts = {
        "documents": len(documents),
        "factors": len(documented),
        "formal": sum(1 for item in documented if item["documentation"].reporting_use == "Formal"),
        "pilot": sum(1 for item in documented if item["documentation"].reporting_use == "Piloto"),
        "reference_cases": len(cases),
    }
    return {
        "version": LIBRARY_VERSION,
        "documents": documents,
        "factors": documented,
        "cases": cases,
        "counts": counts,
        "fuel_options": FUEL_CALCULATOR_FACTORS,
        "limitations": [
            "Los factores del Decreto 926 son equivalencias regulatorias y se mantienen como uso piloto condicionado; no reemplazan una evaluación metodológica del inventario corporativo.",
            "Los valores FECOC B10/E10 fueron transcritos desde una fuente secundaria que cita FECOC 2016 y permanecen en revisión documental.",
            "El cálculo de aguas residuales exige carga orgánica y MCF del sistema real; el consumo de agua por sí solo no basta.",
            "Las emisiones por fertilizantes corresponden a aplicación de nitrógeno al suelo y no a fabricación del producto.",
            "El balance de biogás es una herramienta operativa; las mediciones de planta tienen prioridad.",
        ],
    }


def build_colombia_workbook(summary: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Factores"
    ws.append(["Factor", "Versión", "Gas", "Valor", "Unidad entrada", "Unidad salida", "Uso", "Revisión", "Fuente", "Página/tabla", "Restricciones"])
    for item in summary["factors"]:
        version = item["version"]
        doc = item["documentation"]
        ws.append([item["factor"].name, version.version, item["gas"].code, version.value, version.input_unit, version.output_unit, doc.reporting_use, doc.review_status, item["source"].code, f"{doc.page_reference} {doc.table_reference}".strip(), doc.restriction_notes])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    docs = wb.create_sheet("Fuentes")
    docs.append(["Código", "Título", "Entidad", "Fecha", "Jurisdicción", "Estado", "URL", "Citación", "Notas"])
    for item in summary["documents"]:
        docs.append([item.code, item.title, item.issuing_body, item.publication_date.isoformat() if item.publication_date else "", item.jurisdiction, item.status, item.source_url, item.citation, item.notes])

    refs = wb.create_sheet("Casos patrón")
    refs.append(["Código", "Título", "Categoría", "Actividad", "Unidad", "Factor", "Unidad factor", "Gas", "GWP", "kg CO2e esperados", "Fuente"])
    for case in summary["cases"]:
        refs.append([case.code, case.title, case.category, case.activity_value, case.activity_unit, case.factor_value, case.factor_input_unit, case.gas_code, case.gwp_value, case.expected_co2e_kg, case.source_reference])

    limits = wb.create_sheet("Limitaciones")
    limits.append(["Limitación metodológica"])
    for item in summary["limitations"]:
        limits.append([item])

    for sheet in wb.worksheets:
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(55, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def summary_json(summary: dict[str, Any]) -> str:
    payload = {
        "version": summary["version"],
        "counts": summary["counts"],
        "documents": [{"code": item.code, "title": item.title, "issuing_body": item.issuing_body, "status": item.status, "source_url": item.source_url} for item in summary["documents"]],
        "factors": [{"name": item["factor"].name, "version": item["version"].version, "gas": item["gas"].code, "value": item["version"].value, "input_unit": item["version"].input_unit, "reporting_use": item["documentation"].reporting_use, "review_status": item["documentation"].review_status, "source_code": item["source"].code} for item in summary["factors"]],
        "limitations": summary["limitations"],
    }
    return json.dumps(payload, ensure_ascii=False, indent=2)
