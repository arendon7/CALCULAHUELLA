from __future__ import annotations

from datetime import date
from io import BytesIO
import json

from fastapi import Depends, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from .database import (
    CustomerOnboardingItem,
    Facility,
    Inventory,
    InventoryFacility,
    Organization,
    add_audit,
    get_db,
)
from .guided_onboarding import (
    ASSURANCE_OPTIONS,
    DATA_SYSTEM_OPTIONS,
    FREQUENCIES,
    OBJECTIVES,
    READINESS_OPTIONS,
    SCOPE_AMBITIONS,
    SECTOR_FAMILIES,
    data_checklist,
    decision_plan,
    load_profile,
    profile_completion,
    save_profile,
)
from .inventory_starters import add_starter_sources
from .repositories.organizations import list_active_facilities
from .services.inventories import create_inventory as create_inventory_record


_OBJECTIVE_CODES = {item.code for item in OBJECTIVES}
_SECTOR_CODES = {item.code for item in SECTOR_FAMILIES}
_SCOPE_CODES = {item.code for item in SCOPE_AMBITIONS}
_READINESS_CODES = {item.code for item in READINESS_OPTIONS}
_DATA_SYSTEM_CODES = {item.code for item in DATA_SYSTEM_OPTIONS}


def _latest_inventory(session: Session, organization_id: int) -> Inventory | None:
    return session.scalar(
        select(Inventory)
        .where(Inventory.organization_id == organization_id)
        .options(
            selectinload(Inventory.sources),
            selectinload(Inventory.facility_links),
        )
        .order_by(Inventory.start_date.desc(), Inventory.id.desc())
        .limit(1)
    )


def _parse_iso_date(value: str, field_label: str) -> date:
    try:
        return date.fromisoformat(value)
    except (TypeError, ValueError) as exc:
        raise HTTPException(400, f"{field_label} no tiene una fecha válida") from exc


def _update_onboarding_item(session: Session, organization_id: int, code: str, status: str, actor: str) -> None:
    row = session.scalar(
        select(CustomerOnboardingItem).where(
            CustomerOnboardingItem.organization_id == organization_id,
            CustomerOnboardingItem.code == code,
        )
    )
    if not row:
        return
    row.status = status
    row.updated_by = actor


def _style_workbook(workbook: Workbook) -> None:
    navy = "173C4B"
    green = "2E7D4F"
    light = "EAF3ED"
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[1].height = 28
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index, column in enumerate(sheet.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[get_column_letter(index)].width = min(max(max_length + 2, 12), 42)
        if sheet.max_row >= 2:
            for cell in sheet[2]:
                cell.fill = PatternFill("solid", fgColor=light)
                cell.font = Font(color=green)


def register_guided_onboarding_routes(
    app,
    templates,
    common_context,
    require_user,
    ensure_capability,
    set_flash,
) -> None:
    @app.get("/onboarding/guiado", response_class=HTMLResponse)
    def guided_onboarding_page(
        request: Request,
        session: Session = Depends(get_db),
        user: dict = Depends(require_user),
    ):
        organization = session.get(Organization, int(user["organization_id"]))
        if not organization:
            raise HTTPException(404, "Organización no encontrada")
        profile = load_profile(session, organization)
        inventory = _latest_inventory(session, organization.id)
        plan = decision_plan(profile, organization, inventory=inventory)
        facilities = list_active_facilities(session, organization.id)
        selected_facilities = [link.facility_id for link in inventory.facility_links if link.included] if inventory else []
        return templates.TemplateResponse(
            request=request,
            name="guided_onboarding.html",
            context=common_context(
                request,
                session,
                user,
                "onboarding",
                profile=profile,
                plan=plan,
                inventory=inventory,
                facilities=facilities,
                selected_facilities=selected_facilities,
                objectives=OBJECTIVES,
                sectors=SECTOR_FAMILIES,
                scope_ambitions=SCOPE_AMBITIONS,
                readiness_options=READINESS_OPTIONS,
                data_system_options=DATA_SYSTEM_OPTIONS,
                frequencies=FREQUENCIES,
                assurance_options=ASSURANCE_OPTIONS,
            ),
        )

    @app.post("/onboarding/guiado/guardar")
    def guided_onboarding_save(
        request: Request,
        objective: str = Form(...),
        reporting_driver: str = Form(...),
        success_definition: str = Form(...),
        sector_family: str = Form(...),
        operating_description: str = Form(...),
        scope_ambition: str = Form(...),
        reporting_frequency: str = Form(...),
        assurance_ambition: str = Form(...),
        data_readiness: str = Form(...),
        evidence_readiness: str = Form(...),
        data_systems: list[str] = Form(default=[]),
        inventory_owner: str = Form(...),
        executive_sponsor: str = Form(...),
        period_start: str = Form(...),
        period_end: str = Form(...),
        notes: str = Form(""),
        session: Session = Depends(get_db),
        user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_inventory")
        organization = session.get(Organization, int(user["organization_id"]))
        if not organization:
            raise HTTPException(404, "Organización no encontrada")
        start = _parse_iso_date(period_start, "La fecha inicial")
        end = _parse_iso_date(period_end, "La fecha final")
        if end < start:
            raise HTTPException(400, "La fecha final no puede ser anterior a la fecha inicial")
        if objective not in _OBJECTIVE_CODES:
            raise HTTPException(400, "El objetivo seleccionado no es válido")
        if sector_family not in _SECTOR_CODES:
            raise HTTPException(400, "El perfil sectorial seleccionado no es válido")
        if scope_ambition not in _SCOPE_CODES:
            raise HTTPException(400, "La cobertura seleccionada no es válida")
        if data_readiness not in _READINESS_CODES or evidence_readiness not in _READINESS_CODES:
            raise HTTPException(400, "El nivel de preparación seleccionado no es válido")
        if reporting_frequency not in FREQUENCIES or assurance_ambition not in ASSURANCE_OPTIONS:
            raise HTTPException(400, "La frecuencia o el nivel de revisión no son válidos")
        clean_systems = sorted({item.strip() for item in data_systems if item.strip() in _DATA_SYSTEM_CODES})
        payload = {
            "objective": objective.strip(),
            "reporting_driver": reporting_driver.strip(),
            "success_definition": success_definition.strip(),
            "sector_family": sector_family.strip(),
            "operating_description": operating_description.strip(),
            "scope_ambition": scope_ambition.strip(),
            "reporting_frequency": reporting_frequency.strip(),
            "assurance_ambition": assurance_ambition.strip(),
            "data_readiness": data_readiness.strip(),
            "evidence_readiness": evidence_readiness.strip(),
            "data_systems": clean_systems,
            "inventory_owner": inventory_owner.strip(),
            "executive_sponsor": executive_sponsor.strip(),
            "period_start": start.isoformat(),
            "period_end": end.isoformat(),
            "notes": notes.strip(),
        }
        profile = save_profile(session, organization, payload, actor_email=str(user["email"]))
        _update_onboarding_item(session, organization.id, "ORG-01", "Completado", str(user["email"]))
        add_audit(
            session,
            organization.id,
            str(user["email"]),
            "CONFIGURAR",
            "Asistente inicial",
            organization.name,
            new_value=f"Perfil {profile_completion(profile)}% completo",
        )
        session.commit()
        set_flash(request, "El diagnóstico inicial quedó guardado. Revisa la ruta recomendada antes de aplicarla al inventario.")
        return RedirectResponse("/onboarding/guiado#recomendacion", status_code=303)

    @app.post("/onboarding/guiado/aplicar")
    def guided_onboarding_apply(
        request: Request,
        inventory_id: int | None = Form(None),
        facility_ids: list[int] = Form(default=[]),
        session: Session = Depends(get_db),
        user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_inventory")
        organization = session.get(Organization, int(user["organization_id"]))
        if not organization:
            raise HTTPException(404, "Organización no encontrada")
        profile = load_profile(session, organization)
        plan = decision_plan(profile, organization)
        if not plan["ready_to_apply"]:
            raise HTTPException(400, "Completa al menos el 75% del diagnóstico antes de aplicar la recomendación")
        start = _parse_iso_date(str(profile.get("period_start")), "La fecha inicial")
        end = _parse_iso_date(str(profile.get("period_end")), "La fecha final")
        allowed_facilities = list_active_facilities(session, organization.id)
        allowed_ids = {item.id for item in allowed_facilities}
        selected_ids = [item for item in facility_ids if item in allowed_ids]
        inventory = None
        if inventory_id:
            inventory = session.scalar(
                select(Inventory)
                .where(Inventory.id == inventory_id, Inventory.organization_id == organization.id)
                .options(selectinload(Inventory.sources), selectinload(Inventory.facility_links))
            )
            if not inventory:
                raise HTTPException(404, "Inventario no encontrado")
            if inventory.locked or inventory.status == "Cerrado":
                raise HTTPException(409, "El inventario está cerrado. Crea una nueva versión antes de aplicar cambios.")
            previous = f"{inventory.methodology} · {inventory.objective}"
            inventory.objective = f"{plan['objective']['label']}. {profile.get('success_definition', '')}"[:220]
            inventory.methodology = plan["methodology"]["methodology"]
            inventory.methodology_version = plan["methodology"]["methodology_version"]
            inventory.gwp_version = plan["methodology"]["gwp_version"]
            inventory.consolidation_approach = plan["methodology"]["consolidation_approach"]
            inventory.materiality_threshold = float(plan["methodology"]["materiality_threshold"])
            inventory.current_stage = "Fuentes" if inventory.sources else "Configuración"
            inventory.progress = max(inventory.progress, 20)
            recommendation_note = (
                "Recomendación V0.52 aplicada desde el asistente inicial. "
                f"Perfil: {plan['sector']['label']}; ambición: {plan['scope']['label']}."
            )
            if recommendation_note not in inventory.notes:
                inventory.notes = "\n".join(item for item in (inventory.notes.strip(), recommendation_note) if item)
            for link in inventory.facility_links:
                if selected_ids:
                    link.included = link.facility_id in selected_ids
            existing_ids = {link.facility_id for link in inventory.facility_links}
            for facility_id in selected_ids:
                if facility_id not in existing_ids:
                    session.add(InventoryFacility(inventory_id=inventory.id, facility_id=facility_id, included=True, inclusion_percentage=100))
            add_audit(
                session,
                organization.id,
                str(user["email"]),
                "APLICAR",
                "Recomendación de onboarding",
                inventory.name,
                previous_value=previous,
                new_value=f"{inventory.methodology} · {inventory.objective}",
            )
        else:
            inventory = create_inventory_record(
                session,
                organization.id,
                actor_email=str(user["email"]),
                name=f"Inventario corporativo {start.year}",
                start_date=start,
                end_date=end,
                objective=f"{plan['objective']['label']}. {profile.get('success_definition', '')}"[:220],
                base_year=start.year,
                methodology=plan["methodology"]["methodology"],
                methodology_version=plan["methodology"]["methodology_version"],
                gwp_version=plan["methodology"]["gwp_version"],
                consolidation_approach=plan["methodology"]["consolidation_approach"],
                materiality_threshold=float(plan["methodology"]["materiality_threshold"]),
                notes=(
                    "Creado desde el asistente inicial V0.52. "
                    f"Perfil: {plan['sector']['label']}; ambición: {plan['scope']['label']}."
                ),
                facility_ids=selected_ids,
            )
            session.flush()

        primary_facility_id = selected_ids[0] if selected_ids else None
        created_sources = add_starter_sources(
            session,
            inventory,
            pack_code=plan["starter_pack"]["code"],
            responsible=str(profile.get("inventory_owner") or "Responsable del inventario"),
            actor_email=str(user["email"]),
            facility_id=primary_facility_id,
        )
        _update_onboarding_item(session, organization.id, "MET-01", "Completado", str(user["email"]))
        _update_onboarding_item(session, organization.id, "DAT-01", "En progreso", str(user["email"]))
        session.commit()
        message = "La ruta recomendada fue aplicada"
        if created_sources:
            message += f" y se agregaron {len(created_sources)} fuentes iniciales"
        set_flash(request, message + ". Confirma cada fuente antes de cargar datos.")
        return RedirectResponse(f"/inventarios/{inventory.id}/fuentes", status_code=303)

    @app.get("/onboarding/guiado/checklist.xlsx")
    def guided_onboarding_checklist(
        session: Session = Depends(get_db),
        user: dict = Depends(require_user),
    ):
        organization = session.get(Organization, int(user["organization_id"]))
        if not organization:
            raise HTTPException(404, "Organización no encontrada")
        profile = load_profile(session, organization)
        inventory = _latest_inventory(session, organization.id)
        plan = decision_plan(profile, organization, inventory=inventory)
        checklist = data_checklist(profile, organization)

        workbook = Workbook()
        route_sheet = workbook.active
        route_sheet.title = "Ruta de trabajo"
        route_sheet.append(["Etapa", "Actividad", "Estado", "Resultado esperado", "Ruta en plataforma"])
        for item in plan["route"]:
            route_sheet.append([item["number"], item["title"], item["status"], item["result"], item["route"]])

        data_sheet = workbook.create_sheet("Datos requeridos")
        data_sheet.append(["Fuente", "Alcance", "Categoría", "Frecuencia", "Unidad inicial", "Evidencia esperada", "Responsable", "Prioridad"])
        for item in checklist:
            data_sheet.append([
                item["source"], item["scope"], item["category"], item["frequency"], item["unit"],
                item["evidence"], item["owner"], item["priority"],
            ])

        decision_sheet = workbook.create_sheet("Decisiones iniciales")
        decision_sheet.append(["Decisión", "Recomendación", "Aclaración"])
        decision_sheet.append(["Propósito", plan["objective"]["label"], plan["objective"]["description"]])
        decision_sheet.append(["Perfil sectorial", plan["sector"]["label"], plan["sector"]["description"]])
        decision_sheet.append(["Cobertura", plan["scope"]["label"], "La inclusión definitiva depende de relevancia, límites y disponibilidad de datos."])
        decision_sheet.append(["Metodología", plan["methodology"]["methodology"], plan["methodology"]["review_level"]])
        decision_sheet.append(["GWP", plan["methodology"]["gwp_version"], "Debe mantenerse consistente dentro de la versión del inventario."])
        decision_sheet.append(["Fuentes iniciales", plan["starter_pack"]["name"], "El paquete no asigna factores ni calcula automáticamente."])
        decision_sheet.append(["Advertencia", "Inventario bruto separado de reducción y emisiones evitadas", "No compensar emisiones del inventario con beneficios estimados dentro del mismo total."])
        _style_workbook(workbook)

        buffer = BytesIO()
        workbook.save(buffer)
        safe_name = "".join(character if character.isalnum() else "_" for character in organization.name.casefold())
        safe_name = "_".join(part for part in safe_name.split("_") if part)[:45] or "organizacion"
        filename = f"lista_inicial_datos_{safe_name}.xlsx"
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @app.get("/api/onboarding/guiado")
    def guided_onboarding_api(
        session: Session = Depends(get_db),
        user: dict = Depends(require_user),
    ):
        organization = session.get(Organization, int(user["organization_id"]))
        if not organization:
            raise HTTPException(404, "Organización no encontrada")
        profile = load_profile(session, organization)
        inventory = _latest_inventory(session, organization.id)
        plan = decision_plan(profile, organization, inventory=inventory)
        return JSONResponse({"profile": profile, "plan": plan, "data_checklist": data_checklist(profile, organization)})
