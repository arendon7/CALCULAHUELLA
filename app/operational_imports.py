from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from datetime import UTC, date, datetime
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload

from .calculations import convert_value, recalculate_source
from .database import (
    ActivityData,
    DataImportBatch,
    DataImportRow,
    DataQualityFinding,
    EmissionSource,
    Facility,
    Inventory,
    OperationalImportProfile,
    add_audit,
    refresh_progress,
)
from .period_close import assert_periods_editable

OPERATIONAL_IMPORT_VERSION = "0.45.5"
MAX_IMPORT_ROWS = 50_000
ALLOWED_ORIGINS = {
    "Medición directa",
    "Factura",
    "Registro operativo",
    "Registro contable",
    "Información de proveedor",
    "Certificado",
    "Encuesta",
    "Estimación",
}
MAPPING_FIELDS = (
    "source",
    "facility",
    "period_start",
    "period_end",
    "value",
    "unit",
    "origin",
    "estimated",
    "evidence",
    "notes",
)
FIELD_LABELS = {
    "source": "Fuente o ID de fuente",
    "facility": "Sede",
    "period_start": "Fecha inicial",
    "period_end": "Fecha final",
    "value": "Valor",
    "unit": "Unidad",
    "origin": "Origen del dato",
    "estimated": "Dato estimado",
    "evidence": "Referencia de evidencia",
    "notes": "Notas",
}
HEADER_ALIASES = {
    "source": ("fuente id", "id fuente", "codigo fuente", "código fuente", "fuente", "source", "source id"),
    "facility": ("sede", "instalacion", "instalación", "facility", "planta"),
    "period_start": ("fecha inicio", "fecha inicial", "periodo inicio", "período inicio", "period start", "fecha"),
    "period_end": ("fecha fin", "fecha final", "periodo fin", "período fin", "period end"),
    "value": ("valor", "cantidad", "consumo", "dato actividad", "activity value", "value"),
    "unit": ("unidad", "unit", "uom"),
    "origin": ("origen", "origen del dato", "data origin", "fuente dato"),
    "estimated": ("estimado", "es estimado", "estimated"),
    "evidence": ("evidencia", "referencia evidencia", "soporte", "factura", "evidence"),
    "notes": ("notas", "observaciones", "comentarios", "notes"),
}


def _normalize_text(value: Any) -> str:
    text = str(value or "").strip().casefold()
    return re.sub(r"\s+", " ", text)


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


def _parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_bool(value: Any, default: bool = False) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    return _normalize_text(value) in {"1", "si", "sí", "s", "true", "yes", "x", "estimado"}


def _parse_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace("\u00a0", "").replace(" ", "")
    if not text:
        return None
    # Supports 1.234,56 and 1,234.56 without silently changing simple decimals.
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        parts = text.split(",")
        text = "".join(parts) if len(parts[-1]) == 3 and len(parts) > 1 else text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _quality_level(origin: str, estimated: bool, evidence: str) -> str:
    if origin == "Medición directa" and evidence and not estimated:
        return "A"
    if origin in {"Factura", "Registro operativo", "Registro contable", "Información de proveedor", "Certificado"} and not estimated:
        return "B"
    if estimated or origin in {"Encuesta", "Estimación"}:
        return "C"
    return "D"


def _dedupe_headers(values: Iterable[Any]) -> list[str]:
    result: list[str] = []
    counts: dict[str, int] = {}
    for index, value in enumerate(values, 1):
        base = str(value or "").strip() or f"Columna {index}"
        count = counts.get(base, 0) + 1
        counts[base] = count
        result.append(base if count == 1 else f"{base} ({count})")
    return result


def _csv_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("El CSV no tiene una codificación reconocible.")


def _detect_csv_delimiter(text: str) -> str:
    candidates = (",", ";", "\t", "|")
    non_empty_lines = [line for line in text.splitlines() if line.strip()]
    header = non_empty_lines[0] if non_empty_lines else ""
    header_counts = {candidate: header.count(candidate) for candidate in candidates}
    strongest = max(candidates, key=lambda candidate: header_counts[candidate])
    if header_counts[strongest] > 0:
        return strongest
    sample = "\n".join(non_empty_lines[:20])
    try:
        return csv.Sniffer().sniff(sample, delimiters="".join(candidates)).delimiter
    except csv.Error:
        return ","


def inspect_import_file(
    content: bytes,
    filename: str,
    *,
    sheet_name: str = "",
    delimiter: str = "auto",
    header_row: int = 1,
    preview_rows: int = 10,
) -> dict[str, Any]:
    extension = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if header_row < 1 or header_row > 100:
        raise ValueError("La fila de encabezados debe estar entre 1 y 100.")
    rows: list[dict[str, Any]] = []
    sheets: list[str] = []
    used_delimiter = delimiter

    if extension == "csv":
        text = _csv_text(content)
        sample = text[:8192]
        if delimiter in {"", "auto"}:
            used_delimiter = _detect_csv_delimiter(sample)
        if used_delimiter not in {",", ";", "\t", "|"}:
            raise ValueError("El separador CSV seleccionado no es válido.")
        reader = csv.reader(io.StringIO(text), delimiter=used_delimiter)
        raw_rows = list(reader)
        if len(raw_rows) < header_row:
            raise ValueError("El CSV no contiene la fila de encabezados indicada.")
        headers = _dedupe_headers(raw_rows[header_row - 1])
        for number, values in enumerate(raw_rows[header_row:], header_row + 1):
            if not any(str(value or "").strip() for value in values):
                continue
            payload = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
            rows.append({"row_number": number, "payload": payload})
            if len(rows) >= MAX_IMPORT_ROWS:
                break
        source_format = "CSV"
    elif extension == "xlsx":
        try:
            workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        except Exception as exc:
            raise ValueError("El archivo XLSX está dañado o no puede leerse.") from exc
        sheets = list(workbook.sheetnames)
        selected_sheet = sheet_name if sheet_name in sheets else workbook.active.title
        worksheet = workbook[selected_sheet]
        raw = worksheet.iter_rows(values_only=True)
        headers_raw = None
        for current, values in enumerate(raw, 1):
            if current == header_row:
                headers_raw = values
                break
        if headers_raw is None:
            raise ValueError("El libro no contiene la fila de encabezados indicada.")
        headers = _dedupe_headers(headers_raw)
        for number, values in enumerate(raw, header_row + 1):
            if not any(value not in (None, "") for value in values):
                continue
            payload = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
            rows.append({"row_number": number, "payload": payload})
            if len(rows) >= MAX_IMPORT_ROWS:
                break
        sheet_name = selected_sheet
        source_format = "XLSX"
    else:
        raise ValueError("Solo se admiten archivos CSV o XLSX.")

    if not headers:
        raise ValueError("No se encontraron encabezados.")
    return {
        "source_format": source_format,
        "sheet_name": sheet_name,
        "delimiter": used_delimiter,
        "header_row": header_row,
        "headers": headers,
        "rows": rows,
        "preview": rows[:preview_rows],
        "sheets": sheets,
        "truncated": len(rows) >= MAX_IMPORT_ROWS,
    }


def suggest_mapping(headers: list[str]) -> dict[str, str]:
    normalized = {_normalize_text(header): header for header in headers}
    suggestions: dict[str, str] = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                suggestions[field] = normalized[alias]
                break
        if field in suggestions:
            continue
        for normalized_header, original in normalized.items():
            if any(alias in normalized_header for alias in aliases):
                suggestions[field] = original
                break
    return suggestions


def build_operational_template(inventory: Inventory) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Carga operativa"
    headers = [
        "Fuente ID",
        "Fuente",
        "Sede",
        "Fecha inicio",
        "Fecha fin",
        "Valor",
        "Unidad",
        "Origen",
        "Estimado",
        "Referencia evidencia",
        "Notas",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F6B49")
    for source in sorted(inventory.sources, key=lambda item: ((item.facility.name if item.facility else ""), item.name)):
        sheet.append([
            source.id,
            source.name,
            source.facility.name if source.facility else "",
            inventory.start_date,
            inventory.start_date,
            None,
            source.preferred_unit,
            "Registro operativo",
            "No",
            "",
            "Fila de ejemplo: duplicar y ajustar por periodo.",
        ])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [12, 34, 22, 14, 14, 14, 14, 24, 12, 32, 45]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width

    catalog = workbook.create_sheet("Catálogo de fuentes")
    catalog.append(["Fuente ID", "Sede", "Fuente", "Alcance", "Categoría", "Frecuencia", "Unidad preferida", "Materialidad", "Estado"])
    for source in sorted(inventory.sources, key=lambda item: item.id):
        catalog.append([
            source.id,
            source.facility.name if source.facility else "",
            source.name,
            source.scope,
            source.category,
            source.data_frequency,
            source.preferred_unit,
            source.materiality,
            source.status,
        ])
    for cell in catalog[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="285C4D")
    catalog.freeze_panes = "A2"
    catalog.auto_filter.ref = catalog.dimensions
    for index, width in enumerate([12, 22, 34, 10, 32, 16, 18, 16, 16], 1):
        catalog.column_dimensions[chr(64 + index)].width = width

    instructions = workbook.create_sheet("Instrucciones")
    instructions.append(["Paso", "Instrucción"])
    rows = [
        (1, "No cambies los ID de fuente. También puedes usar un archivo propio y mapear sus columnas en la plataforma."),
        (2, "Usa una fila por fuente y periodo. Los duplicados se rechazan, omiten o actualizan según la política elegida."),
        (3, "Los valores deben ser numéricos y no negativos. Los ceros deben justificarse en notas."),
        (4, "La unidad debe ser compatible con la unidad preferida de la fuente."),
        (5, "Las fuentes de materialidad alta deben incluir una referencia de evidencia."),
        (6, "La validación no modifica el inventario. El lote se aplica únicamente después de revisar los hallazgos."),
    ]
    for row in rows:
        instructions.append(row)
    instructions.column_dimensions["A"].width = 10
    instructions.column_dimensions["B"].width = 110

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _source_lookup(inventory: Inventory) -> tuple[dict[int, EmissionSource], dict[str, list[EmissionSource]]]:
    by_id = {source.id: source for source in inventory.sources}
    by_name: dict[str, list[EmissionSource]] = {}
    for source in inventory.sources:
        by_name.setdefault(_normalize_text(source.name), []).append(source)
    return by_id, by_name


def _resolve_source(
    raw_source: Any,
    raw_facility: Any,
    inventory: Inventory,
    default_source_id: int | None,
) -> tuple[EmissionSource | None, str | None]:
    by_id, by_name = _source_lookup(inventory)
    value = str(raw_source or "").strip()
    if not value and default_source_id:
        source = by_id.get(default_source_id)
        return (source, None) if source else (None, "La fuente predeterminada no pertenece al inventario.")
    if not value:
        return None, "Falta la fuente o su ID."

    numeric_match = re.match(r"^\s*(\d+)(?:\s*[|\-:].*)?$", value)
    if numeric_match:
        source = by_id.get(int(numeric_match.group(1)))
        if source:
            return source, None

    candidates = by_name.get(_normalize_text(value), [])
    if len(candidates) == 1:
        return candidates[0], None
    facility = _normalize_text(raw_facility)
    if facility and candidates:
        narrowed = [source for source in candidates if source.facility and _normalize_text(source.facility.name) == facility]
        if len(narrowed) == 1:
            return narrowed[0], None
    if len(candidates) > 1:
        return None, "La fuente es ambigua; agrega la sede o usa el ID de fuente."
    return None, f"No existe una fuente del inventario que coincida con '{value}'."


def _mapped(payload: dict[str, Any], mapping: dict[str, str], field: str, default: Any = "") -> Any:
    header = mapping.get(field, "")
    if not header:
        return default
    return payload.get(header, default)


def _finding(session: Session, batch: DataImportBatch, row: DataImportRow | None, rule: str, severity: str, message: str) -> None:
    session.add(DataQualityFinding(batch=batch, row=row, rule_code=rule, severity=severity, message=message, status="Abierto"))


def _row_fingerprint(source_id: int | None, start: date | None, end: date | None, value: float | None, unit: str) -> str:
    raw = f"{source_id or ''}|{start or ''}|{end or ''}|{value if value is not None else ''}|{unit.strip().casefold()}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _batch_defaults(batch: DataImportBatch) -> dict[str, Any]:
    try:
        payload = json.loads(batch.mapping_json or "{}")
    except json.JSONDecodeError:
        payload = {}
    defaults = payload.get("defaults", {}) if isinstance(payload, dict) else {}
    if not isinstance(defaults, dict):
        defaults = {}
    return defaults


def _refresh_batch_validation_totals(session: Session, batch: DataImportBatch) -> None:
    rows = list(batch.rows)
    row_errors = sum(row.status == "Error" for row in rows)
    row_warnings = sum(row.status in {"Advertencia", "Omitir"} for row in rows)
    valid_rows = sum(row.status == "Válido" for row in rows)
    global_findings = list(session.scalars(select(DataQualityFinding).where(
        DataQualityFinding.batch_id == batch.id,
        DataQualityFinding.row_id.is_(None),
        DataQualityFinding.status == "Abierto",
    )))
    global_errors = sum(item.severity == "Error" for item in global_findings)
    global_warnings = sum(item.severity != "Error" for item in global_findings)
    batch.total_rows = len(rows)
    batch.valid_rows = valid_rows
    batch.warning_rows = row_warnings + global_warnings
    batch.error_rows = row_errors + global_errors
    denominator = max(len(rows), 1)
    batch.quality_score = max(0, round((valid_rows + row_warnings * 0.6) / denominator * 100))
    batch.status = "Con errores" if batch.error_rows else "Validado"
    batch.validated_at = datetime.now(UTC)


def update_operational_row(
    session: Session,
    *,
    organization_id: int,
    batch_id: int,
    row_id: int,
    user_email: str,
    source_id: Any,
    period_start: Any,
    period_end: Any,
    value: Any,
    unit: Any,
    origin: Any,
    estimated: Any = False,
    evidence: Any = "",
    notes: Any = "",
) -> DataImportRow:
    batch = session.scalar(
        select(DataImportBatch)
        .where(
            DataImportBatch.id == batch_id,
            DataImportBatch.organization_id == organization_id,
            DataImportBatch.execution_id.is_(None),
        )
        .options(
            selectinload(DataImportBatch.rows).selectinload(DataImportRow.source),
            selectinload(DataImportBatch.inventory).selectinload(Inventory.sources),
        )
    )
    if not batch or not batch.inventory:
        raise ValueError("Lote operativo no encontrado.")
    if batch.status == "Aplicado":
        raise ValueError("Un lote aplicado ya no puede corregirse.")
    if batch.inventory.locked or batch.inventory.status == "Cerrado":
        raise ValueError("El inventario está cerrado y no admite correcciones.")
    row = next((item for item in batch.rows if item.id == row_id), None)
    if not row:
        raise ValueError("La fila no pertenece al lote seleccionado.")

    parsed_source_id = int(source_id) if str(source_id or "").isdigit() else None
    source = next((item for item in batch.inventory.sources if item.id == parsed_source_id), None)
    start = _parse_date(period_start)
    end = _parse_date(period_end) or start
    parsed_value = _parse_float(value)
    clean_unit = str(unit or "").strip()
    clean_origin = str(origin or "Registro operativo").strip()
    is_estimated = _parse_bool(estimated, False)
    clean_evidence = str(evidence or "").strip()
    clean_notes = str(notes or "").strip()
    messages: list[dict[str, str]] = []

    if not source:
        messages.append({"rule": "OP-001", "severity": "Error", "message": "Selecciona una fuente válida del inventario."})
    if not start:
        messages.append({"rule": "OP-002", "severity": "Error", "message": "La fecha inicial está vacía o no es válida."})
    if start and end and end < start:
        messages.append({"rule": "OP-003", "severity": "Error", "message": "La fecha final no puede ser anterior a la fecha inicial."})
    if start and end and (start < batch.inventory.start_date or end > batch.inventory.end_date):
        messages.append({"rule": "OP-004", "severity": "Error", "message": "El periodo está por fuera de las fechas del inventario."})
    if parsed_value is None:
        messages.append({"rule": "OP-005", "severity": "Error", "message": "El valor está vacío o no es numérico."})
    elif parsed_value < 0:
        messages.append({"rule": "OP-006", "severity": "Error", "message": "No se admiten valores negativos."})
    elif parsed_value == 0 and not clean_notes:
        messages.append({"rule": "OP-007", "severity": "Advertencia", "message": "El valor cero debe justificarse en notas."})
    if not clean_unit:
        messages.append({"rule": "OP-008", "severity": "Error", "message": "Falta la unidad."})
    if clean_origin not in ALLOWED_ORIGINS:
        messages.append({"rule": "OP-009", "severity": "Advertencia", "message": f"Origen '{clean_origin}' no reconocido; se usará Registro operativo."})
        clean_origin = "Registro operativo"
    if source and clean_unit and parsed_value is not None:
        preferred = source.preferred_unit or clean_unit
        normalized, explanation = convert_value(session, parsed_value, clean_unit, preferred)
        if normalized is None:
            messages.append({"rule": "OP-010", "severity": "Error", "message": explanation or f"No se puede convertir {clean_unit} a {preferred}."})
    if source and source.materiality == "Alta" and not clean_evidence:
        messages.append({"rule": "OP-011", "severity": "Advertencia", "message": "La fuente de materialidad alta no tiene referencia de evidencia."})

    existing: ActivityData | None = None
    defaults = _batch_defaults(batch)
    duplicate_policy = str(defaults.get("duplicate_policy") or "reject").strip().lower()
    if duplicate_policy not in {"reject", "skip", "update"}:
        duplicate_policy = "reject"
    if source and start and end:
        duplicate_row = session.scalar(select(DataImportRow).where(
            DataImportRow.batch_id == batch.id,
            DataImportRow.id != row.id,
            DataImportRow.source_id == source.id,
            DataImportRow.period_start == start,
            DataImportRow.period_end == end,
        ).limit(1))
        if duplicate_row:
            messages.append({"rule": "OP-012", "severity": "Error", "message": "La misma fuente y periodo están repetidos dentro del lote."})
        existing = session.scalar(select(ActivityData).where(
            ActivityData.source_id == source.id,
            ActivityData.period_start == start,
            ActivityData.period_end == end,
        ))
        if existing:
            if duplicate_policy == "reject":
                messages.append({"rule": "OP-013", "severity": "Error", "message": "Ya existe un dato para la fuente y periodo. Cambia la política o corrige el periodo."})
            elif duplicate_policy == "skip":
                messages.append({"rule": "OP-013", "severity": "Advertencia", "message": "El registro existente será omitido al aplicar el lote."})
            else:
                messages.append({"rule": "OP-013", "severity": "Advertencia", "message": "El registro existente será actualizado al aplicar el lote."})
        try:
            assert_periods_editable(session, batch.inventory.id, [(start, end)])
        except ValueError as exc:
            messages.append({"rule": "OP-014", "severity": "Error", "message": str(exc)})

    status = "Válido"
    if any(message["severity"] == "Error" for message in messages):
        status = "Error"
    elif duplicate_policy == "skip" and existing:
        status = "Omitir"
    elif messages:
        status = "Advertencia"

    old_findings = list(session.scalars(select(DataQualityFinding).where(
        DataQualityFinding.batch_id == batch.id,
        DataQualityFinding.row_id == row.id,
        DataQualityFinding.status == "Abierto",
    )))
    for finding in old_findings:
        finding.status = "Resuelto"
        finding.resolution = "Fila corregida y revalidada desde el asistente de carga."
        finding.resolved_at = datetime.now(UTC)

    try:
        raw_payload = json.loads(row.raw_payload_json or "{}")
    except json.JSONDecodeError:
        raw_payload = {}
    if not isinstance(raw_payload, dict):
        raw_payload = {}
    raw_payload["__correction__"] = {
        "source_id": parsed_source_id,
        "period_start": str(start or ""),
        "period_end": str(end or ""),
        "value": parsed_value,
        "unit": clean_unit,
        "origin": clean_origin,
        "estimated": is_estimated,
        "evidence": clean_evidence,
        "notes": clean_notes,
        "corrected_by": user_email,
        "corrected_at": datetime.now(UTC).isoformat(),
    }

    row.requirement_code = str(parsed_source_id or row.requirement_code or "")[:80]
    row.source_id = source.id if source else None
    row.period_start = start
    row.period_end = end
    row.value = parsed_value
    row.unit = clean_unit
    row.evidence_reference = clean_evidence[:300]
    row.data_origin = clean_origin
    row.is_estimated = is_estimated
    row.quality_level = _quality_level(clean_origin, is_estimated, clean_evidence)
    row.status = status
    row.validation_messages = _json(messages)
    row.raw_payload_json = _json(raw_payload)
    row.row_fingerprint = _row_fingerprint(source.id if source else None, start, end, parsed_value, clean_unit)
    row.duplicate_of_activity_id = existing.id if existing else None
    for message in messages:
        _finding(session, batch, row, message["rule"], message["severity"], message["message"])

    _refresh_batch_validation_totals(session, batch)
    add_audit(
        session,
        organization_id,
        user_email,
        "CORREGIR",
        "Fila de lote operativo",
        f"{batch.code} · fila {row.row_number}",
        detail=f"Estado resultante: {row.status} · {len(messages)} hallazgos",
    )
    session.flush()
    return row


def create_operational_batch(
    session: Session,
    *,
    organization_id: int,
    inventory: Inventory,
    filename: str,
    content: bytes,
    user_email: str,
    mapping: dict[str, str],
    defaults: dict[str, Any] | None = None,
    sheet_name: str = "",
    delimiter: str = "auto",
    header_row: int = 1,
    profile_name: str = "",
    profile_id: int | None = None,
    save_profile: bool = False,
) -> DataImportBatch:
    defaults = defaults or {}
    required = {"source", "period_start", "value"}
    missing = [FIELD_LABELS[field] for field in required if not mapping.get(field) and not (field == "source" and defaults.get("source_id"))]
    if missing:
        raise ValueError("Falta mapear: " + ", ".join(missing) + ".")
    if inventory.organization_id != organization_id:
        raise ValueError("El inventario no pertenece a la organización activa.")
    if inventory.locked or inventory.status == "Cerrado":
        raise ValueError("El inventario está cerrado y no admite cargas.")

    inspection = inspect_import_file(content, filename, sheet_name=sheet_name, delimiter=delimiter, header_row=header_row, preview_rows=10)
    digest = hashlib.sha256(content).hexdigest()
    duplicate_batch = session.scalar(select(DataImportBatch).where(DataImportBatch.organization_id == organization_id, DataImportBatch.file_hash == digest))
    if duplicate_batch:
        raise ValueError(f"Este archivo ya fue cargado como {duplicate_batch.code}.")

    profile: OperationalImportProfile | None = None
    if profile_id:
        profile = session.scalar(select(OperationalImportProfile).where(
            OperationalImportProfile.id == profile_id,
            OperationalImportProfile.organization_id == organization_id,
        ))
        if not profile:
            raise ValueError("El perfil de importación no existe.")
    if save_profile:
        clean_name = profile_name.strip() or (profile.name if profile else "")
        if not clean_name:
            raise ValueError("Indica un nombre para guardar el perfil.")
        existing = session.scalar(select(OperationalImportProfile).where(
            OperationalImportProfile.organization_id == organization_id,
            OperationalImportProfile.name == clean_name,
        ))
        profile = existing or OperationalImportProfile(organization_id=organization_id, name=clean_name, created_by=user_email)
        profile.inventory_id = inventory.id
        profile.source_format = inspection["source_format"]
        profile.sheet_name = inspection["sheet_name"] or ""
        profile.delimiter = inspection["delimiter"] or ","
        profile.header_row = header_row
        profile.mapping_json = _json(mapping)
        profile.defaults_json = _json(defaults)
        profile.active = True
        session.add(profile)
        session.flush()

    count = session.scalar(select(func.count(DataImportBatch.id)).where(DataImportBatch.organization_id == organization_id)) or 0
    batch = DataImportBatch(
        organization_id=organization_id,
        inventory_id=inventory.id,
        import_profile_id=profile.id if profile else None,
        execution_id=None,
        code=f"OP-{inventory.start_date.year}-{count + 1:04d}",
        filename=filename[:220],
        file_hash=digest,
        source_format=inspection["source_format"],
        source_sheet=inspection["sheet_name"] or "",
        mapping_json=_json({"mapping": mapping, "defaults": defaults}),
        original_headers_json=_json(inspection["headers"]),
        status="Validando",
        uploaded_by=user_email,
        notes=f"Carga operativa V{OPERATIONAL_IMPORT_VERSION}",
    )
    session.add(batch)
    session.flush()

    duplicate_policy = str(defaults.get("duplicate_policy") or "reject").strip().lower()
    if duplicate_policy not in {"reject", "skip", "update"}:
        duplicate_policy = "reject"
    default_unit = str(defaults.get("unit") or "").strip()
    default_origin = str(defaults.get("origin") or "Registro operativo").strip()
    default_estimated = _parse_bool(defaults.get("estimated"), False)
    default_source_id = int(defaults["source_id"]) if str(defaults.get("source_id") or "").isdigit() else None

    seen_periods: set[tuple[int, date, date]] = set()
    error_rows = 0
    warning_rows = 0
    valid_rows = 0

    for source_row in inspection["rows"]:
        payload = source_row["payload"]
        raw_source = _mapped(payload, mapping, "source")
        raw_facility = _mapped(payload, mapping, "facility")
        source, source_error = _resolve_source(raw_source, raw_facility, inventory, default_source_id)
        start = _parse_date(_mapped(payload, mapping, "period_start"))
        end = _parse_date(_mapped(payload, mapping, "period_end")) or start
        value = _parse_float(_mapped(payload, mapping, "value"))
        unit = str(_mapped(payload, mapping, "unit", default_unit) or default_unit).strip()
        origin = str(_mapped(payload, mapping, "origin", default_origin) or default_origin).strip()
        estimated = _parse_bool(_mapped(payload, mapping, "estimated", default_estimated), default_estimated)
        evidence = str(_mapped(payload, mapping, "evidence") or "").strip()
        notes = str(_mapped(payload, mapping, "notes") or "").strip()
        messages: list[dict[str, str]] = []

        if source_error:
            messages.append({"rule": "OP-001", "severity": "Error", "message": source_error})
        if not start:
            messages.append({"rule": "OP-002", "severity": "Error", "message": "La fecha inicial está vacía o no es válida."})
        if start and end and end < start:
            messages.append({"rule": "OP-003", "severity": "Error", "message": "La fecha final no puede ser anterior a la fecha inicial."})
        if start and end and (start < inventory.start_date or end > inventory.end_date):
            messages.append({"rule": "OP-004", "severity": "Error", "message": "El periodo está por fuera de las fechas del inventario."})
        if value is None:
            messages.append({"rule": "OP-005", "severity": "Error", "message": "El valor está vacío o no es numérico."})
        elif value < 0:
            messages.append({"rule": "OP-006", "severity": "Error", "message": "No se admiten valores negativos."})
        elif value == 0 and not notes:
            messages.append({"rule": "OP-007", "severity": "Advertencia", "message": "El valor cero debe justificarse en notas."})
        if not unit:
            messages.append({"rule": "OP-008", "severity": "Error", "message": "Falta la unidad."})
        if origin not in ALLOWED_ORIGINS:
            messages.append({"rule": "OP-009", "severity": "Advertencia", "message": f"Origen '{origin}' no reconocido; se usará Registro operativo."})
            origin = "Registro operativo"
        if source and unit and value is not None:
            preferred = source.preferred_unit or unit
            normalized, explanation = convert_value(session, value, unit, preferred)
            if normalized is None:
                messages.append({"rule": "OP-010", "severity": "Error", "message": explanation or f"No se puede convertir {unit} a {preferred}."})
        if source and source.materiality == "Alta" and not evidence:
            messages.append({"rule": "OP-011", "severity": "Advertencia", "message": "La fuente de materialidad alta no tiene referencia de evidencia."})

        existing: ActivityData | None = None
        if source and start and end:
            key = (source.id, start, end)
            if key in seen_periods:
                messages.append({"rule": "OP-012", "severity": "Error", "message": "La misma fuente y periodo están repetidos dentro del archivo."})
            else:
                seen_periods.add(key)
            existing = session.scalar(select(ActivityData).where(
                ActivityData.source_id == source.id,
                ActivityData.period_start == start,
                ActivityData.period_end == end,
            ))
            if existing:
                if duplicate_policy == "reject":
                    messages.append({"rule": "OP-013", "severity": "Error", "message": "Ya existe un dato para la fuente y periodo. Cambia la política o corrige el archivo."})
                elif duplicate_policy == "skip":
                    messages.append({"rule": "OP-013", "severity": "Advertencia", "message": "El registro existente será omitido al aplicar el lote."})
                else:
                    messages.append({"rule": "OP-013", "severity": "Advertencia", "message": "El registro existente será actualizado al aplicar el lote."})
            try:
                assert_periods_editable(session, inventory.id, [(start, end)])
            except ValueError as exc:
                messages.append({"rule": "OP-014", "severity": "Error", "message": str(exc)})

        status = "Válido"
        if any(message["severity"] == "Error" for message in messages):
            status = "Error"
            error_rows += 1
        elif duplicate_policy == "skip" and existing:
            status = "Omitir"
            warning_rows += 1
        elif messages:
            status = "Advertencia"
            warning_rows += 1
        else:
            valid_rows += 1

        row = DataImportRow(
            batch=batch,
            row_number=int(source_row["row_number"]),
            requirement_code=str(raw_source or "")[:80],
            source_id=source.id if source else None,
            period_start=start,
            period_end=end,
            value=value,
            unit=unit,
            evidence_reference=evidence[:300],
            data_origin=origin,
            is_estimated=estimated,
            quality_level=_quality_level(origin, estimated, evidence),
            status=status,
            validation_messages=_json(messages),
            raw_payload_json=_json(payload),
            row_fingerprint=_row_fingerprint(source.id if source else None, start, end, value, unit),
            duplicate_of_activity_id=existing.id if existing else None,
        )
        session.add(row)
        session.flush()
        for message in messages:
            _finding(session, batch, row, message["rule"], message["severity"], message["message"])

    total_rows = error_rows + warning_rows + valid_rows
    if inspection["truncated"]:
        _finding(session, batch, None, "OP-015", "Error", f"El archivo supera el máximo de {MAX_IMPORT_ROWS:,} filas por lote.")
        error_rows += 1
    if total_rows == 0:
        _finding(session, batch, None, "OP-016", "Error", "El archivo no contiene filas de datos.")
        error_rows += 1

    batch.total_rows = total_rows
    batch.valid_rows = valid_rows
    batch.warning_rows = warning_rows
    batch.error_rows = error_rows
    denominator = max(total_rows, 1)
    batch.quality_score = max(0, round((valid_rows + warning_rows * 0.6) / denominator * 100))
    batch.status = "Con errores" if error_rows else "Validado"
    batch.validated_at = datetime.now(UTC)
    add_audit(
        session,
        organization_id,
        user_email,
        "VALIDAR",
        "Lote operativo",
        batch.code,
        detail=f"{total_rows} filas · {valid_rows} válidas · {warning_rows} advertencias · {error_rows} errores",
    )
    session.flush()
    return batch


def apply_operational_batch(session: Session, organization_id: int, batch_id: int, user_email: str) -> DataImportBatch:
    batch = session.scalar(
        select(DataImportBatch)
        .where(
            DataImportBatch.id == batch_id,
            DataImportBatch.organization_id == organization_id,
            DataImportBatch.execution_id.is_(None),
        )
        .options(
            selectinload(DataImportBatch.rows).selectinload(DataImportRow.source),
            selectinload(DataImportBatch.inventory).selectinload(Inventory.sources),
        )
    )
    if not batch or not batch.inventory:
        raise ValueError("Lote operativo no encontrado.")
    if batch.status == "Aplicado":
        return batch
    if batch.error_rows:
        raise ValueError("El lote contiene errores y no puede aplicarse.")
    if batch.inventory.locked or batch.inventory.status == "Cerrado":
        raise ValueError("El inventario está cerrado y no admite cargas.")

    defaults: dict[str, Any] = _batch_defaults(batch)
    if not defaults and batch.import_profile and batch.import_profile.defaults_json:
        try:
            defaults = json.loads(batch.import_profile.defaults_json)
        except json.JSONDecodeError:
            defaults = {}
    if not defaults:
        defaults = {"duplicate_policy": "update" if any(row.duplicate_of_activity_id for row in batch.rows) else "reject"}
    duplicate_policy = str(defaults.get("duplicate_policy") or "reject").lower()

    periods = [(row.period_start, row.period_end) for row in batch.rows if row.status in {"Válido", "Advertencia"} and row.period_start and row.period_end]
    assert_periods_editable(session, batch.inventory.id, periods)

    touched: set[int] = set()
    applied = 0
    omitted = 0
    for row in batch.rows:
        if row.status == "Omitir":
            row.status = "Omitido"
            omitted += 1
            continue
        if row.status not in {"Válido", "Advertencia"} or not row.source or row.value is None or not row.period_start or not row.period_end:
            continue
        preferred = row.source.preferred_unit or row.unit
        normalized, explanation = convert_value(session, row.value, row.unit, preferred)
        if normalized is None:
            raise ValueError(explanation or f"No se puede convertir {row.unit} a {preferred}.")

        record: ActivityData | None = None
        if row.duplicate_of_activity_id and duplicate_policy == "update":
            record = session.get(ActivityData, row.duplicate_of_activity_id)
        if record is None:
            record = ActivityData(
                source_id=row.source.id,
                period_start=row.period_start,
                period_end=row.period_end,
                value=normalized,
                unit=preferred,
                created_by=user_email,
            )
            session.add(record)
            session.flush()
        else:
            record.value = normalized
            record.unit = preferred
            record.updated_at = datetime.now(UTC)

        record.data_origin = row.data_origin
        record.quality_level = row.quality_level
        record.is_estimated = row.is_estimated
        record.status = "Provisional" if row.is_estimated or not row.evidence_reference else "Cargado"
        note_parts = [f"Lote operativo {batch.code}", f"Archivo {batch.filename}"]
        if row.evidence_reference:
            note_parts.append(f"Evidencia: {row.evidence_reference}")
        record.notes = " · ".join(note_parts)
        row.activity_data_id = record.id
        row.status = "Aplicado"
        touched.add(row.source.id)
        applied += 1

    # Persist inserts/updates before recalculation reloads source records with
    # populate_existing=True. Otherwise an updated duplicate can be restored
    # from the previous database value.
    session.flush()

    for source_id in touched:
        source = session.get(EmissionSource, source_id)
        if source:
            recalculate_source(session, source)
            for request in source.requests:
                request.status = "Completado" if source.progress >= 100 else "En curso"
                if request.status == "Completado":
                    request.completed_at = datetime.now(UTC)

    refresh_progress(session, batch.inventory)
    batch.applied_rows = applied
    batch.status = "Aplicado"
    batch.applied_at = datetime.now(UTC)
    add_audit(
        session,
        organization_id,
        user_email,
        "APLICAR",
        "Lote operativo",
        batch.code,
        detail=f"{applied} registros aplicados · {omitted} omitidos · {len(touched)} fuentes actualizadas",
    )
    session.flush()
    return batch


def operational_import_summary(session: Session, organization_id: int, inventory_id: int | None = None, batch_id: int | None = None) -> dict[str, Any]:
    inventories = list(session.scalars(
        select(Inventory).where(Inventory.organization_id == organization_id).order_by(Inventory.start_date.desc(), Inventory.id.desc())
    ))
    inventory_query = select(Inventory).where(Inventory.organization_id == organization_id).options(
        selectinload(Inventory.sources).selectinload(EmissionSource.facility),
        selectinload(Inventory.facility_links),
    )
    if inventory_id:
        inventory = session.scalar(inventory_query.where(Inventory.id == inventory_id).limit(1))
    else:
        inventory = session.scalar(
            inventory_query
            .where(Inventory.locked.is_(False), Inventory.status != "Cerrado")
            .order_by(Inventory.start_date.desc(), Inventory.id.desc())
            .limit(1)
        )
        if not inventory:
            inventory = session.scalar(
                inventory_query.order_by(Inventory.start_date.desc(), Inventory.id.desc()).limit(1)
            )
    if not inventory:
        raise ValueError("No existe un inventario para la organización activa.")

    profiles = list(session.scalars(
        select(OperationalImportProfile)
        .where(OperationalImportProfile.organization_id == organization_id, OperationalImportProfile.active.is_(True))
        .order_by(OperationalImportProfile.name)
    ))
    batches = list(session.scalars(
        select(DataImportBatch)
        .where(
            DataImportBatch.organization_id == organization_id,
            DataImportBatch.inventory_id == inventory.id,
            DataImportBatch.execution_id.is_(None),
        )
        .order_by(DataImportBatch.id.desc())
    ))
    selected = None
    target_id = batch_id or (batches[0].id if batches else None)
    if target_id:
        selected = session.scalar(
            select(DataImportBatch)
            .where(
                DataImportBatch.id == target_id,
                DataImportBatch.organization_id == organization_id,
                DataImportBatch.inventory_id == inventory.id,
                DataImportBatch.execution_id.is_(None),
            )
            .options(
                selectinload(DataImportBatch.rows).selectinload(DataImportRow.source).selectinload(EmissionSource.facility),
                selectinload(DataImportBatch.findings).selectinload(DataQualityFinding.row),
                selectinload(DataImportBatch.import_profile),
            )
        )
    if selected:
        for row in selected.rows:
            try:
                messages = json.loads(row.validation_messages or "[]")
            except json.JSONDecodeError:
                messages = []
            row.validation_items = messages
            row.validation_display = " · ".join(item.get("message", "") for item in messages if item.get("message"))
            try:
                raw_payload = json.loads(row.raw_payload_json or "{}")
            except json.JSONDecodeError:
                raw_payload = {}
            correction = raw_payload.get("__correction__", {}) if isinstance(raw_payload, dict) else {}
            row.correction_notes = correction.get("notes", "") if isinstance(correction, dict) else ""

    open_findings = session.scalar(
        select(func.count(DataQualityFinding.id))
        .join(DataImportBatch)
        .where(
            DataImportBatch.organization_id == organization_id,
            DataImportBatch.execution_id.is_(None),
            DataQualityFinding.status == "Abierto",
        )
    ) or 0
    return {
        "inventory": inventory,
        "inventories": inventories,
        "profiles": profiles,
        "batches": batches,
        "selected": selected,
        "metrics": {
            "profiles": len(profiles),
            "batches": len(batches),
            "rows": sum(batch.total_rows for batch in batches),
            "applied": sum(batch.applied_rows for batch in batches),
            "open_findings": open_findings,
        },
        "field_labels": FIELD_LABELS,
        "mapping_fields": MAPPING_FIELDS,
        "origins": sorted(ALLOWED_ORIGINS),
    }


def profile_payload(profile: OperationalImportProfile) -> dict[str, Any]:
    try:
        mapping = json.loads(profile.mapping_json or "{}")
    except json.JSONDecodeError:
        mapping = {}
    try:
        defaults = json.loads(profile.defaults_json or "{}")
    except json.JSONDecodeError:
        defaults = {}
    return {
        "id": profile.id,
        "name": profile.name,
        "inventory_id": profile.inventory_id,
        "source_format": profile.source_format,
        "sheet_name": profile.sheet_name,
        "delimiter": profile.delimiter,
        "header_row": profile.header_row,
        "mapping": mapping,
        "defaults": defaults,
    }


def batch_errors_csv(batch: DataImportBatch) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Lote", "Fila", "Fuente", "Periodo inicio", "Periodo fin", "Estado", "Regla", "Severidad", "Hallazgo"])
    findings = sorted(batch.findings, key=lambda item: ((item.row.row_number if item.row else 0), item.id))
    for finding in findings:
        row = finding.row
        writer.writerow([
            batch.code,
            row.row_number if row else "General",
            row.requirement_code if row else "",
            row.period_start if row else "",
            row.period_end if row else "",
            row.status if row else batch.status,
            finding.rule_code,
            finding.severity,
            finding.message,
        ])
    return output.getvalue().encode("utf-8-sig")
