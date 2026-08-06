from __future__ import annotations

import hashlib
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from .analytics import full_analysis
from .config import settings
from .delivery_readiness import professional_delivery_summary
from .methodology_closure import closure_summary
from .reduction_portfolio import portfolio_summary
from .report_consulting import consulting_report_summary
from .report_docx import generate_editable_consulting_docx
from .storage import storage

from .database import (
    ActivityData,
    EmissionCalculation,
    EmissionFactorVersion,
    EmissionSource,
    Inventory,
    ReportArtifact,
    SupplierCampaign,
    SupplierDataRequest,
    SupplierResponse,
    INSTANCE_DIR,
)

REPORTS_DIR = INSTANCE_DIR / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

NAVY = colors.HexColor("#0F2D4D")
GREEN = colors.HexColor("#2E7D5B")
BLUE = colors.HexColor("#2D7DBD")
LIGHT = colors.HexColor("#F4F7F5")
GRID = colors.HexColor("#D9E3E0")
TEXT = colors.HexColor("#17232B")


def _money_cop(value: float) -> str:
    return "$" + f"{value:,.0f}".replace(",", ".")


def _number(value: float, decimals: int = 1) -> str:
    return f"{value:,.{decimals}f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _pdf_text(value: object) -> str:
    """Normalize glyphs not supported by ReportLab's built-in WinAnsi fonts."""
    return (
        str(value)
        .replace("CO₂e", "CO2e")
        .replace("CO₂", "CO2")
        .replace("tCO₂e", "tCO2e")
        .replace("–", "-")
        .replace("—", "-")
        .replace(" ", " ")
    )


def _report_path(inventory: Inventory, suffix: str, label: str) -> Path:
    folder = REPORTS_DIR / f"org_{inventory.organization_id}" / f"inventory_{inventory.id}"
    folder.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S_%f")
    safe_label = label.lower().replace(" ", "_").replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    safe_name = f"{safe_label}_{inventory.start_date.year}_{inventory.id}_{stamp}.{suffix}"
    return folder / safe_name


def _pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Brand", fontName="Helvetica-Bold", fontSize=18, textColor=NAVY, leading=22))
    styles.add(ParagraphStyle(name="Subtitle", fontName="Helvetica", fontSize=10, textColor=GREEN, leading=14))
    styles.add(ParagraphStyle(name="Section", fontName="Helvetica-Bold", fontSize=13, textColor=NAVY, spaceBefore=8, spaceAfter=6))
    styles.add(ParagraphStyle(name="BodySmall", fontName="Helvetica", fontSize=8.5, textColor=TEXT, leading=12))
    styles.add(ParagraphStyle(name="CenterSmall", fontName="Helvetica", fontSize=8, textColor=TEXT, alignment=TA_CENTER))
    styles.add(ParagraphStyle(name="Metric", fontName="Helvetica-Bold", fontSize=16, textColor=GREEN, alignment=TA_CENTER))
    styles.add(ParagraphStyle(name="MetricLabel", fontName="Helvetica", fontSize=7.5, textColor=NAVY, alignment=TA_CENTER))
    return styles


def _header_footer(canvas, doc):
    canvas.saveState()
    canvas.setStrokeColor(GRID)
    canvas.line(18 * mm, 14 * mm, 192 * mm, 14 * mm)
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#61727B"))
    canvas.drawString(18 * mm, 9 * mm, f"Calcula tu Huella {settings.version} - documento generado desde la plataforma")
    canvas.drawRightString(192 * mm, 9 * mm, f"Página {doc.page}")
    canvas.restoreState()


def _source_bar_chart(rows: list[dict[str, object]], width: float = 165 * mm, height: float = 48 * mm) -> Drawing:
    """Compact horizontal bars for the highest-emitting sources."""
    drawing = Drawing(width, height)
    visible = rows[:6]
    if not visible:
        drawing.add(String(4, height / 2, "Sin fuentes calculadas", fontName="Helvetica", fontSize=8, fillColor=TEXT))
        return drawing
    max_value = max(float(item.get("emissions", 0) or 0) for item in visible) or 1.0
    label_width = width * 0.39
    bar_width = width * 0.47
    row_height = height / max(len(visible), 1)
    for index, item in enumerate(visible):
        y = height - (index + 1) * row_height + row_height * 0.24
        name = _pdf_text(str(item.get("name", "")))
        if len(name) > 34:
            name = name[:31] + "..."
        value = float(item.get("emissions", 0) or 0)
        share = float(item.get("share", 0) or 0)
        drawing.add(String(0, y + 2, name, fontName="Helvetica", fontSize=6.8, fillColor=TEXT))
        drawing.add(Rect(label_width, y, bar_width, row_height * 0.46, fillColor=colors.HexColor("#E7EFEC"), strokeColor=None))
        drawing.add(Rect(label_width, y, bar_width * value / max_value, row_height * 0.46, fillColor=GREEN, strokeColor=None))
        drawing.add(String(label_width + bar_width + 4, y + 1, f"{share:.1f}%", fontName="Helvetica-Bold", fontSize=6.8, fillColor=NAVY))
    return drawing


def _table(data, widths=None, header=True, font_size=8):
    header_style = ParagraphStyle("TableHeader", fontName="Helvetica-Bold", fontSize=font_size, leading=font_size + 2, textColor=colors.white)
    body_style = ParagraphStyle("TableBody", fontName="Helvetica", fontSize=font_size, leading=font_size + 2, textColor=TEXT)
    prepared = []
    for row_index, row in enumerate(data):
        style = header_style if header and row_index == 0 else body_style
        prepared.append([cell if isinstance(cell, Paragraph) else Paragraph(_pdf_text(cell), style) for cell in row])
    table = Table(prepared, colWidths=widths, repeatRows=1 if header else 0, hAlign="LEFT")
    commands = [
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("TEXTCOLOR", (0, 0), (-1, -1), TEXT),
        ("GRID", (0, 0), (-1, -1), 0.35, GRID),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    if header:
        commands.extend([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ])
    for row in range(1 if header else 0, len(data)):
        if row % 2 == 0:
            commands.append(("BACKGROUND", (0, row), (-1, row), LIGHT))
    table.setStyle(TableStyle(commands))
    return table



def generate_decision_brief_pdf(session: Session, inventory: Inventory, output: Path) -> None:
    """Generate a concise decision document with publication control."""
    analysis = full_analysis(session, inventory)
    closure = closure_summary(session, inventory)
    delivery = professional_delivery_summary(session, inventory, analysis=analysis, closure=closure)
    styles = _pdf_styles()
    # The decision brief is intentionally constrained to one page. The tighter
    # margins and compact styles preserve legibility while avoiding orphaned
    # limitations or publication controls on a second page.
    doc = SimpleDocTemplate(
        str(output), pagesize=A4, rightMargin=14 * mm, leftMargin=14 * mm,
        topMargin=11 * mm, bottomMargin=15 * mm,
        title=f"Ficha ejecutiva {inventory.name}", author="Calcula tu Huella",
    )
    brief_section = ParagraphStyle(
        "BriefSection", parent=styles["Section"], fontSize=12,
        leading=14, spaceBefore=5, spaceAfter=3,
    )
    brief_body = ParagraphStyle(
        "BriefBody", parent=styles["BodySmall"], fontSize=8,
        leading=10.2, spaceAfter=0,
    )
    publication = delivery["publication"]
    decision = delivery["decision"]
    story = [
        Paragraph("CALCULA TU HUELLA", styles["Brand"]),
        Paragraph(f"Ficha ejecutiva para decisión · Plataforma {settings.version}", styles["Subtitle"]),
        Spacer(1, 3 * mm),
        Paragraph(_pdf_text(inventory.organization.name), styles["Title"]),
        Paragraph(_pdf_text(f"Inventario: {inventory.name} | Periodo: {inventory.start_date:%d/%m/%Y} - {inventory.end_date:%d/%m/%Y}"), brief_body),
        Paragraph(_pdf_text(f"Control de publicación: {publication['level']} | Estado formal: {inventory.status}"), brief_body),
        Spacer(1, 3 * mm),
    ]
    top = decision["top_source"]
    metric_rows = [
        ["Emisiones", "Alistamiento", "Confianza", "Concentración top 3"],
        [f"{_number(analysis['total'], 1)} tCO2e", f"{delivery['score']}%", f"{decision['confidence_label']} ({decision['confidence_score']}%)", f"{decision['top_three_share']:.1f}%"],
    ]
    story.extend([_table(metric_rows, [43 * mm] * 4, font_size=8), Spacer(1, 3 * mm)])
    story.append(Paragraph("Decisión principal", brief_section))
    story.append(Paragraph(_pdf_text(decision["primary_decision"]), brief_body))
    story.append(Paragraph("Lectura del resultado", brief_section))
    story.append(Paragraph(_pdf_text(delivery["narrative"]["headline"]), brief_body))
    if top:
        story.append(Paragraph(_pdf_text(f"Foco dominante: {top['name']} · alcance {top['scope']} · {top['share']:.1f}% del total."), brief_body))

    hotspot_rows = [["Fuente prioritaria", "Alcance", "Sede", "tCO2e", "%"]]
    for row in analysis["sources_summary"][:5]:
        hotspot_rows.append([row["name"], row["scope"], row["facility"], _number(row["emissions"], 2), f"{row['share']:.1f}%"])
    if len(hotspot_rows) == 1:
        hotspot_rows.append(["Sin resultados", "-", "-", "0", "0%"] )
    story.extend([Paragraph("Focos de emisión", brief_section), _table(hotspot_rows, [55 * mm, 18 * mm, 43 * mm, 28 * mm, 18 * mm], font_size=7.0), Spacer(1, 2 * mm)])

    plan_rows = [["Prioridad", "Actividad", "Responsable", "Criterio de cierre"]]
    for item in delivery["action_plan"][:5]:
        plan_rows.append([item["priority"], item["title"], item["owner"], item["acceptance"]])
    if len(plan_rows) == 1:
        plan_rows.append(["Control", "Mantener expediente y control de versión", "Responsable climático", "Entrega conservada y trazable"])
    story.extend([Paragraph("Plan inmediato de cierre", brief_section), _table(plan_rows, [23 * mm, 45 * mm, 38 * mm, 56 * mm], font_size=6.8), Spacer(1, 2 * mm)])

    story.append(Paragraph("Recomendaciones para dirección", brief_section))
    for item in decision["recommendations"]:
        story.append(Paragraph(_pdf_text(f"- {item}"), brief_body))
    story.append(Paragraph("Regla de comunicación", brief_section))
    story.append(Paragraph(_pdf_text(publication["message"]), brief_body))
    story.append(Paragraph(_pdf_text(publication["notice"]), brief_body))
    story.append(Paragraph("Limitaciones", brief_section))
    for item in delivery["narrative"]["limitations"][:4]:
        story.append(Paragraph(_pdf_text(f"- {item}"), brief_body))
    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)

def generate_executive_pdf(session: Session, inventory: Inventory, output: Path) -> None:
    analysis = full_analysis(session, inventory)
    closure = closure_summary(session, inventory)
    delivery = professional_delivery_summary(session, inventory, analysis=analysis, closure=closure)
    portfolio = portfolio_summary(session, inventory)
    consulting = consulting_report_summary(session, inventory, analysis=analysis, delivery=delivery, closure=closure, portfolio=portfolio)
    styles = _pdf_styles()
    doc = SimpleDocTemplate(
        str(output), pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm,
        topMargin=18 * mm, bottomMargin=20 * mm,
        title=f"Informe ejecutivo {inventory.name}", author="Calcula tu Huella",
    )
    story = [
        Paragraph("CALCULA TU HUELLA", styles["Brand"]),
        Paragraph(f"Informe ejecutivo de huella de carbono · Plataforma {settings.version}", styles["Subtitle"]),
        Spacer(1, 8 * mm),
        Paragraph(inventory.organization.name, styles["Title"]),
        Paragraph(f"Periodo: {inventory.start_date:%d/%m/%Y} - {inventory.end_date:%d/%m/%Y}", styles["BodySmall"]),
        Paragraph(f"Metodologia: {inventory.methodology} | GWP: {inventory.gwp_version}", styles["BodySmall"]),
        Paragraph(_pdf_text(f"Control de publicación: {delivery['publication']['level']} | Alistamiento: {delivery['score']}%"), styles["BodySmall"]),
        Spacer(1, 8 * mm),
    ]
    story.extend([
        Paragraph("Conclusión ejecutiva", styles["Section"]),
        Paragraph(_pdf_text(delivery["narrative"]["headline"]), styles["BodySmall"]),
        Paragraph(_pdf_text(delivery["narrative"]["conclusion"]), styles["BodySmall"]),
        Paragraph(_pdf_text(f"Decisión sugerida: {delivery['decision']['primary_decision']}"), styles["BodySmall"]),
        Paragraph(_pdf_text(f"Confianza del resultado: {delivery['decision']['confidence_label']} ({delivery['decision']['confidence_score']}%)."), styles["BodySmall"]),
        Spacer(1, 5 * mm),
    ])
    history = analysis["history"]
    total_change = history["total_change"]
    metrics_data = [
        [Paragraph("EMISIONES TOTALES", styles["MetricLabel"]), Paragraph("INTENSIDAD PRODUCTIVA", styles["MetricLabel"]), Paragraph("CALIDAD DEL DATO", styles["MetricLabel"]), Paragraph("VARIACION ANUAL", styles["MetricLabel"])],
        [Paragraph(f"{_number(analysis['total'])}<br/><font size=8>tCO2e</font>", styles["Metric"]),
         Paragraph(f"{_number(analysis['intensity_production'] or 0, 4)}<br/><font size=8>tCO2e/t</font>", styles["Metric"]),
         Paragraph(f"{analysis['quality']['score']}%", styles["Metric"]),
         Paragraph("N/D" if total_change is None else f"{total_change:+.1f}%", styles["Metric"])],
    ]
    metric_table = Table(metrics_data, colWidths=[43 * mm] * 4)
    metric_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("BOX", (0, 0), (-1, -1), 0.5, GRID),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, GRID),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.extend([metric_table, Spacer(1, 7 * mm)])
    balance = closure["balance"]
    uncertainty = closure["uncertainty"]
    story.append(Paragraph("Balance metodológico", styles["Section"]))
    methodology_rows = [
        ["Partida", "tCO2e", "Tratamiento"],
        ["Emisiones brutas", _number(balance["gross_emissions"], 3), "Inventario corporativo"],
        ["CO2 biogénico", _number(balance["biogenic_memo"], 3), "Partida informativa"],
        ["Remociones", _number(balance["removals"], 3), "Separadas del bruto"],
        ["Emisiones evitadas", _number(balance["avoided_emissions"], 3), "Fuera del inventario físico"],
        ["Compensaciones", _number(balance["offsets"], 3), "Fuera del inventario bruto"],
    ]
    story.extend([_table(methodology_rows, [65 * mm, 38 * mm, 62 * mm], font_size=7.5)])
    story.append(Paragraph(
        f"Incertidumbre combinada sobre emisiones cubiertas: {uncertainty['combined_percentage']:.2f}%. "
        f"Rango cubierto: {_number(uncertainty['lower_tco2e'], 3)} a {_number(uncertainty['upper_tco2e'], 3)} tCO2e "
        f"({uncertainty['emission_coverage_percentage']:.1f}% del total bruto). "
        f"Preparación metodológica: {closure['readiness_score']}%.",
        styles["BodySmall"],
    ))
    story.extend([Spacer(1, 7 * mm), Paragraph("Resultados por alcance", styles["Section"])])
    scopes = analysis["scopes"]
    scope_rows = [["Alcance", "tCO2e", "Participación"]]
    for scope in (1, 2, 3):
        share = scopes[scope] / analysis["total"] * 100 if analysis["total"] else 0
        scope_rows.append([f"Alcance {scope}", _number(scopes[scope]), f"{share:.1f}%"])
    scope_rows.append(["Total", _number(analysis["total"]), "100,0%"])
    story.extend([_table(scope_rows, [70 * mm, 45 * mm, 45 * mm]), Spacer(1, 7 * mm)])

    story.append(Paragraph("Principales fuentes", styles["Section"]))
    source_rows = [["Fuente", "Alcance", "Sede", "tCO2e", "%"]]
    for row in analysis["sources_summary"][:8]:
        source_rows.append([row["name"], row["scope"], row["facility"], _number(row["emissions"]), f"{row['share']:.1f}%"])
    story.extend([_table(source_rows, [42 * mm, 18 * mm, 45 * mm, 30 * mm, 20 * mm]), Spacer(1, 4 * mm)])
    story.extend([_source_bar_chart(analysis["sources_summary"]), Spacer(1, 5 * mm)])

    story.append(Paragraph("Comparación e intensidades", styles["Section"]))
    comparison = consulting["comparison"]
    story.append(Paragraph(_pdf_text(
        f"Periodo anterior: {comparison['previous_year'] if comparison['available'] else 'no disponible'}. "
        f"Variación absoluta: {comparison['absolute_change']:+.1f}%" if comparison['absolute_change'] is not None else comparison['warning']
    ), styles["BodySmall"]))
    intensity_rows = [["Indicador", "Actual", "Anterior", "Variación", "Lectura"]]
    for item in consulting["intensities"]:
        intensity_rows.append([
            item["name"],
            "N/D" if item["value"] is None else _number(item["value"], 6),
            "N/D" if item["previous_value"] is None else _number(item["previous_value"], 6),
            "N/D" if item["change"] is None else f"{item['change']:+.1f}%",
            item["direction"],
        ])
    story.extend([_table(intensity_rows, [32 * mm, 32 * mm, 32 * mm, 26 * mm, 38 * mm], font_size=7.0), Spacer(1, 6 * mm)])

    story.append(Paragraph("Lectura ejecutiva explicable", styles["Section"]))
    finding_rows = [["Prioridad", "Tema", "Hallazgo", "Recomendación"]]
    for item in consulting["findings"][:6]:
        finding_rows.append([item["level"], item["topic"], item["finding"], item["recommendation"]])
    story.extend([_table(finding_rows, [22 * mm, 27 * mm, 55 * mm, 58 * mm], font_size=6.7), Spacer(1, 5 * mm)])
    story.append(Paragraph("Limitaciones y cautelas", styles["Section"]))
    for item in consulting["limitations"]:
        story.append(Paragraph(_pdf_text(f"- {item['category']}: {item['detail']}"), styles["BodySmall"]))
    story.append(Paragraph("Portafolio de reducción", styles["Section"]))
    story.append(Paragraph(_pdf_text(
        f"Estado: {portfolio['portfolio_status']}. Cobertura de la meta: {portfolio['coverage_percent']:.1f}%. "
        f"Preparación: {portfolio['readiness_score']}%. Decisión: {portfolio['primary_decision']} "
        f"Estado de publicación: {'versión final controlada' if delivery['release_ready'] else 'borrador técnico'}."
    ), styles["BodySmall"]))
    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)


def generate_technical_pdf(session: Session, inventory: Inventory, output: Path) -> None:
    analysis = full_analysis(session, inventory)
    closure = closure_summary(session, inventory)
    delivery = professional_delivery_summary(session, inventory, analysis=analysis, closure=closure)
    portfolio = portfolio_summary(session, inventory)
    consulting = consulting_report_summary(session, inventory, analysis=analysis, delivery=delivery, closure=closure, portfolio=portfolio)
    styles = _pdf_styles()
    doc = SimpleDocTemplate(
        str(output), pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm,
        topMargin=18 * mm, bottomMargin=20 * mm,
        title=f"Informe técnico {inventory.name}", author="Calcula tu Huella",
    )
    story = [
        Paragraph("CALCULA TU HUELLA", styles["Brand"]),
        Paragraph(f"Informe técnico del inventario corporativo de GEI · Plataforma {settings.version}", styles["Subtitle"]),
        Spacer(1, 6 * mm),
        Paragraph(inventory.organization.name, styles["Title"]),
        Paragraph(f"Inventario: {inventory.name} | Version: {inventory.version}", styles["BodySmall"]),
        Paragraph(f"Periodo: {inventory.start_date:%d/%m/%Y} - {inventory.end_date:%d/%m/%Y}", styles["BodySmall"]),
        Paragraph(_pdf_text(f"Control de publicación: {delivery['publication']['level']} | Alistamiento integral: {delivery['score']}%"), styles["BodySmall"]),
        Spacer(1, 8 * mm),
        _table([
            ["Emisiones", "Calidad", "Cobertura documental", "Confianza"],
            [
                f"{_number(analysis['total'], 1)} tCO2e",
                f"{analysis['quality']['score']}%",
                f"{analysis['quality']['evidence_coverage']}%",
                f"{delivery['decision']['confidence_label']} ({delivery['decision']['confidence_score']}%)",
            ],
        ], [43 * mm] * 4, font_size=8),
        Spacer(1, 8 * mm),
        Paragraph("Propósito y lectura", styles["Section"]),
        Paragraph(_pdf_text(delivery["narrative"]["conclusion"]), styles["BodySmall"]),
        Paragraph(_pdf_text(f"Decisión sugerida: {delivery['decision']['primary_decision']}"), styles["BodySmall"]),
        Spacer(1, 6 * mm),
        Paragraph("Control documental", styles["Section"]),
        _table([
            ["Campo", "Valor"],
            ["Documento", "Informe técnico del inventario corporativo de GEI"],
            ["Versión del inventario", inventory.version],
            ["Estado formal", inventory.status],
            ["Nivel de publicación", delivery["publication"]["level"]],
            ["Regla de uso", delivery["publication"]["notice"]],
        ], [48 * mm, 118 * mm], font_size=7.5),
        Spacer(1, 6 * mm),
        Paragraph("Advertencia de uso", styles["Section"]),
        Paragraph(
            _pdf_text("Este informe refleja la información y los factores registrados en la plataforma. "
                      "La revisión interna no equivale a verificación independiente y la comunicación externa "
                      "debe respetar el nivel de publicación indicado."),
            styles["BodySmall"],
        ),
        PageBreak(),
        Paragraph("1. Perfil y objetivo", styles["Section"]),
        Paragraph(
            f"La organizacion pertenece al sector {inventory.organization.sector}, con {inventory.organization.employees} empleados. "
            f"El objetivo declarado del inventario es: {inventory.objective}", styles["BodySmall"]
        ),
        Paragraph("2. Límites y metodología", styles["Section"]),
        Paragraph(
            f"Metodologia: {inventory.methodology}. Version metodologica: {inventory.methodology_version}. "
            f"Consolidacion: {inventory.consolidation_approach}. GWP: {inventory.gwp_version}. "
            f"Umbral de materialidad: {inventory.materiality_threshold:.1f}%.", styles["BodySmall"]
        ),
        Paragraph("3. Fuentes incluidas", styles["Section"]),
    ]
    source_rows = [["Fuente", "Alcance", "Categoría", "Sede", "Materialidad", "tCO2e"]]
    for row in analysis["sources_summary"]:
        source = next(item for item in inventory.sources if item.id == row["id"])
        source_rows.append([row["name"], row["scope"], row["category"], row["facility"], source.materiality, _number(row["emissions"], 3)])
    story.extend([_table(source_rows, [30 * mm, 14 * mm, 38 * mm, 38 * mm, 25 * mm, 24 * mm], font_size=7), Spacer(1, 5 * mm), _source_bar_chart(analysis["sources_summary"]), PageBreak()])

    story.append(Paragraph("4. Resultados consolidados", styles["Section"]))
    scope_rows = [["Alcance", "tCO2e"]] + [[f"Alcance {scope}", _number(value, 3)] for scope, value in analysis["scopes"].items()] + [["Total", _number(analysis["total"], 3)]]
    story.extend([_table(scope_rows, [80 * mm, 50 * mm]), Spacer(1, 6 * mm)])

    balance = closure["balance"]
    uncertainty = closure["uncertainty"]
    scope2 = closure["scope2"]
    story.append(Paragraph("4.1. Tratamiento contable y partidas separadas", styles["Section"]))
    accounting_rows = [
        ["Partida", "tCO2e", "Presentación"],
        ["Emisiones brutas", _number(balance["gross_emissions"], 6), "Total principal por alcances"],
        ["CO2 biogénico", _number(balance["biogenic_memo"], 6), "Memorando informativo"],
        ["Remociones", _number(balance["removals"], 6), "Separadas; no reducen el bruto"],
        ["Emisiones evitadas", _number(balance["avoided_emissions"], 6), "Fuera del inventario físico"],
        ["Compensaciones", _number(balance["offsets"], 6), "Fuera del inventario bruto"],
        ["Neto de referencia tras remociones", _number(balance["net_after_removals"], 6), "Indicador complementario"],
    ]
    story.extend([_table(accounting_rows, [58 * mm, 35 * mm, 77 * mm], font_size=7.2), Spacer(1, 5 * mm)])

    story.append(Paragraph("4.2. Alcance 2 e incertidumbre", styles["Section"]))
    scope2_rows = [
        ["Métrica", "Resultado"],
        ["Alcance 2 location-based", f"{_number(scope2['location_based'], 6)} tCO2e"],
        ["Alcance 2 market-based", f"{_number(scope2['market_based'], 6)} tCO2e"],
        ["Fuentes de alcance 2 sin clasificar", str(len(scope2['unclassified']))],
        ["Incertidumbre combinada", f"{uncertainty['combined_percentage']:.2f}%"],
        ["Rango inferior de emisiones cubiertas", f"{_number(uncertainty['lower_tco2e'], 6)} tCO2e"],
        ["Rango superior de emisiones cubiertas", f"{_number(uncertainty['upper_tco2e'], 6)} tCO2e"],
        ["Cobertura de emisiones brutas", f"{uncertainty['emission_coverage_percentage']}%"],
    ]
    story.extend([_table(scope2_rows, [90 * mm, 70 * mm]), Spacer(1, 5 * mm)])
    readiness_rows = [["Puerta metodológica", "Estado"]] + [[item["name"], item["status"]] for item in closure["readiness"]]
    story.extend([_table(readiness_rows, [105 * mm, 55 * mm], font_size=7.2), Spacer(1, 6 * mm)])

    story.append(Paragraph("5. Indicadores de intensidad", styles["Section"]))
    indicator_rows = [["Indicador", "Valor", "Unidad", "Intensidad"]]
    for name, metric in analysis["indicators"].items():
        if name == "Producción":
            intensity = analysis["intensity_production"]
            intensity_label = f"{_number(intensity or 0, 6)} tCO2e/{metric.unit}"
        elif name == "Empleados":
            intensity_label = f"{_number(analysis['intensity_employee'] or 0, 6)} tCO2e/persona"
        elif name == "Ingresos":
            intensity_label = f"{_number(analysis['intensity_revenue'] or 0, 6)} tCO2e/millón COP"
        else:
            intensity_label = "-"
        indicator_rows.append([name, _number(metric.value, 2), metric.unit, intensity_label])
    story.extend([_table(indicator_rows, [40 * mm, 35 * mm, 35 * mm, 60 * mm]), Spacer(1, 6 * mm)])

    quality = analysis["quality"]
    story.append(Paragraph("6. Calidad de los datos", styles["Section"]))
    quality_rows = [["Criterio", "Resultado"], ["Puntaje consolidado", f"{quality['score']}%"], ["Cobertura documental", f"{quality['evidence_coverage']}%"], ["Datos estimados", f"{quality['estimated_share']}%"]]
    for level, count in quality["counts"].items():
        quality_rows.append([f"Registros nivel {level}", str(count)])
    story.extend([_table(quality_rows, [85 * mm, 65 * mm]), Spacer(1, 6 * mm)])

    story.append(Paragraph("7. Cadena de valor y proveedores", styles["Section"]))
    supplier_responses = list(session.scalars(
        select(SupplierResponse)
        .join(SupplierDataRequest).join(SupplierCampaign)
        .where(SupplierCampaign.inventory_id == inventory.id)
        .options(selectinload(SupplierResponse.request).selectinload(SupplierDataRequest.supplier))
        .order_by(SupplierResponse.calculated_emissions_tco2e.desc())
    ))
    supplier_rows = [["Proveedor", "Producto/servicio", "Método", "Calidad", "Revisión", "tCO2e"]]
    for response in supplier_responses:
        supplier_rows.append([response.request.supplier.name, response.request.product_service, response.method, response.quality_level, response.review_status, _number(response.calculated_emissions_tco2e, 3)])
    if len(supplier_rows) == 1:
        supplier_rows.append(["Sin respuestas", "-", "-", "-", "-", "0"] )
    story.extend([_table(supplier_rows, [38 * mm, 42 * mm, 34 * mm, 18 * mm, 24 * mm, 22 * mm], font_size=6.7), PageBreak()])

    story.append(Paragraph("8. Memoria resumida de cálculos", styles["Section"]))
    calculations = list(
        session.scalars(
            select(EmissionCalculation)
            .join(ActivityData)
            .join(EmissionSource)
            .where(EmissionSource.inventory_id == inventory.id)
            .options(selectinload(EmissionCalculation.activity_data), selectinload(EmissionCalculation.factor_version).selectinload(EmissionFactorVersion.factor))
            .order_by(ActivityData.period_start, EmissionCalculation.id)
        )
    )
    calc_rows = [["Periodo", "Fuente", "Dato", "Factor", "Gas", "kg CO2e", "Estado"]]
    for calc in calculations:
        calc_rows.append([
            calc.activity_data.period_start.strftime("%Y-%m"),
            calc.activity_data.source.name,
            f"{_number(calc.original_value, 3)} {calc.original_unit}",
            calc.factor_version.factor.name,
            calc.gas_code,
            _number(calc.co2e_kg, 3),
            calc.status,
        ])
    story.extend([_table(calc_rows, [21 * mm, 26 * mm, 25 * mm, 47 * mm, 15 * mm, 22 * mm, 20 * mm], font_size=6.2), PageBreak()])

    story.append(Paragraph("9. Portafolio de reducción y abatimiento", styles["Section"]))
    story.append(Paragraph(_pdf_text(
        f"Reducción requerida: {_number(portfolio['required_reduction'])} tCO2e/año; "
        f"reducción estructurada: {_number(portfolio['expected_reduction'])} tCO2e/año; "
        f"cobertura: {portfolio['coverage_percent']:.1f}%; brecha: {_number(portfolio['gap'])} tCO2e/año; "
        f"preparación del portafolio: {portfolio['readiness_score']}%."
    ), styles["BodySmall"]))
    reduction_rows = [["Acción", "Clase", "Reducción", "Costo marginal", "Preparación", "Avance"]]
    for item in portfolio["actions"]:
        marginal = "Ahorro neto" if item["marginal_cost"] <= 0 else _money_cop(item["marginal_cost"]) + "/tCO2e"
        reduction_rows.append([
            item["title"], item["classification"], f"{_number(item['expected_reduction'])} tCO2e",
            marginal, f"{item['readiness_score']}%", f"{item['progress_percent']}%",
        ])
    story.extend([_table(reduction_rows, [48 * mm, 30 * mm, 26 * mm, 35 * mm, 21 * mm, 18 * mm], font_size=6.5), Spacer(1, 6 * mm)])
    story.append(Paragraph("10. Hallazgos y recomendaciones", styles["Section"]))
    finding_rows = [["Prioridad", "Tema", "Hallazgo", "Evidencia", "Recomendación"]]
    for item in consulting["findings"]:
        finding_rows.append([item["level"], item["topic"], item["finding"], item["evidence"], item["recommendation"]])
    story.extend([_table(finding_rows, [20 * mm, 24 * mm, 42 * mm, 42 * mm, 48 * mm], font_size=6.1), Spacer(1, 6 * mm)])

    story.append(Paragraph("11. Puertas de entrega y limitaciones", styles["Section"]))
    delivery_rows = [["Control", "Estado", "Detalle"]]
    for gate in delivery["gates"]:
        delivery_rows.append([gate["name"], gate["status"], gate["detail"]])
    story.extend([_table(delivery_rows, [47 * mm, 26 * mm, 98 * mm], font_size=6.6), Spacer(1, 5 * mm)])
    limitation_rows = [["Categoría", "Limitación"]] + [[item["category"], item["detail"]] for item in consulting["limitations"]]
    story.extend([_table(limitation_rows, [38 * mm, 133 * mm], font_size=6.8), Spacer(1, 5 * mm)])
    story.append(Paragraph("12. Declaración técnica", styles["Section"]))
    story.append(Paragraph(
        f"Estado de publicación: {'versión final controlada' if delivery['release_ready'] else 'borrador técnico'}. "
        "El inventario fue elaborado a partir de la información registrada por la organización y los factores seleccionados. "
        "Los factores demostrativos deben sustituirse por fuentes oficiales o específicas antes de uso externo. "
        "La aprobación dentro de la plataforma corresponde a control interno y no constituye verificación independiente.", styles["BodySmall"]
    ))
    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)


def generate_calculation_workbook(session: Session, inventory: Inventory, output: Path) -> None:
    analysis = full_analysis(session, inventory)
    closure = closure_summary(session, inventory)
    delivery = professional_delivery_summary(session, inventory, analysis=analysis, closure=closure)
    portfolio = portfolio_summary(session, inventory)
    consulting = consulting_report_summary(session, inventory, analysis=analysis, delivery=delivery, closure=closure, portfolio=portfolio)
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["CALCULA TU HUELLA - MEMORIA DE CALCULO"])
    ws.append(["Organización", inventory.organization.name])
    ws.append(["Inventario", inventory.name])
    ws.append(["Periodo", f"{inventory.start_date:%Y-%m-%d} a {inventory.end_date:%Y-%m-%d}"])
    ws.append(["Metodología", inventory.methodology])
    ws.append(["GWP", inventory.gwp_version])
    ws.append([])
    ws.append(["Indicador", "Valor", "Unidad"])
    ws.append(["Emisiones totales", analysis["total"], "tCO2e"])
    for scope, value in analysis["scopes"].items():
        ws.append([f"Alcance {scope}", value, "tCO2e"])
    ws.append(["Calidad de datos", analysis["quality"]["score"], "%"])
    ws.append(["Intensidad productiva", analysis["intensity_production"] or 0, "tCO2e/t"])
    ws.append(["Incertidumbre combinada", closure["uncertainty"]["combined_percentage"], "%"])
    ws.append(["Rango inferior (emisiones cubiertas)", closure["uncertainty"]["lower_tco2e"], "tCO2e"])
    ws.append(["Rango superior (emisiones cubiertas)", closure["uncertainty"]["upper_tco2e"], "tCO2e"])
    ws.append(["Preparación metodológica", closure["readiness_score"], "%"])
    ws.append(["Alistamiento integral para entrega", delivery["score"], "%"])
    ws.append(["Control de publicación", delivery["publication"]["level"], "estado"])
    ws.append(["Confianza para decisión", delivery["decision"]["confidence_score"], "%"])
    ws.append(["Decisión principal", delivery["decision"]["primary_decision"], "orientación"])


    delivery_ws = wb.create_sheet("Control de entrega")
    delivery_ws.append(["Puerta", "Estado", "Responsable", "Detalle", "Criterio de aceptación", "Acción", "Ruta"])
    for gate in delivery["gates"]:
        delivery_ws.append([gate["name"], gate["status"], gate["owner"], gate["detail"], gate["acceptance"], gate["action"], gate["href"]])
    delivery_ws.append([])
    delivery_ws.append(["Control de publicación", delivery["publication"]["level"], delivery["publication"]["audience"], delivery["publication"]["message"], delivery["publication"]["notice"]])
    publication_row = delivery_ws.max_row
    delivery_ws.append(["Decisión principal", delivery["decision"]["primary_decision"], "Dirección", delivery["narrative"]["headline"], f"Confianza {delivery['decision']['confidence_label']} ({delivery['decision']['confidence_score']}%)"])
    decision_row = delivery_ws.max_row
    delivery_ws.append([])
    delivery_ws.append(["Plan priorizado", "Prioridad", "Responsable", "Detalle", "Criterio de cierre", "Acción", "Ruta"])
    plan_header_row = delivery_ws.max_row
    for item in delivery["action_plan"]:
        delivery_ws.append([item["stage"], item["priority"], item["owner"], item["detail"], item["acceptance"], item["title"], item["href"]])
    delivery_ws.append([])
    delivery_ws.append(["Conclusión ejecutiva", delivery["narrative"]["headline"]])
    narrative_start_row = delivery_ws.max_row
    delivery_ws.append(["Conclusión de estado", delivery["narrative"]["conclusion"]])
    for item in delivery["narrative"]["limitations"]:
        delivery_ws.append(["Limitación", item])
    narrative_end_row = delivery_ws.max_row

    consulting_ws = wb.create_sheet("Narrativa consultoría")
    consulting_ws.append(["CAPÍTULOS Y PREPARACIÓN EDITORIAL"])
    consulting_ws.append(["Preparación editorial", consulting["report_score"], "%"])
    consulting_ws.append(["Estado", consulting["status"], "control"])
    consulting_ws.append(["Propósito", consulting["purpose"], "orientación"])
    consulting_ws.append([])
    consulting_ws.append(["Capítulo", "Estado", "Puntaje", "Acción requerida"])
    consulting_header_row = consulting_ws.max_row
    for item in consulting["chapters"]:
        consulting_ws.append([item["name"], item["status"], item["score"], item["action"]])
    consulting_ws.append([])
    consulting_ws.append(["HALLAZGOS EXPLICABLES"])
    consulting_ws.append(["Prioridad", "Tema", "Hallazgo", "Evidencia", "Implicación", "Recomendación"])
    findings_header_row = consulting_ws.max_row
    for item in consulting["findings"]:
        consulting_ws.append([item["level"], item["topic"], item["finding"], item["evidence"], item["implication"], item["recommendation"]])
    consulting_ws.append([])
    consulting_ws.append(["LIMITACIONES Y REGLAS DE COMUNICACIÓN"])
    consulting_ws.append(["Categoría", "Detalle"])
    limitations_header_row = consulting_ws.max_row
    for item in consulting["limitations"]:
        consulting_ws.append([item["category"], item["detail"]])
    consulting_ws.append([])
    consulting_ws.append(["Afirmación", "Permitida", "Orientación"])
    claims_header_row = consulting_ws.max_row
    for item in consulting["claims"]:
        consulting_ws.append([item["label"], "Sí" if item["allowed"] else "No", item["guidance"]])

    data_ws = wb.create_sheet("Datos de actividad")
    data_ws.append(["ID", "Fuente", "Sede", "Periodo inicial", "Periodo final", "Valor", "Unidad", "Origen", "Calidad", "Estimado", "Incertidumbre %", "Base incertidumbre", "Evidencia", "Estado", "Notas"])
    records = list(
        session.scalars(
            select(ActivityData)
            .join(EmissionSource)
            .where(EmissionSource.inventory_id == inventory.id)
            .options(selectinload(ActivityData.source).selectinload(EmissionSource.facility), selectinload(ActivityData.evidence))
            .order_by(ActivityData.period_start, ActivityData.id)
        )
    )
    for record in records:
        data_ws.append([
            record.id, record.source.name, record.source.facility.name if record.source.facility else "Corporativo",
            record.period_start, record.period_end, record.value, record.unit, record.data_origin, record.quality_level,
            "Sí" if record.is_estimated else "No", record.uncertainty_percentage, record.uncertainty_basis,
            record.evidence.name if record.evidence else "", record.status, record.notes,
        ])

    calc_ws = wb.create_sheet("Cálculos")
    calc_ws.append(["ID", "Dato ID", "Fuente", "Periodo", "Valor original", "Unidad", "Valor normalizado", "Unidad normalizada", "Factor", "Versión", "Gas", "Resultado gas kg", "GWP", "kg CO2e", "tCO2e", "Partida", "Incertidumbre %", "Límite inferior kg", "Límite superior kg", "Estado", "Alerta", "Fórmula"])
    calculations = list(
        session.scalars(
            select(EmissionCalculation)
            .join(ActivityData)
            .join(EmissionSource)
            .where(EmissionSource.inventory_id == inventory.id)
            .options(selectinload(EmissionCalculation.activity_data).selectinload(ActivityData.source), selectinload(EmissionCalculation.factor_version).selectinload(EmissionFactorVersion.factor))
            .order_by(ActivityData.period_start, EmissionCalculation.id)
        )
    )
    for calc in calculations:
        calc_ws.append([
            calc.id, calc.activity_data_id, calc.activity_data.source.name, calc.activity_data.period_start,
            calc.original_value, calc.original_unit, calc.normalized_value, calc.normalized_unit,
            calc.factor_version.factor.name, calc.factor_version.version, calc.gas_code, calc.gas_result_kg,
            calc.gwp_value, calc.co2e_kg, calc.co2e_kg / 1000, calc.reporting_bucket, calc.uncertainty_percentage,
            calc.lower_co2e_kg, calc.upper_co2e_kg, calc.status, calc.warning, calc.formula_snapshot,
        ])

    factors_ws = wb.create_sheet("Factores")
    factors_ws.append(["Fuente", "Factor", "Versión", "Valor", "Unidad entrada", "Unidad salida", "Gas", "Incertidumbre %", "Organización fuente", "Documento", "Año", "Estado", "Demostrativo"])
    seen = set()
    for source in inventory.sources:
        for assignment in source.factor_assignments:
            version = assignment.factor_version
            key = (source.id, version.id)
            if key in seen:
                continue
            seen.add(key)
            factors_ws.append([
                source.name, version.factor.name, version.version, version.value, version.input_unit, version.output_unit,
                version.gas.code, version.uncertainty_percentage, version.source_organization, version.source_document, version.publication_year,
                version.status, "Sí" if version.factor.is_demo else "No",
            ])

    method_ws = wb.create_sheet("Cierre metodológico")
    method_ws.append(["Componente", "Valor", "Unidad / estado"])
    balance = closure["balance"]
    method_ws.append(["Emisiones brutas", balance["gross_emissions"], "tCO2e"])
    method_ws.append(["CO2 biogénico informativo", balance["biogenic_memo"], "tCO2e"])
    method_ws.append(["Remociones", balance["removals"], "tCO2e"])
    method_ws.append(["Emisiones evitadas", balance["avoided_emissions"], "tCO2e"])
    method_ws.append(["Compensaciones", balance["offsets"], "tCO2e"])
    method_ws.append(["Alcance 2 location-based", closure["scope2"]["location_based"], "tCO2e"])
    method_ws.append(["Alcance 2 market-based", closure["scope2"]["market_based"], "tCO2e"])
    method_ws.append(["Fuentes alcance 2 sin clasificar", len(closure["scope2"]["unclassified"]), "cantidad"])
    method_ws.append(["Incertidumbre combinada (emisiones cubiertas)", closure["uncertainty"]["combined_percentage"], "%"])
    method_ws.append(["Cobertura de emisiones brutas", closure["uncertainty"]["emission_coverage_percentage"], "%"])
    method_ws.append(["Política metodológica", closure["policy"].get("status", "Borrador"), "estado"])
    method_ws.append([])
    method_ws.append(["Puerta metodológica", "Estado", "Acción requerida"])
    for item in closure["readiness"]:
        method_ws.append([item["name"], item["status"], item["detail"]])

    indicator_ws = wb.create_sheet("Indicadores")
    indicator_ws.append(["Tipo", "Periodo inicial", "Periodo final", "Sede", "Valor", "Unidad", "Fuente", "Estado", "Notas"])
    for indicator in inventory.indicators:
        indicator_ws.append([
            indicator.indicator_type, indicator.period_start, indicator.period_end,
            indicator.facility.name if indicator.facility else "Corporativo", indicator.value, indicator.unit,
            indicator.source_name, indicator.status, indicator.notes,
        ])

    supplier_ws = wb.create_sheet("Proveedores alcance 3")
    supplier_ws.append(["Campaña", "Proveedor", "NIT", "Producto/servicio", "Cantidad", "Unidad", "Gasto COP", "Método", "Factor", "Unidad factor", "tCO2e", "Metodología", "Límite", "Verificada", "Calidad", "Revisión", "SHA256 evidencia"])
    supplier_responses = list(session.scalars(
        select(SupplierResponse)
        .join(SupplierDataRequest).join(SupplierCampaign)
        .where(SupplierCampaign.inventory_id == inventory.id)
        .options(selectinload(SupplierResponse.request).selectinload(SupplierDataRequest.supplier), selectinload(SupplierResponse.request).selectinload(SupplierDataRequest.campaign))
        .order_by(SupplierResponse.id)
    ))
    for response in supplier_responses:
        req = response.request
        supplier_ws.append([req.campaign.name, req.supplier.name, req.supplier.tax_id, req.product_service, req.quantity, req.unit, req.spend_cop, response.method, response.emission_factor, response.factor_unit, response.calculated_emissions_tco2e, response.methodology, response.boundary, "Sí" if response.verified else "No", response.quality_level, response.review_status, response.evidence_sha256])

    reduction_ws = wb.create_sheet("Reducción")
    reduction_ws.append(["Clasificación", "Acción", "Fuente", "Descripción", "Reducción esperada", "Reducción real", "Impacto inventario %", "Inversión COP", "Ahorro anual COP", "Retorno años", "Costo marginal COP/tCO2e", "Preparación %", "Pendientes", "Prioridad", "Responsable", "Fecha objetivo", "Vencida", "Próximos 90 días", "Estado", "Avance %", "Viabilidad", "Riesgo"])
    for item in portfolio["actions"]:
        action = item["action"]
        reduction_ws.append([
            item["classification"], action.title, item["source"], action.description,
            item["expected_reduction"], item["actual_reduction"], item["impact_share"], item["investment"],
            item["annual_savings"], item["payback_years"], item["marginal_cost"], item["readiness_score"],
            "; ".join(item["missing"]), action.priority, item["owner"], item["target_date"],
            "Sí" if item["overdue"] else "No", "Sí" if item["due_soon"] else "No", item["status"],
            item["progress_percent"], item["feasibility"], item["risk_level"],
        ])

    trajectory_ws = wb.create_sheet("Trayectoria reducción")
    trajectory_ws.append(["Año", "Reducción acumulada tCO2e", "Emisiones proyectadas tCO2e", "Valor meta tCO2e"])
    for point in portfolio["timeline"]:
        trajectory_ws.append([point["year"], point["reduction"], point["projected_emissions"], point["target_value"]])

    header_fill = PatternFill("solid", fgColor="0F2D4D")
    header_font = Font(color="FFFFFF", bold=True)
    title_fill = PatternFill("solid", fgColor="EAF3EF")
    warning_fill = PatternFill("solid", fgColor="FFF1ED")
    decision_fill = PatternFill("solid", fgColor="EAF4F7")
    ready_fill = PatternFill("solid", fgColor="E7F4EA")
    progress_fill = PatternFill("solid", fgColor="FFF4D6")
    blocked_fill = PatternFill("solid", fgColor="FDE7E5")
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2" if sheet.title != "Resumen" else "A8"
        header_row = 1 if sheet.title != "Resumen" else 8
        for cell in sheet[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        if sheet.title == "Resumen":
            sheet.merge_cells("A1:C1")
            sheet["A1"].fill = title_fill
            sheet["A1"].font = Font(bold=True, size=16, color="0F2D4D")
            sheet["A1"].alignment = Alignment(vertical="center", horizontal="left")
            sheet.row_dimensions[1].height = 28
        for column_cells in sheet.columns:
            max_len = 0
            for cell in list(column_cells)[:200]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 10), 42)
        sheet.sheet_view.showGridLines = False
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.print_options.horizontalCentered = False

    # Consulting narrative is a controlled editorial working paper.
    consulting_ws.merge_cells("A1:F1")
    consulting_ws["A1"].fill = title_fill
    consulting_ws["A1"].font = Font(bold=True, size=15, color="0F2D4D")
    consulting_ws["A1"].alignment = Alignment(vertical="center", horizontal="left")
    consulting_ws.row_dimensions[1].height = 26
    for header_index in (consulting_header_row, findings_header_row, limitations_header_row, claims_header_row):
        for cell in consulting_ws[header_index]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        consulting_ws.row_dimensions[header_index].height = 28
    for row in range(2, consulting_ws.max_row + 1):
        for column in range(1, 7):
            consulting_ws.cell(row, column).alignment = Alignment(vertical="top", wrap_text=True)
    for column, width in {"A": 25, "B": 24, "C": 38, "D": 42, "E": 42, "F": 45}.items():
        consulting_ws.column_dimensions[column].width = width
    consulting_ws.freeze_panes = f"A{consulting_header_row + 1}"
    consulting_ws.page_setup.orientation = "landscape"
    consulting_ws.print_area = f"A1:F{consulting_ws.max_row}"

    # The delivery control sheet is a working paper, not a raw export. It has
    # explicit sections, visual status cues and print settings for committee use.
    for header_index in (1, plan_header_row):
        for cell in delivery_ws[header_index]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        delivery_ws.row_dimensions[header_index].height = 28
    delivery_widths = {"A": 34, "B": 24, "C": 28, "D": 52, "E": 52, "F": 25, "G": 28}
    for column, width in delivery_widths.items():
        delivery_ws.column_dimensions[column].width = width
    delivery_ws.page_setup.orientation = "landscape"
    delivery_ws.print_title_rows = "1:1"
    delivery_ws.print_area = f"A1:G{delivery_ws.max_row}"
    delivery_ws.auto_filter.ref = f"A1:G{publication_row - 2}"

    for row in range(2, delivery_ws.max_row + 1):
        is_blank = all(delivery_ws.cell(row, column).value in (None, "") for column in range(1, 8))
        delivery_ws.row_dimensions[row].height = 10 if is_blank else 46
        for column in range(1, 8):
            delivery_ws.cell(row, column).alignment = Alignment(vertical="top", wrap_text=True)
        status = str(delivery_ws.cell(row, 2).value or "")
        if status == "Listo":
            delivery_ws.cell(row, 2).fill = ready_fill
        elif status == "En progreso":
            delivery_ws.cell(row, 2).fill = progress_fill
        elif status == "Bloqueado":
            delivery_ws.cell(row, 2).fill = blocked_fill

    for column in range(1, 8):
        delivery_ws.cell(publication_row, column).fill = warning_fill
        delivery_ws.cell(decision_row, column).fill = decision_fill
    delivery_ws.cell(publication_row, 1).font = Font(bold=True, color="9C2E22")
    delivery_ws.cell(decision_row, 1).font = Font(bold=True, color="0F2D4D")
    delivery_ws.row_dimensions[publication_row].height = 58
    delivery_ws.row_dimensions[decision_row].height = 64

    for row in range(narrative_start_row, narrative_end_row + 1):
        delivery_ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        delivery_ws.cell(row, 1).fill = title_fill
        delivery_ws.cell(row, 1).font = Font(bold=True, color="0F2D4D")
        delivery_ws.cell(row, 2).alignment = Alignment(vertical="top", wrap_text=True)
        delivery_ws.row_dimensions[row].height = 48

    wb.save(output)


def create_report_artifact(session: Session, inventory: Inventory, report_type: str, generated_by: str) -> ReportArtifact:
    normalized = report_type.lower()
    if normalized == "ficha":
        output = _report_path(inventory, "pdf", "ficha_ejecutiva")
        generate_decision_brief_pdf(session, inventory, output)
        label = "Ficha ejecutiva"
    elif normalized == "ejecutivo":
        output = _report_path(inventory, "pdf", "informe_ejecutivo")
        generate_executive_pdf(session, inventory, output)
        label = "Informe ejecutivo"
    elif normalized == "tecnico":
        output = _report_path(inventory, "pdf", "informe_tecnico")
        generate_technical_pdf(session, inventory, output)
        label = "Informe técnico"
    elif normalized == "memoria":
        output = _report_path(inventory, "xlsx", "memoria_calculo")
        generate_calculation_workbook(session, inventory, output)
        label = "Memoria de cálculo"
    elif normalized == "editable":
        output = _report_path(inventory, "docx", "informe_consultoria_editable")
        generate_editable_consulting_docx(session, inventory, output)
        label = "Informe de consultoría editable"
    else:
        raise ValueError("Tipo de informe no soportado")
    content = output.read_bytes()
    storage_key = str(output.relative_to(INSTANCE_DIR))
    media_types = {
        ".pdf": "application/pdf",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    }
    storage.put_bytes(storage_key, content, media_types.get(output.suffix.lower(), "application/octet-stream"))
    artifact = ReportArtifact(
        inventory_id=inventory.id,
        report_type=label,
        version=f"{inventory.version}-{datetime.now(UTC):%Y%m%d%H%M%S}",
        status="Generado",
        file_name=output.name,
        stored_name=storage_key,
        file_size=len(content),
        sha256=hashlib.sha256(content).hexdigest(),
        generated_by=generated_by,
    )
    session.add(artifact)
    session.flush()
    return artifact
