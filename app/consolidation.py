from __future__ import annotations

import ast
import json
from collections import Counter
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from .access_control import ROLE_ORDER, permission_matrix
from .database import ConsolidationFinding, JourneyValidation, ReleaseGate
from .product_registry import PRODUCT_MODULES, ROLE_JOURNEYS, product_layers
from .release_candidate import release_candidate_summary

OPEN_FINDING_STATUSES = {"Abierto", "En curso", "Bloqueado"}
READY_GATE_STATUSES = {"Aprobado", "Completado"}


def codebase_metrics(project_dir: Path) -> dict[str, object]:
    app_dir = project_dir / "app"
    tests_dir = project_dir / "tests"
    python_files = list(app_dir.rglob("*.py")) + list(tests_dir.rglob("*.py"))
    template_files = list((app_dir / "templates").glob("*.html"))
    total_lines = 0
    route_count = 0
    files: list[dict[str, object]] = []
    for path in python_files:
        try:
            text = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            continue
        line_count = len(text.splitlines())
        total_lines += line_count
        route_count += sum(1 for line in text.splitlines() if line.lstrip().startswith("@app.") or line.lstrip().startswith("@router."))
        files.append({"path": str(path.relative_to(project_dir)), "lines": line_count})
    files.sort(key=lambda item: int(item["lines"]), reverse=True)
    test_functions = 0
    for path in tests_dir.rglob("test_*.py"):
        try:
            tree = ast.parse(path.read_text(encoding="utf-8"))
            test_functions += sum(isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)) and node.name.startswith("test_") for node in ast.walk(tree))
        except (SyntaxError, UnicodeDecodeError):
            pass
    return {
        "python_files": len(python_files),
        "templates": len(template_files),
        "routes": route_count,
        "tests": test_functions,
        "total_lines": total_lines,
        "largest_files": files[:8],
    }


def consolidation_summary(session: Session, organization_id: int, project_dir: Path) -> dict[str, object]:
    findings = list(session.scalars(
        select(ConsolidationFinding)
        .where(ConsolidationFinding.organization_id == organization_id)
        .order_by(ConsolidationFinding.priority, ConsolidationFinding.code)
    ))
    gates = list(session.scalars(
        select(ReleaseGate)
        .where(ReleaseGate.organization_id == organization_id)
        .order_by(ReleaseGate.category, ReleaseGate.code)
    ))
    validations = list(session.scalars(
        select(JourneyValidation)
        .where(JourneyValidation.organization_id == organization_id)
        .order_by(JourneyValidation.role, JourneyValidation.journey_code)
    ))
    validation_by_code = {item.journey_code: item for item in validations}
    journeys = []
    for journey in ROLE_JOURNEYS:
        item = dict(journey)
        item["validation"] = validation_by_code.get(str(journey["code"]))
        journeys.append(item)

    open_findings = [item for item in findings if item.status in OPEN_FINDING_STATUSES]
    critical_open = [item for item in open_findings if item.priority == "Crítica"]
    approved_gates = [item for item in gates if item.status in READY_GATE_STATUSES]
    validated_journeys = [item for item in validations if item.status == "Aprobado"]
    gate_score = round(len(approved_gates) / max(len(gates), 1) * 100)
    journey_score = round(len(validated_journeys) / max(len(ROLE_JOURNEYS), 1) * 100)
    debt_penalty = min(50, len(critical_open) * 10 + max(0, len(open_findings) - len(critical_open)) * 2)
    readiness_score = max(0, round(gate_score * 0.65 + journey_score * 0.35 - debt_penalty))
    candidate = release_candidate_summary(
        project_dir,
        critical_open=len(critical_open),
        approved_gates=len(approved_gates),
        gate_count=len(gates),
        validated_journeys=len(validated_journeys),
        journey_count=len(ROLE_JOURNEYS),
    )

    return {
        "metrics": codebase_metrics(project_dir),
        "findings": findings,
        "gates": gates,
        "journeys": journeys,
        "layers": product_layers(),
        "permission_matrix": permission_matrix(),
        "roles": ROLE_ORDER,
        "module_count": len(PRODUCT_MODULES),
        "open_findings": len(open_findings),
        "critical_open": len(critical_open),
        "approved_gates": len(approved_gates),
        "gate_count": len(gates),
        "validated_journeys": len(validated_journeys),
        "journey_count": len(ROLE_JOURNEYS),
        "readiness_score": readiness_score,
        "finding_counts": dict(Counter(item.status for item in findings)),
        "gate_counts": dict(Counter(item.status for item in gates)),
        "release_candidate": candidate,
    }


def build_consolidation_workbook(summary: dict[str, object]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Indicador", "Valor"])
    for label, key in [
        ("Preparación V1.0", "readiness_score"),
        ("Módulos registrados", "module_count"),
        ("Hallazgos abiertos", "open_findings"),
        ("Hallazgos críticos", "critical_open"),
        ("Puertas aprobadas", "approved_gates"),
        ("Recorridos aprobados", "validated_journeys"),
    ]:
        ws.append([label, summary[key]])

    findings_ws = wb.create_sheet("Deuda y hallazgos")
    findings_ws.append(["Código", "Área", "Título", "Prioridad", "Estado", "Responsable", "Versión objetivo", "Detalle", "Evidencia"])
    for item in summary["findings"]:
        findings_ws.append([item.code, item.area, item.title, item.priority, item.status, item.owner, item.target_version, item.detail, item.evidence])

    candidate_ws = wb.create_sheet("Liberación V1")
    candidate_ws.append(["Versión", summary["release_candidate"]["version"]])
    candidate_ws.append(["Estado", summary["release_candidate"]["status"]])
    candidate_ws.append(["Paquete interno", "Aprobado" if summary["release_candidate"]["package_ready"] else "Pendiente"])
    candidate_ws.append(["Gobierno de release", "Aprobado" if summary["release_candidate"]["governance_ready"] else "Pendiente"])
    candidate_ws.append(["Aceptación interna", "Aprobado" if summary["release_candidate"]["internal_ready"] else "Pendiente"])
    candidate_ws.append(["Despliegue controlado", "Aprobado" if summary["release_candidate"]["controlled_release_ready"] else "Pendiente"])
    candidate_ws.append(["Producción pública", "Aprobado" if summary["release_candidate"]["production_ready"] else "Pendiente"])
    candidate_ws.append([])
    candidate_ws.append(["Código", "Grupo", "Control", "Estado", "Detalle"])
    for item in summary["release_candidate"]["checks"]:
        candidate_ws.append([item["code"], item["group"], item["label"], "Aprobado" if item["ok"] else "Pendiente", item["detail"]])

    gates_ws = wb.create_sheet("Puertas V1")
    gates_ws.append(["Código", "Categoría", "Puerta", "Estado", "Responsable", "Evidencia", "Notas"])
    for item in summary["gates"]:
        gates_ws.append([item.code, item.category, item.name, item.status, item.responsible, item.evidence, item.notes])

    journeys_ws = wb.create_sheet("Recorridos")
    journeys_ws.append(["Código", "Recorrido", "Rol", "Objetivo", "Estado", "Probado por", "Notas"])
    for journey in summary["journeys"]:
        validation = journey["validation"]
        journeys_ws.append([
            journey["code"], journey["name"], journey["role"], journey["objective"],
            validation.status if validation else "No probado",
            validation.tested_by if validation else "",
            validation.notes if validation else "",
        ])

    permissions_ws = wb.create_sheet("Permisos")
    permissions_ws.append(["Capacidad", *summary["roles"]])
    for row in summary["permission_matrix"]:
        permissions_ws.append([row["label"], *["Sí" if row["roles"][role] else "No" for role in summary["roles"]]])

    metrics_ws = wb.create_sheet("Arquitectura")
    metrics_ws.append(["Métrica", "Valor"])
    for key, label in [("python_files", "Archivos Python"), ("templates", "Plantillas"), ("routes", "Rutas"), ("tests", "Pruebas"), ("total_lines", "Líneas Python")]:
        metrics_ws.append([label, summary["metrics"][key]])
    metrics_ws.append([])
    metrics_ws.append(["Archivo", "Líneas"])
    for item in summary["metrics"]["largest_files"]:
        metrics_ws.append([item["path"], item["lines"]])

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def summary_json(summary: dict[str, object]) -> str:
    serializable = {
        "readiness_score": summary["readiness_score"],
        "module_count": summary["module_count"],
        "open_findings": summary["open_findings"],
        "critical_open": summary["critical_open"],
        "approved_gates": summary["approved_gates"],
        "gate_count": summary["gate_count"],
        "validated_journeys": summary["validated_journeys"],
        "journey_count": summary["journey_count"],
        "metrics": summary["metrics"],
        "release_candidate": summary["release_candidate"],
    }
    return json.dumps(serializable, ensure_ascii=False, default=str)
