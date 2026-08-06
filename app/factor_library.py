from __future__ import annotations

import json
from datetime import date
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from .database import (
    ActivityData,
    EmissionFactorVersion,
    EmissionSource,
    FactorDocumentation,
    MethodologySourceDocument,
    UnitConversion,
    UnitDefinition,
)
from .factor_advisor import advise_factor

QUALITY_RANK = {"A": 4, "B": 3, "C": 2, "D": 1, "N/A": 0}
FORMAL_REVIEW_STATES = {"Aprobado documentalmente", "Aprobado"}
PRELIMINARY_SOURCE_MARKERS = {"preliminar", "borrador", "consulta", "proyecto", "no incorporado"}
HIERARCHY_LEVELS = {
    1: ("Específico verificado", "Dato o factor primario de proveedor con verificación y límites comparables."),
    2: ("Oficial nacional", "Factor oficial colombiano pertinente al periodo y uso del inventario."),
    3: ("Sectorial reconocido", "Factor sectorial o metodología reconocida, documentada y aplicable."),
    4: ("Internacional por gas", "Factor IPCC u otra referencia internacional con conversión y GWP explícitos."),
    5: ("Secundario o piloto", "Fuente secundaria, transcripción o factor condicionado que exige justificación reforzada."),
    6: ("Demostrativo o retirado", "Dato sintético, retirado o no apto para inventarios formales."),
}


def _clean(value: object) -> str:
    return str(value or "").strip()


def _contains(value: object, query: str) -> bool:
    return query.lower() in _clean(value).lower()


def _document_completeness(documentation: FactorDocumentation | None) -> tuple[int, list[str]]:
    if not documentation:
        return 0, ["No existe una ficha documental asociada a esta versión."]
    checks = [
        (bool(documentation.source_document_id), "fuente documental"),
        (bool(documentation.page_reference or documentation.table_reference), "referencia de página o tabla"),
        (bool(documentation.source_unit), "unidad de la fuente"),
        (documentation.source_value is not None, "valor transcrito"),
        (bool(documentation.conversion_expression) or documentation.source_unit == documentation.factor_version.input_unit, "memoria de conversión"),
        (bool(documentation.restriction_notes), "restricciones de uso"),
        (bool(documentation.reviewer), "revisor documental"),
        (documentation.reviewed_at is not None, "fecha de revisión"),
    ]
    completed = sum(1 for ok, _ in checks if ok)
    missing = [label for ok, label in checks if not ok]
    return round(completed / len(checks) * 100), missing


def _effective_state(version: EmissionFactorVersion, reference_date: date | None) -> tuple[str, bool, str]:
    if reference_date is None:
        if version.effective_to and version.effective_to < date.today():
            return "Vigencia vencida", False, f"La versión dejó de estar vigente el {version.effective_to.isoformat()}."
        if version.effective_from and version.effective_from > date.today():
            return "Vigencia futura", False, f"La versión entra en vigencia el {version.effective_from.isoformat()}."
        return "Vigencia disponible", True, "La aptitud temporal debe confirmarse contra el periodo del dato."
    if version.effective_from and reference_date < version.effective_from:
        return "Fuera de vigencia", False, f"El dato es anterior al inicio de vigencia ({version.effective_from.isoformat()})."
    if version.effective_to and reference_date > version.effective_to:
        return "Fuera de vigencia", False, f"El dato es posterior al fin de vigencia ({version.effective_to.isoformat()})."
    if not version.effective_from and not version.effective_to:
        return "Sin ventana explícita", True, "La versión no declara fechas de vigencia; compara el año del dato y la fuente."
    return "Vigente para el dato", True, "El periodo del dato se encuentra dentro de la ventana declarada."


def _geographic_state(version: EmissionFactorVersion, factor_country: str, desired_country: str | None) -> tuple[str, str]:
    geography = _clean(version.geographic_scope or factor_country)
    if not desired_country:
        return "Por confirmar", f"Cobertura declarada: {geography or 'sin definir'}."
    desired = desired_country.lower()
    current = geography.lower()
    if desired in current or current in {"global", "internacional"}:
        return "Compatible", f"La cobertura {geography} conversa con {desired_country}."
    return "Requiere justificación", f"La cobertura declarada ({geography}) no coincide directamente con {desired_country}."


def _review_due_state(documentation: FactorDocumentation | None) -> tuple[str, bool]:
    if not documentation or not documentation.next_review_date:
        return "Sin fecha programada", False
    if documentation.next_review_date < date.today():
        return f"Revisión vencida desde {documentation.next_review_date.isoformat()}", True
    return f"Próxima revisión {documentation.next_review_date.isoformat()}", False


def _source_label(documentation: FactorDocumentation | None) -> str:
    if not documentation or not documentation.source_document:
        return "Sin fuente documental vinculada"
    return f"{documentation.source_document.code} · {documentation.source_document.issuing_body}"


def _source_governance_state(documentation: FactorDocumentation | None) -> tuple[str, bool, str]:
    if not documentation or not documentation.source_document:
        return "Sin fuente controlada", True, "No existe un documento controlado que permita reproducir el factor."
    source_document = documentation.source_document
    status = _clean(source_document.status) or "Sin estado"
    normalized = status.lower()
    preliminary = any(marker in normalized for marker in PRELIMINARY_SOURCE_MARKERS)
    if preliminary:
        return status, True, "La fuente está en condición preliminar o de revisión y no debe promoverse automáticamente a cálculo formal."
    return status, False, "La fuente no presenta una marca automática de preliminariedad; conserva la revisión profesional obligatoria."


def _hierarchy_state(
    version: EmissionFactorVersion,
    documentation: FactorDocumentation | None,
) -> tuple[int, str, str]:
    use = _clean(documentation.reporting_use if documentation else "").lower()
    kind = _clean(documentation.factor_kind if documentation else "").lower()
    source = documentation.source_document if documentation else None
    issuer = _clean(source.issuing_body if source else version.source_organization).lower()
    status = _clean(source.status if source else "").lower()

    if version.factor.is_demo or use in {"demostrativo", "retirado"}:
        level = 6
    elif any(marker in status for marker in PRELIMINARY_SOURCE_MARKERS) or use == "piloto" or "secund" in kind or "transcri" in kind:
        level = 5
    elif "específico" in kind and ("proveedor" in kind or "primario" in kind):
        level = 1
    elif "oficial nacional" in kind or (version.factor.country == "Colombia" and any(token in issuer for token in ("upme", "ministerio", "ideam", "xm")) and use == "formal"):
        level = 2
    elif "sector" in kind or "reconocido" in kind:
        level = 3
    elif "ipcc" in kind or "internacional" in version.factor.country.lower() or "ipcc" in issuer:
        level = 4
    else:
        level = 5
    label, rationale = HIERARCHY_LEVELS[level]
    return level, label, rationale


def _temporal_alignment(
    version: EmissionFactorVersion,
    documentation: FactorDocumentation | None,
    record: ActivityData | None,
) -> tuple[str, str, int | None, int | None, int | None]:
    reference_year = record.period_start.year if record else date.today().year
    source_year = documentation.data_year if documentation and documentation.data_year else version.publication_year
    if not source_year:
        return "Sin año fuente", "No es posible contrastar temporalmente la versión.", reference_year, None, None
    gap = abs(reference_year - source_year)
    electricity = "electric" in version.factor.activity_type.lower() or "electric" in version.factor.name.lower()
    if gap == 0:
        state = "Mismo periodo"
        note = f"El año fuente {source_year} coincide con el periodo evaluado."
    elif gap == 1:
        state = "Periodo próximo"
        note = f"Existe una diferencia de un año entre el dato ({reference_year}) y la fuente ({source_year})."
    elif gap <= 3:
        state = "Requiere justificación"
        note = f"La fuente está separada {gap} años del periodo evaluado; documenta representatividad y ausencia de una alternativa más reciente."
    else:
        state = "Desactualizado"
        note = f"La fuente está separada {gap} años del periodo evaluado y requiere sustitución o justificación excepcional."
    if electricity and gap > 0:
        note += " Para electricidad del SIN se debe priorizar el factor oficial correspondiente al año del consumo cuando esté disponible."
    return state, note, reference_year, source_year, gap


def _decision_readiness(
    version: EmissionFactorVersion,
    documentation: FactorDocumentation | None,
    *,
    effective_ok: bool,
    completeness: int,
    review_overdue: bool,
    source_preliminary: bool,
) -> tuple[str, list[str]]:
    blockers: list[str] = []
    if version.status != "Aprobado":
        blockers.append("La versión del factor no está aprobada.")
    if not documentation:
        blockers.append("No existe documentación metodológica vinculada.")
    else:
        if documentation.reporting_use == "Demostrativo":
            blockers.append("El factor solo está autorizado para demostración.")
        if documentation.reporting_use == "Retirado":
            blockers.append("El factor está retirado para nuevos usos.")
        if documentation.review_status not in FORMAL_REVIEW_STATES:
            blockers.append(f"La revisión documental está en estado {documentation.review_status}.")
        if completeness < 70:
            blockers.append("La ficha documental tiene menos de 70% de completitud.")
    if source_preliminary:
        blockers.append("La fuente documental está en condición preliminar, en revisión o no está controlada.")
    if not effective_ok:
        blockers.append("La versión no es temporalmente aplicable al periodo evaluado.")
    if review_overdue:
        blockers.append("La revisión documental está vencida.")
    if blockers:
        if documentation and documentation.reporting_use == "Demostrativo":
            return "Solo demostración", blockers
        return "Requiere revisión", blockers
    return "Listo para evaluación", []

def load_factor_versions(session: Session) -> list[EmissionFactorVersion]:
    return list(session.scalars(
        select(EmissionFactorVersion)
        .options(
            selectinload(EmissionFactorVersion.factor),
            selectinload(EmissionFactorVersion.gas),
        )
        .order_by(EmissionFactorVersion.id)
    ))


def documentation_map(session: Session) -> dict[int, FactorDocumentation]:
    return {
        item.factor_version_id: item
        for item in session.scalars(
            select(FactorDocumentation).options(selectinload(FactorDocumentation.source_document))
        )
    }


def factor_passport(
    session: Session,
    version: EmissionFactorVersion,
    *,
    documentation: FactorDocumentation | None = None,
    source: EmissionSource | None = None,
    record: ActivityData | None = None,
    desired_country: str | None = "Colombia",
) -> dict[str, Any]:
    documentation = documentation or session.scalar(
        select(FactorDocumentation)
        .where(FactorDocumentation.factor_version_id == version.id)
        .options(selectinload(FactorDocumentation.source_document))
    )
    reference_date = record.period_start if record else None
    effective_state, effective_ok, effective_note = _effective_state(version, reference_date)
    geography_state, geography_note = _geographic_state(version, version.factor.country, desired_country)
    completeness, missing_fields = _document_completeness(documentation)
    review_due, review_overdue = _review_due_state(documentation)
    source_status, source_preliminary, source_governance_note = _source_governance_state(documentation)
    hierarchy_tier, hierarchy_label, hierarchy_rationale = _hierarchy_state(version, documentation)
    temporal_alignment, temporal_note, reference_year, source_data_year, year_gap = _temporal_alignment(version, documentation, record)
    readiness, blockers = _decision_readiness(
        version,
        documentation,
        effective_ok=effective_ok,
        completeness=completeness,
        review_overdue=review_overdue,
        source_preliminary=source_preliminary,
    )
    advice = advise_factor(session, source, record, version) if source and record else None
    warnings: list[str] = []
    if documentation and documentation.aggregated_co2e:
        warnings.append("El valor ya está agregado en CO₂e; no debe combinarse con factores por gas para la misma fracción de actividad.")
    if documentation and documentation.gwp_embedded:
        warnings.append(f"La fuente incorpora un GWP embebido ({documentation.gwp_embedded}); evita aplicar un GWP adicional.")
    if documentation and documentation.methane_origin not in {"", "No aplica"}:
        warnings.append(f"Origen de metano declarado: {documentation.methane_origin}.")
    if version.uncertainty_percentage <= 0:
        warnings.append("La versión no declara incertidumbre cuantitativa; documenta el supuesto o asigna un valor conservador.")
    if geography_state != "Compatible" and desired_country:
        warnings.append(geography_note)
    if source_preliminary:
        warnings.append(source_governance_note)
    if temporal_alignment in {"Requiere justificación", "Desactualizado"}:
        warnings.append(temporal_note)
    if advice:
        warnings.extend(item for item in advice["blockers"] if item not in warnings)
    return {
        "version": version,
        "factor": version.factor,
        "gas": version.gas,
        "documentation": documentation,
        "source_document": documentation.source_document if documentation else None,
        "source_label": _source_label(documentation),
        "source_status": source_status,
        "source_preliminary": source_preliminary,
        "source_governance_note": source_governance_note,
        "hierarchy_tier": hierarchy_tier,
        "hierarchy_label": hierarchy_label,
        "hierarchy_rationale": hierarchy_rationale,
        "temporal_alignment": temporal_alignment,
        "temporal_note": temporal_note,
        "reference_year": reference_year,
        "source_data_year": source_data_year,
        "year_gap": year_gap,
        "effective_state": effective_state,
        "effective_ok": effective_ok,
        "effective_note": effective_note,
        "geography_state": geography_state,
        "geography_note": geography_note,
        "document_completeness": completeness,
        "missing_document_fields": missing_fields,
        "review_due": review_due,
        "review_overdue": review_overdue,
        "decision_readiness": readiness,
        "decision_blockers": blockers,
        "warnings": warnings,
        "advice": advice,
        "quality_rank": QUALITY_RANK.get(documentation.quality_grade, 0) if documentation else 0,
        "formal": bool(
            documentation
            and documentation.reporting_use == "Formal"
            and documentation.review_status in FORMAL_REVIEW_STATES
            and version.status == "Aprobado"
            and not source_preliminary
        ),
    }

def factor_catalog(
    session: Session,
    *,
    query: str = "",
    sector: str = "",
    country: str = "",
    gas: str = "",
    unit: str = "",
    reporting_use: str = "",
    quality: str = "",
    readiness: str = "",
    hierarchy: int | None = None,
    temporal_status: str = "",
    data_year: int | None = None,
    source: EmissionSource | None = None,
    record: ActivityData | None = None,
) -> dict[str, Any]:
    versions = load_factor_versions(session)
    documents = documentation_map(session)
    passports = [
        factor_passport(
            session,
            version,
            documentation=documents.get(version.id),
            source=source,
            record=record,
            desired_country=country or (source.inventory.organization.country if source and source.inventory and source.inventory.organization else "Colombia"),
        )
        for version in versions
    ]
    normalized_query = query.strip().lower()
    result: list[dict[str, Any]] = []
    for item in passports:
        version = item["version"]
        factor = item["factor"]
        doc = item["documentation"]
        source_doc = item["source_document"]
        haystack = " ".join([
            factor.name, factor.activity_type, factor.sector, factor.country,
            version.geographic_scope, version.technology_scope, version.source_organization,
            version.source_document, version.notes, item["hierarchy_label"], item["temporal_alignment"],
            item["gas"].code, item["gas"].name,
            doc.factor_kind if doc else "", doc.restriction_notes if doc else "",
            source_doc.code if source_doc else "", source_doc.title if source_doc else "",
            source_doc.issuing_body if source_doc else "", item["source_status"],
        ]).lower()
        if normalized_query and normalized_query not in haystack:
            continue
        if sector and factor.sector != sector:
            continue
        if country and not (_contains(version.geographic_scope, country) or _contains(factor.country, country)):
            continue
        if gas and item["gas"].code != gas:
            continue
        if unit and version.input_unit != unit:
            continue
        if reporting_use and (not doc or doc.reporting_use != reporting_use):
            continue
        if quality and (not doc or doc.quality_grade != quality):
            continue
        if readiness and item["decision_readiness"] != readiness:
            continue
        if hierarchy and item["hierarchy_tier"] != hierarchy:
            continue
        if temporal_status and item["temporal_alignment"] != temporal_status:
            continue
        if data_year and (not doc or doc.data_year != data_year):
            continue
        result.append(item)
    if record:
        result.sort(
            key=lambda item: (
                bool(item["advice"] and item["advice"]["calculable"]),
                -item["hierarchy_tier"],
                item["temporal_alignment"] == "Mismo periodo",
                int(item["advice"]["score"]) if item["advice"] else 0,
                item["formal"],
                item["quality_rank"],
            ),
            reverse=True,
        )
    else:
        result.sort(
            key=lambda item: (
                item["decision_readiness"] == "Listo para evaluación",
                -item["hierarchy_tier"],
                item["formal"],
                item["quality_rank"],
                item["version"].publication_year,
            ),
            reverse=True,
        )
    all_docs = [item["documentation"] for item in passports if item["documentation"]]
    preliminary_documents = list(session.scalars(select(MethodologySourceDocument)))
    preliminary_source_ids = {
        item.id for item in preliminary_documents
        if any(marker in _clean(item.status).lower() for marker in PRELIMINARY_SOURCE_MARKERS)
    }
    filters = {
        "sectors": sorted({item["factor"].sector for item in passports if item["factor"].sector}),
        "countries": sorted({item["version"].geographic_scope for item in passports if item["version"].geographic_scope}),
        "gases": sorted({item["gas"].code for item in passports if item["gas"].code}),
        "units": sorted({item["version"].input_unit for item in passports if item["version"].input_unit}),
        "uses": sorted({item.reporting_use for item in all_docs if item.reporting_use}),
        "qualities": sorted({item.quality_grade for item in all_docs if item.quality_grade}, key=lambda value: -QUALITY_RANK.get(value, 0)),
        "readiness": ["Listo para evaluación", "Requiere revisión", "Solo demostración"],
        "hierarchy": [{"tier": tier, "label": value[0]} for tier, value in HIERARCHY_LEVELS.items()],
        "temporal": ["Mismo periodo", "Periodo próximo", "Requiere justificación", "Desactualizado", "Sin año fuente"],
        "years": sorted({item.data_year for item in all_docs if item.data_year}, reverse=True),
    }
    metrics = {
        "total": len(passports),
        "visible": len(result),
        "formal": sum(1 for item in passports if item["formal"]),
        "ready": sum(1 for item in passports if item["decision_readiness"] == "Listo para evaluación"),
        "review": sum(1 for item in passports if item["decision_readiness"] == "Requiere revisión"),
        "demo": sum(1 for item in passports if item["decision_readiness"] == "Solo demostración"),
        "overdue": sum(1 for item in passports if item["review_overdue"]),
        "documented": sum(1 for item in passports if item["document_completeness"] >= 70),
        "official_national": sum(1 for item in passports if item["hierarchy_tier"] == 2),
        "preliminary_sources": len(preliminary_source_ids),
        "temporally_stale": sum(1 for item in passports if item["temporal_alignment"] == "Desactualizado"),
        "same_period": sum(1 for item in passports if item["temporal_alignment"] == "Mismo periodo"),
    }
    return {
        "items": result,
        "all_items": passports,
        "filters": filters,
        "metrics": metrics,
        "hierarchy_policy": [{"tier": tier, "label": value[0], "rationale": value[1]} for tier, value in HIERARCHY_LEVELS.items()],
        "context": {"source": source, "record": record},
        "query": {
            "q": query, "sector": sector, "country": country, "gas": gas, "unit": unit,
            "reporting_use": reporting_use, "quality": quality, "readiness": readiness,
            "hierarchy": hierarchy, "temporal_status": temporal_status, "data_year": data_year,
        },
    }

def compare_factors(
    session: Session,
    version_ids: Iterable[int],
    *,
    source: EmissionSource | None = None,
    record: ActivityData | None = None,
) -> dict[str, Any]:
    ids = list(dict.fromkeys(int(item) for item in version_ids))[:6]
    versions = [version for version in load_factor_versions(session) if version.id in ids]
    documents = documentation_map(session)
    items = [
        factor_passport(session, version, documentation=documents.get(version.id), source=source, record=record)
        for version in versions
    ]
    alerts: list[dict[str, str]] = []
    if len(items) < 2:
        alerts.append({"level": "warning", "message": "Selecciona al menos dos versiones para una comparación útil."})
    gases = [item["gas"].code for item in items]
    duplicate_gases = sorted({gas for gas in gases if gases.count(gas) > 1})
    if duplicate_gases:
        alerts.append({"level": "danger", "message": "Posible doble conteo: más de una versión representa " + ", ".join(duplicate_gases) + "."})
    aggregate = [item for item in items if item["documentation"] and item["documentation"].aggregated_co2e]
    specific = [item for item in items if not (item["documentation"] and item["documentation"].aggregated_co2e)]
    if aggregate and specific:
        alerts.append({"level": "danger", "message": "No mezcles un factor agregado en CO₂e con factores por gas para la misma porción del dato."})
    units = sorted({item["version"].input_unit for item in items})
    if len(units) > 1:
        alerts.append({"level": "warning", "message": "Las versiones usan unidades de entrada distintas: " + ", ".join(units) + ". Compara la cadena de conversión antes de decidir."})
    uses = sorted({item["documentation"].reporting_use if item["documentation"] else "Sin clasificar" for item in items})
    if len(uses) > 1:
        alerts.append({"level": "warning", "message": "La comparación mezcla aptitudes de uso: " + ", ".join(uses) + "."})
    if any(item["review_overdue"] for item in items):
        alerts.append({"level": "warning", "message": "Al menos una ficha tiene revisión documental vencida."})
    if any(item["source_preliminary"] for item in items):
        alerts.append({"level": "danger", "message": "Al menos una versión depende de una fuente preliminar, en revisión o no controlada."})
    if len({item["hierarchy_tier"] for item in items}) > 1:
        alerts.append({"level": "warning", "message": "La comparación mezcla niveles de jerarquía. Debe justificarse por qué no se adopta el candidato de mayor prioridad metodológica."})
    if any(item["temporal_alignment"] in {"Requiere justificación", "Desactualizado"} for item in items):
        alerts.append({"level": "warning", "message": "Al menos una versión presenta separación temporal relevante frente al periodo evaluado."})
    if any(item["decision_readiness"] == "Solo demostración" for item in items):
        alerts.append({"level": "danger", "message": "Los factores demostrativos no deben sustentar inventarios formales ni declaraciones públicas."})
    if record and any(not item["advice"]["calculable"] for item in items if item["advice"]):
        alerts.append({"level": "danger", "message": "Al menos una versión no puede convertir la unidad del dato mediante una regla aprobada."})
    if not alerts:
        alerts.append({"level": "success", "message": "No se detectaron incompatibilidades estructurales automáticas. La decisión profesional sigue siendo obligatoria."})
    items.sort(key=lambda item: (bool(item["advice"] and item["advice"]["calculable"]), -item["hierarchy_tier"], int(item["advice"]["score"]) if item["advice"] else 0, item["formal"], item["quality_rank"]), reverse=True)
    return {"items": items, "alerts": alerts, "context": {"source": source, "record": record}}


def build_factor_comparison_workbook(comparison: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparación"
    headers = [
        "ID", "Factor", "Versión", "Actividad", "Sector", "Gas", "Valor", "Unidad entrada", "Unidad salida",
        "Geografía", "Tecnología", "Año publicación", "Vigencia", "Jerarquía", "Alineación temporal", "Estado fuente",
        "Tipo", "Uso", "Calidad", "Revisión", "Completitud documental %", "Preparación", "Compatibilidad %",
        "Calculable", "Incertidumbre %", "Fuente", "Referencia", "Restricciones",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for item in comparison["items"]:
        version = item["version"]
        factor = item["factor"]
        documentation = item["documentation"]
        source_doc = item["source_document"]
        advice = item["advice"]
        ws.append([
            version.id, factor.name, version.version, factor.activity_type, factor.sector, item["gas"].code,
            version.value, version.input_unit, version.output_unit, version.geographic_scope, version.technology_scope,
            version.publication_year, item["effective_state"], f"{item['hierarchy_tier']} · {item['hierarchy_label']}",
            item["temporal_alignment"], item["source_status"], documentation.factor_kind if documentation else "",
            documentation.reporting_use if documentation else "", documentation.quality_grade if documentation else "",
            documentation.review_status if documentation else "", item["document_completeness"], item["decision_readiness"],
            advice["score"] if advice else None, "Sí" if advice and advice["calculable"] else "No" if advice else "N/A",
            version.uncertainty_percentage, source_doc.code if source_doc else version.source_organization,
            f"{documentation.page_reference} {documentation.table_reference}".strip() if documentation else version.source_document,
            documentation.restriction_notes if documentation else version.notes,
        ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        letter = column[0].column_letter
        ws.column_dimensions[letter].width = min(52, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws = wb.create_sheet("Alertas")
    ws.append(["Nivel", "Alerta"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for alert in comparison["alerts"]:
        ws.append([alert["level"], alert["message"]])
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 100
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws = wb.create_sheet("Contexto")
    ws.append(["Campo", "Valor"])
    ws["A1"].font = ws["B1"].font = Font(bold=True)
    source = comparison["context"].get("source")
    record = comparison["context"].get("record")
    context_rows = [
        ("Fuente", source.name if source else "Comparación general"),
        ("Dato", f"{record.value} {record.unit}" if record else "Sin dato específico"),
        ("Periodo", f"{record.period_start.isoformat()} a {record.period_end.isoformat()}" if record else "No aplica"),
        ("Regla", "El puntaje organiza candidatos; no reemplaza la revisión y aprobación metodológica."),
    ]
    for row in context_rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def catalog_json(summary: dict[str, Any]) -> str:
    payload = {
        "version": "1.0.0",
        "governance_version": "1.1.0",
        "metrics": summary["metrics"],
        "hierarchy_policy": summary.get("hierarchy_policy", []),
        "context": {
            "source_id": summary["context"]["source"].id if summary["context"]["source"] else None,
            "activity_data_id": summary["context"]["record"].id if summary["context"]["record"] else None,
        },
        "items": [
            {
                "id": item["version"].id,
                "factor": item["factor"].name,
                "activity_type": item["factor"].activity_type,
                "sector": item["factor"].sector,
                "gas": item["gas"].code,
                "value": item["version"].value,
                "input_unit": item["version"].input_unit,
                "output_unit": item["version"].output_unit,
                "geographic_scope": item["version"].geographic_scope,
                "publication_year": item["version"].publication_year,
                "reporting_use": item["documentation"].reporting_use if item["documentation"] else "Sin clasificar",
                "quality_grade": item["documentation"].quality_grade if item["documentation"] else "N/A",
                "document_completeness": item["document_completeness"],
                "decision_readiness": item["decision_readiness"],
                "hierarchy_tier": item["hierarchy_tier"],
                "hierarchy_label": item["hierarchy_label"],
                "temporal_alignment": item["temporal_alignment"],
                "reference_year": item["reference_year"],
                "source_data_year": item["source_data_year"],
                "year_gap": item["year_gap"],
                "source_status": item["source_status"],
                "source_preliminary": item["source_preliminary"],
                "compatibility_score": item["advice"]["score"] if item["advice"] else None,
                "calculable": item["advice"]["calculable"] if item["advice"] else None,
                "warnings": item["warnings"],
            }
            for item in summary["items"]
        ],
    }
    return json.dumps(payload, ensure_ascii=False, default=str)
