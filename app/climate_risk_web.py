from __future__ import annotations

from io import BytesIO

from fastapi import Depends, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from .climate_risk import (
    assessment_summary,
    calculate_risk_scores,
    refresh_assessment_status,
    risk_level,
    synchronize_control_effectiveness,
)
from .database import add_audit, get_db
from .db.models import (
    ClimateRisk,
    ClimateRiskAssessment,
    ClimateRiskControl,
    ClimateTransitionAction,
    ClimateTransitionRoadmap,
    Inventory,
)


def _require_climate_risk_view(user: dict[str, object]) -> None:
    capabilities = user["capabilities"]
    if "view_climate_risk" not in capabilities and "manage_climate_risk" not in capabilities:
        raise HTTPException(403, "Tu rol no tiene acceso a riesgos climáticos")


def register_climate_risk_routes(
    app, templates, common_context, require_user, ensure_capability, set_flash, parse_date
) -> None:
    @app.get("/riesgos-climaticos", response_class=HTMLResponse)
    def climate_risk_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
        _require_climate_risk_view(user)
        organization_id = int(user["organization_id"])
        summary = assessment_summary(session, organization_id)
        inventories = list(session.scalars(select(Inventory).where(Inventory.organization_id == organization_id).order_by(Inventory.start_date.desc())))
        return templates.TemplateResponse(
            request=request, name="climate_risk.html",
            context=common_context(request, session, user, "climate_risk", summary=summary, inventories=inventories, risk_level=risk_level),
        )

    @app.post("/riesgos-climaticos/evaluacion")
    def climate_assessment_save(
        request: Request, name: str = Form(...), inventory_id: str = Form(""), methodology: str = Form("Análisis corporativo de escenarios"),
        scenario: str = Form("Escenario central"), base_year: int = Form(...), short_horizon: int = Form(...), medium_horizon: int = Form(...),
        long_horizon: int = Form(...), currency: str = Form("COP"), owner: str = Form(...), status: str = Form("En evaluación"), notes: str = Form(""),
        session: Session = Depends(get_db), user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_climate_risk")
        organization_id = int(user["organization_id"])
        if not (base_year <= short_horizon <= medium_horizon <= long_horizon):
            raise HTTPException(400, "Los horizontes deben estar ordenados desde el año base hasta el largo plazo")
        summary = assessment_summary(session, organization_id)
        assessment = summary["assessment"] or ClimateRiskAssessment(organization_id=organization_id, name=name.strip(), created_by=str(user["email"]))
        if not summary["assessment"]:
            session.add(assessment)
        assessment.name = name.strip(); assessment.inventory_id = int(inventory_id) if inventory_id else None
        assessment.methodology = methodology.strip(); assessment.scenario = scenario.strip(); assessment.base_year = base_year
        assessment.short_horizon = short_horizon; assessment.medium_horizon = medium_horizon; assessment.long_horizon = long_horizon
        assessment.currency = currency.strip().upper()[:20]; assessment.owner = owner.strip(); assessment.status = status; assessment.notes = notes.strip()
        refresh_assessment_status(session, assessment, str(user["email"])) if assessment.id else None
        add_audit(session, organization_id, str(user["email"]), "ACTUALIZAR", "Evaluación climática", assessment.name, new_value=assessment.status)
        session.commit(); set_flash(request, "Evaluación climática guardada.")
        return RedirectResponse("/riesgos-climaticos", status_code=303)

    @app.post("/riesgos-climaticos/riesgos/nuevo")
    def climate_risk_create(
        request: Request, risk_type: str = Form(...), category: str = Form(...), hazard: str = Form(...), description: str = Form(""),
        location: str = Form("Corporativo"), value_chain_stage: str = Form("Operación propia"), time_horizon: str = Form("Mediano plazo"),
        scenario: str = Form("Escenario central"), likelihood: int = Form(...), financial_impact: int = Form(...), operational_impact: int = Form(...),
        reputational_impact: int = Form(...), control_effectiveness: int = Form(0), financial_exposure: float = Form(0), owner: str = Form(...),
        response_strategy: str = Form("Mitigar"), response_detail: str = Form(""), status: str = Form("Abierto"), source_reference: str = Form(""),
        session: Session = Depends(get_db), user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_climate_risk")
        organization_id = int(user["organization_id"]); summary = assessment_summary(session, organization_id)
        assessment = summary["assessment"]
        if not assessment: raise HTTPException(409, "Primero crea la evaluación climática")
        if risk_type not in {"Físico", "Transición", "Oportunidad"}: raise HTTPException(400, "Tipo de riesgo inválido")
        likelihood = max(1, min(5, likelihood)); financial_impact = max(1, min(5, financial_impact))
        operational_impact = max(1, min(5, operational_impact)); reputational_impact = max(1, min(5, reputational_impact))
        inherent, residual = calculate_risk_scores(likelihood, financial_impact, operational_impact, reputational_impact, control_effectiveness)
        risk = ClimateRisk(
            assessment_id=assessment.id, organization_id=organization_id, risk_type=risk_type, category=category.strip(), hazard=hazard.strip(),
            description=description.strip(), location=location.strip(), value_chain_stage=value_chain_stage.strip(), time_horizon=time_horizon,
            scenario=scenario.strip(), likelihood=likelihood, financial_impact=financial_impact, operational_impact=operational_impact,
            reputational_impact=reputational_impact, inherent_score=inherent, control_effectiveness=max(0, min(100, control_effectiveness)),
            residual_score=residual, financial_exposure=max(0, financial_exposure), owner=owner.strip(), response_strategy=response_strategy,
            response_detail=response_detail.strip(), status=status, source_reference=source_reference.strip(), created_by=str(user["email"]),
        )
        session.add(risk); add_audit(session, organization_id, str(user["email"]), "CREAR", "Riesgo climático", risk.hazard, new_value=f"{risk_type} · {risk_level(residual)}")
        session.commit(); set_flash(request, "Riesgo climático registrado.")
        return RedirectResponse("/riesgos-climaticos", status_code=303)

    @app.post("/riesgos-climaticos/riesgos/{risk_id}/actualizar")
    def climate_risk_update(
        risk_id: int, request: Request, likelihood: int = Form(...), financial_impact: int = Form(...), operational_impact: int = Form(...),
        reputational_impact: int = Form(...), financial_exposure: float = Form(0), owner: str = Form(...), response_strategy: str = Form(...),
        response_detail: str = Form(""), status: str = Form(...), source_reference: str = Form(""),
        session: Session = Depends(get_db), user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_climate_risk")
        risk = session.get(ClimateRisk, risk_id)
        if not risk or risk.organization_id != int(user["organization_id"]): raise HTTPException(404, "Riesgo no encontrado")
        risk.likelihood=max(1,min(5,likelihood)); risk.financial_impact=max(1,min(5,financial_impact)); risk.operational_impact=max(1,min(5,operational_impact)); risk.reputational_impact=max(1,min(5,reputational_impact))
        risk.financial_exposure=max(0, financial_exposure); risk.owner=owner.strip(); risk.response_strategy=response_strategy; risk.response_detail=response_detail.strip(); risk.status=status; risk.source_reference=source_reference.strip()
        synchronize_control_effectiveness(session, risk)
        add_audit(session, risk.organization_id, str(user["email"]), "ACTUALIZAR", "Riesgo climático", risk.hazard, new_value=f"Residual {risk.residual_score}")
        session.commit(); set_flash(request, "Riesgo actualizado.")
        return RedirectResponse("/riesgos-climaticos", status_code=303)

    @app.post("/riesgos-climaticos/controles/nuevo")
    def climate_control_create(
        request: Request, risk_id: int = Form(...), name: str = Form(...), control_type: str = Form(...), owner: str = Form(...),
        status: str = Form(...), effectiveness: int = Form(...), implementation_date: str = Form(""), next_review: str = Form(""),
        annual_cost: float = Form(0), evidence: str = Form(""), session: Session = Depends(get_db), user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_climate_risk")
        organization_id = int(user["organization_id"]); risk = session.get(ClimateRisk, risk_id)
        if not risk or risk.organization_id != organization_id: raise HTTPException(404, "Riesgo no encontrado")
        control = ClimateRiskControl(risk_id=risk.id, organization_id=organization_id, name=name.strip(), control_type=control_type,
            owner=owner.strip(), status=status, effectiveness=max(0, min(100, effectiveness)),
            implementation_date=parse_date(implementation_date) if implementation_date else None, next_review=parse_date(next_review) if next_review else None,
            annual_cost=max(0, annual_cost), evidence=evidence.strip(), created_by=str(user["email"]))
        session.add(control); session.flush(); synchronize_control_effectiveness(session, risk)
        add_audit(session, organization_id, str(user["email"]), "CREAR", "Control climático", control.name, new_value=f"Efectividad {control.effectiveness}%")
        session.commit(); set_flash(request, "Control registrado y riesgo residual recalculado.")
        return RedirectResponse("/riesgos-climaticos", status_code=303)

    @app.post("/riesgos-climaticos/hoja-ruta")
    def climate_roadmap_save(
        request: Request, name: str = Form(...), baseline_year: int = Form(...), target_year: int = Form(...), owner: str = Form(...),
        governance: str = Form(""), approved_budget: float = Form(0), status: str = Form(...), notes: str = Form(""),
        session: Session = Depends(get_db), user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_climate_risk")
        organization_id = int(user["organization_id"]); summary=assessment_summary(session, organization_id); assessment=summary["assessment"]
        if target_year < baseline_year: raise HTTPException(400, "El año objetivo no puede ser anterior al año base")
        if not assessment: raise HTTPException(409, "Primero crea la evaluación climática")
        roadmap = summary["roadmap"] or ClimateTransitionRoadmap(organization_id=organization_id, assessment_id=assessment.id, name=name.strip(), created_by=str(user["email"]))
        if not summary["roadmap"]: session.add(roadmap)
        roadmap.name=name.strip(); roadmap.baseline_year=baseline_year; roadmap.target_year=target_year; roadmap.owner=owner.strip(); roadmap.governance=governance.strip(); roadmap.approved_budget=max(0, approved_budget); roadmap.status=status; roadmap.notes=notes.strip()
        add_audit(session, organization_id, str(user["email"]), "ACTUALIZAR", "Hoja de ruta climática", roadmap.name, new_value=roadmap.status)
        session.commit(); set_flash(request, "Hoja de ruta guardada.")
        return RedirectResponse("/riesgos-climaticos", status_code=303)

    @app.post("/riesgos-climaticos/acciones/nueva")
    def climate_action_create(
        request: Request, risk_id: str = Form(""), category: str = Form(...), title: str = Form(...), description: str = Form(""), owner: str = Form(...),
        start_date: str = Form(""), end_date: str = Form(""), priority: str = Form("Media"), status: str = Form("Planeada"), progress: int = Form(0),
        expected_reduction_tco2e: float = Form(0), capex: float = Form(0), annual_opex: float = Form(0), annual_savings: float = Form(0),
        avoided_loss: float = Form(0), indicator: str = Form(""), target_value: float = Form(0), current_value: float = Form(0), unit: str = Form(""),
        dependencies: str = Form(""), evidence_note: str = Form(""), session: Session = Depends(get_db), user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_climate_risk")
        organization_id=int(user["organization_id"]); summary=assessment_summary(session, organization_id); roadmap=summary["roadmap"]
        parsed_start = parse_date(start_date) if start_date else None; parsed_end = parse_date(end_date) if end_date else None
        if parsed_start and parsed_end and parsed_end < parsed_start: raise HTTPException(400, "La fecha final no puede ser anterior a la fecha inicial")
        if not roadmap: raise HTTPException(409, "Primero crea la hoja de ruta")
        linked_risk = session.get(ClimateRisk, int(risk_id)) if risk_id else None
        if linked_risk and linked_risk.organization_id != organization_id: raise HTTPException(404, "Riesgo no encontrado")
        action=ClimateTransitionAction(roadmap_id=roadmap.id, organization_id=organization_id, risk_id=linked_risk.id if linked_risk else None,
            category=category.strip(), title=title.strip(), description=description.strip(), owner=owner.strip(), start_date=parsed_start,
            end_date=parsed_end, priority=priority, status=status, progress=max(0,min(100,progress)), expected_reduction_tco2e=max(0,expected_reduction_tco2e),
            capex=max(0,capex), annual_opex=max(0,annual_opex), annual_savings=max(0,annual_savings), avoided_loss=max(0,avoided_loss), indicator=indicator.strip(),
            target_value=target_value, current_value=current_value, unit=unit.strip(), dependencies=dependencies.strip(), evidence_note=evidence_note.strip(), created_by=str(user["email"]))
        session.add(action); add_audit(session, organization_id, str(user["email"]), "CREAR", "Acción climática", action.title, new_value=action.status)
        session.commit(); set_flash(request, "Acción añadida a la hoja de ruta.")
        return RedirectResponse("/riesgos-climaticos", status_code=303)

    @app.post("/riesgos-climaticos/acciones/{action_id}/estado")
    def climate_action_update(
        action_id: int, request: Request, status: str = Form(...), progress: int = Form(...), current_value: float = Form(0), evidence_note: str = Form(""),
        session: Session = Depends(get_db), user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_climate_risk")
        action=session.get(ClimateTransitionAction, action_id)
        if not action or action.organization_id != int(user["organization_id"]): raise HTTPException(404, "Acción no encontrada")
        action.status=status; action.progress=max(0,min(100,progress)); action.current_value=current_value; action.evidence_note=evidence_note.strip()
        add_audit(session, action.organization_id, str(user["email"]), "ACTUALIZAR", "Acción climática", action.title, new_value=f"{status} · {action.progress}%")
        session.commit(); set_flash(request, "Avance de la acción actualizado.")
        return RedirectResponse("/riesgos-climaticos", status_code=303)

    @app.get("/riesgos-climaticos/exportar.xlsx")
    def climate_risk_export(session: Session = Depends(get_db), user: dict = Depends(require_user)):
        _require_climate_risk_view(user); organization_id=int(user["organization_id"]); summary=assessment_summary(session, organization_id)
        wb=Workbook(); ws=wb.active; ws.title="Resumen"
        ws.append(["Evaluación", summary["assessment"].name if summary["assessment"] else ""]); ws.append(["Riesgos", summary["counts"]["total"]]); ws.append(["Exposición bruta", summary["financial"]["gross_exposure"]]); ws.append(["Exposición residual", summary["financial"]["residual_exposure"]]); ws.append(["Exposición evitada", summary["financial"]["avoided_exposure"]]); ws.append(["Valor de oportunidades", summary["financial"]["opportunity_value"]]); ws.append(["Costo anual de controles", summary["financial"]["control_cost"]]); ws.append(["Preparación", summary["readiness_score"]])
        ws2=wb.create_sheet("Riesgos"); ws2.append(["Tipo","Categoría","Riesgo u oportunidad","Ubicación","Horizonte","Probabilidad","Impacto máximo","Inherente","Controles %","Residual","Nivel","Exposición","Responsable","Estrategia","Estado","Fuente"])
        for risk in summary["risks"]: ws2.append([risk.risk_type,risk.category,risk.hazard,risk.location,risk.time_horizon,risk.likelihood,max(risk.financial_impact,risk.operational_impact,risk.reputational_impact),risk.inherent_score,risk.control_effectiveness,risk.residual_score,risk_level(risk.residual_score),risk.financial_exposure,risk.owner,risk.response_strategy,risk.status,risk.source_reference])
        ws3=wb.create_sheet("Controles"); ws3.append(["Riesgo","Control","Tipo","Responsable","Estado","Efectividad %","Costo anual","Próxima revisión","Evidencia"])
        risk_map={risk.id:risk.hazard for risk in summary["risks"]}
        for control in summary["controls"]: ws3.append([risk_map.get(control.risk_id,""),control.name,control.control_type,control.owner,control.status,control.effectiveness,control.annual_cost,control.next_review,control.evidence])
        ws4=wb.create_sheet("Hoja de ruta"); ws4.append(["Categoría","Acción","Riesgo vinculado","Responsable","Inicio","Fin","Prioridad","Estado","Avance %","Reducción tCO2e","CAPEX","OPEX anual","Ahorro anual","Pérdida evitada","Indicador","Meta","Actual","Unidad","Dependencias"])
        for action in summary["actions"]: ws4.append([action.category,action.title,risk_map.get(action.risk_id,""),action.owner,action.start_date,action.end_date,action.priority,action.status,action.progress,action.expected_reduction_tco2e,action.capex,action.annual_opex,action.annual_savings,action.avoided_loss,action.indicator,action.target_value,action.current_value,action.unit,action.dependencies])
        buffer=BytesIO(); wb.save(buffer)
        return Response(content=buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="riesgos_climaticos_{organization_id}.xlsx"'})
