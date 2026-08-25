from __future__ import annotations

import hashlib
import secrets
from calendar import monthrange
from datetime import UTC, date, datetime
from io import BytesIO
from pathlib import Path

from fastapi import Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from .calculations import recalculate_inventory, recalculate_source
from .capture_guidance import capture_summary
from .config import settings
from .data_request_status import open_data_requests
from .database import (
    ActivityData, DataRequest, EmissionSource, EvidenceDocument, Inventory,
    add_audit, get_db, refresh_progress,
)
from .period_close import assert_periods_editable
from .security import validate_upload_bytes
from .storage import StorageError, storage


def register_information_routes(
    app,
    templates,
    common_context,
    require_user,
    ensure_capability,
    set_flash,
    parse_date,
    get_inventory,
    ensure_inventory_editable,
    quality_from,
    safe_filename,
    format_bytes,
    allowed_upload_extensions,
    max_upload_size,
    allowed_units,
    data_origins,
    parse_excel_period,
) -> None:
    ALLOWED_UPLOAD_EXTENSIONS = allowed_upload_extensions
    MAX_UPLOAD_SIZE = max_upload_size
    ALLOWED_UNITS = allowed_units
    DATA_ORIGINS = data_origins
    _parse_excel_period = parse_excel_period

    @app.get("/informacion", response_class=HTMLResponse)
    def information_page(request: Request, show_all: bool = False, session: Session = Depends(get_db), user: dict = Depends(require_user)):
        inventory = get_inventory(session, user)
        requests = list(session.scalars(select(DataRequest).where(DataRequest.inventory_id == inventory.id).order_by(DataRequest.due_date)))
        documents = list(session.scalars(select(EvidenceDocument).where(EvidenceDocument.inventory_id == inventory.id).order_by(EvidenceDocument.uploaded_at.desc())))
        records = list(
            session.scalars(
                select(ActivityData)
                .join(EmissionSource)
                .where(EmissionSource.inventory_id == inventory.id)
                .options(selectinload(ActivityData.source).selectinload(EmissionSource.facility), selectinload(ActivityData.evidence))
                .order_by(ActivityData.period_start.desc(), ActivityData.created_at.desc())
            )
        )
        open_requests = open_data_requests(requests)
        role = str(user.get("role", "Cliente"))
        if role == "Cliente" and open_requests:
            information_focus = {
                "task": "solicitudes",
                "eyebrow": "RESPONDE LO PENDIENTE",
                "title": "Atiende las solicitudes de información abiertas",
                "detail": f"Tienes {len(open_requests)} requerimiento(s) activo(s). Revisa qué dato o soporte falta y actualiza su estado cuando quede entregado.",
                "href": "#solicitudes",
                "action": "Abrir solicitudes",
            }
        elif bool(user.get("can_provide_data")):
            information_focus = {
                "task": "datos",
                "eyebrow": "CONTINÚA LA CAPTURA",
                "title": "Registra el siguiente dato verificable",
                "detail": "Selecciona una fuente y periodo pendientes. Prioriza valores respaldados por factura, certificado o registro operativo.",
                "href": "/captura-guiada",
                "action": "Ir a captura guiada",
            }
        elif bool(user.get("can_review")):
            information_focus = {
                "task": "evidencias",
                "eyebrow": "REVISA LA EVIDENCIA",
                "title": "Comprueba soportes, periodos e integridad",
                "detail": f"Hay {len(documents)} evidencia(s) en el expediente. Revisa que el soporte corresponda al dato, periodo y fuente antes de aprobarlo.",
                "href": "#evidencias",
                "action": "Revisar evidencias",
            }
        elif open_requests:
            information_focus = {
                "task": "solicitudes",
                "eyebrow": "GESTIONA EL FLUJO",
                "title": "Resuelve las solicitudes abiertas",
                "detail": f"Quedan {len(open_requests)} requerimiento(s) activos antes de completar la recolección del periodo.",
                "href": "#solicitudes",
                "action": "Ver solicitudes",
            }
        else:
            information_focus = {
                "task": "datos",
                "eyebrow": "REVISA LA COBERTURA",
                "title": "Confirma que cada fuente tenga datos y soporte",
                "detail": "No hay solicitudes abiertas. Revisa el historial y completa los periodos o evidencias que aún falten.",
                "href": "#datos",
                "action": "Revisar registros",
            }
        quality_counts = {level: sum(1 for item in records if item.quality_level == level) for level in ("A", "B", "C", "D")}
        visible_records = records if show_all else records[:12]
        first_period_start = inventory.start_date
        first_period_end = min(
            inventory.end_date,
            date(first_period_start.year, first_period_start.month, monthrange(first_period_start.year, first_period_start.month)[1]),
        )
        return templates.TemplateResponse(
            request=request,
            name="information.html",
            context=common_context(
                request,
                session,
                user,
                "information",
                inventory=inventory,
                requests=requests,
                open_requests=open_requests,
                information_focus=information_focus,
                documents=documents,
                records=records,
                visible_records=visible_records,
                show_all=show_all,
                sources=inventory.sources,
                quality_counts=quality_counts,
                allowed_units=ALLOWED_UNITS,
                data_origins=DATA_ORIGINS,
                default_period_start=first_period_start.isoformat(),
                default_period_end=first_period_end.isoformat(),
            ),
        )

    @app.post("/informacion/solicitudes/nueva")
    def request_create(
        request: Request,
        title: str = Form(...),
        source_id: int | None = Form(None),
        source_name: str = Form(""),
        requested_to: str = Form(...),
        due_date: str = Form(...),
        instructions: str = Form(""),
        session: Session = Depends(get_db),
        user: dict = Depends(require_user),
    ):
        ensure_capability(user, "manage_sources")
        inventory = get_inventory(session, user)
        ensure_inventory_editable(inventory)
        source = session.get(EmissionSource, source_id) if source_id else None
        if source and source.inventory_id != inventory.id:
            raise HTTPException(400, "Fuente inválida")
        data_request = DataRequest(
            inventory_id=inventory.id,
            source_id=source.id if source else None,
            title=title.strip(),
            source_name=source.name if source else source_name.strip(),
            requested_to=requested_to.strip(),
            due_date=parse_date(due_date),
            status="Pendiente",
            instructions=instructions.strip(),
        )
        session.add(data_request)
        add_audit(session, int(user["organization_id"]), str(user["email"]), "SOLICITAR", "Información", data_request.title, f"Responsable: {data_request.requested_to}")
        session.commit()
        set_flash(request, "La solicitud de información fue creada.")
        return RedirectResponse("/informacion", status_code=303)

    @app.post("/informacion/solicitudes/{request_id}/estado")
    def request_status_update(
        request_id: int,
        request: Request,
        status: str = Form(...),
        session: Session = Depends(get_db),
        user: dict = Depends(require_user),
    ):
        if not (user["can_manage_sources"] or user["can_provide_data"]):
            raise HTTPException(403, "Tu rol no puede actualizar solicitudes")
        item = session.scalar(select(DataRequest).join(Inventory).where(DataRequest.id == request_id, Inventory.organization_id == int(user["organization_id"])).options(selectinload(DataRequest.inventory)))
        if not item:
            raise HTTPException(404, "Solicitud no encontrada")
        ensure_inventory_editable(item.inventory)
        allowed = {"Pendiente", "En preparación", "Cargado", "En revisión", "Completado", "Devuelto"}
        if status not in allowed:
            raise HTTPException(400, "Estado inválido")
        item.status = status
        item.completed_at = datetime.now(UTC) if status == "Completado" else None
        add_audit(session, int(user["organization_id"]), str(user["email"]), "ACTUALIZAR", "Solicitud", item.title, f"Estado: {status}")
        session.commit()
        set_flash(request, "El estado de la solicitud fue actualizado.")
        return RedirectResponse("/informacion", status_code=303)

    @app.post("/informacion/datos/nuevo")
    def activity_data_create(
        request: Request,
        source_id: int = Form(...),
        period_start: str = Form(...),
        period_end: str = Form(...),
        value: float = Form(...),
        unit: str = Form(...),
        data_origin: str = Form(...),
        evidence_id: int | None = Form(None),
        is_estimated: str | None = Form(None),
        uncertainty_percentage: float = Form(0),
        uncertainty_basis: str = Form(""),
        notes: str = Form(""),
        session: Session = Depends(get_db),
        user: dict = Depends(require_user),
    ):
        if not user["can_provide_data"]:
            raise HTTPException(403, "Tu rol no puede cargar datos")
        inventory = get_inventory(session, user)
        ensure_inventory_editable(inventory)
        source = session.scalar(select(EmissionSource).where(EmissionSource.id == source_id, EmissionSource.inventory_id == inventory.id))
        if not source:
            raise HTTPException(400, "Fuente inválida")
        if unit not in ALLOWED_UNITS:
            raise HTTPException(400, "Unidad no autorizada")
        if data_origin not in DATA_ORIGINS:
            raise HTTPException(400, "Origen del dato inválido")
        start_date = parse_date(period_start)
        end_date = parse_date(period_end)
        if start_date > end_date or start_date < inventory.start_date or end_date > inventory.end_date:
            raise HTTPException(400, "El periodo no corresponde al inventario")
        try:
            assert_periods_editable(session, inventory.id, [(start_date, end_date)])
        except ValueError as exc:
            raise HTTPException(409, str(exc)) from exc
        duplicate = session.scalar(select(ActivityData).where(ActivityData.source_id == source.id, ActivityData.period_start == start_date, ActivityData.period_end == end_date))
        if duplicate:
            set_flash(request, "Ya existe un registro para esa fuente y periodo.", "error")
            return RedirectResponse("/informacion", status_code=303)
        evidence = session.get(EvidenceDocument, evidence_id) if evidence_id else None
        if evidence and evidence.inventory_id != inventory.id:
            raise HTTPException(400, "Evidencia inválida")
        estimated = is_estimated == "on"
        record = ActivityData(
            source_id=source.id,
            evidence_id=evidence.id if evidence else None,
            period_start=start_date,
            period_end=end_date,
            value=max(value, 0),
            unit=unit,
            data_origin=data_origin,
            quality_level=quality_from(data_origin, estimated, evidence is not None),
            is_estimated=estimated,
            uncertainty_percentage=max(0, uncertainty_percentage),
            uncertainty_basis=uncertainty_basis.strip(),
            notes=notes.strip(),
            status="Provisional" if estimated else "Cargado",
            created_by=str(user["email"]),
        )
        session.add(record)
        session.flush()
        refresh_progress(session, inventory)
        calculation_result = recalculate_source(session, source)
        add_audit(session, int(user["organization_id"]), str(user["email"]), "CARGAR", "Dato de actividad", f"{source.name} · {start_date:%Y-%m}", f"{record.value} {record.unit} · calidad {record.quality_level} · {calculation_result['calculations']} cálculos")
        session.commit()
        set_flash(request, "El dato de actividad fue registrado.")
        return RedirectResponse("/informacion", status_code=303)

    @app.post("/informacion/datos/{record_id}/editar")
    def activity_data_update(
        record_id: int,
        request: Request,
        value: float = Form(...),
        unit: str = Form(...),
        data_origin: str = Form(...),
        status: str = Form("Cargado"),
        evidence_id: int | None = Form(None),
        is_estimated: str | None = Form(None),
        uncertainty_percentage: float = Form(0),
        uncertainty_basis: str = Form(""),
        notes: str = Form(""),
        session: Session = Depends(get_db),
        user: dict = Depends(require_user),
    ):
        if not user["can_provide_data"] and not user["can_review"]:
            raise HTTPException(403, "Tu rol no puede editar datos")
        record = session.scalar(
            select(ActivityData)
            .join(EmissionSource)
            .join(Inventory)
            .where(ActivityData.id == record_id, Inventory.organization_id == int(user["organization_id"]))
            .options(selectinload(ActivityData.source).selectinload(EmissionSource.inventory))
        )
        if not record:
            raise HTTPException(404, "Dato no encontrado")
        ensure_inventory_editable(record.source.inventory)
        try:
            assert_periods_editable(session, record.source.inventory_id, [(record.period_start, record.period_end)])
        except ValueError as exc:
            raise HTTPException(409, str(exc)) from exc
        if unit not in ALLOWED_UNITS or data_origin not in DATA_ORIGINS:
            raise HTTPException(400, "Unidad u origen inválidos")
        evidence = session.get(EvidenceDocument, evidence_id) if evidence_id else None
        if evidence and evidence.inventory_id != record.source.inventory_id:
            raise HTTPException(400, "Evidencia inválida")
        estimated = is_estimated == "on"
        record.value = max(value, 0)
        record.unit = unit
        record.data_origin = data_origin
        record.evidence_id = evidence.id if evidence else None
        record.is_estimated = estimated
        record.uncertainty_percentage = max(0, uncertainty_percentage)
        record.uncertainty_basis = uncertainty_basis.strip()
        record.quality_level = quality_from(data_origin, estimated, evidence is not None)
        record.notes = notes.strip()
        record.status = status if status in {"Cargado", "En revisión", "Aprobado", "Devuelto", "Provisional"} else "Cargado"
        # SessionLocal usa autoflush=False. Persistir antes de recargar la fuente evita
        # recalcular con los valores anteriores y perder incertidumbre/ediciones.
        session.flush()
        refresh_progress(session, record.source.inventory)
        calculation_result = recalculate_source(session, record.source)
        add_audit(session, int(user["organization_id"]), str(user["email"]), "EDITAR", "Dato de actividad", f"{record.source.name} · {record.period_start:%Y-%m}", f"Estado {record.status} · calidad {record.quality_level} · {calculation_result['calculations']} cálculos")
        session.commit()
        set_flash(request, "El dato fue actualizado.")
        return RedirectResponse(f"/fuentes/{record.source_id}", status_code=303)

    @app.post("/informacion/evidencias/nueva")
    async def evidence_upload(
        request: Request,
        file: UploadFile = File(...),
        document_type: str = Form(...),
        source_id: int | None = Form(None),
        period_label: str = Form(""),
        notes: str = Form(""),
        session: Session = Depends(get_db),
        user: dict = Depends(require_user),
    ):
        if not user["can_provide_data"]:
            raise HTTPException(403, "Tu rol no puede cargar evidencias")
        inventory = get_inventory(session, user)
        ensure_inventory_editable(inventory)
        original_name = safe_filename(file.filename or "archivo")
        extension = Path(original_name).suffix.lower()
        if extension not in ALLOWED_UPLOAD_EXTENSIONS:
            set_flash(request, "Formato no permitido. Usa PDF, Excel, CSV, JPG o PNG.", "error")
            return RedirectResponse("/informacion", status_code=303)
        content = await file.read(MAX_UPLOAD_SIZE + 1)
        if len(content) > MAX_UPLOAD_SIZE:
            set_flash(request, f"El archivo supera el límite de {settings.max_upload_mb} MB.", "error")
            return RedirectResponse("/informacion", status_code=303)
        valid_file, validation_message, detected_mime = validate_upload_bytes(
            original_name, content, file.content_type, ALLOWED_UPLOAD_EXTENSIONS
        )
        if not valid_file:
            set_flash(request, validation_message, "error")
            return RedirectResponse("/informacion", status_code=303)
        source = session.get(EmissionSource, source_id) if source_id else None
        if source and source.inventory_id != inventory.id:
            raise HTTPException(400, "Fuente inválida")
        stored_file = f"{datetime.now():%Y%m%d_%H%M%S}_{secrets.token_hex(4)}_{original_name}"
        storage_key = f"uploads/org_{user['organization_id']}/inventory_{inventory.id}/{stored_file}"
        try:
            storage.put_bytes(storage_key, content, detected_mime)
        except StorageError as exc:
            raise HTTPException(500, f"No fue posible almacenar la evidencia: {exc}") from exc
        document = EvidenceDocument(
            inventory_id=inventory.id,
            source_id=source.id if source else None,
            name=original_name,
            stored_name=storage_key,
            document_type=document_type.strip(),
            source_name=source.name if source else "General",
            period_label=period_label.strip(),
            status="Cargado",
            uploaded_by=str(user["name"]),
            file_size=len(content),
            sha256=hashlib.sha256(content).hexdigest(),
            notes=notes.strip(),
        )
        session.add(document)
        add_audit(session, int(user["organization_id"]), str(user["email"]), "CARGAR", "Evidencia", document.name, f"{document.document_type} · {format_bytes(document.file_size)}")
        session.commit()
        set_flash(request, "La evidencia fue cargada y protegida con huella SHA-256.")
        return RedirectResponse("/informacion", status_code=303)

    @app.get("/evidencias/{document_id}/descargar")
    def evidence_download(document_id: int, session: Session = Depends(get_db), user: dict = Depends(require_user)):
        document = session.scalar(select(EvidenceDocument).join(Inventory).where(EvidenceDocument.id == document_id, Inventory.organization_id == int(user["organization_id"])))
        if not document:
            raise HTTPException(404, "Documento no encontrado")
        if not document.stored_name or not storage.exists(document.stored_name):
            raise HTTPException(404, "El archivo físico no está disponible")
        add_audit(session, int(user["organization_id"]), str(user["email"]), "DESCARGAR", "Evidencia", document.name, "Descarga autorizada")
        session.commit()
        local_path = storage.local_path(document.stored_name)
        if local_path:
            return FileResponse(local_path, filename=document.name)
        return RedirectResponse(storage.presigned_url(document.stored_name), status_code=302)

    @app.post("/evidencias/{document_id}/estado")
    def evidence_status_update(
        document_id: int,
        request: Request,
        status: str = Form(...),
        session: Session = Depends(get_db),
        user: dict = Depends(require_user),
    ):
        if not user["can_review"]:
            raise HTTPException(403, "Tu rol no puede revisar evidencias")
        document = session.scalar(select(EvidenceDocument).join(Inventory).where(EvidenceDocument.id == document_id, Inventory.organization_id == int(user["organization_id"])))
        if not document:
            raise HTTPException(404, "Documento no encontrado")
        inventory = session.get(Inventory, document.inventory_id)
        ensure_inventory_editable(inventory)
        if status not in {"Cargado", "En revisión", "Aprobado", "Devuelto"}:
            raise HTTPException(400, "Estado inválido")
        document.status = status
        add_audit(session, int(user["organization_id"]), str(user["email"]), "REVISAR", "Evidencia", document.name, f"Estado: {status}")
        session.commit()
        set_flash(request, "La evidencia fue revisada.")
        return RedirectResponse("/informacion", status_code=303)

    @app.get("/informacion/plantilla.xlsx")
    def activity_template(session: Session = Depends(get_db), user: dict = Depends(require_user)):
        inventory = get_inventory(session, user)
        summary = capture_summary(inventory)
        workbook = Workbook()
        plan = workbook.active
        plan.title = "Plan de captura"
        plan_headers = ["Prioridad", "Fuente", "Alcance", "Categoría", "Frecuencia", "Responsable", "Unidad esperada", "Soporte recomendado", "Próximo inicio", "Próximo final", "Cobertura %", "Estado"]
        plan.append(plan_headers)
        for position, item in enumerate(summary["cards"], 1):
            plan.append([
                position,
                item["source"].name,
                item["source"].scope,
                item["source"].category,
                item["source"].data_frequency,
                item["source"].responsible,
                item["profile"]["unit"],
                item["profile"]["evidence"],
                item["next_start"].isoformat() if item["next_start"] else "",
                item["next_end"].isoformat() if item["next_end"] else "",
                item["coverage"],
                item["status"],
            ])
        data = workbook.create_sheet("Datos")
        headers = ["Fuente", "Periodo", "Valor", "Unidad", "Origen", "Estimado", "Incertidumbre %", "Base incertidumbre", "Observaciones"]
        data.append(headers)
        data.append(["Electricidad", f"{inventory.start_date:%Y-%m}", 18450, "kWh", "Factura", "No", 5, "Facturación medida", "Fila de ejemplo: reemplázala o elimínala antes de importar"])
        catalogs = workbook.create_sheet("Catálogos")
        catalogs.append(["Fuentes", "Unidades", "Orígenes"])
        max_rows = max(len(inventory.sources), len(ALLOWED_UNITS), len(DATA_ORIGINS))
        for index in range(max_rows):
            catalogs.append([
                inventory.sources[index].name if index < len(inventory.sources) else "",
                ALLOWED_UNITS[index] if index < len(ALLOWED_UNITS) else "",
                DATA_ORIGINS[index] if index < len(DATA_ORIGINS) else "",
            ])
        instructions = workbook.create_sheet("Instrucciones")
        instructions.append(["Paso", "Qué hacer", "Control de calidad"])
        instruction_rows = [
            (1, "Revisa el Plan de captura y prioriza las fuentes pendientes.", "La prioridad combina materialidad, periodos faltantes y soportes."),
            (2, "Completa la hoja Datos sin cambiar los encabezados.", "Usa una fila por fuente y periodo; no combines unidades."),
            (3, "Conserva facturas, certificados o registros del mismo periodo.", "El archivo Excel no reemplaza las evidencias."),
            (4, "Marca Estimado cuando el valor no provenga de medición o soporte directo.", "Documenta la base de incertidumbre y los supuestos."),
            (5, "Importa el archivo y resuelve todos los errores antes de aplicar.", "La plataforma evita importaciones parciales."),
        ]
        for row in instruction_rows:
            instructions.append(row)
        header_fill = PatternFill("solid", fgColor="1F5B45")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet in (plan, data, catalogs, instructions):
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            sheet.auto_filter.ref = sheet.dimensions
        widths = {
            "Plan de captura": [10, 30, 10, 28, 14, 24, 16, 38, 14, 14, 13, 18],
            "Datos": [30, 14, 15, 14, 26, 12, 18, 30, 50],
            "Catálogos": [30, 20, 30],
            "Instrucciones": [10, 58, 58],
        }
        for sheet_name, values in widths.items():
            sheet = workbook[sheet_name]
            for index, width in enumerate(values, 1):
                sheet.column_dimensions[chr(64 + index)].width = width
        source_validation = DataValidation(type="list", formula1=f"'Catálogos'!$A$2:$A${len(inventory.sources)+1}", allow_blank=False)
        unit_validation = DataValidation(type="list", formula1=f"'Catálogos'!$B$2:$B${len(ALLOWED_UNITS)+1}", allow_blank=False)
        origin_validation = DataValidation(type="list", formula1=f"'Catálogos'!$C$2:$C${len(DATA_ORIGINS)+1}", allow_blank=False)
        estimated_validation = DataValidation(type="list", formula1='"Sí,No"', allow_blank=False)
        for validation in (source_validation, unit_validation, origin_validation, estimated_validation):
            data.add_data_validation(validation)
        source_validation.add("A2:A1000")
        unit_validation.add("D2:D1000")
        origin_validation.add("E2:E1000")
        estimated_validation.add("F2:F1000")
        output = BytesIO()
        workbook.save(output)
        filename = f"Plantilla_sectorial_datos_{inventory.base_year}.xlsx"
        return Response(
            output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @app.get("/informacion/importar", response_class=HTMLResponse)
    def import_page(request: Request, session: Session = Depends(get_db), user: dict = Depends(require_user)):
        inventory = get_inventory(session, user)
        return templates.TemplateResponse(
            request=request,
            name="import_data.html",
            context=common_context(request, session, user, "information", inventory=inventory, sources=inventory.sources, errors=[], preview=[]),
        )

    @app.post("/informacion/importar", response_class=HTMLResponse)
    async def import_activity_data(
        request: Request,
        file: UploadFile = File(...),
        session: Session = Depends(get_db),
        user: dict = Depends(require_user),
    ):
        if not user["can_provide_data"]:
            raise HTTPException(403, "Tu rol no puede importar datos")
        inventory = get_inventory(session, user)
        ensure_inventory_editable(inventory)
        if Path(file.filename or "").suffix.lower() != ".xlsx":
            return templates.TemplateResponse(request=request, name="import_data.html", context=common_context(request, session, user, "information", inventory=inventory, sources=inventory.sources, errors=["El archivo debe estar en formato .xlsx"], preview=[]), status_code=400)
        content = await file.read(MAX_UPLOAD_SIZE + 1)
        if len(content) > MAX_UPLOAD_SIZE:
            return templates.TemplateResponse(request=request, name="import_data.html", context=common_context(request, session, user, "information", inventory=inventory, sources=inventory.sources, errors=[f"El archivo supera {settings.max_upload_mb} MB"], preview=[]), status_code=400)
        valid_file, validation_message, _ = validate_upload_bytes(
            file.filename or "datos.xlsx", content, file.content_type, {".xlsx"}
        )
        if not valid_file:
            return templates.TemplateResponse(request=request, name="import_data.html", context=common_context(request, session, user, "information", inventory=inventory, sources=inventory.sources, errors=[validation_message], preview=[]), status_code=400)
        try:
            workbook = load_workbook(BytesIO(content), data_only=True)
            sheet = workbook["Datos"]
        except Exception as exc:
            return templates.TemplateResponse(request=request, name="import_data.html", context=common_context(request, session, user, "information", inventory=inventory, sources=inventory.sources, errors=[f"No fue posible leer la hoja Datos: {exc}"], preview=[]), status_code=400)
        headers = [str(cell.value or "").strip() for cell in sheet[1]]
        legacy_headers = ["Fuente", "Periodo", "Valor", "Unidad", "Origen", "Estimado", "Observaciones"]
        v034_headers = ["Fuente", "Periodo", "Valor", "Unidad", "Origen", "Estimado", "Incertidumbre %", "Base incertidumbre", "Observaciones"]
        if headers[: len(v034_headers)] == v034_headers:
            template_version = "current"
        elif headers[: len(legacy_headers)] == legacy_headers:
            template_version = "legacy"
        else:
            return templates.TemplateResponse(request=request, name="import_data.html", context=common_context(request, session, user, "information", inventory=inventory, sources=inventory.sources, errors=["Las columnas no coinciden con la plantilla oficial."], preview=[]), status_code=400)
        source_map = {item.name.casefold(): item for item in inventory.sources}
        errors: list[str] = []
        prepared: list[dict[str, object]] = []
        seen: set[tuple[int, date]] = set()
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            if not any(value not in (None, "") for value in row):
                continue
            if template_version == "current":
                source_name, period_value, raw_value, unit, origin, estimated_text, uncertainty_raw, uncertainty_basis, notes = row[:9]
            else:
                source_name, period_value, raw_value, unit, origin, estimated_text, notes = row[:7]
                uncertainty_raw, uncertainty_basis = 0, ""
            source = source_map.get(str(source_name or "").strip().casefold())
            if not source:
                errors.append(f"Fila {row_number}: fuente no reconocida ({source_name}).")
                continue
            try:
                start, end = _parse_excel_period(period_value, inventory)
                numeric_value = float(raw_value)
                if numeric_value < 0:
                    raise ValueError("valor negativo")
            except (TypeError, ValueError) as exc:
                errors.append(f"Fila {row_number}: {exc}.")
                continue
            unit = str(unit or "").strip()
            origin = str(origin or "").strip()
            if unit not in ALLOWED_UNITS:
                errors.append(f"Fila {row_number}: unidad no autorizada ({unit}).")
                continue
            if origin not in DATA_ORIGINS:
                errors.append(f"Fila {row_number}: origen no autorizado ({origin}).")
                continue
            key = (source.id, start)
            existing = session.scalar(select(ActivityData.id).where(ActivityData.source_id == source.id, ActivityData.period_start == start))
            if key in seen or existing:
                errors.append(f"Fila {row_number}: ya existe un dato para {source.name} en {start:%Y-%m}.")
                continue
            seen.add(key)
            estimated = str(estimated_text or "").strip().casefold() in {"sí", "si", "s", "yes", "true", "1"}
            try:
                uncertainty = max(0.0, float(uncertainty_raw or 0))
            except (TypeError, ValueError):
                errors.append(f"Fila {row_number}: incertidumbre inválida ({uncertainty_raw}).")
                continue
            prepared.append({"source": source, "start": start, "end": end, "value": numeric_value, "unit": unit, "origin": origin, "estimated": estimated, "uncertainty": uncertainty, "uncertainty_basis": str(uncertainty_basis or "").strip(), "notes": str(notes or "").strip()})
        if errors:
            return templates.TemplateResponse(request=request, name="import_data.html", context=common_context(request, session, user, "information", inventory=inventory, sources=inventory.sources, errors=errors, preview=prepared[:10]), status_code=400)
        if not prepared:
            return templates.TemplateResponse(request=request, name="import_data.html", context=common_context(request, session, user, "information", inventory=inventory, sources=inventory.sources, errors=["El archivo no contiene filas para importar."], preview=[]), status_code=400)
        for item in prepared:
            source = item["source"]
            session.add(ActivityData(
                source_id=source.id,
                period_start=item["start"],
                period_end=item["end"],
                value=item["value"],
                unit=item["unit"],
                data_origin=item["origin"],
                quality_level=quality_from(item["origin"], item["estimated"], False),
                is_estimated=item["estimated"],
                uncertainty_percentage=item["uncertainty"],
                uncertainty_basis=item["uncertainty_basis"],
                notes=item["notes"],
                status="Provisional" if item["estimated"] else "Cargado",
                created_by=str(user["email"]),
            ))
        session.flush()
        refresh_progress(session, inventory)
        calculation_result = recalculate_inventory(session, inventory)
        add_audit(session, int(user["organization_id"]), str(user["email"]), "IMPORTAR", "Datos de actividad", file.filename or "Excel", f"{len(prepared)} registros validados · {calculation_result['calculations']} cálculos")
        session.commit()
        set_flash(request, f"Se importaron {len(prepared)} registros sin errores.")
        return RedirectResponse("/informacion", status_code=303)
