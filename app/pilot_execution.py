from __future__ import annotations

from datetime import UTC, date, datetime, timedelta
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload

from .calculations import recalculate_source
from .database import (
    ActivityData,
    DataRequest,
    EmissionFactor,
    EmissionFactorVersion,
    EmissionSource,
    Facility,
    Inventory,
    InventoryFacility,
    PilotExecution,
    PilotExecutionSourceLink,
    PilotIssue,
    PilotSourceComparison,
    PilotProject,
    PilotSourceRequirement,
    SourceFactorAssignment,
    add_audit,
    refresh_progress,
)

PILOT_ENGINE_VERSION = "1.0.0"

SITE_NAMES = {
    "Yarumal": ("Planta Yarumal", "Planta de aprovechamiento", "Yarumal"),
    "Támesis": ("Planta Támesis", "Planta de aprovechamiento", "Támesis"),
    "Corporativo": ("Sede corporativa Medellín", "Oficina", "Medellín"),
    "Cadena de valor": ("Sede corporativa Medellín", "Oficina", "Medellín"),
}

FACTOR_ACTIVITY_MAP = {
    "Electricidad": "Electricidad adquirida",
    "Compostaje": "Compostaje de residuos orgánicos",
    "Digestión anaerobia": "Digestión anaerobia de residuos orgánicos",
    "Refrigerantes": "Emisiones fugitivas de refrigerante",
}


def _pilot(session: Session, organization_id: int) -> PilotProject | None:
    return session.scalar(
        select(PilotProject)
        .where(PilotProject.organization_id == organization_id, PilotProject.code == "GREENATICS-2026")
        .options(selectinload(PilotProject.requirements))
    )


def get_pilot_execution(session: Session, organization_id: int) -> PilotExecution | None:
    return session.scalar(
        select(PilotExecution)
        .join(PilotProject)
        .where(PilotProject.organization_id == organization_id, PilotProject.code == "GREENATICS-2026")
        .options(
            selectinload(PilotExecution.inventory)
            .selectinload(Inventory.sources)
            .selectinload(EmissionSource.activity_records),
            selectinload(PilotExecution.source_links).selectinload(PilotExecutionSourceLink.requirement),
            selectinload(PilotExecution.source_links)
            .selectinload(PilotExecutionSourceLink.source)
            .selectinload(EmissionSource.factor_assignments),
            selectinload(PilotExecution.source_links)
            .selectinload(PilotExecutionSourceLink.source)
            .selectinload(EmissionSource.evidence_documents),
            selectinload(PilotExecution.source_links).selectinload(PilotExecutionSourceLink.request),
            selectinload(PilotExecution.issues),
            selectinload(PilotExecution.comparisons).selectinload(PilotSourceComparison.requirement),
        )
        .execution_options(populate_existing=True)
    )


def _facility_for_site(session: Session, organization_id: int, site: str) -> Facility:
    name, facility_type, city = SITE_NAMES.get(site, (site, "Operación", site))
    facility = session.scalar(
        select(Facility).where(Facility.organization_id == organization_id, Facility.name == name)
    )
    if not facility:
        facility = Facility(
            organization_id=organization_id,
            name=name,
            facility_type=facility_type,
            city=city,
            address="Por documentar",
            employees=0,
            operational_control=True,
            financial_control=True,
            ownership_percentage=100,
            active=True,
        )
        session.add(facility)
        session.flush()
    return facility


def _factor_for_requirement(session: Session, requirement: PilotSourceRequirement) -> EmissionFactorVersion | None:
    activity_type = None
    for key, mapped in FACTOR_ACTIVITY_MAP.items():
        if key.lower() in requirement.category.lower() or key.lower() in requirement.source_name.lower():
            activity_type = mapped
            break
    if not activity_type:
        return None
    versions = list(
        session.scalars(
            select(EmissionFactorVersion)
            .join(EmissionFactor)
            .where(
                EmissionFactor.activity_type == activity_type,
                EmissionFactor.is_demo.is_(False),
                EmissionFactorVersion.status == "Aprobado",
            )
            .order_by(EmissionFactorVersion.publication_year.desc(), EmissionFactorVersion.id.desc())
        )
    )
    return next((item for item in versions if item.input_unit == requirement.activity_unit), versions[0] if versions else None)


def start_pilot_execution(
    session: Session,
    organization_id: int,
    user_email: str,
    user_name: str,
    target_date: date | None = None,
) -> PilotExecution:
    existing = get_pilot_execution(session, organization_id)
    if existing:
        return existing
    pilot = _pilot(session, organization_id)
    if not pilot:
        raise ValueError("No existe la matriz del piloto Greenatics para esta organización.")

    inventory = Inventory(
        organization_id=organization_id,
        name="Inventario corporativo Greenatics 2026 · piloto controlado",
        start_date=date(pilot.reporting_year, 1, 1),
        end_date=date(pilot.reporting_year, 12, 31),
        objective="Validar en operación real el motor, la biblioteca metodológica y los recorridos de usuario.",
        base_year=pilot.reporting_year,
        methodology="GHG Protocol + ISO 14064-1",
        methodology_version="Piloto operativo Greenatics V0.45",
        gwp_version="IPCC AR6 · 100 años",
        consolidation_approach=pilot.consolidation_approach,
        materiality_threshold=5.0,
        status="Borrador",
        progress=0,
        current_stage="Preparación del piloto",
        notes="Inventario creado desde la matriz Greenatics. No constituye declaración externa hasta cerrar contraste independiente.",
        version="0.45",
        locked=False,
    )
    session.add(inventory)
    session.flush()

    facilities: dict[str, Facility] = {}
    for site in {item.site for item in pilot.requirements}:
        facility = _facility_for_site(session, organization_id, site)
        facilities[site] = facility
        if not session.scalar(
            select(InventoryFacility).where(
                InventoryFacility.inventory_id == inventory.id,
                InventoryFacility.facility_id == facility.id,
            )
        ):
            session.add(
                InventoryFacility(
                    inventory_id=inventory.id,
                    facility_id=facility.id,
                    included=True,
                    inclusion_percentage=100,
                    exclusion_reason="",
                )
            )

    execution = PilotExecution(
        pilot_id=pilot.id,
        inventory_id=inventory.id,
        status="En ejecución",
        started_by=user_email,
        started_at=datetime.now(UTC),
        target_date=target_date or (date.today() + timedelta(days=60)),
        comparison_status="Pendiente",
    )
    session.add(execution)
    session.flush()

    due = date.today() + timedelta(days=30)
    created_links = 0
    for requirement in sorted(pilot.requirements, key=lambda item: (item.site, item.code)):
        if requirement.scope not in {1, 2, 3}:
            continue
        facility = facilities.get(requirement.site)
        source = EmissionSource(
            inventory_id=inventory.id,
            facility_id=facility.id if facility else None,
            name=f"{requirement.site} · {requirement.source_name}"[:100],
            scope=requirement.scope,
            category=requirement.category,
            responsible=requirement.data_owner,
            materiality=requirement.materiality,
            data_frequency=requirement.frequency,
            preferred_unit=requirement.activity_unit,
            included=True,
            progress=0,
            status="Pendiente",
            emissions=0,
            unit="tCO₂e",
            icon="activity",
        )
        session.add(source)
        session.flush()

        request = DataRequest(
            inventory_id=inventory.id,
            source_id=source.id,
            title=f"{requirement.code} · {requirement.source_name}",
            source_name=requirement.source_name,
            requested_to=requirement.data_owner,
            due_date=due,
            status="Pendiente",
            instructions=f"Cargar dato {requirement.frequency.lower()} en {requirement.activity_unit}. Evidencia esperada: {requirement.evidence_expected}",
        )
        session.add(request)
        session.flush()

        factor_version = _factor_for_requirement(session, requirement)
        if factor_version:
            session.add(
                SourceFactorAssignment(
                    source_id=source.id,
                    factor_version_id=factor_version.id,
                    active=True,
                    assigned_by=user_email,
                    notes=f"Asignación automática controlada desde {requirement.factor_reference}.",
                )
            )

        session.add(
            PilotExecutionSourceLink(
                execution_id=execution.id,
                requirement_id=requirement.id,
                source_id=source.id,
                request_id=request.id,
            )
        )
        created_links += 1
        requirement.status = "Solicitado"
        requirement.updated_by = user_email

    default_issues = [
        (
            "PIL-025-001",
            "Metodología",
            "Cerrar factores todavía pendientes",
            "Los combustibles, transporte, aguas residuales y fertilizantes requieren parametrización y aprobación antes del cierre.",
            "Alta",
        ),
        (
            "PIL-025-002",
            "Cálculo",
            "Contrastar con memoria independiente",
            "Repetir el cálculo en una memoria externa y documentar la variación frente a la plataforma.",
            "Alta",
        ),
        (
            "PIL-025-003",
            "Límites",
            "Revisar doble conteo de biogás y tratamiento",
            "Confirmar que recuperación, uso, quema, venteo y fugas de metano no se contabilicen dos veces.",
            "Crítica",
        ),
    ]
    for code, category, title, description, severity in default_issues:
        session.add(
            PilotIssue(
                execution_id=execution.id,
                code=code,
                category=category,
                title=title,
                description=description,
                severity=severity,
                status="Abierto",
                owner="Equipo piloto Greenatics",
                due_date=execution.target_date,
                created_by=user_email,
            )
        )

    pilot.status = "En ejecución"
    refresh_progress(session, inventory)
    # La preparación termina al crear las fuentes y solicitudes controladas.
    # Desde aquí la autoridad del lifecycle del piloto es la recolección de datos.
    inventory.current_stage = "Recolección"
    add_audit(
        session,
        organization_id,
        user_email,
        "INICIAR",
        "Piloto Greenatics",
        pilot.code,
        detail=f"Inventario #{inventory.id} creado con {created_links} fuentes y solicitudes controladas.",
        new_value=f"En ejecución · responsable {user_name}",
    )
    session.flush()
    return execution


MONTH_LABELS = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

GREENATICS_OPERATIONAL_CONTEXT = [
    {"site": "Yarumal", "indicator": "Capacidad de fertilizante sólido", "value": 14.4, "unit": "t/mes", "status": "Referencia operativa", "use": "No alimenta emisiones automáticamente"},
    {"site": "Yarumal", "indicator": "Capacidad de fertilizante líquido", "value": 4000, "unit": "L/mes", "status": "Referencia operativa", "use": "Indicador de intensidad"},
    {"site": "Yarumal", "indicator": "Residuos recibidos enero–13 junio 2026", "value": 124.64, "unit": "t", "status": "Por conciliar con pesajes", "use": "No se aplica hasta cargar soporte mensual"},
    {"site": "Yarumal", "indicator": "Material aprovechable enero–13 junio 2026", "value": 119.56, "unit": "t", "status": "Por conciliar con bitácora", "use": "No se aplica hasta cargar soporte mensual"},
    {"site": "Támesis", "indicator": "Capacidad de fertilizante sólido", "value": 8.0, "unit": "t/mes", "status": "Referencia operativa", "use": "No alimenta emisiones automáticamente"},
    {"site": "Támesis", "indicator": "Capacidad de fertilizante líquido", "value": 8000, "unit": "L/mes", "status": "Referencia operativa", "use": "Indicador de intensidad"},
]


def _expected_records(frequency: str) -> int:
    normalized = (frequency or "Mensual").strip().lower()
    if "anual" in normalized:
        return 1
    if "semes" in normalized:
        return 2
    if "trimes" in normalized:
        return 4
    if "bimes" in normalized:
        return 6
    return 12


def _ensure_source_comparisons(
    session: Session,
    execution: PilotExecution,
    links: list[PilotExecutionSourceLink],
) -> dict[int, PilotSourceComparison]:
    existing = {item.requirement_id: item for item in execution.comparisons}
    for link in links:
        if not link.source or link.requirement_id in existing:
            continue
        comparison = PilotSourceComparison(
            execution_id=execution.id,
            requirement_id=link.requirement_id,
            source_id=link.source_id,
            platform_tco2e=round(link.source.emissions or 0, 6),
            status="Pendiente",
        )
        session.add(comparison)
        session.flush()
        existing[link.requirement_id] = comparison
    return existing


def _refresh_source_comparison(comparison: PilotSourceComparison, platform_tco2e: float) -> None:
    comparison.platform_tco2e = round(platform_tco2e or 0, 6)
    if comparison.independent_tco2e is None:
        comparison.absolute_difference_tco2e = None
        comparison.variance_percent = None
        comparison.status = "Pendiente"
        return
    difference = abs(comparison.platform_tco2e - comparison.independent_tco2e)
    denominator = max(abs(comparison.independent_tco2e), 0.000001)
    comparison.absolute_difference_tco2e = round(difference, 6)
    comparison.variance_percent = round(difference / denominator * 100, 2)
    comparison.status = "Conforme" if comparison.variance_percent <= 2 else "Revisar"


def _source_control_row(link: PilotExecutionSourceLink, comparison: PilotSourceComparison | None, year: int) -> dict[str, Any]:
    source = link.source
    records = [item for item in (source.activity_records if source else []) if item.period_start.year == year]
    expected = _expected_records(link.requirement.frequency)
    distinct_periods = {(item.period_start.year, item.period_start.month) for item in records}
    evidence_ids = {item.evidence_id for item in records if item.evidence_id}
    direct_evidence = len(source.evidence_documents) if source else 0
    evidence_count = max(len(evidence_ids), direct_evidence)
    coverage = min(100, round(100 * len(distinct_periods) / max(expected, 1)))
    return {
        "link": link,
        "comparison": comparison,
        "expected_records": expected,
        "actual_records": len(distinct_periods),
        "coverage": coverage,
        "months": {month for _, month in distinct_periods},
        "evidence_count": evidence_count,
        "estimated_records": sum(1 for item in records if item.is_estimated),
        "quality_levels": sorted({item.quality_level for item in records}),
    }


def pilot_execution_summary(session: Session, organization_id: int) -> dict[str, Any]:
    session.flush()
    execution = get_pilot_execution(session, organization_id)
    if not execution:
        return {
            "execution": None,
            "inventory": None,
            "links": [],
            "source_controls": [],
            "issues": [],
            "metrics": {
                "readiness": 0, "source_count": 0, "data_complete": 0, "factor_covered": 0,
                "evidence_covered": 0, "open_issues": 0, "critical_issues": 0,
                "platform_total": 0, "variance": None, "comparison_coverage": 0,
            },
            "blockers": ["El piloto todavía no ha sido iniciado."],
            "next_actions": [{"title": "Iniciar piloto controlado", "detail": "Crear inventario, fuentes y solicitudes a partir de la matriz validada.", "href": "/piloto-greenatics"}],
            "by_site": {}, "site_metrics": {}, "monthly_coverage": [],
            "operational_context": GREENATICS_OPERATIONAL_CONTEXT,
        }

    inventory = execution.inventory
    links = sorted(execution.source_links, key=lambda item: (item.requirement.site, item.requirement.code))
    source_links = [item for item in links if item.source]
    comparisons = _ensure_source_comparisons(session, execution, source_links)
    source_controls: list[dict[str, Any]] = []
    for link in source_links:
        comparison = comparisons.get(link.requirement_id)
        if comparison:
            _refresh_source_comparison(comparison, link.source.emissions)
        source_controls.append(_source_control_row(link, comparison, execution.pilot.reporting_year))

    completed_controls = [item for item in source_controls if item["coverage"] >= 100]
    factor_covered = [item for item in source_links if any(assignment.active and assignment.factor_version_id for assignment in item.source.factor_assignments)]
    evidence_covered = [item for item in source_controls if item["evidence_count"] > 0]
    open_issues = [item for item in execution.issues if item.status not in {"Cerrado", "Resuelto"}]
    critical_issues = [item for item in open_issues if item.severity in {"Crítica", "Alta"}]
    platform_total = round(sum(item.source.emissions for item in source_links), 6)
    execution.platform_total_tco2e = platform_total

    legacy_independent_total = execution.independent_total_tco2e
    compared = [item for item in comparisons.values() if item.independent_tco2e is not None]
    conforming = [item for item in compared if item.status == "Conforme"]
    comparison_coverage = round(100 * len(compared) / max(len(source_links), 1))
    compared_platform_total = round(sum(item.platform_tco2e for item in compared), 6)
    independent_total = round(sum(item.independent_tco2e or 0 for item in compared), 6) if compared else None
    execution.independent_total_tco2e = independent_total
    if compared:
        denominator = max(abs(independent_total or 0), 0.000001)
        execution.variance_percent = round(abs(compared_platform_total - (independent_total or 0)) / denominator * 100, 2)
        if len(compared) < len(source_links):
            execution.comparison_status = "Parcial"
        elif len(conforming) == len(source_links):
            execution.comparison_status = "Conforme"
        else:
            execution.comparison_status = "Revisar"
    else:
        legacy_total = legacy_independent_total
        if legacy_total is not None:
            denominator = max(abs(legacy_total), 0.000001)
            execution.independent_total_tco2e = legacy_total
            execution.variance_percent = round(abs(platform_total - legacy_total) / denominator * 100, 2)
            execution.comparison_status = "Conforme" if execution.variance_percent <= 2 else "Revisar"
            independent_total = legacy_total
            compared_platform_total = platform_total
        else:
            execution.variance_percent = None
            execution.comparison_status = "Pendiente"

    data_score = round(sum(item["coverage"] for item in source_controls) / max(len(source_controls), 1))
    factor_score = round(100 * len(factor_covered) / max(len(source_links), 1))
    evidence_score = round(100 * len(evidence_covered) / max(len(source_links), 1))
    issue_score = 100 if not open_issues else max(0, 100 - len(open_issues) * 12 - len(critical_issues) * 10)
    comparison_score = round(0.65 * comparison_coverage + 0.35 * (100 * len(conforming) / max(len(source_links), 1)))
    readiness = round(data_score * 0.35 + factor_score * 0.2 + evidence_score * 0.15 + issue_score * 0.15 + comparison_score * 0.15)

    blockers: list[str] = []
    missing_data = len(source_controls) - len(completed_controls)
    missing_factors = len(source_links) - len(factor_covered)
    missing_evidence = len(source_links) - len(evidence_covered)
    if missing_data:
        blockers.append(f"{missing_data} fuente(s) no tienen la frecuencia anual esperada completa.")
    if missing_factors:
        blockers.append(f"{missing_factors} fuente(s) no tienen factor aprobado asignado.")
    if missing_evidence:
        blockers.append(f"{missing_evidence} fuente(s) no tienen evidencia vinculada.")
    if critical_issues:
        blockers.append(f"{len(critical_issues)} incidencia(s) alta(s) o crítica(s) siguen abiertas.")
    if comparison_coverage < 100:
        blockers.append(f"El contraste independiente por fuente cubre {comparison_coverage}% del piloto.")
    elif execution.comparison_status != "Conforme":
        blockers.append(f"El contraste por fuente no es conforme; variación agregada {execution.variance_percent or 0:.2f}%.")

    by_site: dict[str, list[PilotExecutionSourceLink]] = {}
    site_metrics: dict[str, dict[str, Any]] = {}
    for link in links:
        by_site.setdefault(link.requirement.site, []).append(link)
    for site, site_links in by_site.items():
        controls = [item for item in source_controls if item["link"].requirement.site == site]
        site_comparisons = [item["comparison"] for item in controls if item["comparison"] and item["comparison"].independent_tco2e is not None]
        site_metrics[site] = {
            "sources": len(controls),
            "data_coverage": round(sum(item["coverage"] for item in controls) / max(len(controls), 1)),
            "evidence_coverage": round(100 * sum(1 for item in controls if item["evidence_count"] > 0) / max(len(controls), 1)),
            "factor_coverage": round(100 * sum(1 for item in site_links if item.source and any(a.active and a.factor_version_id for a in item.source.factor_assignments)) / max(len(controls), 1)),
            "comparison_coverage": round(100 * len(site_comparisons) / max(len(controls), 1)),
            "emissions": round(sum(item.source.emissions for item in site_links if item.source), 6),
        }

    monthly_coverage: list[dict[str, Any]] = []
    for month, label in enumerate(MONTH_LABELS, start=1):
        row = {"month": month, "label": label, "sites": {}}
        for site in by_site:
            monthly_controls = [item for item in source_controls if item["link"].requirement.site == site and _expected_records(item["link"].requirement.frequency) == 12]
            present = sum(1 for item in monthly_controls if month in item["months"])
            expected = len(monthly_controls)
            row["sites"][site] = {"present": present, "expected": expected, "coverage": round(100 * present / max(expected, 1))}
        monthly_coverage.append(row)

    next_actions: list[dict[str, str]] = []
    if missing_data:
        next_actions.append({"title": "Completar cobertura mensual", "detail": "Carga los periodos pendientes por sede y fuente.", "href": "/piloto-greenatics/ejecucion#cobertura"})
    if missing_evidence:
        next_actions.append({"title": "Vincular evidencias", "detail": "Adjunta facturas, pesajes, bitácoras y soportes a cada fuente.", "href": f"/inventarios/{inventory.id}/documentos" if inventory else "/documentos"})
    if missing_factors:
        next_actions.append({"title": "Cerrar cobertura de factores", "detail": "Asigna únicamente versiones aprobadas y aplicables al periodo.", "href": f"/inventarios/{inventory.id}/fuentes" if inventory else "/metodologia/nucleo"})
    if comparison_coverage < 100 or execution.comparison_status != "Conforme":
        next_actions.append({"title": "Completar contraste por fuente", "detail": "Diligencia la memoria independiente y explica cada diferencia.", "href": "/piloto-greenatics/ejecucion#contraste"})
    if open_issues:
        next_actions.append({"title": "Resolver incidencias", "detail": "Cierra hallazgos de datos, límites y metodología.", "href": "/piloto-greenatics/ejecucion#incidencias"})
    if not next_actions:
        next_actions.append({"title": "Solicitar aprobación del piloto", "detail": "Todas las puertas de salida están completas.", "href": "/piloto-greenatics/ejecucion#cierre"})

    session.flush()
    return {
        "execution": execution,
        "inventory": inventory,
        "links": links,
        "source_controls": source_controls,
        "comparisons": sorted(comparisons.values(), key=lambda item: (item.requirement.site, item.requirement.code)),
        "issues": sorted(execution.issues, key=lambda item: (item.status in {"Cerrado", "Resuelto"}, item.severity, item.code)),
        "metrics": {
            "readiness": readiness,
            "source_count": len(source_links),
            "data_complete": len(completed_controls),
            "factor_covered": len(factor_covered),
            "evidence_covered": len(evidence_covered),
            "open_issues": len(open_issues),
            "critical_issues": len(critical_issues),
            "platform_total": platform_total,
            "compared_platform_total": compared_platform_total,
            "independent_total": independent_total,
            "variance": execution.variance_percent,
            "comparison_coverage": comparison_coverage,
            "comparison_conforming": len(conforming),
            "data_score": data_score,
            "factor_score": factor_score,
            "evidence_score": evidence_score,
        },
        "blockers": blockers,
        "next_actions": next_actions,
        "by_site": by_site,
        "site_metrics": site_metrics,
        "monthly_coverage": monthly_coverage,
        "operational_context": GREENATICS_OPERATIONAL_CONTEXT,
    }

def build_pilot_execution_workbook(summary: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Control de ejecución"
    execution = summary["execution"]
    inventory = summary["inventory"]
    rows = [
        ("Estado", execution.status if execution else "No iniciado"),
        ("Inventario", inventory.name if inventory else ""),
        ("Fecha objetivo", execution.target_date if execution else ""),
        ("Preparación", summary["metrics"]["readiness"]),
        ("Total plataforma tCO2e", summary["metrics"]["platform_total"]),
        ("Total memoria independiente tCO2e", execution.independent_total_tco2e if execution else ""),
        ("Variación %", summary["metrics"]["variance"] if summary["metrics"]["variance"] is not None else ""),
        ("Estado contraste", execution.comparison_status if execution else ""),
    ]
    ws.append(["Campo", "Valor"])
    for row in rows:
        ws.append(list(row))

    ws = wb.create_sheet("Fuentes vinculadas")
    ws.append(["Código fuente", "Sede", "Alcance", "Categoría", "Fuente", "Unidad", "Responsable", "Progreso %", "Emisiones tCO2e", "Factor asignado", "Solicitud"])
    for link in summary["links"]:
        source = link.source
        ws.append([
            link.requirement.code,
            link.requirement.site,
            link.requirement.scope,
            link.requirement.category,
            link.requirement.source_name,
            link.requirement.activity_unit,
            link.requirement.data_owner,
            source.progress if source else 0,
            source.emissions if source else 0,
            "Sí" if source and source.factor_assignments else "No",
            link.request.status if link.request else "",
        ])

    ws = wb.create_sheet("Datos mensuales")
    ws.append(["Código fuente", "Periodo inicio", "Periodo fin", "Valor", "Unidad", "Documento soporte", "Responsable", "Calidad", "Observaciones"])
    for link in summary["links"]:
        requirement = link.requirement
        if requirement.scope not in {1, 2, 3}:
            continue
        for month in range(1, 13):
            start = date(2026, month, 1)
            end = date(2026, 12, 31) if month == 12 else date(2026, month + 1, 1) - timedelta(days=1)
            ws.append([requirement.code, start, end, "", requirement.activity_unit, "", requirement.data_owner, "B", ""])

    ws = wb.create_sheet("Contraste por fuente")
    ws.append(["Código fuente", "Sede", "Fuente", "Plataforma tCO2e", "Memoria independiente tCO2e", "Diferencia absoluta tCO2e", "Variación %", "Estado", "Explicación"])
    for item in summary.get("source_controls", []):
        link = item["link"]
        comparison = item.get("comparison")
        ws.append([
            link.requirement.code, link.requirement.site, link.requirement.source_name,
            comparison.platform_tco2e if comparison else (link.source.emissions if link.source else 0),
            comparison.independent_tco2e if comparison and comparison.independent_tco2e is not None else "",
            comparison.absolute_difference_tco2e if comparison and comparison.absolute_difference_tco2e is not None else "",
            comparison.variance_percent if comparison and comparison.variance_percent is not None else "",
            comparison.status if comparison else "Pendiente", comparison.notes if comparison else "",
        ])

    ws = wb.create_sheet("Cobertura mensual")
    sites = list(summary.get("by_site", {}).keys())
    ws.append(["Mes"] + [f"{site} registros" for site in sites] + [f"{site} cobertura %" for site in sites])
    for row in summary.get("monthly_coverage", []):
        ws.append([row["label"]] + [row["sites"][site]["present"] for site in sites] + [row["sites"][site]["coverage"] for site in sites])

    ws = wb.create_sheet("Contexto operativo")
    ws.append(["Sede", "Indicador", "Valor", "Unidad", "Estado", "Uso en el piloto"])
    for item in summary.get("operational_context", []):
        ws.append([item["site"], item["indicator"], item["value"], item["unit"], item["status"], item["use"]])

    ws = wb.create_sheet("Incidencias")
    ws.append(["Código", "Categoría", "Título", "Severidad", "Estado", "Responsable", "Vencimiento", "Resolución"])
    for issue in summary["issues"]:
        ws.append([issue.code, issue.category, issue.title, issue.severity, issue.status, issue.owner, issue.due_date, issue.resolution])

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 52)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _as_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return datetime.strptime(value.strip(), "%Y-%m-%d").date()
    raise ValueError("Fecha vacía o inválida")


def import_pilot_workbook(session: Session, organization_id: int, content: bytes, user_email: str) -> dict[str, Any]:
    execution = get_pilot_execution(session, organization_id)
    if not execution or not execution.inventory:
        raise ValueError("Primero debes iniciar la ejecución del piloto.")
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet_name = "Datos mensuales" if "Datos mensuales" in workbook.sheetnames else "Plantilla de datos"
    if sheet_name not in workbook.sheetnames:
        raise ValueError("El archivo no contiene la hoja 'Datos mensuales' o 'Plantilla de datos'.")
    sheet = workbook[sheet_name]
    headers = {str(cell.value or "").strip(): index for index, cell in enumerate(sheet[1], start=1)}
    required = {"Código fuente", "Periodo inicio", "Periodo fin", "Valor", "Unidad"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"Faltan columnas obligatorias: {', '.join(missing)}")

    links = {link.requirement.code: link for link in execution.source_links if link.source}
    imported = 0
    updated = 0
    errors: list[str] = []
    touched_sources: set[int] = set()

    for row_number in range(2, sheet.max_row + 1):
        code = str(sheet.cell(row_number, headers["Código fuente"]).value or "").strip()
        raw_value = sheet.cell(row_number, headers["Valor"]).value
        if raw_value in {None, ""}:
            continue
        if not code:
            errors.append(f"Fila {row_number}: el código de fuente está vacío.")
            continue
        link = links.get(code)
        if not link:
            errors.append(f"Fila {row_number}: código {code or '(vacío)'} no pertenece a esta ejecución.")
            continue
        try:
            start = _as_date(sheet.cell(row_number, headers["Periodo inicio"]).value)
            end = _as_date(sheet.cell(row_number, headers["Periodo fin"]).value)
            value = float(raw_value)
            unit = str(sheet.cell(row_number, headers["Unidad"]).value or "").strip()
            if start > end:
                raise ValueError("la fecha inicial es posterior a la final")
            if start.year != execution.inventory.base_year or end.year != execution.inventory.base_year:
                raise ValueError(f"el periodo debe corresponder a {execution.inventory.base_year}")
            if not unit:
                raise ValueError("la unidad está vacía")
            if value < 0:
                raise ValueError("el valor no puede ser negativo")
        except (TypeError, ValueError) as exc:
            errors.append(f"Fila {row_number}: {exc}.")
            continue

        existing = session.scalar(
            select(ActivityData).where(
                ActivityData.source_id == link.source_id,
                ActivityData.period_start == start,
                ActivityData.period_end == end,
            )
        )
        quality = str(sheet.cell(row_number, headers.get("Calidad", 0)).value or "B").strip().upper() if headers.get("Calidad") else "B"
        quality = quality if quality in {"A", "B", "C", "D"} else "B"
        notes = str(sheet.cell(row_number, headers.get("Observaciones", 0)).value or "").strip() if headers.get("Observaciones") else ""
        if existing:
            existing.value = value
            existing.unit = unit
            existing.quality_level = quality
            existing.notes = notes
            existing.updated_at = datetime.now(UTC)
            updated += 1
        else:
            session.add(
                ActivityData(
                    source_id=link.source_id,
                    period_start=start,
                    period_end=end,
                    value=value,
                    unit=unit,
                    data_origin="Registro operativo",
                    quality_level=quality,
                    is_estimated=False,
                    notes=notes,
                    status="Cargado",
                    created_by=user_email,
                )
            )
            imported += 1
        touched_sources.add(link.source_id)
        link.requirement.status = "Disponible"
        link.requirement.updated_by = user_email
        if link.request:
            link.request.status = "En revisión"

    session.flush()
    for source_id in touched_sources:
        source = session.scalar(
            select(EmissionSource)
            .where(EmissionSource.id == source_id)
            .options(
                selectinload(EmissionSource.activity_records),
                selectinload(EmissionSource.factor_assignments),
            )
            .execution_options(populate_existing=True)
        )
        if source:
            recalculate_source(session, source)
    # Las relaciones pudieron estar cargadas antes de insertar los registros. Se
    # expiran y recargan para calcular el avance contra el estado real de la BD.
    session.flush()
    session.expire_all()
    inventory = session.scalar(
        select(Inventory)
        .where(Inventory.id == execution.inventory_id)
        .options(
            selectinload(Inventory.sources).selectinload(EmissionSource.activity_records),
            selectinload(Inventory.sources).selectinload(EmissionSource.factor_assignments),
        )
        .execution_options(populate_existing=True)
    )
    if not inventory:
        raise ValueError("No fue posible recargar el inventario del piloto.")
    refresh_progress(session, inventory)
    session.flush()
    add_audit(
        session,
        organization_id,
        user_email,
        "IMPORTAR",
        "Datos piloto Greenatics",
        execution.pilot.code,
        detail=f"{imported} registros creados, {updated} actualizados y {len(errors)} errores.",
    )
    return {"imported": imported, "updated": updated, "errors": errors, "sheet": sheet_name}



def import_pilot_comparison_workbook(session: Session, organization_id: int, content: bytes, user_email: str) -> dict[str, Any]:
    execution = get_pilot_execution(session, organization_id)
    if not execution or not execution.inventory:
        raise ValueError("Primero debes iniciar la ejecución del piloto.")
    workbook = load_workbook(BytesIO(content), data_only=True)
    if "Contraste por fuente" not in workbook.sheetnames:
        raise ValueError("El archivo no contiene la hoja 'Contraste por fuente'.")
    sheet = workbook["Contraste por fuente"]
    headers = {str(cell.value or "").strip(): index for index, cell in enumerate(sheet[1], start=1)}
    required = {"Código fuente", "Memoria independiente tCO2e"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"Faltan columnas obligatorias: {', '.join(missing)}")

    links = {link.requirement.code: link for link in execution.source_links if link.source}
    comparisons = _ensure_source_comparisons(session, execution, list(links.values()))
    updated = 0
    errors: list[str] = []
    for row_number in range(2, sheet.max_row + 1):
        code = str(sheet.cell(row_number, headers["Código fuente"]).value or "").strip()
        raw_value = sheet.cell(row_number, headers["Memoria independiente tCO2e"]).value
        if raw_value in {None, ""}:
            continue
        link = links.get(code)
        if not link:
            errors.append(f"Fila {row_number}: código {code or '(vacío)'} no pertenece al piloto.")
            continue
        try:
            value = float(raw_value)
            if value < 0:
                raise ValueError("el valor no puede ser negativo")
        except (TypeError, ValueError) as exc:
            errors.append(f"Fila {row_number}: valor independiente inválido ({exc}).")
            continue
        comparison = comparisons[link.requirement_id]
        comparison.independent_tco2e = value
        comparison.notes = str(sheet.cell(row_number, headers.get("Explicación", 0)).value or "").strip() if headers.get("Explicación") else comparison.notes
        comparison.reviewed_by = user_email
        comparison.reviewed_at = datetime.now(UTC)
        _refresh_source_comparison(comparison, link.source.emissions)
        updated += 1

    summary = pilot_execution_summary(session, organization_id)
    add_audit(
        session, organization_id, user_email, "IMPORTAR", "Contraste piloto Greenatics", execution.pilot.code,
        detail=f"{updated} fuente(s) contrastadas; {len(errors)} fila(s) rechazadas; cobertura {summary['metrics']['comparison_coverage']}%.",
    )
    session.flush()
    return {"updated": updated, "errors": errors, "summary": summary}


def update_pilot_source_comparison(
    session: Session,
    organization_id: int,
    comparison_id: int,
    independent_tco2e: float,
    notes: str,
    user_email: str,
) -> PilotSourceComparison:
    execution = get_pilot_execution(session, organization_id)
    if not execution:
        raise ValueError("Ejecución del piloto no encontrada.")
    comparison = next((item for item in execution.comparisons if item.id == comparison_id), None)
    if not comparison:
        raise ValueError("Contraste por fuente no encontrado.")
    if independent_tco2e < 0:
        raise ValueError("El resultado independiente no puede ser negativo.")
    link = next((item for item in execution.source_links if item.requirement_id == comparison.requirement_id), None)
    comparison.independent_tco2e = independent_tco2e
    comparison.notes = notes.strip()
    comparison.reviewed_by = user_email
    comparison.reviewed_at = datetime.now(UTC)
    _refresh_source_comparison(comparison, link.source.emissions if link and link.source else comparison.platform_tco2e)
    add_audit(
        session, organization_id, user_email, "CONTRASTAR", "Fuente piloto Greenatics", comparison.requirement.code,
        new_value=f"{independent_tco2e} tCO2e · {comparison.status}", detail=comparison.notes,
    )
    session.flush()
    return comparison

def guided_workspace(session: Session, user: dict[str, Any], inventory: Inventory) -> dict[str, Any]:
    sources = [source for source in inventory.sources if source.included]
    source_count = len(sources)
    complete_sources = sum(1 for source in sources if source.progress >= 100)
    factor_sources = sum(1 for source in sources if any(item.active for item in source.factor_assignments))
    pending_requests = session.scalar(
        select(func.count(DataRequest.id)).where(
            DataRequest.inventory_id == inventory.id,
            DataRequest.status.notin_(["Completada", "Cerrada"]),
        )
    ) or 0
    open_observations = sum(1 for item in inventory.observations if item.status != "Cerrada")
    reports = len(inventory.reports)
    reduction_actions = list(inventory.reduction_actions)
    reduction_expected = sum(float(item.expected_reduction or 0) for item in reduction_actions)
    configured = bool(inventory.methodology and inventory.gwp_version and inventory.facility_links)
    milestones = [
        {"name": "Configurar", "done": configured, "detail": "Metodología, periodo y sedes", "href": f"/inventarios/{inventory.id}"},
        {"name": "Recolectar", "done": source_count > 0 and complete_sources == source_count, "detail": f"{complete_sources}/{source_count} fuentes completas", "href": "/captura-guiada"},
        {"name": "Calcular", "done": source_count > 0 and factor_sources == source_count, "detail": f"{factor_sources}/{source_count} con factor", "href": "/calculos"},
        {"name": "Revisar", "done": open_observations == 0 and inventory.status in {"En revisión", "Aprobado", "Cerrado"}, "detail": f"{open_observations} observaciones abiertas", "href": "/control"},
        {"name": "Reducir", "done": bool(reduction_actions), "detail": f"{len(reduction_actions)} acción(es) · {reduction_expected:.1f} tCO₂e/año", "href": "/reduccion"},
        {"name": "Reportar", "done": reports > 0, "detail": f"{reports} documento(s) generado(s)", "href": "/entrega-profesional"},
    ]
    completed = sum(1 for item in milestones if item["done"])
    score = round(completed / len(milestones) * 100)
    role = str(user.get("role", ""))
    actions: list[dict[str, str]] = []
    if role == "Cliente":
        if pending_requests:
            actions.append({"title": "Responder solicitudes pendientes", "detail": f"Hay {pending_requests} solicitud(es) activas.", "href": "/captura-guiada", "priority": "Alta"})
        if complete_sources < source_count:
            actions.append({"title": "Completar datos y evidencias", "detail": f"Faltan {source_count - complete_sources} fuente(s).", "href": "/captura-guiada", "priority": "Alta"})
    elif role == "Revisor":
        actions.append({"title": "Revisar puertas de aprobación", "detail": f"Existen {open_observations} observación(es) abiertas.", "href": "/control", "priority": "Alta" if open_observations else "Media"})
        actions.append({"title": "Comprobar trazabilidad metodológica", "detail": "Verifica factores, conversiones y evidencia.", "href": "/calculos", "priority": "Media"})
    else:
        for milestone in milestones:
            if not milestone["done"]:
                actions.append({"title": f"Completar etapa: {milestone['name']}", "detail": milestone["detail"], "href": milestone["href"], "priority": "Alta"})
                break
        pilot = pilot_execution_summary(session, int(user["organization_id"]))
        if pilot["execution"] and pilot["metrics"]["readiness"] < 100:
            actions.append({"title": "Avanzar piloto Greenatics", "detail": f"Preparación actual: {pilot['metrics']['readiness']}%.", "href": "/piloto-greenatics/ejecucion", "priority": "Alta"})
        elif not pilot["execution"]:
            actions.append({"title": "Iniciar piloto Greenatics", "detail": "Crea el inventario controlado desde la matriz sectorial.", "href": "/piloto-greenatics", "priority": "Media"})
    return {
        "score": score,
        "completed": completed,
        "total": len(milestones),
        "milestones": milestones,
        "actions": actions[:3],
        "pending_requests": pending_requests,
        "open_observations": open_observations,
    }
