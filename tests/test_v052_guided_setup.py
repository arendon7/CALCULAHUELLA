from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook
from fastapi.testclient import TestClient
from sqlalchemy import func, select

from app.database import (
    Base,
    EmissionSource,
    ENGINE,
    Inventory,
    OrganizationCarbonProfile,
    PlatformSetting,
    SessionLocal,
    init_db,
)
from app.guided_onboarding import SETTING_KEY
from app.main import app

ROOT = Path(__file__).resolve().parents[1]


@pytest.fixture(autouse=True)
def fresh_database_v052():
    # La base semilla aislada se restaura desde tests/conftest.py.
    yield


def login(client: TestClient, email: str = "consultor@calculatuhuella.local") -> None:
    response = client.post("/login", data={"email": email, "password": "Demo2026!"}, follow_redirects=False)
    assert response.status_code == 303


def payload(**overrides):
    data = {
        "objective": "management",
        "reporting_driver": "Definir una línea base y un plan de reducción para decisiones de gerencia.",
        "success_definition": "Inventario trazable aprobado y portafolio inicial de medidas priorizadas.",
        "sector_family": "waste",
        "operating_description": "Recolección, transporte, pesaje, tratamiento biológico, valorización y disposición de rechazos en dos plantas.",
        "scope_ambition": "prioritized",
        "reporting_frequency": "Mensual",
        "assurance_ambition": "Revisión técnica dirigida",
        "data_readiness": "medium",
        "evidence_readiness": "medium",
        "data_systems": ["spreadsheets", "meters", "providers"],
        "inventory_owner": "Coordinación ambiental",
        "executive_sponsor": "Gerencia general",
        "period_start": "2026-01-01",
        "period_end": "2026-12-31",
        "notes": "Mantener emisiones evitadas separadas del inventario bruto.",
    }
    data.update(overrides)
    return data


def test_v052_guided_page_connects_decision_flow_and_demo_profile_is_ready():
    with TestClient(app) as client:
        login(client)
        page = client.get("/onboarding/guiado")
        assert page.status_code == 200
        assert "Define el inventario antes de pedir datos" in page.text
        assert "Propósito" in page.text
        assert "Perfil" in page.text
        assert "Factores" in page.text
        assert "Guardar no modifica el inventario" in page.text
        assert "100% completo" in page.text
        assert 'disabled aria-disabled="true"' not in page.text
        assert "Perfil sintético preparado para la demostración multiempresa" in page.text


def test_v052_save_persists_profile_and_updates_carbon_profile():
    with TestClient(app) as client:
        login(client)
        response = client.post("/onboarding/guiado/guardar", data=payload(), follow_redirects=False)
        assert response.status_code == 303
        assert response.headers["location"] == "/onboarding/guiado#recomendacion"
        api = client.get("/api/onboarding/guiado")
        assert api.status_code == 200
        body = api.json()
        assert body["plan"]["completion"] == 100
        assert body["plan"]["ready_to_apply"] is True
        assert body["plan"]["starter_pack"]["code"] == "waste"
        assert "Tratamiento" in " ".join(item["source"] for item in body["data_checklist"])

    with SessionLocal() as session:
        setting = session.scalar(select(PlatformSetting).where(PlatformSetting.key == SETTING_KEY))
        assert setting is not None
        carbon_profile = session.scalar(select(OrganizationCarbonProfile).where(OrganizationCarbonProfile.source == "Asistente inicial V0.52"))
        assert carbon_profile is not None
        assert carbon_profile.profile_completion == 100
        assert carbon_profile.inventory_owner == "Coordinación ambiental"


def test_v052_apply_updates_methodology_without_deleting_sources_and_is_idempotent():
    with SessionLocal() as session:
        inventory = session.scalar(select(Inventory).order_by(Inventory.id))
        assert inventory is not None
        inventory_id = inventory.id
        before_ids = set(session.scalars(select(EmissionSource.id).where(EmissionSource.inventory_id == inventory_id)))

    with TestClient(app) as client:
        login(client)
        assert client.post("/onboarding/guiado/guardar", data=payload(), follow_redirects=False).status_code == 303
        first = client.post(
            "/onboarding/guiado/aplicar",
            data={"inventory_id": str(inventory_id)},
            follow_redirects=False,
        )
        assert first.status_code == 303
        assert first.headers["location"] == f"/inventarios/{inventory_id}/fuentes"
        second = client.post(
            "/onboarding/guiado/aplicar",
            data={"inventory_id": str(inventory_id)},
            follow_redirects=False,
        )
        assert second.status_code == 303

    with SessionLocal() as session:
        inventory = session.get(Inventory, inventory_id)
        after_ids = set(session.scalars(select(EmissionSource.id).where(EmissionSource.inventory_id == inventory_id)))
        assert before_ids <= after_ids
        assert inventory.methodology == "GHG Protocol + ISO 14064-1"
        assert inventory.gwp_version == "IPCC AR6 · 100 años"
        assert "Recomendación V0.52 aplicada" in inventory.notes
        names = list(session.scalars(select(EmissionSource.name).where(EmissionSource.inventory_id == inventory_id)))
        assert len(names) == len(set(names))


def test_v052_checklist_is_a_valid_multisheet_workbook():
    with TestClient(app) as client:
        login(client)
        client.post("/onboarding/guiado/guardar", data=payload(), follow_redirects=False)
        response = client.get("/onboarding/guiado/checklist.xlsx")
        assert response.status_code == 200
        assert response.headers["content-type"].startswith("application/vnd.openxmlformats")
        assert response.content.startswith(b"PK")
        assert len(response.content) > 5000
        workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
        assert workbook.sheetnames == ["Ruta de trabajo", "Datos requeridos", "Decisiones iniciales"]
        assert workbook["Ruta de trabajo"]["A1"].value == "Etapa"
        assert workbook["Datos requeridos"].max_row >= 2
        assert workbook["Decisiones iniciales"]["A2"].value == "Propósito"


def test_v052_client_cannot_change_or_apply_guided_configuration():
    with TestClient(app) as client:
        login(client, "cliente@calculatuhuella.local")
        assert client.get("/onboarding/guiado").status_code == 200
        assert client.post("/onboarding/guiado/guardar", data=payload(), follow_redirects=False).status_code == 403
        assert client.post("/onboarding/guiado/aplicar", data={}, follow_redirects=False).status_code == 403


def test_v052_navigation_and_release_artifacts_are_present():
    onboarding = (ROOT / "app/templates/onboarding.html").read_text(encoding="utf-8")
    assert "/onboarding/guiado" in onboarding
    assert (ROOT / "app/templates/guided_onboarding.html").is_file()
    assert (ROOT / "app/guided_onboarding.py").is_file()
    assert (ROOT / "app/guided_onboarding_web.py").is_file()


def test_v052_agricultural_profile_has_a_specific_starter_pack():
    with TestClient(app) as client:
        login(client)
        data = payload(
            sector_family="agro",
            operating_description="Cultivos permanentes, fertilización, maquinaria, manejo de suelos y producción pecuaria.",
        )
        assert client.post("/onboarding/guiado/guardar", data=data, follow_redirects=False).status_code == 303
        body = client.get("/api/onboarding/guiado").json()
        assert body["plan"]["starter_pack"]["code"] == "agro"
        sources = {item["source"] for item in body["data_checklist"]}
        assert "Fertilización nitrogenada y enmiendas" in sources
        assert "Fermentación entérica y manejo de estiércol" in sources


def test_v052_rejects_unknown_decision_codes():
    with TestClient(app) as client:
        login(client)
        response = client.post(
            "/onboarding/guiado/guardar",
            data=payload(sector_family="inventado"),
            follow_redirects=False,
        )
        assert response.status_code == 400
        assert "perfil sectorial" in response.text
