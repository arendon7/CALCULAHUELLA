from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.config import settings
from app.database import Base, ENGINE, init_db
from app.main import app
from app.release_candidate import release_candidate_summary

ROOT = Path(__file__).resolve().parents[1]


@pytest.fixture(autouse=True)
def fresh_database_v1():
    # La base semilla aislada se restaura desde tests/conftest.py.
    yield


def login(client: TestClient, email: str = "consultor@calculatuhuella.local") -> None:
    response = client.post("/login", data={"email": email, "password": "Demo2026!"}, follow_redirects=False)
    assert response.status_code == 303


def _write_internal_acceptance_evidence(root: Path, test_count: int = 386) -> None:
    (root / "release").mkdir(parents=True, exist_ok=True)
    (root / "release" / "FINAL_TEST_EVIDENCE.json").write_text(
        json.dumps({"status": "passed", "test_count": test_count}), encoding="utf-8"
    )
    for name in (
        "ACTA_CIERRE_V1_0_0.md",
        "APROBACION_METODOLOGICA_V1_CARLOS_URIBE.md",
        "APROBACION_JURIDICA_V1_AGUSTIN_RENDON.md",
        "INFORME_PILOTO_INTERNO_GREENATICS_V1.md",
        "INFORME_PILOTO_INTERNO_SEGUNDO_SECTOR_V1.md",
        "REVISION_SEGURIDAD_INTERNA_OWASP_ASVS_V1.md",
    ):
        (root / name).write_text("evidencia interna", encoding="utf-8")


def test_v1_version_is_final_and_internal_acceptance_is_explicit():
    assert settings.version == "1.0.0"
    assert settings.final_methodology_internal_approved is True
    assert settings.final_legal_internal_approved is True
    assert settings.final_greenatics_internal_pilot_approved is True
    assert settings.final_second_sector_internal_pilot_approved is True
    assert settings.rc_windows_10_approved is False
    assert settings.rc_windows_11_approved is False
    assert settings.rc_security_review_approved is False


def test_v1_summary_authorizes_controlled_release_without_faking_public_production(tmp_path: Path):
    _write_internal_acceptance_evidence(tmp_path)
    summary = release_candidate_summary(
        tmp_path,
        critical_open=0,
        approved_gates=9,
        gate_count=9,
        validated_journeys=5,
        journey_count=5,
    )
    assert summary["package_ready"] is True
    assert summary["governance_ready"] is True
    assert summary["internal_ready"] is True
    assert summary["controlled_release_ready"] is True
    assert summary["external_ready"] is False
    assert summary["production_ready"] is False
    assert summary["status"] == "V1.0 final · despliegue controlado"


def test_v1_consolidation_page_exposes_controlled_and_public_gates():
    with TestClient(app) as client:
        login(client)
        response = client.get("/consolidacion")
        assert response.status_code == 200
        assert "V1.0 final" in response.text
        assert "despliegue controlado" in response.text.lower()
        assert "Producción pública" in response.text


def test_v1_consolidation_api_is_machine_readable_and_conservative():
    with TestClient(app) as client:
        login(client)
        response = client.get("/api/consolidacion/resumen")
        assert response.status_code == 200
        release = response.json()["release_candidate"]
        assert release["version"] == "1.0.0"
        assert release["controlled_release_ready"] is True
        assert release["production_ready"] is False
        assert any(item["code"] == "V1-WIN10" and item["ok"] is False for item in release["external_checks"])
        assert any(item["code"] == "V1-CARLOS" and item["ok"] is True for item in release["internal_checks"])


def test_v1_workbook_includes_release_matrix():
    with TestClient(app) as client:
        login(client)
        response = client.get("/consolidacion/exportar.xlsx")
        assert response.status_code == 200
        workbook = load_workbook(BytesIO(response.content))
        assert "Liberación V1" in workbook.sheetnames
        sheet = workbook["Liberación V1"]
        assert sheet["A1"].value == "Versión"
        assert sheet["B1"].value == "1.0.0"
        assert sheet["A6"].value == "Despliegue controlado"
        assert sheet["A7"].value == "Producción pública"


def test_v1_structural_validator_runs_successfully():
    # Se evalúa la función pura dentro de la sesión para no competir con la
    # base SQLite abierta por pytest. El ejecutable CLI se cubre en la
    # certificación de instalación de la Iteración 10.
    from scripts.validate_release_candidate import inspect_candidate

    payload = inspect_candidate()
    assert payload["status"] == "passed"
    assert all(item["ok"] for item in payload["checks"])


def test_v1_internal_approval_and_launch_documents_are_distributed_in_canonical_tree():
    required = (
        "docs/gobierno/ACTA_CIERRE_V1_0_0.md",
        "docs/gobierno/APROBACION_METODOLOGICA_V1_CARLOS_URIBE.md",
        "docs/gobierno/APROBACION_JURIDICA_V1_AGUSTIN_RENDON.md",
        "docs/guias/INFORME_PILOTO_INTERNO_GREENATICS_V1.md",
        "docs/guias/INFORME_PILOTO_INTERNO_SEGUNDO_SECTOR_V1.md",
        "docs/gobierno/REVISION_SEGURIDAD_INTERNA_OWASP_ASVS_V1.md",
        "docs/guias/GUIA_LANZAMIENTO_CONTROLADO_V1.md",
    )
    assert all((ROOT / name).is_file() for name in required)


@pytest.mark.parametrize("path,title", [
    ("/legal/terminos", "Términos"),
    ("/legal/privacidad", "Privacidad"),
    ("/legal/dpa", "tratamiento de datos"),
    ("/legal/sla", "nivel de servicio"),
    ("/legal/metodologia", "metodológico"),
])
def test_v1_public_legal_documents_are_available(path: str, title: str):
    with TestClient(app) as client:
        response = client.get(path)
        assert response.status_code == 200
        assert title.lower() in response.text.lower()
        assert "GREENATICS S.A.S." in response.text


def test_v1_public_contact_requires_privacy_consent():
    with TestClient(app) as client:
        response = client.post(
            "/contacto",
            data={
                "company_name": "Empresa de prueba",
                "contact_name": "Ana Pérez",
                "email": "ana@example.com",
                "message": "Solicitamos una evaluación inicial de nuestra huella corporativa.",
            },
            follow_redirects=False,
        )
        assert response.status_code == 400


def test_v1_public_production_remains_blocked_without_real_identity_and_external_evidence():
    summary = release_candidate_summary(
        ROOT,
        critical_open=0,
        approved_gates=9,
        gate_count=9,
        validated_journeys=5,
        journey_count=5,
    )
    assert summary["controlled_release_ready"] is True
    assert summary["production_ready"] is False
    assert all(item["ok"] is False for item in summary["external_checks"])
