from datetime import date
from io import BytesIO
import zipfile

import pytest
from openpyxl import load_workbook
from fastapi.testclient import TestClient
from sqlalchemy import delete, func, select

from app.calculations import convert_value, recalculate_source
from app.database import (
    ActivityData,
    Base,
    EmissionCalculation,
    EmissionSource,
    ENGINE,
    EvidenceDocument,
    Inventory,
    InventoryDecision,
    ActivityIndicator,
    ReductionAction,
    EmissionTarget,
    AppUser,
    SectorTemplate,
    ReportArtifact,
    ReviewObservation,
    ReductionScenario,
    ReductionScenarioAction,
    VerificationFinding,
    Supplier,
    SupplierCampaign,
    SupplierDataRequest,
    SupplierResponse,
    Notification,
    NotificationPreference,
    PlatformSetting,
    Organization,
    OrganizationMembership,
    AutomationRun,
    IntegrationEvent,
    MethodologyRelease,
    InventoryMethodologySnapshot,
    ComplianceAssessment,
    DocumentControlRecord,
    CommercialReadinessItem,
    ServicePlan,
    OrganizationSubscription,
    CustomerOnboardingItem,
    SupportTicket,
    BillingInvoice,
    CommercialLead,
    CommercialProposal,
    PaymentTransaction,
    ServiceContract,
    ServiceOrder,
    CollectionAction,
    BillingDocumentRecord,
    CustomerSuccessProfile,
    AccountHealthSnapshot,
    ValueMilestone,
    SuccessCommitment,
    RenewalOpportunity,
    BenchmarkReference,
    ImpactSnapshot,
    ClimateRiskAssessment,
    ClimateRisk,
    ClimateRiskControl,
    ClimateTransitionRoadmap,
    ClimateTransitionAction,
    ClimateScenarioDefinition,
    ClimateDisclosureStatement,
    ClimateDisclosureRequirement,
    ClimateBoardBriefing,
    ClimateBoardDecision,
    SessionLocal,
    init_db,
)
from app.main import app
from app.scenarios import scenario_summary
from app.impact_intelligence import impact_metrics, compare_benchmarks
from app.climate_risk import assessment_summary, calculate_risk_scores, risk_level
from app.climate_disclosure import scenario_comparison, disclosure_summary, board_summary


@pytest.fixture(autouse=True)
def fresh_database():
    # La base semilla aislada se restaura desde tests/conftest.py.
    yield


def login(client: TestClient, email: str = "consultor@calculatuhuella.local"):
    response = client.post("/login", data={"email": email, "password": "Demo2026!"}, follow_redirects=False)
    assert response.status_code == 303


def make_inventory_approvable():
    with SessionLocal() as session:
        inventory = session.scalar(select(Inventory).where(Inventory.id == 1))
        for source in inventory.sources:
            source.progress = 100
            source.status = "Completado"
        for observation in inventory.observations:
            if observation.status != "Cerrada":
                observation.status = "Cerrada"
                observation.resolution = "Cerrada por prueba automática"
                observation.resolved_by = "revisor@calculatuhuella.local"
                observation.closed_by = "revisor@calculatuhuella.local"
        session.commit()


def approve_and_close(client: TestClient):
    make_inventory_approvable()
    login(client, "revisor@calculatuhuella.local")
    response = client.post(
        "/control/inventario/recomendar",
        data={"inventory_id": 1, "comments": "Revisión técnica favorable"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    client.post("/logout")
    login(client, "admin@calculatuhuella.local")
    response = client.post(
        "/control/inventario/aprobar",
        data={"inventory_id": 1, "comments": "Aprobación independiente"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    response = client.post(
        "/control/inventario/cerrar",
        data={"inventory_id": 1, "comments": "Cierre definitivo"},
        follow_redirects=False,
    )
    assert response.status_code == 303


def test_health():
    with TestClient(app) as client:
        response = client.get("/api/health")
        assert response.status_code == 200
        assert response.json()["version"] == "1.0.0"


def test_core_pages_load():
    with TestClient(app) as client:
        login(client)
        paths = [
            "/dashboard", "/recorrido-inventario", "/inventarios", "/inventarios/1", "/inventarios/1/editar",
            "/inventarios/1/fuentes", "/fuentes/1", "/organizacion", "/informacion",
            "/informacion/importar", "/cargas-operativas", "/calculos", "/metodologia", "/metodologia/nucleo", "/piloto-greenatics", "/analisis", "/reduccion",
            "/control", "/entrega-profesional", "/reportes", "/modulos", "/sectorizacion", "/escenarios", "/verificacion", "/cadena-valor", "/notificaciones",
            "/direccion-ejecutiva", "/cumplimiento", "/gobierno-metodologico", "/centro-documental", "/consolidacion",
            "/cuenta-servicio", "/onboarding", "/soporte", "/comercial", "/operacion-comercial", "/exito-cliente", "/inteligencia-impacto",
        ]
        for path in paths:
            response = client.get(path)
            assert response.status_code == 200, path
            assert "Calcula tu Huella" in response.text


def test_excel_template_downloads():
    with TestClient(app) as client:
        login(client)
        response = client.get("/informacion/plantilla.xlsx")
        assert response.status_code == 200
        assert response.content[:2] == b"PK"


def test_activity_data_can_be_created_and_calculated():
    with SessionLocal() as session:
        source = session.scalar(select(EmissionSource).where(EmissionSource.name == "Transporte contratado"))
        session.execute(delete(ActivityData).where(ActivityData.source_id == source.id, ActivityData.period_start == date(2025, 4, 1)))
        session.commit()
        source_id = source.id
    with TestClient(app) as client:
        login(client, "cliente@calculatuhuella.local")
        response = client.post(
            "/informacion/datos/nuevo",
            data={
                "source_id": source_id,
                "period_start": "2025-04-01",
                "period_end": "2025-04-30",
                "value": "14500",
                "unit": "t·km",
                "data_origin": "Registro operativo",
                "notes": "Prueba automática",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        record = session.scalar(select(ActivityData).where(ActivityData.source_id == source_id, ActivityData.period_start == date(2025, 4, 1)))
        assert record is not None
        assert record.quality_level == "B"
        calculation_count = session.scalar(select(func.count()).select_from(EmissionCalculation).where(EmissionCalculation.activity_data_id == record.id))
        assert calculation_count == 1


def test_evidence_upload_and_download():
    with TestClient(app) as client:
        login(client, "cliente@calculatuhuella.local")
        response = client.post(
            "/informacion/evidencias/nueva",
            data={"document_type": "Certificado", "period_label": "Prueba", "notes": "Archivo de prueba"},
            files={"file": ("soporte_prueba.csv", b"fuente,valor\nElectricidad,10\n", "text/csv")},
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        document = session.scalar(select(EvidenceDocument).where(EvidenceDocument.name == "soporte_prueba.csv").order_by(EvidenceDocument.id.desc()))
        assert document is not None
        document_id = document.id
    with TestClient(app) as client:
        login(client)
        response = client.get(f"/evidencias/{document_id}/descargar")
        assert response.status_code == 200
        assert b"Electricidad" in response.content


def test_role_restriction():
    with TestClient(app) as client:
        login(client, "cliente@calculatuhuella.local")
        response = client.post(
            "/organizacion/editar",
            data={"name": "No autorizado", "tax_id": "1", "sector": "X", "city": "X"},
        )
        assert response.status_code == 403
        methodology = client.get("/metodologia")
        assert methodology.status_code == 403


def test_unit_conversion_engine():
    with SessionLocal() as session:
        converted, note = convert_value(session, 10, "gal", "L")
        assert converted is not None
        assert round(converted, 6) == round(37.85411784, 6)
        assert "gal" in note and "L" in note
        incompatible, warning = convert_value(session, 10, "kg", "kWh")
        assert incompatible is None
        assert "incompatibles" in warning


def test_diesel_is_calculated_by_gas_components():
    with SessionLocal() as session:
        source = session.scalar(select(EmissionSource).where(EmissionSource.name == "Diésel"))
        result = recalculate_source(session, source)
        session.commit()
        assert result["calculations"] == 36
        assert 40 < result["emissions"] < 42
        gases = set(session.scalars(select(EmissionCalculation.gas_code).join(ActivityData).where(ActivityData.source_id == source.id)))
        assert gases == {"CO2", "CH4", "N2O"}


def test_formula_snapshot_and_factor_traceability():
    with SessionLocal() as session:
        calculation = session.scalar(select(EmissionCalculation).order_by(EmissionCalculation.id))
        assert calculation is not None
        assert "GWP" in calculation.formula_snapshot
        assert calculation.factor_version_id is not None
        assert calculation.engine_version == "1.1.0"


def test_observation_response_and_closure_workflow():
    with TestClient(app) as client:
        login(client, "revisor@calculatuhuella.local")
        response = client.post(
            "/control/observaciones/nueva",
            data={
                "inventory_id": 1,
                "entity_type": "Inventario",
                "title": "Prueba de revisión",
                "description": "Validar la trazabilidad del dato",
                "severity": "Menor",
                "assigned_to": "Gestión ambiental",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
        client.post("/logout")
        login(client, "cliente@calculatuhuella.local")
        with SessionLocal() as session:
            observation = session.scalar(select(ReviewObservation).where(ReviewObservation.title == "Prueba de revisión"))
            observation_id = observation.id
        assert client.post(f"/control/observaciones/{observation_id}/responder", data={"response": "Dato corregido y documentado"}, follow_redirects=False).status_code == 303
        assert client.post(f"/control/observaciones/{observation_id}/enviar", follow_redirects=False).status_code == 303
        client.post("/logout")
        login(client, "revisor@calculatuhuella.local")
        response = client.post(
            f"/control/observaciones/{observation_id}/cerrar",
            data={"resolution": "Corrección satisfactoria", "decision": "Cerrar"},
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        observation = session.get(ReviewObservation, observation_id)
        assert observation.status == "Cerrada"
        assert observation.closed_by == "revisor@calculatuhuella.local"


def test_approval_is_blocked_when_quality_gates_fail():
    with TestClient(app) as client:
        login(client, "revisor@calculatuhuella.local")
        response = client.post(
            "/control/inventario/recomendar",
            data={"inventory_id": 1, "comments": "Intento prematuro"},
            follow_redirects=False,
        )
        assert response.status_code == 409


def test_approval_requires_independent_user_and_closure_is_immutable():
    with TestClient(app) as client:
        make_inventory_approvable()
        login(client, "revisor@calculatuhuella.local")
        response = client.post(
            "/control/inventario/recomendar",
            data={"inventory_id": 1, "comments": "Revisión favorable"},
            follow_redirects=False,
        )
        assert response.status_code == 303
        same_user = client.post(
            "/control/inventario/aprobar",
            data={"inventory_id": 1, "comments": "Autaprobación no válida"},
            follow_redirects=False,
        )
        assert same_user.status_code == 409
        client.post("/logout")
        login(client, "admin@calculatuhuella.local")
        assert client.post("/control/inventario/aprobar", data={"inventory_id": 1, "comments": "Aprobación independiente"}, follow_redirects=False).status_code == 303
        assert client.post("/control/inventario/cerrar", data={"inventory_id": 1, "comments": "Cierre definitivo"}, follow_redirects=False).status_code == 303
        client.post("/logout")
        login(client, "cliente@calculatuhuella.local")
        blocked = client.post(
            "/informacion/datos/nuevo",
            data={
                "source_id": 1,
                "period_start": "2025-01-01",
                "period_end": "2025-01-31",
                "value": "10",
                "unit": "kWh",
                "data_origin": "Factura",
            },
            follow_redirects=False,
        )
        assert blocked.status_code == 409
    with SessionLocal() as session:
        inventory = session.get(Inventory, 1)
        assert inventory.locked is True
        assert inventory.status == "Cerrado"
        assert session.scalar(select(func.count()).select_from(InventoryDecision).where(InventoryDecision.inventory_id == 1)) >= 3


def test_reopen_creates_new_version_and_preserves_original():
    with TestClient(app) as client:
        approve_and_close(client)
        response = client.post(
            "/control/inventario/reabrir",
            data={"inventory_id": 1, "reason": "Se identificó una factura omitida"},
            follow_redirects=False,
        )
        assert response.status_code == 303
        assert response.headers["location"].startswith("/inventarios/")
    with SessionLocal() as session:
        original = session.get(Inventory, 1)
        new_inventory = session.scalar(select(Inventory).where(Inventory.parent_inventory_id == 1))
        assert original.locked is True
        assert new_inventory is not None
        assert new_inventory.locked is False
        assert new_inventory.version.endswith("-r1")
        assert len(new_inventory.sources) == len(original.sources)


def test_v06_indicators_and_reduction_are_persistent():
    with TestClient(app) as client:
        login(client)
        response = client.post(
            "/analisis/indicadores/nuevo",
            data={
                "inventory_id": 1, "indicator_type": "Servicios", "value": "2500", "unit": "servicios",
                "period_start": "2025-01-01", "period_end": "2025-12-31", "source_name": "ERP",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
        response = client.post(
            "/reduccion/acciones/nueva",
            data={
                "inventory_id": 1, "title": "Acción de prueba", "description": "Validación automática",
                "source_id": 1, "expected_reduction": "5", "investment_cost": "1000000",
                "annual_savings": "500000", "priority": "Media", "responsible": "Operaciones",
                "target_date": "2026-12-31", "status": "Identificada",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ActivityIndicator).where(ActivityIndicator.indicator_type == "Servicios")) == 1
        assert session.scalar(select(func.count()).select_from(ReductionAction).where(ReductionAction.title == "Acción de prueba")) == 1


def test_v06_reports_generate_and_download():
    with TestClient(app) as client:
        login(client)
        for report_type, signature in [("ejecutivo", b"%PDF"), ("tecnico", b"%PDF"), ("memoria", b"PK")]:
            response = client.post(
                "/reportes/generar",
                data={"inventory_id": 1, "report_type": report_type},
                follow_redirects=False,
            )
            assert response.status_code == 303
            with SessionLocal() as session:
                artifact = session.scalar(select(ReportArtifact).order_by(ReportArtifact.id.desc()))
                artifact_id = artifact.id
            download = client.get(f"/reportes/{artifact_id}/descargar")
            assert download.status_code == 200
            assert download.content.startswith(signature)


def test_v06_analysis_uses_historical_inventory():
    with TestClient(app) as client:
        login(client)
        response = client.get("/analisis")
        assert response.status_code == 200
        assert "2024" in response.text
        assert "2025" in response.text
        assert "COMPARACIÓN HISTÓRICA" in response.text


def test_v07_sector_template_applies_sources_without_duplicates():
    with TestClient(app) as client:
        login(client)
        with SessionLocal() as session:
            template = session.scalar(select(SectorTemplate).where(SectorTemplate.sector == "Servicios y oficinas"))
            before = session.scalar(select(func.count()).select_from(EmissionSource).where(EmissionSource.inventory_id == 1))
            template_id = template.id
        response = client.post("/sectorizacion/aplicar", data={"inventory_id": 1, "template_id": template_id, "facility_id": 1, "include_optional": "true"}, follow_redirects=False)
        assert response.status_code == 303
        with SessionLocal() as session:
            after = session.scalar(select(func.count()).select_from(EmissionSource).where(EmissionSource.inventory_id == 1))
            assert after > before
        second = client.post("/sectorizacion/aplicar", data={"inventory_id": 1, "template_id": template_id, "facility_id": 1, "include_optional": "true"}, follow_redirects=False)
        assert second.status_code == 303
        with SessionLocal() as session:
            final_count = session.scalar(select(func.count()).select_from(EmissionSource).where(EmissionSource.inventory_id == 1))
            assert final_count == after


def test_v07_user_governance_and_verifier_login():
    with TestClient(app) as client:
        login(client, "admin@calculatuhuella.local")
        assert client.get("/usuarios").status_code == 200
        response = client.post("/usuarios/nuevo", data={"name": "Verificador Externo", "email": "externo@example.com", "role": "Verificador", "password": "ClaveSegura2026!"}, follow_redirects=False)
        assert response.status_code == 303
        with SessionLocal() as session:
            created = session.scalar(select(AppUser).where(AppUser.email == "externo@example.com"))
            assert created is not None and created.role == "Verificador" and created.active is True
            created_id = created.id
        response = client.post(f"/usuarios/{created_id}/estado", data={"active": "false"}, follow_redirects=False)
        assert response.status_code == 303
        with SessionLocal() as session:
            membership = session.scalar(select(OrganizationMembership).where(OrganizationMembership.user_id == created_id, OrganizationMembership.organization_id == 1))
            assert membership is not None and membership.active is False
            assert session.get(AppUser, created_id).active is True


def test_v07_climate_target_lifecycle():
    with TestClient(app) as client:
        login(client)
        response = client.post("/reduccion/metas/nueva", data={"inventory_id": 1, "name": "Meta de prueba", "metric_type": "Absoluta", "baseline_year": 2025, "target_year": 2030, "baseline_value": 250, "target_value": 180, "unit": "tCO₂e", "notes": "Prueba automática"}, follow_redirects=False)
        assert response.status_code == 303
        with SessionLocal() as session:
            target = session.scalar(select(EmissionTarget).where(EmissionTarget.name == "Meta de prueba"))
            assert target is not None
            target_id = target.id
        response = client.post(f"/reduccion/metas/{target_id}/actualizar", data={"current_value": 210, "status": "Activa", "notes": "Actualizada"}, follow_redirects=False)
        assert response.status_code == 303
        with SessionLocal() as session:
            target = session.get(EmissionTarget, target_id)
            assert target.current_value == 210
            assert target.progress_percent > 0



def test_v08_seeded_scenario_has_financial_summary():
    with SessionLocal() as session:
        scenario = session.scalar(select(ReductionScenario).order_by(ReductionScenario.id))
        assert scenario is not None
        # Load through the route-oriented helper so relationships are available.
        from app.scenarios import get_scenario
        scenario = get_scenario(session, scenario.id, 1)
        summary = scenario_summary(scenario)
        assert summary["total_reduction"] > 0
        assert summary["investment"] > 0
        assert len(summary["macc"]) == 3
        assert summary["projected_emissions"] < summary["baseline"]


def test_v08_scenario_creation_and_configuration():
    with TestClient(app) as client:
        login(client)
        response = client.post(
            "/escenarios/nuevo",
            data={"inventory_id": 1, "name": "Escenario prueba", "description": "Prueba", "start_year": 2026, "target_year": 2032, "discount_rate": 8.5},
            follow_redirects=False,
        )
        assert response.status_code == 303
        with SessionLocal() as session:
            scenario = session.scalar(select(ReductionScenario).where(ReductionScenario.name == "Escenario prueba"))
            assert scenario is not None
            links = list(session.scalars(select(ReductionScenarioAction).where(ReductionScenarioAction.scenario_id == scenario.id)))
            assert len(links) == 3
            scenario_id = scenario.id
            first_action = links[0].action_id
        response = client.post(
            f"/escenarios/{scenario_id}/configurar",
            data={"status": "En evaluación", "discount_rate": "9", f"include_{first_action}": "on", f"adoption_{first_action}": "75", f"year_{first_action}": "2027"},
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        scenario = session.get(ReductionScenario, scenario_id)
        link = session.scalar(select(ReductionScenarioAction).where(ReductionScenarioAction.scenario_id == scenario_id, ReductionScenarioAction.action_id == first_action))
        assert scenario.status == "En evaluación"
        assert scenario.discount_rate == 9
        assert link.included is True
        assert link.adoption_percent == 75
        assert link.implementation_year == 2027


def test_v08_verifier_finding_workflow():
    with TestClient(app) as client:
        response = client.post("/login", data={"email": "verificador@calculatuhuella.local", "password": "Demo2026!"}, follow_redirects=False)
        assert response.status_code == 303
        assert response.headers["location"] == "/verificacion"
        response = client.post(
            "/verificacion/hallazgos/nuevo",
            data={"inventory_id": 1, "title": "Hallazgo prueba", "description": "Revisar soporte", "finding_type": "Solicitud de información", "severity": "Menor", "source_id": "1"},
            follow_redirects=False,
        )
        assert response.status_code == 303
        client.post("/logout")
        login(client, "cliente@calculatuhuella.local")
        with SessionLocal() as session:
            finding = session.scalar(select(VerificationFinding).where(VerificationFinding.title == "Hallazgo prueba"))
            finding_id = finding.id
        assert client.post(f"/verificacion/hallazgos/{finding_id}/responder", data={"management_response": "Soporte cargado"}, follow_redirects=False).status_code == 303
        client.post("/logout")
        login(client, "verificador@calculatuhuella.local")
        assert client.post(f"/verificacion/hallazgos/{finding_id}/cerrar", data={"conclusion": "Respuesta suficiente", "decision": "Cerrar"}, follow_redirects=False).status_code == 303
    with SessionLocal() as session:
        finding = session.get(VerificationFinding, finding_id)
        assert finding.status == "Cerrado"
        assert finding.management_response == "Soporte cargado"
        assert finding.closed_by == "verificador@calculatuhuella.local"


def test_v08_verification_package_is_downloadable_zip():
    with TestClient(app) as client:
        login(client, "verificador@calculatuhuella.local")
        response = client.post("/verificacion/paquete", data={"inventory_id": 1}, follow_redirects=False)
        assert response.status_code == 303
        with SessionLocal() as session:
            artifact = session.scalar(select(ReportArtifact).where(ReportArtifact.report_type == "Paquete de verificación").order_by(ReportArtifact.id.desc()))
            artifact_id = artifact.id
        download = client.get(f"/reportes/{artifact_id}/descargar")
        assert download.status_code == 200
        assert download.content.startswith(b"PK")
        with zipfile.ZipFile(BytesIO(download.content)) as archive:
            names = archive.namelist()
            assert "00_manifiesto.json" in names
            assert "03_calculos.csv" in names
            assert "07_hallazgos_verificacion.csv" in names


def test_v09_supply_chain_demo_is_consolidated():
    with SessionLocal() as session:
        source = session.scalar(select(EmissionSource).where(EmissionSource.category == "Datos específicos de proveedores"))
        assert source is not None
        assert round(source.emissions, 1) == 340.8
        assert session.scalar(select(func.count()).select_from(Supplier)) == 4
        assert session.scalar(select(func.count()).select_from(SupplierCampaign)) == 1


def test_v09_supplier_public_response_and_review():
    with SessionLocal() as session:
        data_request = session.scalar(
            select(SupplierDataRequest).join(Supplier).where(Supplier.name == "Químicos Andinos Ltda.")
        )
        token = data_request.access_token
        request_id = data_request.id
    with TestClient(app) as client:
        portal = client.get(f"/proveedor/responder/{token}")
        assert portal.status_code == 200
        assert "Químicos Andinos" in portal.text
        response = client.post(
            f"/proveedor/responder/{token}",
            data={
                "method": "Factor por unidad",
                "activity_value": "96",
                "activity_unit": "t",
                "emission_factor": "950",
                "factor_unit": "kg CO2e/t",
                "reported_emissions_tco2e": "0",
                "methodology": "Huella de producto específica",
                "boundary": "Cradle-to-gate",
                "notes": "Respuesta automática de prueba",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        supplier_response = session.scalar(select(SupplierResponse).where(SupplierResponse.request_id == request_id))
        assert round(supplier_response.calculated_emissions_tco2e, 1) == 91.2
        assert supplier_response.quality_level == "C"
        response_id = supplier_response.id
    with TestClient(app) as client:
        login(client, "revisor@calculatuhuella.local")
        response = client.post(
            f"/cadena-valor/respuestas/{response_id}/revisar",
            data={"decision": "Aprobado", "reviewer_comments": "Factor aceptado para prueba"},
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        source = session.scalar(select(EmissionSource).where(EmissionSource.category == "Datos específicos de proveedores"))
        assert round(source.emissions, 1) == 432.0


def test_v09_supply_chain_excel_downloads():
    with TestClient(app) as client:
        login(client)
        response = client.get("/cadena-valor/plantilla.xlsx")
        assert response.status_code == 200
        assert response.content[:2] == b"PK"


def test_v09_invalid_supplier_token_is_not_exposed():
    with TestClient(app) as client:
        response = client.get("/proveedor/responder/token-inexistente")
        assert response.status_code == 404


def test_readiness_endpoint_and_security_headers():
    with TestClient(app) as client:
        response = client.get("/api/ready")
        assert response.status_code == 200
        payload = response.json()
        assert payload["database_ok"] is True
        assert payload["storage_ok"] is True
        assert payload["database_backend"] == "SQLite"
        health = client.get("/api/health")
        assert health.headers["x-content-type-options"] == "nosniff"
        assert health.headers["x-frame-options"] == "DENY"
        assert "frame-ancestors 'none'" in health.headers["content-security-policy"]


def test_password_hashing_is_salted_and_verifiable():
    from app.security import hash_password as secure_hash_password, verify_password

    first = secure_hash_password("PruebaSegura2026!")
    second = secure_hash_password("PruebaSegura2026!")
    assert first.startswith("pbkdf2_sha256$")
    assert first != second
    assert verify_password("PruebaSegura2026!", first)
    assert not verify_password("incorrecta", first)


def test_legacy_password_is_upgraded_on_login():
    import hashlib

    with SessionLocal() as session:
        user = session.scalar(select(AppUser).where(AppUser.email == "consultor@calculatuhuella.local"))
        user.password_hash = hashlib.sha256(b"Demo2026!").hexdigest()
        session.commit()
    with TestClient(app) as client:
        login(client)
    with SessionLocal() as session:
        user = session.scalar(select(AppUser).where(AppUser.email == "consultor@calculatuhuella.local"))
        assert user.password_hash.startswith("pbkdf2_sha256$")


def test_admin_can_generate_and_download_backup():
    from app.operations import BACKUP_DIR

    for item in BACKUP_DIR.glob("*.zip"):
        item.unlink()
    with TestClient(app) as client:
        login(client, "admin@calculatuhuella.local")
        page = client.get("/operacion")
        assert page.status_code == 200
        assert "Preparación productiva, seguridad y continuidad" in page.text
        response = client.post("/operacion/respaldos", data={"label": "prueba"}, follow_redirects=False)
        assert response.status_code == 303
        archive = next(BACKUP_DIR.glob("*prueba.zip"))
        download = client.get(f"/operacion/respaldos/{archive.name}")
        assert download.status_code == 200
        assert download.content[:2] == b"PK"


def test_non_admin_cannot_access_operations():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.get("/operacion")
        assert response.status_code == 403


def test_v011_notifications_preferences_and_read_flow():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        page = client.get("/notificaciones")
        assert page.status_code == 200
        assert "Notificaciones y alertas" in page.text
        with SessionLocal() as session:
            user = session.scalar(select(AppUser).where(AppUser.email == "consultor@calculatuhuella.local"))
            notification = session.scalar(select(Notification).where(Notification.user_id == user.id).order_by(Notification.id))
            notification_id = notification.id
        opened = client.post(f"/notificaciones/{notification_id}/leer", follow_redirects=False)
        assert opened.status_code == 303
        preference = client.post("/notificaciones/preferencias", data={"in_app_enabled": "on", "email_enabled": "on", "digest_frequency": "Diario"}, follow_redirects=False)
        assert preference.status_code == 303
    with SessionLocal() as session:
        user = session.scalar(select(AppUser).where(AppUser.email == "consultor@calculatuhuella.local"))
        notification = session.get(Notification, notification_id)
        pref = session.scalar(select(NotificationPreference).where(NotificationPreference.user_id == user.id))
        assert notification.read_at is not None
        assert pref.digest_frequency == "Diario"


def test_v011_admin_platform_and_notification_queue():
    with TestClient(app) as client:
        login(client, "admin@calculatuhuella.local")
        page = client.get("/administracion-plataforma")
        assert page.status_code == 200
        assert "Configuración y operación avanzada" in page.text
        created = client.post("/administracion-plataforma/notificaciones/prueba", data={"role": "Consultor", "title": "Prueba automática", "message": "Mensaje V0.11"}, follow_redirects=False)
        assert created.status_code == 303
        processed = client.post("/administracion-plataforma/notificaciones/procesar", follow_redirects=False)
        assert processed.status_code == 303
        config = client.post("/administracion-plataforma/configuracion", data={"key": "test_setting", "value": "activo", "description": "Prueba"}, follow_redirects=False)
        assert config.status_code == 303
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(Notification).where(Notification.title == "Prueba automática")) >= 1
        row = session.scalar(select(PlatformSetting).where(PlatformSetting.key == "test_setting"))
        assert row is not None and row.value == "activo"


def test_v011_non_admin_cannot_access_platform_admin():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        assert client.get("/administracion-plataforma").status_code == 403



def test_portfolio_switches_organization_context():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.get("/portafolio")
        assert response.status_code == 200
        assert "Transportes Horizonte Demo" in response.text
        response = client.post("/portafolio/cambiar/2", follow_redirects=False)
        assert response.status_code == 303
        response = client.get("/dashboard")
        assert response.status_code == 200
        assert "Transportes Horizonte" in response.text


def test_admin_can_create_new_organization_and_membership():
    with TestClient(app) as client:
        login(client, "admin@calculatuhuella.local")
        response = client.post(
            "/portafolio/nueva",
            data={"name": "Servicios Climáticos Demo S.A.S.", "trade_name": "Servicios Climáticos", "tax_id": "901999000-1", "sector": "Servicios y oficinas", "city": "Bogotá"},
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        organization = session.scalar(select(Organization).where(Organization.name == "Servicios Climáticos Demo S.A.S."))
        assert organization is not None
        membership = session.scalar(select(OrganizationMembership).where(OrganizationMembership.organization_id == organization.id))
        assert membership is not None
        assert membership.role == "Administrador"


def test_manual_automation_execution_creates_run():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post("/automatizaciones/1/ejecutar", follow_redirects=False)
        assert response.status_code == 303
    with SessionLocal() as session:
        run = session.scalar(select(AutomationRun).where(AutomationRun.automation_id == 1).order_by(AutomationRun.id.desc()))
        assert run is not None
        assert run.status in {"Ejecutado", "Error"}
        assert run.finished_at is not None


def test_api_integration_ingests_activity_data():
    with TestClient(app) as client:
        response = client.post(
            "/api/v1/activity-data",
            headers={"X-API-Key": "cth_demo_1_2026"},
            json={
                "source_id": 1,
                "period_start": "2025-12-01",
                "period_end": "2025-12-31",
                "value": 1111.0,
                "unit": "kWh",
                "external_reference": "TEST-ERP-001",
            },
        )
        assert response.status_code == 200
        assert response.json()["ok"] is True
    with SessionLocal() as session:
        event = session.scalar(select(IntegrationEvent).where(IntegrationEvent.external_reference == "TEST-ERP-001"))
        assert event is not None
        assert event.status == "Recibido"


def test_api_rejects_invalid_key():
    with TestClient(app) as client:
        response = client.get("/api/v1/sources", headers={"X-API-Key": "invalid"})
        assert response.status_code == 401



def test_api_external_reference_is_idempotent():
    payload = {
        "source_id": 1,
        "period_start": "2025-11-01",
        "period_end": "2025-11-30",
        "value": 987.0,
        "unit": "kWh",
        "external_reference": "IDEMPOTENT-001",
    }
    with TestClient(app) as client:
        first = client.post("/api/v1/activity-data", headers={"X-API-Key": "cth_demo_1_2026"}, json=payload)
        second = client.post("/api/v1/activity-data", headers={"X-API-Key": "cth_demo_1_2026"}, json=payload)
        assert first.status_code == 200
        assert second.status_code == 200
        assert second.json()["duplicate"] is True
        assert second.json()["activity_data_id"] == first.json()["activity_data_id"]


def test_v013_compliance_update_and_verifier_read_only():
    with SessionLocal() as session:
        assessment = session.scalar(select(ComplianceAssessment).where(ComplianceAssessment.inventory_id == 1).order_by(ComplianceAssessment.id))
        assessment_id = assessment.id
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            f"/cumplimiento/{assessment_id}/actualizar",
            data={"status": "Cumple", "owner": "Carlos Uribe", "notes": "Control validado"},
            follow_redirects=False,
        )
        assert response.status_code == 303
        client.post("/logout")
        login(client, "verificador@calculatuhuella.local")
        assert client.get("/cumplimiento").status_code == 200
        denied = client.post(
            f"/cumplimiento/{assessment_id}/actualizar",
            data={"status": "No cumple", "owner": "Verificador"},
            follow_redirects=False,
        )
        assert denied.status_code == 403
    with SessionLocal() as session:
        assessment = session.get(ComplianceAssessment, assessment_id)
        assert assessment.status == "Cumple"
        assert assessment.updated_by == "consultor@calculatuhuella.local"


def test_v013_methodology_release_and_snapshot():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            "/gobierno-metodologico/versiones/nueva",
            data={
                "name": "Protocolo de prueba", "version": "2.0", "issuing_body": "Equipo metodológico",
                "publication_date": "2026-07-31", "effective_from": "2026-08-01",
                "source_reference": "Documento interno", "notes": "Versión de prueba",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        release = session.scalar(select(MethodologyRelease).where(MethodologyRelease.name == "Protocolo de prueba"))
        assert release is not None and release.status == "Borrador"
        release_id = release.id
    with TestClient(app) as client:
        login(client, "revisor@calculatuhuella.local")
        assert client.post(f"/gobierno-metodologico/versiones/{release_id}/aprobar", follow_redirects=False).status_code == 303
        assert client.post(
            "/gobierno-metodologico/snapshots/nuevo",
            data={"inventory_id": 1, "methodology_release_id": release_id, "snapshot_name": "Snapshot prueba", "policy_notes": "Regla congelada"},
            follow_redirects=False,
        ).status_code == 303
    with SessionLocal() as session:
        release = session.get(MethodologyRelease, release_id)
        snapshot = session.scalar(select(InventoryMethodologySnapshot).where(InventoryMethodologySnapshot.snapshot_name == "Snapshot prueba"))
        assert release.status == "Aprobado"
        assert snapshot is not None and snapshot.methodology_release_id == release_id


def test_v013_document_control_register():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            "/centro-documental/registros/nuevo",
            data={
                "document_code": "POL-TEST-001", "title": "Política de prueba", "category": "Política",
                "version": "1.0", "owner": "Dirección", "confidentiality": "Interno",
                "retention_years": 8, "review_due": "2027-07-31", "inventory_id": 1, "notes": "Registro automático",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        row = session.scalar(select(DocumentControlRecord).where(DocumentControlRecord.document_code == "POL-TEST-001"))
        assert row is not None
        assert row.retention_years == 8
        assert row.inventory_id == 1


def test_v013_readiness_workflow():
    with SessionLocal() as session:
        item = session.scalar(select(CommercialReadinessItem).where(CommercialReadinessItem.organization_id == 1).order_by(CommercialReadinessItem.id))
        item_id = item.id
    with TestClient(app) as client:
        login(client, "admin@calculatuhuella.local")
        assert client.get("/alistamiento").status_code == 200
        response = client.post(
            f"/alistamiento/{item_id}/actualizar",
            data={"status": "Completado", "owner": "Dirección general", "due_date": "2026-12-31", "notes": "Listo"},
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        item = session.get(CommercialReadinessItem, item_id)
        assert item.status == "Completado"
        assert item.updated_by == "admin@calculatuhuella.local"


def test_v013_executive_portfolio_contains_both_organizations():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.get("/direccion-ejecutiva")
        assert response.status_code == 200
        assert "Industrias Andinas" in response.text
        assert "Transportes Horizonte" in response.text
        assert "Cumplimiento promedio" in response.text


def test_v013_seeded_governance_data_exists_per_organization():
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(MethodologyRelease).where(MethodologyRelease.organization_id == 1)) >= 2
        assert session.scalar(select(func.count()).select_from(ComplianceAssessment).where(ComplianceAssessment.inventory_id == 1)) >= 10
        assert session.scalar(select(func.count()).select_from(DocumentControlRecord).where(DocumentControlRecord.organization_id == 1)) >= 2
        assert session.scalar(select(func.count()).select_from(CommercialReadinessItem).where(CommercialReadinessItem.organization_id == 1)) >= 10


def test_service_account_has_seeded_subscription_and_plans():
    with TestClient(app) as client:
        login(client, "admin@calculatuhuella.local")
        response = client.get("/cuenta-servicio")
        assert response.status_code == 200
        assert "Gestión de Carbono" in response.text
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ServicePlan)) >= 3
        assert session.scalar(select(func.count()).select_from(OrganizationSubscription)) >= 1
        assert session.scalar(select(func.count()).select_from(BillingInvoice)) >= 1


def test_onboarding_item_can_be_completed():
    with SessionLocal() as session:
        consultor = session.scalar(select(AppUser).where(AppUser.email == "consultor@calculatuhuella.local"))
        assert consultor is not None
        item = session.scalar(
            select(CustomerOnboardingItem)
            .where(CustomerOnboardingItem.organization_id == consultor.organization_id)
            .order_by(CustomerOnboardingItem.id)
        )
        assert item is not None
        item.status = "Pendiente"
        item.completed_at = None
        session.commit()
        item_id = item.id
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(f"/onboarding/{item_id}/actualizar", data={"status": "Completado", "owner": "Consultoría", "due_date": ""}, follow_redirects=False)
        assert response.status_code == 303
    with SessionLocal() as session:
        item = session.get(CustomerOnboardingItem, item_id)
        assert item.status == "Completado"
        assert item.completed_at is not None


def test_client_can_create_support_ticket():
    with TestClient(app) as client:
        login(client, "cliente@calculatuhuella.local")
        response = client.post("/soporte/nuevo", data={
            "subject": "Ayuda con evidencia",
            "description": "No encuentro el soporte del consumo eléctrico.",
            "category": "Soporte funcional",
            "priority": "Normal",
        }, follow_redirects=False)
        assert response.status_code == 303
    with SessionLocal() as session:
        ticket = session.scalar(select(SupportTicket).where(SupportTicket.subject == "Ayuda con evidencia"))
        assert ticket is not None
        assert ticket.status == "Abierto"


def test_client_cannot_close_support_ticket():
    with SessionLocal() as session:
        ticket = session.scalar(select(SupportTicket).order_by(SupportTicket.id))
        ticket_id = ticket.id
    with TestClient(app) as client:
        login(client, "cliente@calculatuhuella.local")
        response = client.post(f"/soporte/{ticket_id}/actualizar", data={"status": "Cerrado", "assigned_to": "Cliente", "resolution": ""})
        assert response.status_code == 403


def test_saas_admin_can_create_plan():
    with TestClient(app) as client:
        login(client, "admin@calculatuhuella.local")
        response = client.post("/administracion-saas/planes/nuevo", data={
            "code": "PILOTO", "name": "Plan Piloto", "description": "Prueba",
            "monthly_fee": "250000", "annual_fee": "2500000", "max_users": "3",
            "max_facilities": "1", "max_inventories": "1", "max_storage_mb": "500",
        }, follow_redirects=False)
        assert response.status_code == 303
    with SessionLocal() as session:
        plan = session.scalar(select(ServicePlan).where(ServicePlan.code == "PILOTO"))
        assert plan is not None
        assert plan.monthly_fee == 250000


def test_public_site_and_diagnostic_flow():
    with TestClient(app) as client:
        home = client.get("/")
        assert home.status_code == 200
        assert "Mide." in home.text
        response = client.post(
            "/diagnostico",
            data={
                "company_name": "Empresa Prospecto Prueba S.A.S.",
                "contact_name": "Juliana Pérez",
                "email": "juliana@prospecto.test",
                "phone": "3000000000",
                "sector": "Manufactura",
                "city": "Medellín",
                "employees_band": "51 a 200",
                "facilities_count": "4",
                "has_previous_inventory": "on",
                "desired_scopes": "Alcances 1, 2 y 3 priorizado",
                "objective": "Preparación para verificación",
                "urgency": "Alta",
                "notes": "Prueba funcional",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
        assert response.headers["location"].startswith("/diagnostico/gracias/")
    with SessionLocal() as session:
        lead = session.scalar(select(CommercialLead).where(CommercialLead.email == "juliana@prospecto.test"))
        assert lead is not None
        assert lead.complexity_score >= 10
        assert lead.recommended_plan_code in {"EMPRESARIAL", "CORPORATIVO"}


def test_commercial_proposal_acceptance_and_demo_payment():
    with TestClient(app) as client:
        login(client)
        with SessionLocal() as session:
            lead = session.scalar(select(CommercialLead).order_by(CommercialLead.id))
            plan = session.scalar(select(ServicePlan).where(ServicePlan.code == "EMPRESARIAL"))
            lead_id, plan_id = lead.id, plan.id
        response = client.post(
            "/comercial/propuestas/nueva",
            data={
                "lead_id": str(lead_id), "plan_id": str(plan_id),
                "title": "Propuesta de prueba", "implementation_fee": "1000000",
                "recurring_fee": "2000000", "discount_amount": "0", "tax_rate": "19",
                "billing_cycle": "Anual", "valid_until": "2027-12-31",
                "scope": "Alcances 1 y 2\nDashboard",
                "deliverables": "Informe técnico\nMemoria de cálculo", "terms": "Condiciones de prueba",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
        client.post("/logout")
        with SessionLocal() as session:
            proposal = session.scalar(select(CommercialProposal).where(CommercialProposal.title == "Propuesta de prueba"))
            assert proposal is not None
            token = proposal.public_token
            assert proposal.first_year_total == 3570000
        public = client.get(f"/propuesta/{token}")
        assert public.status_code == 200
        accepted = client.post(
            f"/propuesta/{token}/aceptar",
            data={"accepted_by": "Representante Legal", "accepted_email": "legal@empresa.test", "accept_terms": "on"},
            follow_redirects=False,
        )
        assert accepted.status_code == 303
        assert accepted.headers["location"].startswith("/pago/")
        payment_path = accepted.headers["location"]
        payment_page = client.get(payment_path)
        assert payment_page.status_code == 200
        confirmed = client.post(
            payment_path + "/confirmar",
            data={"payer_name": "Representante Legal", "payer_email": "legal@empresa.test", "method": "Transferencia demostrativa"},
            follow_redirects=False,
        )
        assert confirmed.status_code == 303
    with SessionLocal() as session:
        proposal = session.scalar(select(CommercialProposal).where(CommercialProposal.title == "Propuesta de prueba"))
        payment = session.scalar(select(PaymentTransaction).where(PaymentTransaction.proposal_id == proposal.id))
        assert proposal.status == "Aceptada"
        assert proposal.acceptance_hash
        assert proposal.organization_id is not None
        assert payment.status == "Pagada"
        assert payment.invoice_id is not None



def test_v016_seeded_revenue_operations_exist():
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ServiceContract)) >= 1
        assert session.scalar(select(func.count()).select_from(ServiceOrder)) >= 1
        assert session.scalar(select(func.count()).select_from(CollectionAction)) >= 1
        assert session.scalar(select(func.count()).select_from(BillingDocumentRecord)) >= 1
        contract = session.scalar(select(ServiceContract).where(ServiceContract.reference == "CTR-DEMO-2026-001"))
        assert contract is not None
        assert contract.status == "Vigente"
        assert len(contract.signature_hash) == 64


def test_v016_contract_creation_and_signature_are_traceable():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            "/operacion-comercial/contratos/nuevo",
            data={
                "organization_id": "1", "proposal_id": "", "reference": "CTR-TEST-2026-001",
                "title": "Contrato de prueba automatizada", "start_date": "2026-08-01", "end_date": "2027-07-31",
                "renewal_type": "Anual", "auto_renew": "1", "notice_days": "45",
                "contract_value": "12000000", "billing_cycle": "Anual", "owner": "Dirección comercial",
                "terms_snapshot": "Condiciones congeladas para la prueba.",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
        with SessionLocal() as session:
            contract = session.scalar(select(ServiceContract).where(ServiceContract.reference == "CTR-TEST-2026-001"))
            contract_id = contract.id
            assert contract.status == "Borrador"
        signed = client.post(
            f"/operacion-comercial/contratos/{contract_id}/firmar",
            data={"signed_by": "Representante Legal", "signed_email": "legal@empresa.test"},
            follow_redirects=False,
        )
        assert signed.status_code == 303
    with SessionLocal() as session:
        contract = session.get(ServiceContract, contract_id)
        assert contract.status == "Vigente"
        assert contract.signed_email == "legal@empresa.test"
        assert len(contract.signature_hash) == 64


def test_v016_contract_renewal_creates_new_version():
    with SessionLocal() as session:
        contract = session.scalar(select(ServiceContract).where(ServiceContract.reference == "CTR-DEMO-2026-001"))
        contract_id = contract.id
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            f"/operacion-comercial/contratos/{contract_id}/renovar",
            data={"start_date": "2027-07-31", "end_date": "2028-07-30", "contract_value": "11000000"},
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        original = session.get(ServiceContract, contract_id)
        renewed = session.scalar(select(ServiceContract).where(ServiceContract.parent_contract_id == contract_id))
        assert original.status == "Renovado"
        assert renewed is not None
        assert renewed.reference.endswith("-R1")
        assert renewed.status == "Borrador"
        assert renewed.contract_value == 11000000


def test_v016_service_order_lifecycle():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            "/operacion-comercial/ordenes/nueva",
            data={
                "organization_id": "1", "contract_id": "", "reference": "OS-TEST-2026-001",
                "title": "Entrega técnica de prueba", "service_type": "Informe", "description": "Validación automática",
                "planned_start": "2026-08-01", "planned_end": "2026-08-31", "owner": "Consultoría",
                "acceptance_criteria": "Informe aprobado", "notes": "Prueba",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
        with SessionLocal() as session:
            order = session.scalar(select(ServiceOrder).where(ServiceOrder.reference == "OS-TEST-2026-001"))
            order_id = order.id
        response = client.post(
            f"/operacion-comercial/ordenes/{order_id}/estado",
            data={"status": "Aceptada", "notes": "Aceptada por el cliente"},
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        order = session.get(ServiceOrder, order_id)
        assert order.status == "Aceptada"
        assert order.delivered_at is not None
        assert order.accepted_at is not None


def test_v016_recurring_charge_creates_billing_document():
    with SessionLocal() as session:
        subscription = session.scalar(select(OrganizationSubscription).where(OrganizationSubscription.organization_id == 1))
        subscription_id = subscription.id
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            "/operacion-comercial/cobros/recurrente",
            data={
                "subscription_id": str(subscription_id), "period_start": "2027-08-01", "period_end": "2028-07-31",
                "due_date": "2027-08-15", "reference": "REC-TEST-2027-001", "notes": "Renovación anual de prueba",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        invoice = session.scalar(select(BillingInvoice).where(BillingInvoice.reference == "REC-TEST-2027-001"))
        assert invoice is not None
        assert invoice.status == "Pendiente"
        assert invoice.amount > 0
        document = session.scalar(select(BillingDocumentRecord).where(BillingDocumentRecord.invoice_id == invoice.id))
        assert document is not None
        assert document.status == "Pendiente de integración"
        assert document.internal_reference == "DOC-REC-TEST-2027-001"


def test_v016_collection_action_can_close_and_mark_invoice_paid():
    with SessionLocal() as session:
        invoice = session.scalar(select(BillingInvoice).where(BillingInvoice.organization_id == 1).order_by(BillingInvoice.id))
        invoice_id = invoice.id
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            "/operacion-comercial/cartera/nueva",
            data={
                "invoice_id": str(invoice_id), "action_type": "Acuerdo de pago", "channel": "Correo",
                "recipient": "tesoreria@empresa.test", "due_at": "2026-08-10", "notes": "Confirmar transferencia",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
        with SessionLocal() as session:
            action = session.scalar(select(CollectionAction).where(CollectionAction.invoice_id == invoice_id).order_by(CollectionAction.id.desc()))
            action_id = action.id
        response = client.post(
            f"/operacion-comercial/cartera/{action_id}/completar",
            data={"result": "Pago confirmado y conciliado", "invoice_status": "Pagada"},
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        action = session.get(CollectionAction, action_id)
        invoice = session.get(BillingInvoice, invoice_id)
        assert action.status == "Completada"
        assert action.completed_at is not None
        assert invoice.status == "Pagada"
        assert invoice.paid_at is not None


def test_v017_seeded_customer_success_data_exists():
    with SessionLocal() as session:
        profile = session.scalar(select(CustomerSuccessProfile).where(CustomerSuccessProfile.organization_id == 1))
        snapshot = session.scalar(select(AccountHealthSnapshot).where(AccountHealthSnapshot.organization_id == 1))
        renewal = session.scalar(select(RenewalOpportunity).where(RenewalOpportunity.organization_id == 1))
        assert profile is not None
        assert profile.lifecycle_stage == "Adopción"
        assert snapshot is not None
        assert 0 <= snapshot.overall_score <= 100
        assert snapshot.risk_level in {"Sano", "Atención", "Riesgo", "Crítico"}
        assert renewal is not None
        assert 0 <= renewal.probability <= 100
        assert session.scalar(select(func.count()).select_from(ValueMilestone).where(ValueMilestone.organization_id == 1)) >= 3
        assert session.scalar(select(func.count()).select_from(SuccessCommitment).where(SuccessCommitment.organization_id == 1)) >= 3


def test_v017_customer_success_page_is_visible_to_client_but_read_only():
    with TestClient(app) as client:
        login(client, "cliente@calculatuhuella.local")
        page = client.get("/exito-cliente")
        assert page.status_code == 200
        assert "Salud y éxito de la cuenta" in page.text
        denied = client.post("/exito-cliente/salud/recalcular", follow_redirects=False)
        assert denied.status_code == 403


def test_v017_profile_update_and_health_recalculation():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            "/exito-cliente/perfil",
            data={
                "lifecycle_stage": "Valor", "owner": "Consultor principal",
                "executive_sponsor": "Gerencia general", "sponsor_email": "gerencia@empresa.test",
                "primary_objective": "Reducir intensidad de carbono", "success_plan": "Cerrar inventario y priorizar medidas.",
                "risk_override": "", "risk_reason": "", "last_business_review": "2026-07-30",
                "next_business_review": "2026-10-30", "satisfaction_score": "4.5", "nps_score": "9",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
        recalculated = client.post("/exito-cliente/salud/recalcular", follow_redirects=False)
        assert recalculated.status_code == 303
    with SessionLocal() as session:
        profile = session.scalar(select(CustomerSuccessProfile).where(CustomerSuccessProfile.organization_id == 1))
        assert profile.lifecycle_stage == "Valor"
        assert profile.satisfaction_score == 4.5
        assert profile.nps_score == 9
        count = session.scalar(select(func.count()).select_from(AccountHealthSnapshot).where(AccountHealthSnapshot.organization_id == 1))
        assert count >= 2


def test_v017_value_milestone_lifecycle():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            "/exito-cliente/hitos/nuevo",
            data={
                "title": "Ahorro energético validado", "category": "Valor financiero", "inventory_id": "1",
                "owner": "Operaciones", "target_date": "2026-12-15", "expected_value": "25000000",
                "realized_value": "0", "unit": "COP/año", "status": "En progreso",
                "evidence_note": "Validación en curso.",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
        with SessionLocal() as session:
            milestone = session.scalar(select(ValueMilestone).where(ValueMilestone.title == "Ahorro energético validado"))
            milestone_id = milestone.id
        response = client.post(
            f"/exito-cliente/hitos/{milestone_id}/estado",
            data={"status": "Completado", "realized_value": "23000000", "evidence_note": "Ahorro conciliado."},
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        milestone = session.get(ValueMilestone, milestone_id)
        assert milestone.status == "Completado"
        assert milestone.realized_value == 23000000
        assert milestone.completed_at is not None


def test_v017_commitment_lifecycle():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            "/exito-cliente/compromisos/nuevo",
            data={
                "title": "Presentar resultados al comité", "description": "Revisión ejecutiva trimestral",
                "owner": "Dirección ambiental", "due_date": "2026-11-01", "priority": "Alta", "source": "QBR",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
        with SessionLocal() as session:
            commitment = session.scalar(select(SuccessCommitment).where(SuccessCommitment.title == "Presentar resultados al comité"))
            commitment_id = commitment.id
        response = client.post(
            f"/exito-cliente/compromisos/{commitment_id}/estado",
            data={"status": "Completado"}, follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        commitment = session.get(SuccessCommitment, commitment_id)
        assert commitment.status == "Completado"
        assert commitment.completed_at is not None


def test_v017_renewal_strategy_is_editable_and_traced():
    with SessionLocal() as session:
        renewal = session.scalar(select(RenewalOpportunity).where(RenewalOpportunity.organization_id == 1))
        renewal_id = renewal.id
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            f"/exito-cliente/renovacion/{renewal_id}/actualizar",
            data={
                "status": "Propuesta enviada", "probability": "85",
                "strategy": "Renovar y ampliar alcance 3.", "blockers": "Confirmación presupuestal.",
                "next_action": "Reunión con patrocinador ejecutivo", "next_action_date": "2027-05-15",
                "decision_notes": "Cliente reconoce valor del primer ciclo.",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        renewal = session.get(RenewalOpportunity, renewal_id)
        assert renewal.status == "Propuesta enviada"
        assert renewal.probability == 85
        assert "alcance 3" in renewal.strategy



def test_v018_seeded_impact_intelligence_exists():
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(BenchmarkReference).where(BenchmarkReference.organization_id == 1)) >= 4
        snapshot = session.scalar(select(ImpactSnapshot).where(ImpactSnapshot.organization_id == 1))
        assert snapshot is not None
        assert 0 <= snapshot.impact_score <= 100
        assert snapshot.total_emissions >= 0


def test_v018_impact_page_is_visible_to_client_but_read_only():
    with TestClient(app) as client:
        login(client, "cliente@calculatuhuella.local")
        page = client.get("/inteligencia-impacto")
        assert page.status_code == 200
        assert "Inteligencia de impacto" in page.text
        denied = client.post("/inteligencia-impacto/recalcular", follow_redirects=False)
        assert denied.status_code == 403


def test_v018_snapshot_recalculation_is_traced():
    with SessionLocal() as session:
        before = session.scalar(select(func.count()).select_from(ImpactSnapshot).where(ImpactSnapshot.organization_id == 1))
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post("/inteligencia-impacto/recalcular", follow_redirects=False)
        assert response.status_code == 303
    with SessionLocal() as session:
        after = session.scalar(select(func.count()).select_from(ImpactSnapshot).where(ImpactSnapshot.organization_id == 1))
        assert after == before + 1


def test_v018_benchmark_creation_and_comparison():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            "/inteligencia-impacto/benchmarks/nuevo",
            data={
                "name": "Benchmark de prueba", "metric_code": "quality_score", "metric_name": "Calidad de datos",
                "period_label": "2026", "unit": "%", "median_value": "70", "top_quartile_value": "90",
                "lower_is_better": "false", "source_type": "Meta interna", "source_reference": "Comité climático",
                "confidence_level": "Alta", "notes": "Prueba automática",
            }, follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        reference = session.scalar(select(BenchmarkReference).where(BenchmarkReference.name == "Benchmark de prueba"))
        assert reference is not None
        metrics = impact_metrics(session, 1)
        comparison = compare_benchmarks(metrics, [reference])[0]
        assert comparison["status"] in {"Cuartil superior", "Mejor que mediana", "Brecha", "Sin dato"}


def test_v018_benchmark_can_be_archived():
    with SessionLocal() as session:
        reference = session.scalar(select(BenchmarkReference).where(BenchmarkReference.organization_id == 1))
        reference_id = reference.id
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(f"/inteligencia-impacto/benchmarks/{reference_id}/estado", data={"status": "Archivado"}, follow_redirects=False)
        assert response.status_code == 303
    with SessionLocal() as session:
        assert session.get(BenchmarkReference, reference_id).status == "Archivado"


def test_v018_impact_export_is_valid_excel():
    with TestClient(app) as client:
        login(client, "cliente@calculatuhuella.local")
        response = client.get("/inteligencia-impacto/exportar.xlsx")
        assert response.status_code == 200
        assert response.content[:2] == b"PK"


def test_v019_seeded_climate_risk_register_exists():
    with SessionLocal() as session:
        assessment = session.scalar(select(ClimateRiskAssessment).where(ClimateRiskAssessment.organization_id == 1))
        assert assessment is not None
        assert assessment.status == "En tratamiento"
        assert session.scalar(select(func.count()).select_from(ClimateRisk).where(ClimateRisk.organization_id == 1)) >= 5
        assert session.scalar(select(func.count()).select_from(ClimateRiskControl).where(ClimateRiskControl.organization_id == 1)) >= 4
        assert session.scalar(select(func.count()).select_from(ClimateTransitionAction).where(ClimateTransitionAction.organization_id == 1)) >= 4
        summary = assessment_summary(session, 1)
        assert summary["counts"]["physical"] >= 2
        assert summary["counts"]["transition"] >= 2
        assert summary["financial"]["gross_exposure"] > 0
        assert 0 <= summary["readiness_score"] <= 100


def test_v019_climate_risk_page_is_visible_to_client_but_read_only():
    with TestClient(app) as client:
        login(client, "cliente@calculatuhuella.local")
        page = client.get("/riesgos-climaticos")
        assert page.status_code == 200
        assert "Riesgos climáticos y hoja de ruta" in page.text
        denied = client.post(
            "/riesgos-climaticos/riesgos/nuevo",
            data={
                "risk_type": "Físico", "category": "Agudo", "hazard": "Prueba denegada",
                "owner": "Operaciones", "likelihood": "3", "financial_impact": "3",
                "operational_impact": "3", "reputational_impact": "2",
            }, follow_redirects=False,
        )
        assert denied.status_code == 403


def test_v019_risk_creation_calculates_inherent_and_residual_scores():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            "/riesgos-climaticos/riesgos/nuevo",
            data={
                "risk_type": "Transición", "category": "Regulatorio", "hazard": "Costo de carbono de prueba",
                "description": "Escenario de prueba", "location": "Corporativo", "value_chain_stage": "Operación propia",
                "time_horizon": "Mediano plazo", "scenario": "Escenario de prueba", "likelihood": "4",
                "financial_impact": "5", "operational_impact": "3", "reputational_impact": "4",
                "control_effectiveness": "25", "financial_exposure": "100000000", "owner": "Finanzas",
                "response_strategy": "Mitigar", "response_detail": "Plan financiero", "status": "Abierto",
                "source_reference": "Supuesto de prueba",
            }, follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        risk = session.scalar(select(ClimateRisk).where(ClimateRisk.hazard == "Costo de carbono de prueba"))
        assert risk is not None
        assert risk.inherent_score == 20
        assert risk.residual_score == 15
        assert risk_level(risk.residual_score) == "Alto"


def test_v019_control_recalculates_combined_residual_risk():
    with SessionLocal() as session:
        risk = session.scalar(select(ClimateRisk).where(ClimateRisk.organization_id == 1).order_by(ClimateRisk.id))
        risk_id = risk.id
        prior_residual = risk.residual_score
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            "/riesgos-climaticos/controles/nuevo",
            data={
                "risk_id": str(risk_id), "name": "Control complementario de prueba", "control_type": "Preventivo",
                "owner": "Operaciones", "status": "Operando", "effectiveness": "30",
                "implementation_date": "2026-08-01", "next_review": "2027-02-01",
                "annual_cost": "5000000", "evidence": "Evidencia de prueba",
            }, follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        risk = session.get(ClimateRisk, risk_id)
        control = session.scalar(select(ClimateRiskControl).where(ClimateRiskControl.name == "Control complementario de prueba"))
        assert control is not None
        assert risk.control_effectiveness > 45
        assert risk.residual_score < prior_residual


def test_v019_transition_action_lifecycle_is_traced():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            "/riesgos-climaticos/acciones/nueva",
            data={
                "risk_id": "", "category": "Gobierno", "title": "Comité climático mensual",
                "description": "Seguimiento ejecutivo", "owner": "Gerencia", "start_date": "2026-08-01",
                "end_date": "2026-12-31", "priority": "Alta", "status": "Planeada", "progress": "0",
                "expected_reduction_tco2e": "0", "capex": "0", "annual_opex": "12000000",
                "annual_savings": "0", "avoided_loss": "50000000", "indicator": "Sesiones realizadas",
                "target_value": "5", "current_value": "0", "unit": "sesiones", "dependencies": "Agenda de gerencia",
                "evidence_note": "",
            }, follow_redirects=False,
        )
        assert response.status_code == 303
        with SessionLocal() as session:
            action = session.scalar(select(ClimateTransitionAction).where(ClimateTransitionAction.title == "Comité climático mensual"))
            action_id = action.id
        response = client.post(
            f"/riesgos-climaticos/acciones/{action_id}/estado",
            data={"status": "En ejecución", "progress": "40", "current_value": "2", "evidence_note": "Dos sesiones realizadas"},
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        action = session.get(ClimateTransitionAction, action_id)
        assert action.status == "En ejecución"
        assert action.progress == 40
        assert action.current_value == 2


def test_v019_climate_risk_export_is_valid_excel():
    with TestClient(app) as client:
        login(client, "verificador@calculatuhuella.local")
        response = client.get("/riesgos-climaticos/exportar.xlsx")
        assert response.status_code == 200
        assert response.content[:2] == b"PK"
        workbook = load_workbook(BytesIO(response.content))
        assert {"Resumen", "Riesgos", "Controles", "Hoja de ruta"}.issubset(workbook.sheetnames)



def test_v020_seeded_scenarios_disclosure_and_board_pack_exist():
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ClimateScenarioDefinition).where(ClimateScenarioDefinition.organization_id == 1)) >= 3
        comparison = scenario_comparison(session, 1)
        assert comparison["probability_total"] == 100
        assert comparison["weighted_total"] > 0
        assert comparison["worst"] is not None
        disclosure = disclosure_summary(session, 1)
        assert disclosure["statement"] is not None
        assert len(disclosure["requirements"]) >= 10
        assert 0 < disclosure["score"] <= 100
        board = board_summary(session, 1)
        assert board["briefing"] is not None
        assert len(board["decisions"]) >= 3


def test_v020_disclosure_page_is_visible_to_client_but_read_only():
    with TestClient(app) as client:
        login(client, "cliente@calculatuhuella.local")
        page = client.get("/divulgacion-climatica")
        assert page.status_code == 200
        assert "Escenarios, divulgación y comité directivo" in page.text
        denied = client.post(
            "/divulgacion-climatica/escenarios/nuevo",
            data={
                "name": "Escenario denegado", "code": "DEN", "scenario_type": "Combinado",
                "physical_multiplier": "1", "transition_multiplier": "1", "opportunity_multiplier": "1",
                "probability_weight": "0",
            }, follow_redirects=False,
        )
        assert denied.status_code == 403


def test_v020_consultant_can_create_scenario_with_validated_assumptions():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            "/divulgacion-climatica/escenarios/nuevo",
            data={
                "name": "Escenario de prueba", "code": "TST-2040", "scenario_type": "Combinado",
                "temperature_pathway": "2,5 °C", "physical_multiplier": "4.5",
                "transition_multiplier": "1.4", "opportunity_multiplier": "0.7",
                "carbon_price_2030": "250000", "energy_cost_change_pct": "15",
                "demand_change_pct": "-10", "probability_weight": "120",
                "narrative": "Sensibilidad de prueba", "source_reference": "Supuesto de prueba", "status": "Activo",
            }, follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        scenario = session.scalar(select(ClimateScenarioDefinition).where(ClimateScenarioDefinition.code == "TST-2040"))
        assert scenario is not None
        assert scenario.physical_multiplier == 3.0
        assert scenario.probability_weight == 100
        assert scenario_comparison(session, 1)["results"]


def test_v020_disclosure_requirement_update_recalculates_score():
    with SessionLocal() as session:
        requirement = session.scalar(
            select(ClimateDisclosureRequirement)
            .where(ClimateDisclosureRequirement.organization_id == 1, ClimateDisclosureRequirement.status == "Parcial")
            .order_by(ClimateDisclosureRequirement.id)
        )
        requirement_id = requirement.id
        prior_score = disclosure_summary(session, 1)["score"]
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            f"/divulgacion-climatica/requisitos/{requirement_id}/actualizar",
            data={
                "response": "Respuesta completada y revisada", "status": "Completo",
                "evidence_reference": "Acta y soporte de prueba", "owner": "Dirección ambiental",
                "due_date": "2026-12-31",
            }, follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        requirement = session.get(ClimateDisclosureRequirement, requirement_id)
        assert requirement.status == "Completo"
        assert disclosure_summary(session, 1)["score"] >= prior_score


def test_v020_board_decision_lifecycle_is_traced():
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post(
            "/divulgacion-climatica/decisiones/nueva",
            data={
                "topic": "Decisión climática de prueba", "decision": "Aprobar piloto",
                "owner": "Gerencia", "due_date": "2026-11-30", "status": "Pendiente",
                "rationale": "Validar retorno", "evidence_reference": "Caso de negocio",
            }, follow_redirects=False,
        )
        assert response.status_code == 303
        with SessionLocal() as session:
            item = session.scalar(select(ClimateBoardDecision).where(ClimateBoardDecision.topic == "Decisión climática de prueba"))
            item_id = item.id
        response = client.post(
            f"/divulgacion-climatica/decisiones/{item_id}/estado",
            data={
                "decision": "Piloto aprobado", "owner": "Operaciones", "due_date": "2026-12-15",
                "status": "En ejecución", "rationale": "Aprobación condicionada", "evidence_reference": "Acta 01",
            }, follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        item = session.get(ClimateBoardDecision, item_id)
        assert item.status == "En ejecución"
        assert item.owner == "Operaciones"


def test_v020_disclosure_export_is_valid_excel():
    with TestClient(app) as client:
        login(client, "verificador@calculatuhuella.local")
        response = client.get("/divulgacion-climatica/exportar.xlsx")
        assert response.status_code == 200
        assert response.content[:2] == b"PK"
        workbook = load_workbook(BytesIO(response.content))
        assert {"Comparación", "Supuestos", "Divulgación", "Decisiones"}.issubset(workbook.sheetnames)


def test_v020_board_pack_pdf_has_hash_and_valid_pdf_signature():
    with TestClient(app) as client:
        login(client, "revisor@calculatuhuella.local")
        response = client.get("/divulgacion-climatica/comite.pdf")
        assert response.status_code == 200
        assert response.content.startswith(b"%PDF")
        assert len(response.headers["x-document-sha256"]) == 64
    with SessionLocal() as session:
        briefing = session.scalar(select(ClimateBoardBriefing).where(ClimateBoardBriefing.organization_id == 1))
        assert len(briefing.document_hash) == 64


def test_v020_requirement_can_be_created_and_duplicate_code_is_blocked():
    payload = {
        "pillar": "Estrategia", "code": "EST-99", "requirement": "Requisito adicional de prueba",
        "response": "Respuesta inicial", "status": "Parcial", "evidence_reference": "Soporte inicial",
        "owner": "Planeación", "due_date": "2027-01-31",
    }
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        response = client.post("/divulgacion-climatica/requisitos/nuevo", data=payload, follow_redirects=False)
        assert response.status_code == 303
        duplicate = client.post("/divulgacion-climatica/requisitos/nuevo", data=payload, follow_redirects=False)
        assert duplicate.status_code == 409
    with SessionLocal() as session:
        item = session.scalar(select(ClimateDisclosureRequirement).where(ClimateDisclosureRequirement.code == "EST-99"))
        assert item is not None
        assert item.status == "Parcial"


def test_v020_disclosure_approval_requires_approval_capability():
    with SessionLocal() as session:
        statement = session.scalar(select(ClimateDisclosureStatement).where(ClimateDisclosureStatement.organization_id == 1))
        payload = {
            "title": statement.title, "inventory_id": str(statement.inventory_id or ""),
            "framework": statement.framework, "reporting_period": statement.reporting_period,
            "scope_description": statement.scope_description, "materiality_basis": statement.materiality_basis,
            "owner": statement.owner, "status": "Aprobada", "notes": statement.notes,
        }
    with TestClient(app) as client:
        login(client, "consultor@calculatuhuella.local")
        denied = client.post("/divulgacion-climatica/declaracion", data=payload, follow_redirects=False)
        assert denied.status_code == 403
        client.post("/logout")
        login(client, "revisor@calculatuhuella.local")
        approved = client.post("/divulgacion-climatica/declaracion", data=payload, follow_redirects=False)
        assert approved.status_code == 303
    with SessionLocal() as session:
        statement = session.scalar(select(ClimateDisclosureStatement).where(ClimateDisclosureStatement.organization_id == 1))
        assert statement.status == "Aprobada"
        assert statement.approved_by == "revisor@calculatuhuella.local"
