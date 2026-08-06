from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from app.data_quality import (
    apply_import_batch,
    build_data_template,
    create_import_batch,
    data_quality_summary,
)
from app.database import (
    ActivityData,
    Base,
    DataImportBatch,
    DataQualityFinding,
    ENGINE,
    SessionLocal,
    init_db,
)
from app.main import app
from app.pilot_execution import start_pilot_execution


@pytest.fixture(autouse=True)
def fresh_database_v026():
    # La base semilla aislada se restaura desde tests/conftest.py.
    yield


def _login(client: TestClient, email: str = "consultor@calculatuhuella.local") -> None:
    response = client.post("/login", data={"email": email, "password": "Demo2026!"}, follow_redirects=False)
    assert response.status_code == 303


def _workbook_with_values(session, rows: list[tuple[str, float, str, str]]):
    content = build_data_template(session, 1)
    workbook = load_workbook(BytesIO(content))
    sheet = workbook["Carga de datos"]
    for code, value, unit, evidence in rows:
        target = None
        for row_number in range(2, sheet.max_row + 1):
            if sheet.cell(row_number, 1).value == code and sheet.cell(row_number, 4).value in (None, ""):
                target = row_number
                break
        assert target is not None
        sheet.cell(target, 4).value = value
        sheet.cell(target, 5).value = unit
        sheet.cell(target, 8).value = evidence
    # Remove all blank data rows so validation only processes the supplied records.
    for row_number in range(sheet.max_row, 1, -1):
        if sheet.cell(row_number, 4).value in (None, ""):
            sheet.delete_rows(row_number)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_v026_health_version():
    with TestClient(app) as client:
        response = client.get("/api/health")
        assert response.status_code == 200
        assert response.json()["version"] == "1.0.0"


def test_v026_template_contains_controlled_sheets():
    with SessionLocal() as session:
        start_pilot_execution(session, 1, "consultor@test", "Consultor prueba")
        session.commit()
        content = build_data_template(session, 1)
        workbook = load_workbook(BytesIO(content), data_only=True)
        assert {"Carga de datos", "Instrucciones", "Catálogo de fuentes"}.issubset(workbook.sheetnames)
        assert workbook["Carga de datos"].max_row > 20


def test_v026_valid_batch_is_created_without_modifying_inventory():
    with SessionLocal() as session:
        start_pilot_execution(session, 1, "consultor@test", "Consultor prueba")
        session.commit()
        content = _workbook_with_values(session, [("YAR-ELEC", 1000, "kWh", "Factura enero")])
        before = session.scalar(select(func.count(ActivityData.id)))
        batch = create_import_batch(session, 1, "enero.xlsx", content, "consultor@test")
        session.commit()
        after = session.scalar(select(func.count(ActivityData.id)))
        assert batch.status in {"Validado", "Con errores"}
        assert batch.error_rows == 0
        assert before == after
        assert batch.total_rows == 1


def test_v026_negative_value_creates_blocking_finding():
    with SessionLocal() as session:
        start_pilot_execution(session, 1, "consultor@test", "Consultor prueba")
        session.commit()
        content = _workbook_with_values(session, [("YAR-ELEC", -10, "kWh", "Factura")])
        batch = create_import_batch(session, 1, "negativo.xlsx", content, "consultor@test")
        session.commit()
        assert batch.status == "Con errores"
        assert batch.error_rows == 1
        assert session.scalar(select(func.count(DataQualityFinding.id)).where(DataQualityFinding.rule_code == "DQ-005")) == 1
        with pytest.raises(ValueError):
            apply_import_batch(session, 1, batch.id, "consultor@test")


def test_v026_duplicate_file_is_rejected_by_hash():
    with SessionLocal() as session:
        start_pilot_execution(session, 1, "consultor@test", "Consultor prueba")
        session.commit()
        content = _workbook_with_values(session, [("YAR-ELEC", 1000, "kWh", "Factura")])
        create_import_batch(session, 1, "enero.xlsx", content, "consultor@test")
        session.commit()
        with pytest.raises(ValueError, match="ya fue cargado"):
            create_import_batch(session, 1, "copia.xlsx", content, "consultor@test")


def test_v026_apply_batch_creates_activity_data_and_audit_state():
    with SessionLocal() as session:
        start_pilot_execution(session, 1, "consultor@test", "Consultor prueba")
        session.commit()
        content = _workbook_with_values(session, [("YAR-ELEC", 1000, "kWh", "Factura enero")])
        batch = create_import_batch(session, 1, "enero.xlsx", content, "consultor@test")
        session.commit()
        applied = apply_import_batch(session, 1, batch.id, "consultor@test")
        session.commit()
        assert applied.status == "Aplicado"
        assert applied.applied_rows == 1
        assert session.scalar(select(func.count(ActivityData.id))) >= 1
        summary = data_quality_summary(session, 1, batch.id)
        assert summary["metrics"]["applied"] == 1


def test_v026_page_and_download_are_available():
    with TestClient(app) as client:
        _login(client)
        client.post("/piloto-greenatics/ejecucion/iniciar", data={}, follow_redirects=False)
        page = client.get("/calidad-datos")
        assert page.status_code == 200
        assert "Centro de calidad de datos" in page.text
        download = client.get("/calidad-datos/plantilla.xlsx")
        assert download.status_code == 200
        assert download.headers["content-type"].startswith("application/vnd.openxmlformats")


def test_v026_client_can_view_but_cannot_upload():
    with TestClient(app) as client:
        _login(client, "cliente@calculatuhuella.local")
        page = client.get("/calidad-datos")
        assert page.status_code == 200
        response = client.post(
            "/calidad-datos/cargar",
            files={"file": ("datos.xlsx", b"not-an-xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 403
