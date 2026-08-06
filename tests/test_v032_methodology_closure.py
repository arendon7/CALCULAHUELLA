from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook
from fastapi.testclient import TestClient
from sqlalchemy import select

from app.analytics import inventory_total
from app.calculations import recalculate_source
from app.database import (
    ActivityData,
    Base,
    BaseYearRecalculation,
    EmissionSource,
    ENGINE,
    Inventory,
    InventoryMethodologySnapshot,
    SessionLocal,
    init_db,
)
from app.main import app
from app.methodology_closure import calculation_uncertainty, closure_summary
from app.reporting import generate_calculation_workbook


@pytest.fixture(autouse=True)
def fresh_database():
    # La base semilla aislada se restaura desde tests/conftest.py.
    yield


def login(client: TestClient, email: str = "consultor@calculatuhuella.local") -> None:
    response = client.post("/login", data={"email": email, "password": "Demo2026!"}, follow_redirects=False)
    assert response.status_code == 303


def test_health_and_closure_page_load() -> None:
    with TestClient(app) as client:
        login(client)
        assert client.get("/api/health").json()["version"] == "1.0.0"
        page = client.get("/metodologia/cierre")
        assert page.status_code == 200
        assert "Cierre metodológico" in page.text
        assert "Emisiones brutas" in page.text
        assert "Evaluaciones de recalculo" in page.text


def test_policy_can_be_approved_and_is_versioned() -> None:
    with TestClient(app) as client:
        login(client, "admin@calculatuhuella.local")
        response = client.post(
            "/metodologia/cierre/politica",
            data={
                "base_year_recalculation_threshold": "4.5",
                "base_year_triggers": "Cambios estructurales y errores materiales",
                "biogenic_co2_policy": "Separado",
                "removals_policy": "Separadas",
                "avoided_emissions_policy": "Fuera del inventario",
                "offsets_policy": "Fuera del bruto",
                "scope2_policy": "Reporte dual",
                "uncertainty_method": "RSS Approach 1",
                "action": "approve",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        snapshot = session.scalar(select(InventoryMethodologySnapshot).where(InventoryMethodologySnapshot.snapshot_name == "Cierre metodológico V0.32"))
        assert snapshot is not None
        assert snapshot.status == "Aprobado"
        assert snapshot.approved_by == "admin@calculatuhuella.local"


def test_source_treatment_separates_gross_inventory() -> None:
    with SessionLocal() as session:
        inventory = session.get(Inventory, 1)
        source = inventory.sources[0]
        baseline = inventory_total(inventory)
        source.accounting_treatment = "Emisión evitada"
        session.commit()
        assert inventory_total(inventory) == pytest.approx(baseline - source.emissions)


def test_scope2_source_can_be_classified_location_based() -> None:
    with SessionLocal() as session:
        source = session.scalar(select(EmissionSource).where(EmissionSource.scope == 2))
        source_id = source.id
    with TestClient(app) as client:
        login(client, "admin@calculatuhuella.local")
        response = client.post(
            f"/metodologia/cierre/fuentes/{source_id}",
            data={"accounting_treatment": "Emisión bruta", "scope2_method": "Location-based", "biogenic_origin": "No aplica"},
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        source = session.get(EmissionSource, source_id)
        assert source.scope2_method == "Location-based"


def test_uncertainty_is_combined_in_calculation_engine() -> None:
    assert calculation_uncertainty(3, 4) == pytest.approx(5)
    with SessionLocal() as session:
        source = session.scalar(select(EmissionSource).where(EmissionSource.name == "Electricidad"))
        record = session.scalar(select(ActivityData).where(ActivityData.source_id == source.id))
        record.uncertainty_percentage = 10
        for assignment in source.factor_assignments:
            assignment.factor_version.uncertainty_percentage = 5
        session.flush()
        recalculate_source(session, source)
        session.commit()
        session.refresh(record)
        calc = record.calculations[0]
        assert calc.uncertainty_percentage == pytest.approx((10**2 + 5**2) ** 0.5)
        assert calc.lower_co2e_kg < calc.co2e_kg < calc.upper_co2e_kg
        assert calc.reporting_bucket == "Emisión bruta"


def test_base_year_recalculation_uses_threshold_and_review() -> None:
    with TestClient(app) as client:
        login(client, "admin@calculatuhuella.local")
        created = client.post(
            "/metodologia/cierre/recalculos",
            data={
                "event_date": date.today().isoformat(),
                "trigger_type": "Cambio estructural",
                "description": "Adquisición de una instalación",
                "previous_total_tco2e": "100",
                "recalculated_total_tco2e": "108",
                "threshold_percentage": "5",
            },
            follow_redirects=False,
        )
        assert created.status_code == 303
    with SessionLocal() as session:
        item = session.scalar(select(BaseYearRecalculation))
        assert item.decision == "Recalcular"
        item_id = item.id
    with TestClient(app) as client:
        login(client, "revisor@calculatuhuella.local")
        reviewed = client.post(
            f"/metodologia/cierre/recalculos/{item_id}/revisar",
            data={"status": "Aprobado", "decision": "Recalcular"},
            follow_redirects=False,
        )
        assert reviewed.status_code == 303
    with SessionLocal() as session:
        item = session.get(BaseYearRecalculation, item_id)
        assert item.status == "Aprobado"
        assert item.reviewed_by == "revisor@calculatuhuella.local"


def test_closure_api_returns_separate_balance() -> None:
    with TestClient(app) as client:
        login(client)
        response = client.get("/api/metodologia/cierre")
        assert response.status_code == 200
        payload = response.json()
        assert "gross_emissions" in payload["balance"]
        assert "combined_percentage" in payload["uncertainty"]
        assert payload["scope2"]["unclassified"] >= 0


def test_uncertainty_reports_partial_coverage_explicitly() -> None:
    with SessionLocal() as session:
        inventory = session.get(Inventory, 1)
        summary = closure_summary(session, inventory)["uncertainty"]
        assert summary["total_gross_tco2e"] >= summary["covered_tco2e"]
        assert 0 <= summary["emission_coverage_percentage"] <= 100
        if summary["total_gross_tco2e"] > summary["covered_tco2e"]:
            assert summary["complete"] is False
            assert summary["uncovered_sources"]


def test_source_reclassification_recalculates_reporting_bucket() -> None:
    with SessionLocal() as session:
        source = session.scalar(select(EmissionSource).where(EmissionSource.name == "Diésel"))
        source_id = source.id
    with TestClient(app) as client:
        login(client, "admin@calculatuhuella.local")
        response = client.post(
            f"/metodologia/cierre/fuentes/{source_id}",
            data={"accounting_treatment": "Emisión evitada", "scope2_method": "No aplica", "biogenic_origin": "No aplica"},
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        source = session.get(EmissionSource, source_id)
        buckets = {calc.reporting_bucket for record in source.activity_records for calc in record.calculations}
        assert buckets == {"Emisión evitada"}


def test_calculation_workbook_contains_methodological_closure(tmp_path: Path) -> None:
    output = tmp_path / "memoria_v032.xlsx"
    with SessionLocal() as session:
        inventory = session.get(Inventory, 1)
        generate_calculation_workbook(session, inventory, output)
    workbook = load_workbook(output, read_only=True)
    assert "Cierre metodológico" in workbook.sheetnames
    headers = [cell.value for cell in next(workbook["Cálculos"].iter_rows(min_row=1, max_row=1))]
    assert "Partida" in headers
    assert "Incertidumbre %" in headers
