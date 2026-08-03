from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app.database import (
    Base,
    EmissionFactor,
    EmissionFactorVersion,
    FactorDocumentation,
    Gas,
    GWPValue,
    ENGINE,
    MethodologySourceDocument,
    MethodologyValidationRun,
    ReferenceCalculationCase,
    SessionLocal,
    init_db,
)
from app.main import app
from app.methodology_core import run_reference_suite, select_factor_candidates


@pytest.fixture(autouse=True)
def fresh_database_v022():
    Base.metadata.drop_all(ENGINE)
    init_db()
    yield


def login(client: TestClient, email: str = "consultor@calculatuhuella.local") -> None:
    response = client.post("/login", data={"email": email, "password": "Demo2026!"}, follow_redirects=False)
    assert response.status_code == 303


def test_v022_methodology_defaults_and_sources_are_seeded():
    with SessionLocal() as session:
        documents = list(session.scalars(select(MethodologySourceDocument)))
        cases = list(session.scalars(select(ReferenceCalculationCase)))
        assert len(documents) >= 13
        assert len(cases) >= 12
        assert {item.code for item in documents} >= {
            "GHGP-CORP-2004",
            "GHGP-S2-2015",
            "GHGP-S3-2011",
            "GHGP-GWP-2024",
            "IPCC-2006-GL",
            "IPCC-2019-RF",
            "UPME-R085-2026",
        }


def test_v022_official_upme_factor_is_formal_and_traceable():
    with SessionLocal() as session:
        version = session.scalar(
            select(EmissionFactorVersion)
            .join(EmissionFactor)
            .where(EmissionFactor.name == "Electricidad SIN Colombia · inventarios 2024")
        )
        assert version is not None
        documentation = session.scalar(
            select(FactorDocumentation).where(FactorDocumentation.factor_version_id == version.id)
        )
        assert version.value == pytest.approx(0.220)
        assert version.input_unit == "kWh"
        assert version.output_unit == "kg CO2e"
        assert documentation is not None
        assert documentation.factor_kind == "Oficial nacional"
        assert documentation.reporting_use == "Formal"
        assert documentation.quality_grade == "A"
        assert documentation.data_year == 2024
        assert documentation.review_status == "Aprobado documentalmente"
        assert documentation.source_document.code == "UPME-R085-2026"


def test_v022_ar6_gwp_values_distinguish_methane_origin():
    with SessionLocal() as session:
        values = {
            (gas.code, gwp.assessment): gwp.value
            for gas, gwp in session.execute(
                select(Gas, GWPValue).join(GWPValue, GWPValue.gas_id == Gas.id)
            ).all()
        }
        assert values[("CH4", "AR6")] == pytest.approx(27.0)
        assert values[("CH4-FOSSIL", "AR6")] == pytest.approx(29.8)
        assert values[("N2O", "AR6")] == pytest.approx(273.0)
        assert values[("HFC-134a", "AR6")] == pytest.approx(1530.0)
        assert values[("SF6", "AR6")] == pytest.approx(24300.0)


def test_v022_reference_suite_passes_all_active_cases():
    with SessionLocal() as session:
        run = run_reference_suite(session, "prueba@calculatuhuella.local")
        session.commit()
        assert run.engine_version == "0.45.0"
        assert run.total_cases >= 12
        assert run.passed_cases == run.total_cases
        assert run.failed_cases == 0
        assert run.status == "Aprobado"
        assert all(result.passed for result in run.results)


def test_v022_page_api_and_role_access():
    with TestClient(app) as client:
        login(client)
        page = client.get("/metodologia/nucleo")
        assert page.status_code == 200
        assert "Núcleo metodológico" in page.text
        api = client.get("/api/metodologia/nucleo")
        assert api.status_code == 200
        assert api.json()["engine_version"] == "0.45.0"
        assert api.json()["metrics"]["formal_factors"] >= 1
        client.post("/logout")

        login(client, "cliente@calculatuhuella.local")
        assert client.get("/metodologia/nucleo").status_code == 403
        client.post("/logout")

        login(client, "verificador@calculatuhuella.local")
        assert client.get("/metodologia/nucleo").status_code == 200
        assert client.post("/metodologia/nucleo/validar", follow_redirects=False).status_code == 403


def test_v022_consultant_can_execute_validation_from_web():
    with TestClient(app) as client:
        login(client)
        with SessionLocal() as session:
            before = len(list(session.scalars(select(MethodologyValidationRun))))
        response = client.post("/metodologia/nucleo/validar", follow_redirects=False)
        assert response.status_code == 303
        with SessionLocal() as session:
            runs = list(session.scalars(select(MethodologyValidationRun)))
            assert len(runs) == before + 1
            latest = max(runs, key=lambda item: item.id)
            assert latest.status == "Aprobado"
            assert latest.executed_by == "consultor@calculatuhuella.local"


def test_v022_reviewer_can_update_factor_documentation():
    with SessionLocal() as session:
        demo_documentation = session.scalar(
            select(FactorDocumentation).where(FactorDocumentation.reporting_use == "Demostrativo")
        )
        assert demo_documentation is not None
        documentation_id = demo_documentation.id
    with TestClient(app) as client:
        login(client, "revisor@calculatuhuella.local")
        response = client.post(
            f"/metodologia/nucleo/factores/{documentation_id}/revisar",
            data={
                "review_status": "En revisión",
                "quality_grade": "C",
                "reporting_use": "Piloto",
                "restriction_notes": "Uso restringido al piloto interno.",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        documentation = session.get(FactorDocumentation, documentation_id)
        assert documentation.review_status == "En revisión"
        assert documentation.reporting_use == "Piloto"
        assert documentation.quality_grade == "C"
        assert documentation.reviewer == "revisor@calculatuhuella.local"
        assert documentation.reviewed_at is not None


def test_v022_methodology_excel_contains_expected_sheets():
    with TestClient(app) as client:
        login(client)
        response = client.get("/metodologia/nucleo/exportar.xlsx")
        assert response.status_code == 200
        workbook = load_workbook(BytesIO(response.content))
        assert {"Fuentes", "Factores", "GWP", "Reglas", "Casos patrón", "Última validación"}.issubset(workbook.sheetnames)
        assert workbook["Factores"]["A1"].value == "Factor"


def test_v022_factor_selection_prioritizes_formal_over_demo():
    with SessionLocal() as session:
        candidates = select_factor_candidates(
            session,
            activity_type="Electricidad adquirida",
            country="Colombia",
            input_unit="kWh",
            data_year=2024,
        )
        assert len(candidates) >= 2
        assert candidates[0]["version"].factor.name == "Electricidad SIN Colombia · inventarios 2024"
        assert candidates[0]["documentation"].reporting_use == "Formal"
        assert candidates[0]["score"] > candidates[-1]["score"]
