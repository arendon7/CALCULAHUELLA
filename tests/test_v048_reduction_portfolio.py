from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy import select

from app.database import Base, ENGINE, Inventory, ReductionAction, SessionLocal, init_db
from app.delivery_readiness import professional_delivery_summary
from app.main import app
from app.reduction_portfolio import action_readiness, build_portfolio_workbook, portfolio_summary
from app.reporting import generate_technical_pdf


@pytest.fixture(autouse=True)
def fresh_database():
    # La base semilla aislada se restaura desde tests/conftest.py.
    yield


def login(client: TestClient) -> None:
    response = client.post(
        "/login",
        data={"email": "consultor@calculatuhuella.local", "password": "Demo2026!"},
        follow_redirects=False,
    )
    assert response.status_code == 303


def test_v048_portfolio_quantifies_target_gap_and_decision_classes():
    with SessionLocal() as session:
        inventory = session.scalar(select(Inventory).where(Inventory.id == 1))
        summary = portfolio_summary(session, inventory, today=date(2026, 8, 4))
        assert summary["baseline"] > 0
        assert summary["required_reduction"] > summary["expected_reduction"] > 0
        assert 50 <= summary["coverage_percent"] <= 55
        assert summary["gap"] > 0
        assert summary["portfolio_status"] == "Brecha material"
        assert summary["decision_ready"] is False
        classes = {item["classification"] for item in summary["actions"]}
        assert "Ganancia rápida" in classes
        assert "Apuesta estratégica" in classes
        assert len(summary["overdue"]) == 1
        assert all(item["readiness_score"] >= 0 for item in summary["actions"])


def test_v048_action_readiness_explains_missing_definition():
    action = ReductionAction(
        inventory_id=1,
        title="Medida sin estructurar",
        description="",
        expected_reduction=0,
        investment_cost=0,
        annual_savings=0,
        priority="Media",
        responsible="",
        status="Identificada",
        progress_percent=0,
        useful_life_years=5,
        feasibility="Media",
        risk_level="Medio",
        created_by="test@example.com",
    )
    result = action_readiness(action, today=date(2026, 8, 4))
    assert result["score"] < 65
    assert result["level"] == "Información insuficiente"
    assert "Impacto cuantificado" in result["missing"]
    assert "Responsable asignado" in result["missing"]
    assert "Fecha objetivo" in result["missing"]


def test_v048_page_api_and_export_present_directed_portfolio():
    with TestClient(app) as client:
        login(client)
        page = client.get("/reduccion")
        assert page.status_code == 200
        assert "PORTAFOLIO PRIORIZADO" in page.text
        assert "COBERTURA DEL PORTAFOLIO" in page.text
        assert "DECISIÓN PRINCIPAL" in page.text
        assert "Exportar control" in page.text

        api = client.get("/api/reduccion/resumen")
        assert api.status_code == 200
        payload = api.json()
        assert payload["coverage_percent"] > 0
        assert payload["gap"] > 0
        assert len(payload["actions"]) == 3
        assert all("readiness_score" in item and "classification" in item for item in payload["actions"])

        exported = client.get("/reduccion/exportar.xlsx")
        assert exported.status_code == 200
        workbook = load_workbook(BytesIO(exported.content), read_only=True)
        assert workbook.sheetnames == ["Dirección", "Acciones", "Metas", "Trayectoria", "Responsables"]
        assert workbook["Acciones"]["A1"].value == "Clasificación"
        assert workbook["Dirección"]["A15"].value == "Estado"


def test_v048_delivery_gate_uses_portfolio_coverage_and_readiness():
    with SessionLocal() as session:
        inventory = session.scalar(select(Inventory).where(Inventory.id == 1))
        delivery = professional_delivery_summary(session, inventory)
        gate = next(item for item in delivery["gates"] if item["code"] == "delivery")
        assert gate["status"] == "En progreso"
        assert "52.7% de cobertura" in gate["detail"]
        assert delivery["metrics"]["reduction_coverage"] > 0
        assert delivery["metrics"]["reduction_readiness"] == 100
        assert delivery["metrics"]["reduction_gap"] > 0


def test_v048_workbook_and_technical_report_include_portfolio_control(tmp_path: Path):
    with SessionLocal() as session:
        inventory = session.scalar(select(Inventory).where(Inventory.id == 1))
        summary = portfolio_summary(session, inventory, today=date(2026, 8, 4))
        payload = build_portfolio_workbook(inventory, summary)
        workbook = load_workbook(BytesIO(payload), read_only=True)
        assert workbook["Dirección"]["B15"].value == "Brecha material"
        assert workbook["Acciones"].max_row == 4

        output = tmp_path / "informe_tecnico.pdf"
        generate_technical_pdf(session, inventory, output)
        text = "\n".join(page.extract_text() or "" for page in PdfReader(output).pages)
        assert "Portafolio de reducción y abatimiento" in text
        assert "Reducción requerida" in text
        assert "Cobertura" in text or "cobertura" in text
