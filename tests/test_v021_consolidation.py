from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app.access_control import ROLE_CAPABILITIES, permission_matrix
from app.database import (
    Base,
    ConsolidationFinding,
    ENGINE,
    JourneyValidation,
    ReleaseGate,
    SessionLocal,
    init_db,
)
from app.main import app
from app.product_registry import PRODUCT_MODULES, ROLE_JOURNEYS


@pytest.fixture(autouse=True)
def fresh_database_v021():
    # La base semilla aislada se restaura desde tests/conftest.py.
    yield


def login(client: TestClient, email: str = "consultor@calculatuhuella.local") -> None:
    response = client.post("/login", data={"email": email, "password": "Demo2026!"}, follow_redirects=False)
    assert response.status_code == 303


def test_v021_defaults_are_seeded_for_consolidation():
    with SessionLocal() as session:
        findings = list(session.scalars(select(ConsolidationFinding).where(ConsolidationFinding.organization_id == 1)))
        gates = list(session.scalars(select(ReleaseGate).where(ReleaseGate.organization_id == 1)))
        journeys = list(session.scalars(select(JourneyValidation).where(JourneyValidation.organization_id == 1)))
        assert len(findings) >= 13
        assert len(gates) == 9
        assert len(journeys) == len(ROLE_JOURNEYS)
        assert any(item.priority == "Crítica" for item in findings)
        assert all(item.status == "Resuelto" for item in findings)


def test_v021_consolidation_page_and_api_load():
    with TestClient(app) as client:
        login(client)
        response = client.get("/consolidacion")
        assert response.status_code == 200
        assert "V1.0 final" in response.text
        api = client.get("/api/consolidacion/resumen")
        assert api.status_code == 200
        payload = api.json()
        assert payload["module_count"] == len(PRODUCT_MODULES)
        assert payload["critical_open"] == 0
        assert payload["release_candidate"]["controlled_release_ready"] is True
        assert payload["release_candidate"]["production_ready"] is False
        assert payload["metrics"]["routes"] >= 190


def test_v021_client_cannot_view_internal_consolidation():
    with TestClient(app) as client:
        login(client, "cliente@calculatuhuella.local")
        assert client.get("/consolidacion").status_code == 403


def test_v021_verifier_has_read_only_consolidation_access():
    with TestClient(app) as client:
        login(client, "verificador@calculatuhuella.local")
        assert client.get("/consolidacion").status_code == 200
        with SessionLocal() as session:
            finding_id = session.scalar(select(ConsolidationFinding.id).where(ConsolidationFinding.organization_id == 1))
        response = client.post(
            f"/consolidacion/hallazgos/{finding_id}",
            data={"status": "Resuelto", "owner": "Verificador", "target_version": "V1.0", "evidence": "Intento"},
        )
        assert response.status_code == 403


def test_v021_consultant_can_update_finding_gate_and_journey():
    with TestClient(app) as client:
        login(client)
        with SessionLocal() as session:
            finding = session.scalar(select(ConsolidationFinding).where(ConsolidationFinding.code == "UX-001", ConsolidationFinding.organization_id == 1))
            gate = session.scalar(select(ReleaseGate).where(ReleaseGate.code == "GATE-UX", ReleaseGate.organization_id == 1))
            journey = session.scalar(select(JourneyValidation).where(JourneyValidation.journey_code == "JRN-CONSULTOR", JourneyValidation.organization_id == 1))
            finding_id, gate_id, journey_id = finding.id, gate.id, journey.id
        assert client.post(
            f"/consolidacion/hallazgos/{finding_id}",
            data={"status": "En curso", "owner": "Producto", "target_version": "V0.23", "evidence": "Sesión 1 ejecutada"},
            follow_redirects=False,
        ).status_code == 303
        assert client.post(
            f"/consolidacion/puertas/{gate_id}",
            data={"status": "En revisión", "responsible": "Producto", "evidence": "Pruebas E2E", "notes": "Pendiente cliente"},
            follow_redirects=False,
        ).status_code == 303
        assert client.post(
            f"/consolidacion/recorridos/{journey_id}",
            data={"status": "Aprobado", "notes": "Flujo completado sin bloqueos"},
            follow_redirects=False,
        ).status_code == 303
    with SessionLocal() as session:
        finding = session.get(ConsolidationFinding, finding_id)
        gate = session.get(ReleaseGate, gate_id)
        journey = session.get(JourneyValidation, journey_id)
        assert finding.evidence == "Sesión 1 ejecutada"
        assert gate.status == "En revisión"
        assert journey.status == "Aprobado"
        assert journey.tested_at is not None


def test_v021_export_contains_governance_sheets():
    with TestClient(app) as client:
        login(client)
        response = client.get("/consolidacion/exportar.xlsx")
        assert response.status_code == 200
        workbook = load_workbook(BytesIO(response.content))
        assert {"Resumen", "Deuda y hallazgos", "Puertas V1", "Recorridos", "Permisos", "Arquitectura"}.issubset(workbook.sheetnames)
        assert workbook["Resumen"]["A1"].value == "Indicador"


def test_v021_access_policy_is_centralized_and_complete():
    assert "manage_consolidation" in ROLE_CAPABILITIES["Administrador"]
    assert "manage_consolidation" in ROLE_CAPABILITIES["Consultor"]
    assert "view_consolidation" in ROLE_CAPABILITIES["Verificador"]
    assert "view_consolidation" not in ROLE_CAPABILITIES["Cliente"]
    matrix = permission_matrix()
    assert any(row["capability"] == "manage_consolidation" for row in matrix)


def test_v021_product_registry_has_three_layers_and_no_duplicate_routes():
    layers = {str(module["layer"]) for module in PRODUCT_MODULES}
    assert layers == {"Núcleo", "Avanzado", "Administración interna"}
    routes = [str(module["route"]) for module in PRODUCT_MODULES]
    assert len(routes) == len(set(routes))
    assert any(module["version"] == "V0.21" for module in PRODUCT_MODULES)
