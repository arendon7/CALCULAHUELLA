from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app.database import (
    Base,
    EmissionFactor,
    EmissionFactorVersion,
    ENGINE,
    FactorDocumentation,
    MethodologySourceDocument,
    PilotProject,
    PilotSourceRequirement,
    ReferenceCalculationCase,
    SessionLocal,
    init_db,
)
from app.main import app
from app.methodology_core import run_reference_suite
from app.sector_library import pilot_summary


@pytest.fixture(autouse=True)
def fresh_database_v023():
    # La base semilla aislada se restaura desde tests/conftest.py.
    yield


def login(client: TestClient, email: str = "consultor@calculatuhuella.local") -> None:
    response = client.post("/login", data={"email": email, "password": "Demo2026!"}, follow_redirects=False)
    assert response.status_code == 303


def test_v023_sources_and_sector_factors_are_seeded():
    with SessionLocal() as session:
        codes = set(session.scalars(select(MethodologySourceDocument.code)))
        assert {
            "IPCC-WASTE-2006-CH4-N2O",
            "IPCC-WASTE-CORRIGENDA",
            "IPCC-WASTEWATER-2019",
            "UPME-FECOC-2016",
            "UPME-FECOCPLUS-3-2023",
        }.issubset(codes)
        names = set(session.scalars(select(EmissionFactor.name)))
        assert {
            "Compostaje de residuos orgánicos húmedos · CH4 Tier 1",
            "Compostaje de residuos orgánicos húmedos · N2O Tier 1",
            "Digestión anaerobia en instalación de biogás · CH4 Tier 1",
            "Liberación directa de HFC-134a · balance de masa",
        }.issubset(names)


def test_v023_ipcc_factor_values_and_documentation():
    with SessionLocal() as session:
        compost_ch4 = session.scalar(
            select(EmissionFactorVersion)
            .join(EmissionFactor)
            .where(EmissionFactor.name == "Compostaje de residuos orgánicos húmedos · CH4 Tier 1")
        )
        compost_n2o = session.scalar(
            select(EmissionFactorVersion)
            .join(EmissionFactor)
            .where(EmissionFactor.name == "Compostaje de residuos orgánicos húmedos · N2O Tier 1")
        )
        digestion = session.scalar(
            select(EmissionFactorVersion)
            .join(EmissionFactor)
            .where(EmissionFactor.name == "Digestión anaerobia en instalación de biogás · CH4 Tier 1")
        )
        assert compost_ch4.value == pytest.approx(4.0)
        assert compost_n2o.value == pytest.approx(0.24)
        assert digestion.value == pytest.approx(0.8)
        for version in (compost_ch4, compost_n2o, digestion):
            doc = session.scalar(select(FactorDocumentation).where(FactorDocumentation.factor_version_id == version.id))
            assert doc is not None
            assert doc.reporting_use == "Formal"
            assert doc.review_status == "Aprobado documentalmente"
            assert doc.source_document_id is not None
            assert doc.restriction_notes


def test_v023_reference_suite_has_twelve_passing_cases():
    with SessionLocal() as session:
        cases = list(session.scalars(select(ReferenceCalculationCase).where(ReferenceCalculationCase.active.is_(True))))
        assert len(cases) >= 12
        run = run_reference_suite(session, "prueba-v023")
        session.commit()
        assert run.engine_version == "1.1.0"
        assert run.total_cases >= 12
        assert run.passed_cases == run.total_cases
        assert run.failed_cases == 0


def test_v023_greenatics_pilot_is_seeded_by_site():
    with SessionLocal() as session:
        pilot = session.scalar(select(PilotProject).where(PilotProject.code == "GREENATICS-2026"))
        assert pilot is not None
        requirements = list(session.scalars(select(PilotSourceRequirement).where(PilotSourceRequirement.pilot_id == pilot.id)))
        assert len(requirements) >= 20
        assert {item.site for item in requirements} == {"Yarumal", "Támesis", "Corporativo"}
        assert any(item.code == "YAR-COMP" and item.factor_status == "Formal condicionado" for item in requirements)
        assert any(item.code == "TAM-AD" and "IPCC" in item.factor_reference for item in requirements)


def test_v023_pilot_page_permissions_update_and_export():
    with TestClient(app) as client:
        login(client)
        page = client.get("/piloto-greenatics")
        assert page.status_code == 200
        assert "Piloto Greenatics 2026" in page.text
        export = client.get("/piloto-greenatics/exportar.xlsx")
        assert export.status_code == 200
        workbook = load_workbook(BytesIO(export.content))
        assert {"Plan piloto", "Fuentes requeridas", "Cobertura metodológica", "Plantilla de datos"}.issubset(workbook.sheetnames)
        with SessionLocal() as session:
            item = session.scalar(select(PilotSourceRequirement).where(PilotSourceRequirement.code == "YAR-ELEC"))
            item_id = item.id
        response = client.post(
            f"/piloto-greenatics/fuentes/{item_id}",
            data={"status": "Disponible", "data_owner": "Administración Yarumal", "notes": "Facturas localizadas."},
            follow_redirects=False,
        )
        assert response.status_code == 303
        with SessionLocal() as session:
            item = session.get(PilotSourceRequirement, item_id)
            assert item.status == "Disponible"
            assert item.data_owner == "Administración Yarumal"

        client.post("/logout")
        login(client, "cliente@calculatuhuella.local")
        assert client.get("/piloto-greenatics").status_code == 200


def test_v023_pilot_readiness_summary_is_explainable():
    with SessionLocal() as session:
        pilot = session.scalar(select(PilotProject).where(PilotProject.code == "GREENATICS-2026"))
        summary = pilot_summary(session, pilot.organization_id)
        assert summary["metrics"]["requirements"] >= 20
        assert 0 <= summary["metrics"]["readiness_score"] <= 100
        assert summary["metrics"]["factor_score"] > 0
        assert summary["metrics"]["high_pending"] > 0
