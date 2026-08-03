from io import BytesIO
from datetime import date

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import inspect, select

from app.database import (
    ActivityData,
    Base,
    ENGINE,
    PilotSourceComparison,
    SessionLocal,
    init_db,
)
from app.main import app
from app.pilot_execution import (
    build_pilot_execution_workbook,
    import_pilot_comparison_workbook,
    pilot_execution_summary,
    start_pilot_execution,
    update_pilot_source_comparison,
)


@pytest.fixture(autouse=True)
def fresh_database_v033():
    Base.metadata.drop_all(ENGINE)
    init_db()
    yield


def login(client: TestClient, email: str = "consultor@calculatuhuella.local") -> None:
    response = client.post("/login", data={"email": email, "password": "Demo2026!"}, follow_redirects=False)
    assert response.status_code == 303


def test_v033_health_and_schema_are_current():
    with TestClient(app) as client:
        assert client.get("/api/health").json()["version"] == "0.45.5"
    assert "pilot_source_comparisons" in inspect(ENGINE).get_table_names()


def test_v033_summary_creates_source_level_comparisons_and_site_metrics():
    with SessionLocal() as session:
        execution = start_pilot_execution(session, 1, "consultor@test", "Consultor")
        session.commit()
        summary = pilot_execution_summary(session, 1)
        session.commit()
        assert summary["metrics"]["source_count"] >= 20
        assert len(summary["comparisons"]) == summary["metrics"]["source_count"]
        assert {"Yarumal", "Támesis", "Corporativo"}.issubset(summary["site_metrics"])
        assert execution.inventory.version == "0.45"


def test_v033_monthly_coverage_uses_real_activity_periods():
    with SessionLocal() as session:
        start_pilot_execution(session, 1, "consultor@test", "Consultor")
        summary = pilot_execution_summary(session, 1)
        target = next(item for item in summary["source_controls"] if item["link"].requirement.code == "YAR-ELEC")
        session.add(ActivityData(
            source_id=target["link"].source_id,
            period_start=date(2026, 1, 1),
            period_end=date(2026, 1, 31),
            value=1000,
            unit="kWh",
            data_origin="Prueba V0.33",
            quality_level="A",
            created_by="consultor@test",
        ))
        session.commit()
        refreshed = pilot_execution_summary(session, 1)
        january = refreshed["monthly_coverage"][0]["sites"]["Yarumal"]
        assert january["present"] >= 1
        assert january["expected"] >= january["present"]


def test_v033_workbook_includes_reconciliation_and_context_sheets():
    with SessionLocal() as session:
        start_pilot_execution(session, 1, "consultor@test", "Consultor")
        summary = pilot_execution_summary(session, 1)
        session.commit()
        workbook = load_workbook(BytesIO(build_pilot_execution_workbook(summary)))
        assert {"Contraste por fuente", "Cobertura mensual", "Contexto operativo"}.issubset(workbook.sheetnames)
        contrast = workbook["Contraste por fuente"]
        assert contrast.max_row == summary["metrics"]["source_count"] + 1


def test_v033_imports_source_level_comparison_and_calculates_variance():
    with SessionLocal() as session:
        start_pilot_execution(session, 1, "consultor@test", "Consultor")
        summary = pilot_execution_summary(session, 1)
        workbook = load_workbook(BytesIO(build_pilot_execution_workbook(summary)))
        sheet = workbook["Contraste por fuente"]
        sheet.cell(2, 5).value = sheet.cell(2, 4).value
        sheet.cell(2, 9).value = "Coincidencia independiente"
        stream = BytesIO(); workbook.save(stream)
        result = import_pilot_comparison_workbook(session, 1, stream.getvalue(), "revisor@test")
        session.commit()
        assert result["updated"] == 1
        assert not result["errors"]
        comparison = session.scalar(select(PilotSourceComparison).where(PilotSourceComparison.reviewed_by == "revisor@test"))
        assert comparison.status == "Conforme"
        assert comparison.variance_percent == 0


def test_v033_manual_source_comparison_is_auditable():
    with SessionLocal() as session:
        start_pilot_execution(session, 1, "consultor@test", "Consultor")
        summary = pilot_execution_summary(session, 1)
        comparison = summary["comparisons"][0]
        updated = update_pilot_source_comparison(session, 1, comparison.id, comparison.platform_tco2e + 1, "Diferencia revisada", "revisor@test")
        session.commit()
        assert updated.reviewed_by == "revisor@test"
        assert updated.status in {"Conforme", "Revisar"}
        assert updated.notes == "Diferencia revisada"


def test_v033_approval_requires_evidence_and_full_source_comparison():
    with SessionLocal() as session:
        start_pilot_execution(session, 1, "consultor@test", "Consultor")
        summary = pilot_execution_summary(session, 1)
        assert any("evidencia" in item.lower() for item in summary["blockers"])
        assert any("contraste independiente por fuente" in item.lower() for item in summary["blockers"])


def test_v033_control_tower_is_visible_in_web_page():
    with TestClient(app) as client:
        login(client)
        client.post("/piloto-greenatics/ejecucion/iniciar", data={}, follow_redirects=False)
        page = client.get("/piloto-greenatics/ejecucion")
        assert page.status_code == 200
        assert "CONTROL POR SEDE" in page.text
        assert "COBERTURA MENSUAL" in page.text
        assert "Diferencias fuente a fuente" in page.text
        assert "Contexto operativo conocido" in page.text
