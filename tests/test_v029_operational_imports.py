from __future__ import annotations

from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select

from app.database import (
    ActivityData,
    Base,
    DataImportBatch,
    ENGINE,
    EmissionSource,
    Inventory,
    OperationalImportProfile,
    SessionLocal,
    init_db,
)
from app.main import app
from app.operational_imports import (
    apply_operational_batch,
    build_operational_template,
    create_operational_batch,
    inspect_import_file,
    operational_import_summary,
    suggest_mapping,
)


@pytest.fixture(autouse=True)
def fresh_database_v029():
    # La base semilla aislada se restaura desde tests/conftest.py.
    yield


def _login(client: TestClient, email: str = "consultor@calculatuhuella.local") -> None:
    response = client.post("/login", data={"email": email, "password": "Demo2026!"}, follow_redirects=False)
    assert response.status_code == 303


def _inventory_and_source(session):
    inventory = session.scalar(
        select(Inventory)
        .where(
            Inventory.organization_id == 1,
            Inventory.locked.is_(False),
            Inventory.status != "Cerrado",
        )
        .order_by(Inventory.start_date.desc(), Inventory.id.desc())
        .limit(1)
    )
    assert inventory
    source = EmissionSource(
        inventory_id=inventory.id,
        facility_id=inventory.facility_links[0].facility_id if inventory.facility_links else None,
        name="Electricidad prueba V029",
        scope=2,
        category="Electricidad adquirida",
        responsible="Pruebas",
        materiality="Alta",
        data_frequency="Mensual",
        preferred_unit="kWh",
        included=True,
    )
    session.add(source)
    session.flush()
    return inventory, source


def _csv(
    source_id: int,
    value: str = "1200",
    evidence: str = "Factura enero",
    period_start: str = "2025-01-01",
    period_end: str = "2025-01-31",
) -> bytes:
    return (
        "Fuente ID;Fecha inicio;Fecha fin;Consumo;Unidad;Origen;Estimado;Evidencia;Notas\n"
        f"{source_id};{period_start};{period_end};{value};kWh;Factura;No;{evidence};Dato de prueba\n"
    ).encode("utf-8")


def _mapping():
    return {
        "source": "Fuente ID",
        "facility": "",
        "period_start": "Fecha inicio",
        "period_end": "Fecha fin",
        "value": "Consumo",
        "unit": "Unidad",
        "origin": "Origen",
        "estimated": "Estimado",
        "evidence": "Evidencia",
        "notes": "Notas",
    }


def test_v029_health_page_and_product_module():
    with TestClient(app) as client:
        assert client.get("/api/health").json()["version"] == "1.0.0"
        _login(client)
        page = client.get("/cargas-operativas")
        assert page.status_code == 200
        assert "Cargas operativas configurables" in page.text
        modules = client.get("/modulos")
        assert modules.status_code == 200
        assert "Cargas operativas configurables" in modules.text


def test_v029_csv_inspection_detects_delimiter_and_mapping():
    content = _csv(99)
    result = inspect_import_file(content, "consumos.csv")
    assert result["source_format"] == "CSV"
    assert result["delimiter"] == ";"
    assert result["rows"][0]["payload"]["Consumo"] == "1200"
    mapping = suggest_mapping(result["headers"])
    assert mapping["source"] == "Fuente ID"
    assert mapping["period_start"] == "Fecha inicio"
    assert mapping["value"] == "Consumo"


def test_v029_template_contains_inventory_catalog():
    with SessionLocal() as session:
        inventory = session.scalar(
            select(Inventory)
            .where(Inventory.organization_id == 1, Inventory.locked.is_(False))
            .order_by(Inventory.start_date.desc(), Inventory.id.desc())
            .limit(1)
        )
        content = build_operational_template(inventory)
        workbook = load_workbook(BytesIO(content), data_only=True)
        assert {"Carga operativa", "Catálogo de fuentes", "Instrucciones"}.issubset(workbook.sheetnames)
        assert workbook["Catálogo de fuentes"].max_row > 1


def test_v029_batch_validation_does_not_modify_inventory_and_saves_profile():
    with SessionLocal() as session:
        inventory, source = _inventory_and_source(session)
        before = session.scalar(select(func.count(ActivityData.id)))
        batch = create_operational_batch(
            session,
            organization_id=1,
            inventory=inventory,
            filename="consumos.csv",
            content=_csv(source.id),
            user_email="consultor@test",
            mapping=_mapping(),
            defaults={"duplicate_policy": "reject"},
            profile_name="Energía mensual",
            save_profile=True,
        )
        session.commit()
        after = session.scalar(select(func.count(ActivityData.id)))
        assert before == after
        assert batch.status == "Validado"
        assert batch.error_rows == 0
        assert session.scalar(select(OperationalImportProfile).where(OperationalImportProfile.name == "Energía mensual"))


def test_v029_apply_creates_activity_data_and_updates_batch():
    with SessionLocal() as session:
        inventory, source = _inventory_and_source(session)
        batch = create_operational_batch(
            session,
            organization_id=1,
            inventory=inventory,
            filename="consumos.csv",
            content=_csv(source.id),
            user_email="consultor@test",
            mapping=_mapping(),
            defaults={"duplicate_policy": "reject"},
        )
        session.commit()
        applied = apply_operational_batch(session, 1, batch.id, "consultor@test")
        session.commit()
        record = session.scalar(select(ActivityData).where(ActivityData.source_id == source.id))
        assert record and record.value == 1200
        assert record.unit == "kWh"
        assert applied.status == "Aplicado"
        assert applied.applied_rows == 1


def test_v029_duplicate_policies_reject_skip_and_update():
    with SessionLocal() as session:
        inventory, source = _inventory_and_source(session)
        existing = ActivityData(
            source_id=source.id,
            period_start=inventory.start_date,
            period_end=inventory.start_date,
            value=100,
            unit="kWh",
            created_by="seed",
        )
        # Use a period inside the active demo inventory.
        existing.period_start = __import__("datetime").date(2025, 1, 1)
        existing.period_end = __import__("datetime").date(2025, 1, 31)
        session.add(existing)
        session.commit()

        reject = create_operational_batch(
            session,
            organization_id=1,
            inventory=inventory,
            filename="reject.csv",
            content=_csv(source.id, "200"),
            user_email="consultor@test",
            mapping=_mapping(),
            defaults={"duplicate_policy": "reject"},
        )
        session.commit()
        assert reject.error_rows == 1

        skip_content = _csv(source.id, "300") + b" "  # distinct file hash
        skip = create_operational_batch(
            session,
            organization_id=1,
            inventory=inventory,
            filename="skip.csv",
            content=skip_content,
            user_email="consultor@test",
            mapping=_mapping(),
            defaults={"duplicate_policy": "skip"},
        )
        session.commit()
        assert skip.rows[0].status == "Omitir"
        apply_operational_batch(session, 1, skip.id, "consultor@test")
        session.commit()
        session.refresh(existing)
        assert existing.value == 100

        update_content = _csv(source.id, "450") + b"  "
        update = create_operational_batch(
            session,
            organization_id=1,
            inventory=inventory,
            filename="update.csv",
            content=update_content,
            user_email="consultor@test",
            mapping=_mapping(),
            defaults={"duplicate_policy": "update"},
        )
        session.commit()
        assert update.error_rows == 0
        apply_operational_batch(session, 1, update.id, "consultor@test")
        session.commit()
        session.refresh(existing)
        assert existing.value == 450


def test_v029_invalid_rows_are_blocked_and_reported():
    with SessionLocal() as session:
        inventory, source = _inventory_and_source(session)
        batch = create_operational_batch(
            session,
            organization_id=1,
            inventory=inventory,
            filename="negative.csv",
            content=_csv(source.id, "-1", ""),
            user_email="consultor@test",
            mapping=_mapping(),
            defaults={"duplicate_policy": "reject"},
        )
        session.commit()
        assert batch.status == "Con errores"
        assert batch.error_rows == 1
        with pytest.raises(ValueError, match="contiene errores"):
            apply_operational_batch(session, 1, batch.id, "consultor@test")


def test_v029_summary_separates_operational_from_pilot_batches():
    with SessionLocal() as session:
        inventory, source = _inventory_and_source(session)
        batch = create_operational_batch(
            session,
            organization_id=1,
            inventory=inventory,
            filename="summary.csv",
            content=_csv(source.id),
            user_email="consultor@test",
            mapping=_mapping(),
            defaults={"duplicate_policy": "reject"},
        )
        session.commit()
        summary = operational_import_summary(session, 1, inventory.id, batch.id)
        assert summary["metrics"]["batches"] == 1
        assert summary["selected"].id == batch.id
        assert summary["selected"].rows[0].validation_display == ""


def test_v029_browser_preview_and_template_download():
    with TestClient(app) as client:
        _login(client)
        with SessionLocal() as session:
            inventory, source = _inventory_and_source(session)
            session.commit()
            inventory_id = inventory.id
            source_id = source.id
        download = client.get(f"/cargas-operativas/plantilla.xlsx?inventory_id={inventory_id}")
        assert download.status_code == 200
        assert download.content.startswith(b"PK")
        response = client.post(
            "/cargas-operativas/previsualizar",
            data={"inventory_id": str(inventory_id), "profile_id": ""},
            files={"file": ("consumos.csv", _csv(source_id), "text/csv")},
            follow_redirects=True,
        )
        assert response.status_code == 200
        assert "Mapear columnas de consumos.csv" in response.text
        assert "Consumo" in response.text


def test_v029_verifier_can_view_but_cannot_upload():
    with TestClient(app) as client:
        _login(client, "verificador@calculatuhuella.local")
        page = client.get("/cargas-operativas")
        assert page.status_code == 200
        with SessionLocal() as session:
            inventory = session.scalar(
                select(Inventory)
                .where(Inventory.organization_id == 1, Inventory.locked.is_(False))
                .order_by(Inventory.start_date.desc(), Inventory.id.desc())
                .limit(1)
            )
        response = client.post(
            "/cargas-operativas/previsualizar",
            data={"inventory_id": str(inventory.id), "profile_id": ""},
            files={"file": ("consumos.csv", b"Fuente ID,Valor\n1,2", "text/csv")},
        )
        assert response.status_code == 403
