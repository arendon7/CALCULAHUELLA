from __future__ import annotations

from datetime import UTC, date, datetime
from io import BytesIO
import secrets

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from app.database import (
    Base,
    ENGINE,
    EmissionSource,
    Inventory,
    Scope3CategoryAssessment,
    SessionLocal,
    Supplier,
    SupplierCampaign,
    SupplierDataRequest,
    SupplierResponse,
    init_db,
)
from app.main import app
from app.scope3_catalog import SCOPE3_CATEGORIES, canonical_category_label, category_from_value
from app.supply_chain import inventory_supply_chain_summary, validate_supplier_response


@pytest.fixture(autouse=True)
def fresh_database_iteration6():
    # La base semilla aislada se restaura desde tests/conftest.py.
    yield


def login(client: TestClient, email: str = "admin@calculatuhuella.local") -> None:
    response = client.post(
        "/login",
        data={"email": email, "password": "Demo2026!"},
        follow_redirects=False,
    )
    assert response.status_code == 303


def test_scope3_catalog_has_all_categories_and_directions():
    assert len(SCOPE3_CATEGORIES) == 15
    assert len([item for item in SCOPE3_CATEGORIES if item.direction == "Aguas arriba"]) == 8
    assert len([item for item in SCOPE3_CATEGORIES if item.direction == "Aguas abajo"]) == 7
    assert category_from_value("Categoría 11 · Uso de productos vendidos").code == "C11"
    assert canonical_category_label("Transporte aguas arriba").startswith("C4 ·")


def test_scope3_screening_is_created_without_changing_supplier_total():
    with SessionLocal() as session:
        inventory = session.scalar(select(Inventory).order_by(Inventory.id))
        source_before = session.scalar(select(EmissionSource).where(EmissionSource.category == "Datos específicos de proveedores"))
        summary = inventory_supply_chain_summary(session, inventory)
        session.commit()
        assert session.scalar(select(func.count()).select_from(Scope3CategoryAssessment).where(
            Scope3CategoryAssessment.inventory_id == inventory.id
        )) == 15
        assert summary["assessed_category_count"] == 1
        assert summary["material_category_count"] == 1
        assert len(summary["categories"]) == 15
        c1 = next(item for item in summary["categories"] if item["code"] == "C1")
        assert c1["assessment_status"] == "Material"
        assert round(source_before.emissions, 1) == 340.8


def test_scope3_page_api_and_assessment_workflow():
    with TestClient(app) as client:
        login(client)
        page = client.get("/cadena-valor")
        assert page.status_code == 200
        assert "Las 15 categorías de Alcance 3" in page.text
        assert "C15 · Inversiones" in page.text
        api = client.get("/api/cadena-valor/resumen")
        assert api.status_code == 200
        payload = api.json()
        assert len(payload["categories"]) == 15
        assert payload["assessed_category_count"] == 1
        response = client.post(
            "/cadena-valor/categorias/C5/evaluar",
            data={
                "status": "No aplica",
                "relevance_score": "1",
                "rationale": "La organización no entrega residuos a gestores externos en el periodo.",
                "owner": "Gestión ambiental",
                "data_strategy": "Declaración operativa anual",
            },
            follow_redirects=False,
        )
        assert response.status_code == 303
    with SessionLocal() as session:
        assessment = session.scalar(select(Scope3CategoryAssessment).where(
            Scope3CategoryAssessment.inventory_id == 1,
            Scope3CategoryAssessment.category_code == "C5",
        ))
        assert assessment.status == "No aplica"
        assert assessment.relevance_score == 1
        assert "residuos" in assessment.rationale.lower()


def test_supplier_factor_unit_mismatch_is_rejected():
    with SessionLocal() as session:
        data_request = session.scalar(
            select(SupplierDataRequest).join(Supplier).where(Supplier.name == "Químicos Andinos Ltda.")
        )
        validation = validate_supplier_response(
            data_request,
            method="Factor por unidad",
            activity_value=96,
            activity_unit="t",
            emission_factor=950,
            factor_unit="kg CO2e/kg",
            reported_emissions_tco2e=0,
            methodology="Huella de producto específica",
            boundary="Cradle-to-gate",
        )
        assert validation["errors"]
        token = data_request.access_token
    with TestClient(app) as client:
        response = client.post(
            f"/proveedor/responder/{token}",
            data={
                "method": "Factor por unidad",
                "activity_value": "96",
                "activity_unit": "t",
                "emission_factor": "950",
                "factor_unit": "kg CO2e/kg",
                "reported_emissions_tco2e": "0",
                "methodology": "Huella de producto específica",
                "boundary": "Cradle-to-gate",
            },
            follow_redirects=False,
        )
        assert response.status_code == 400
        assert "no coincide" in response.text


def test_duplicate_supplier_response_cannot_be_approved():
    with SessionLocal() as session:
        inventory = session.scalar(select(Inventory).order_by(Inventory.id))
        supplier = session.scalar(select(Supplier).where(Supplier.name == "Acero Circular S.A.S."))
        campaign = SupplierCampaign(
            inventory_id=inventory.id,
            name="Duplicado de control",
            category="C1 · Bienes y servicios adquiridos",
            due_date=date(2026, 11, 30),
            status="En curso",
            methodology="GHG Protocol Scope 3",
            created_by="admin@calculatuhuella.local",
        )
        session.add(campaign)
        session.flush()
        request = SupplierDataRequest(
            campaign_id=campaign.id,
            supplier_id=supplier.id,
            product_service="Acero laminado",
            quantity=420,
            unit="t",
            spend_cop=2_450_000_000,
            status="Respondida",
            due_date=date(2026, 11, 30),
            access_token=secrets.token_urlsafe(24),
            token_expires_at=datetime(2026, 11, 30, tzinfo=UTC),
        )
        session.add(request)
        session.flush()
        response = SupplierResponse(
            request_id=request.id,
            method="Factor por unidad",
            activity_value=420,
            activity_unit="t",
            emission_factor=710,
            factor_unit="kg CO2e/t",
            calculated_emissions_tco2e=298.2,
            methodology="EPD del producto",
            boundary="Cradle-to-gate",
            quality_level="C",
            review_status="Pendiente",
        )
        session.add(response)
        session.commit()
        response_id = response.id
    with TestClient(app) as client:
        login(client, "revisor@calculatuhuella.local")
        result = client.post(
            f"/cadena-valor/respuestas/{response_id}/revisar",
            data={"decision": "Aprobado", "reviewer_comments": "Prueba"},
            follow_redirects=False,
        )
        assert result.status_code == 409
        assert "otra respuesta aprobada" in result.text


def test_scope3_workbook_contains_screening_sheet():
    with TestClient(app) as client:
        login(client)
        response = client.get("/cadena-valor/plantilla.xlsx")
        assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["Solicitudes proveedores", "Screening 15 categorías"]
    screening = workbook["Screening 15 categorías"]
    assert screening.max_row == 16
    assert screening["A2"].value == "C1"
    assert screening["D2"].value == "Material"
