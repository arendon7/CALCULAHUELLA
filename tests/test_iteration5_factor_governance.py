from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import selectinload

from app.colombia_library import build_colombia_workbook, colombia_library_summary
from app.database import (
    Base,
    ENGINE,
    EmissionFactor,
    EmissionFactorVersion,
    FactorDocumentation,
    MethodologySourceDocument,
    SessionLocal,
    init_db,
)
from app.factor_library import factor_catalog, factor_passport
from app.main import app


@pytest.fixture(autouse=True)
def fresh_database_iteration5():
    # La base semilla aislada se restaura desde tests/conftest.py.
    yield


def _login(client: TestClient) -> None:
    response = client.post(
        "/login",
        data={"email": "consultor@calculatuhuella.local", "password": "Demo2026!"},
        follow_redirects=False,
    )
    assert response.status_code == 303


def _official_electricity_version(session):
    return session.scalar(
        select(EmissionFactorVersion)
        .where(EmissionFactorVersion.version == "UPME-2024-R085")
        .options(
            selectinload(EmissionFactorVersion.factor),
            selectinload(EmissionFactorVersion.gas),
        )
    )


def test_iteration5_registers_xm_preliminary_source_without_promoting_factor():
    with SessionLocal() as session:
        source = session.scalar(
            select(MethodologySourceDocument).where(
                MethodologySourceDocument.code == "XM-SIN-2025-PRELIM"
            )
        )
        assert source is not None
        assert "preliminar" in source.status.lower()
        linked = session.scalar(
            select(func.count(FactorDocumentation.id)).where(
                FactorDocumentation.source_document_id == source.id
            )
        )
        assert linked == 0


def test_iteration5_upme_factor_is_level_two_and_source_is_controlled():
    with SessionLocal() as session:
        version = _official_electricity_version(session)
        passport = factor_passport(session, version)
        assert passport["hierarchy_tier"] == 2
        assert passport["hierarchy_label"] == "Oficial nacional"
        assert passport["source_status"] == "Vigente"
        assert passport["source_preliminary"] is False
        assert passport["formal"] is True


def test_iteration5_catalog_exposes_governance_metrics_and_filters():
    with SessionLocal() as session:
        summary = factor_catalog(session)
        assert summary["metrics"]["official_national"] >= 1
        assert summary["metrics"]["preliminary_sources"] >= 1
        assert len(summary["hierarchy_policy"]) == 6
        filtered = factor_catalog(session, hierarchy=2)
        assert filtered["items"]
        assert all(item["hierarchy_tier"] == 2 for item in filtered["items"])


def test_iteration5_api_and_pages_show_governance_without_changing_api_version():
    with TestClient(app) as client:
        _login(client)
        page = client.get("/metodologia/biblioteca-factores")
        assert page.status_code == 200
        assert "Jerarquía metodológica" in page.text
        assert "Fuentes preliminares" in page.text
        response = client.get("/api/metodologia/biblioteca-factores")
        assert response.status_code == 200
        payload = response.json()
        assert payload["version"] == "1.0.0"
        assert payload["governance_version"] == "1.1.0"
        official = next(item for item in payload["items"] if item["hierarchy_tier"] == 2)
        assert "temporal_alignment" in official
        assert official["source_status"] == "Vigente"


def test_iteration5_colombia_registry_marks_preliminary_document_as_not_in_calculation():
    with SessionLocal() as session:
        summary = colombia_library_summary(session)
        xm = next(item for item in summary["documents"] if item.code == "XM-SIN-2025-PRELIM")
        assert summary["document_usage"][xm.id] is False
        assert summary["counts"]["preliminary_documents"] >= 1
        assert summary["counts"]["official_national"] >= 1
        workbook = load_workbook(BytesIO(build_colombia_workbook(summary)))
        assert workbook.sheetnames == ["Factores", "Fuentes", "Casos patrón", "Limitaciones", "Gobierno"]
        rows = list(workbook["Fuentes"].iter_rows(values_only=True))
        xm_row = next(row for row in rows if row[0] == "XM-SIN-2025-PRELIM")
        assert xm_row[6] == "No"


def test_iteration5_does_not_add_or_modify_calculable_factor_values():
    with SessionLocal() as session:
        version = _official_electricity_version(session)
        assert version.value == 0.220
        total_versions = session.scalar(select(func.count(EmissionFactorVersion.id)))
        total_factors = session.scalar(select(func.count(EmissionFactor.id)))
        assert total_versions == 25
        assert total_factors == 25
