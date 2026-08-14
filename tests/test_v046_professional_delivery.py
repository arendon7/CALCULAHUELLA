from __future__ import annotations

from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app.database import Base, ENGINE, Inventory, ReportArtifact, SessionLocal, init_db
from app.delivery_readiness import professional_delivery_summary
from app.main import app
from app.product_experience import CORE_SECTIONS, navigation_for
from app.reporting import generate_calculation_workbook, generate_decision_brief_pdf, generate_executive_pdf, generate_technical_pdf


@pytest.fixture(autouse=True)
def fresh_database():
    # La base semilla aislada se restaura desde tests/conftest.py.
    yield


def login(client: TestClient, email: str = "consultor@calculatuhuella.local") -> None:
    response = client.post(
        "/login",
        data={"email": email, "password": "Demo2026!"},
        follow_redirects=False,
    )
    assert response.status_code == 303


def test_v046_delivery_summary_is_explainable_and_complete():
    with SessionLocal() as session:
        inventory = session.scalar(select(Inventory).where(Inventory.id == 1))
        summary = professional_delivery_summary(session, inventory)
        assert 0 <= summary["score"] <= 100
        assert len(summary["gates"]) == 8
        assert {item["status"] for item in summary["gates"]} <= {"Listo", "En progreso", "Bloqueado"}
        assert summary["narrative"]["headline"]
        assert summary["narrative"]["findings"]
        assert len(summary["deliverables"]) == 5
        assert summary["metrics"]["calculation_count"] > 0


def test_v046_delivery_page_api_and_reports_page_load():
    with TestClient(app) as client:
        login(client)
        response = client.get("/entrega-profesional")
        assert response.status_code == 200
        assert "NIVEL DE PUBLICACIÓN" in response.text
        assert "OCHO PUERTAS DE CONTROL" in response.text
        assert "Control de publicación" in response.text

        payload = client.get("/api/entrega-profesional/resumen")
        assert payload.status_code == 200
        assert payload.json()["inventory_id"] == 1
        assert len(payload.json()["gates"]) == 8

        reports = client.get("/reportes")
        assert reports.status_code == 200
        assert "Informe controlado · expediente trazable" in reports.text
        assert "ALISTAMIENTO" in reports.text
        assert "PUBLICACIÓN" in reports.text


def test_v046_working_versions_remain_available_before_final_publication():
    with SessionLocal() as session:
        inventory = session.scalar(select(Inventory).where(Inventory.id == 1))
        inventory.status = "En preparación"
        session.commit()

    with TestClient(app) as client:
        login(client)
        delivery_page = client.get("/entrega-profesional")
        reports_page = client.get("/reportes")
        assert delivery_page.status_code == 200
        assert reports_page.status_code == 200
        assert delivery_page.text.count('data-document-mode="working"') == 5
        assert reports_page.text.count('data-document-mode="working"') == 5
        assert "Cinco documentos autocontenidos y trazables" in delivery_page.text
        assert "Completa los controles requeridos" not in delivery_page.text
        assert "no quedan habilitadas para emisión final" in reports_page.text
        assert "no sustituye la aprobación o cierre del inventario" in reports_page.text

        generated = client.post(
            "/reportes/generar",
            data={"inventory_id": 1, "report_type": "ficha"},
            follow_redirects=False,
        )
        assert generated.status_code == 303

    with SessionLocal() as session:
        artifact = session.scalar(select(ReportArtifact).order_by(ReportArtifact.id.desc()))
        assert artifact is not None
        assert artifact.status == "Borrador"


@pytest.mark.parametrize(
    "email",
    ["cliente@calculatuhuella.local", "verificador@calculatuhuella.local"],
)
def test_v046_read_only_roles_receive_consultation_not_generation_language(email: str):
    with SessionLocal() as session:
        inventory = session.scalar(select(Inventory).where(Inventory.id == 1))
        inventory.status = "En preparación"
        session.commit()

    with TestClient(app) as client:
        login(client, email)
        reports = client.get("/reportes")
        assert reports.status_code == 200
        assert "Consulta el estado de publicación, las versiones disponibles y su trazabilidad" in reports.text
        assert "Entregables para consulta" in reports.text
        assert "Genera la ficha ejecutiva" not in reports.text
        assert 'action="/reportes/generar"' not in reports.text
        assert 'href="/control"' not in reports.text


def test_v046_navigation_exposes_professional_delivery():
    delivery_items = [
        item
        for section in CORE_SECTIONS
        for item in section["items"]
        if item["href"] == "/entrega-profesional"
    ]
    assert len(delivery_items) == 1
    navigation = navigation_for(
        {
            "role": "Consultor",
            "capabilities": {
                "manage_org", "view_methodology", "manage_portfolio", "view_consolidation",
                "provide_data", "manage_sources", "review", "approve", "external_audit",
            },
        },
        "essential",
    )
    assert any(
        item["href"] == "/entrega-profesional"
        for section in navigation["core"]
        for item in section["items"]
    )


def test_v046_documents_include_delivery_control(tmp_path: Path):
    with SessionLocal() as session:
        inventory = session.scalar(select(Inventory).where(Inventory.id == 1))
        brief = tmp_path / "brief.pdf"
        executive = tmp_path / "executive.pdf"
        technical = tmp_path / "technical.pdf"
        workbook = tmp_path / "calculation.xlsx"
        generate_decision_brief_pdf(session, inventory, brief)
        generate_executive_pdf(session, inventory, executive)
        generate_technical_pdf(session, inventory, technical)
        generate_calculation_workbook(session, inventory, workbook)

    assert brief.read_bytes().startswith(b"%PDF")
    assert executive.read_bytes().startswith(b"%PDF")
    assert technical.read_bytes().startswith(b"%PDF")
    assert brief.stat().st_size > 4_000
    assert executive.stat().st_size > 5_000
    assert technical.stat().st_size > 5_000
    book = load_workbook(workbook, read_only=True, data_only=True)
    assert "Control de entrega" in book.sheetnames
    rows = list(book["Control de entrega"].iter_rows(values_only=True))
    assert rows[0] == ("Puerta", "Estado", "Responsable", "Detalle", "Criterio de aceptación", "Acción", "Ruta")
    assert len(rows) >= 11
