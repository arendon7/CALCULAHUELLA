from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from app.database import (
    ActivityData,
    Base,
    DataRequest,
    ENGINE,
    Inventory,
    PilotExecution,
    PilotExecutionSourceLink,
    PilotIssue,
    SessionLocal,
    init_db,
)
from app.main import app
from app.pilot_execution import (
    build_pilot_execution_workbook,
    guided_workspace,
    import_pilot_workbook,
    pilot_execution_summary,
    start_pilot_execution,
)


@pytest.fixture(autouse=True)
def fresh_database_v025():
    # La base semilla aislada se restaura desde tests/conftest.py.
    yield


def _login(client: TestClient, email: str = "consultor@calculatuhuella.local") -> None:
    response = client.post("/login", data={"email": email, "password": "Demo2026!"}, follow_redirects=False)
    assert response.status_code == 303


def test_v025_health_version():
    with TestClient(app) as client:
        response = client.get("/api/health")
        assert response.status_code == 200
        assert response.json()["version"] == "1.0.0"


def test_v025_start_pilot_creates_inventory_sources_requests_and_issues():
    with SessionLocal() as session:
        execution = start_pilot_execution(session, 1, "consultor@test", "Consultor prueba")
        session.commit()
        assert execution.inventory_id is not None
        assert session.scalar(select(func.count(PilotExecutionSourceLink.id)).where(PilotExecutionSourceLink.execution_id == execution.id)) >= 15
        assert session.scalar(select(func.count(DataRequest.id)).where(DataRequest.inventory_id == execution.inventory_id)) >= 15
        assert session.scalar(select(func.count(PilotIssue.id)).where(PilotIssue.execution_id == execution.id)) == 3
        inventory = session.get(Inventory, execution.inventory_id)
        assert inventory.version == "0.45"
        assert inventory.current_stage == "Recolección"


def test_v025_start_is_idempotent():
    with SessionLocal() as session:
        first = start_pilot_execution(session, 1, "consultor@test", "Consultor prueba")
        session.commit()
        second = start_pilot_execution(session, 1, "consultor@test", "Consultor prueba")
        session.commit()
        assert first.id == second.id
        assert session.scalar(select(func.count(PilotExecution.id))) == 1


def test_v025_execution_page_and_guided_dashboard_are_available():
    with TestClient(app) as client:
        _login(client)
        start = client.post("/piloto-greenatics/ejecucion/iniciar", data={"target_date": "2026-10-31"}, follow_redirects=False)
        assert start.status_code == 303
        page = client.get("/piloto-greenatics/ejecucion")
        assert page.status_code == 200
        assert "Ejecución controlada del piloto" in page.text
        assert "Datos mensuales del piloto" in page.text
        dashboard = client.get("/dashboard")
        assert dashboard.status_code == 200
        assert "TU SIGUIENTE ACCIÓN" in dashboard.text
        assert "Seis etapas, una ruta clara" in dashboard.text


def test_v025_workbook_import_creates_activity_data_and_calculates_electricity():
    with SessionLocal() as session:
        execution = start_pilot_execution(session, 1, "consultor@test", "Consultor prueba")
        session.commit()
        summary = pilot_execution_summary(session, 1)
        content = build_pilot_execution_workbook(summary)
        workbook = load_workbook(BytesIO(content))
        sheet = workbook["Datos mensuales"]
        target_row = None
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 1).value == "YAR-ELEC":
                target_row = row
                break
        assert target_row is not None
        sheet.cell(target_row, 4).value = 1000
        stream = BytesIO()
        workbook.save(stream)
        result = import_pilot_workbook(session, 1, stream.getvalue(), "consultor@test")
        session.commit()
        assert result["imported"] == 1
        assert not result["errors"]
        refreshed = pilot_execution_summary(session, 1)
        electricity = next(link for link in refreshed["links"] if link.requirement.code == "YAR-ELEC")
        assert session.scalar(select(func.count(ActivityData.id)).where(ActivityData.source_id == electricity.source_id)) == 1
        assert electricity.source.progress > 0
        assert electricity.source.emissions > 0


def test_v025_independent_comparison_marks_conformity():
    with SessionLocal() as session:
        execution = start_pilot_execution(session, 1, "consultor@test", "Consultor prueba")
        session.commit()
        execution.independent_total_tco2e = 0
        summary = pilot_execution_summary(session, 1)
        session.commit()
        assert summary["metrics"]["variance"] == 0
        assert execution.comparison_status == "Conforme"


def test_v025_guided_workspace_returns_role_specific_actions():
    with SessionLocal() as session:
        inventory = session.scalar(select(Inventory).where(Inventory.organization_id == 1).order_by(Inventory.id.desc()))
        user = {
            "role": "Consultor",
            "organization_id": 1,
            "capabilities": {"manage_inventory", "view_methodology"},
        }
        workspace = guided_workspace(session, user, inventory)
        assert workspace["total"] == 6
        assert 0 <= workspace["score"] <= 100
        assert workspace["actions"]
        assert any("piloto Greenatics" in action["title"] for action in workspace["actions"])


def test_v025_approval_remains_blocked_until_pilot_is_complete():
    with TestClient(app) as client:
        _login(client, "revisor@calculatuhuella.local")
        client.post("/piloto-greenatics/ejecucion/iniciar", data={}, follow_redirects=False)
        response = client.post("/piloto-greenatics/ejecucion/aprobar", data={}, follow_redirects=False)
        assert response.status_code == 409
        assert "No es posible aprobar" in response.text
