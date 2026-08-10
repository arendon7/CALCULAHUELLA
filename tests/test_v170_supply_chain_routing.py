from __future__ import annotations

from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app

ROOT = Path(__file__).resolve().parents[1]

SUPPLY_CHAIN_ROUTES = {
    ("GET", "/cadena-valor"),
    ("GET", "/api/cadena-valor/resumen"),
    ("POST", "/cadena-valor/categorias/{category_code}/evaluar"),
    ("POST", "/cadena-valor/proveedores/nuevo"),
    ("POST", "/cadena-valor/campanas/nueva"),
    ("POST", "/cadena-valor/solicitudes/nueva"),
    ("POST", "/cadena-valor/solicitudes/{request_id}/renovar"),
    ("GET", "/proveedor/responder/{token}"),
    ("POST", "/proveedor/responder/{token}"),
    ("POST", "/cadena-valor/respuestas/{response_id}/revisar"),
    ("GET", "/cadena-valor/respuestas/{response_id}/evidencia"),
    ("GET", "/cadena-valor/plantilla.xlsx"),
}


def _login(client: TestClient) -> None:
    response = client.post(
        "/login",
        data={"email": "admin@calculatuhuella.local", "password": "Demo2026!"},
        follow_redirects=False,
    )
    assert response.status_code == 303


def test_v170_supply_chain_has_dedicated_http_authority():
    main_source = (ROOT / "app/main.py").read_text(encoding="utf-8")
    module_source = (ROOT / "app/supply_chain_web.py").read_text(encoding="utf-8")
    assert '@app.get("/cadena-valor"' not in main_source
    assert '@app.post("/cadena-valor/' not in main_source
    assert '@app.get("/proveedor/responder/' not in main_source
    assert '@app.post("/proveedor/responder/' not in main_source
    assert "register_supply_chain_routes(" in main_source
    assert module_source.count("@app.") == 12
    assert "from .supply_chain import" in module_source
    assert "def calculate_supplier_response(" not in module_source
    assert "def validate_supplier_response(" not in module_source


def test_v170_supply_chain_route_contract_is_unique_and_complete():
    actual = []
    relevant_paths = {path for _, path in SUPPLY_CHAIN_ROUTES}
    for route in app.routes:
        path = getattr(route, "path", None)
        methods = getattr(route, "methods", set()) or set()
        if path in relevant_paths:
            actual.extend((method, path) for method in methods if method in {"GET", "POST"})
    assert set(actual) == SUPPLY_CHAIN_ROUTES
    assert len(actual) == len(SUPPLY_CHAIN_ROUTES)


def test_v170_supply_chain_page_api_and_workbook_remain_available():
    with TestClient(app) as client:
        _login(client)
        page = client.get("/cadena-valor")
        api = client.get("/api/cadena-valor/resumen")
        workbook_response = client.get("/cadena-valor/plantilla.xlsx")
        assert page.status_code == 200
        assert api.status_code == 200
        assert len(api.json()["categories"]) == 15
        assert workbook_response.status_code == 200
        assert workbook_response.content[:2] == b"PK"
        workbook = load_workbook(BytesIO(workbook_response.content), data_only=True)
        assert workbook.sheetnames == ["Solicitudes proveedores", "Screening 15 categorías"]
