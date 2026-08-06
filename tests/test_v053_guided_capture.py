from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from app.database import ActivityData, Base, ENGINE, EvidenceDocument, SessionLocal, init_db
from app.main import app

ROOT = Path(__file__).resolve().parents[1]


@pytest.fixture(autouse=True)
def fresh_database_v053():
    # La base semilla aislada se restaura desde tests/conftest.py.
    yield


def login(client: TestClient, email: str = "cliente@calculatuhuella.local") -> None:
    response = client.post("/login", data={"email": email, "password": "Demo2026!"}, follow_redirects=False)
    assert response.status_code == 303


def next_pending(client: TestClient) -> dict:
    body = client.get("/api/captura-guiada").json()
    return next(item for item in body["sources"] if item["next_start"])


def test_v053_guided_capture_prioritizes_periods_and_explains_evidence():
    with TestClient(app) as client:
        login(client)
        response = client.get("/captura-guiada")
        assert response.status_code == 200
        assert "Captura guiada de información" in response.text
        assert "PLAN DE CAPTURA" in response.text
        assert "SOPORTE RECOMENDADO" in response.text
        assert "Una copia del periodo anterior" in response.text
        api = client.get("/api/captura-guiada")
        assert api.status_code == 200
        body = api.json()
        assert 0 <= body["coverage"] <= 100
        assert body["sources"]
        assert all("expected_evidence" in item for item in body["sources"])


def test_v053_integrated_capture_creates_data_and_evidence_atomically():
    with TestClient(app) as client:
        login(client)
        pending = next_pending(client)
        response = client.post(
            "/captura-guiada/registrar",
            data={
                "source_id": str(pending["id"]),
                "period_start": pending["next_start"],
                "period_end": pending["next_end"],
                "value": "123.45",
                "unit": pending["expected_unit"],
                "data_origin": "Factura",
                "document_type": "Factura",
                "uncertainty_percentage": "3",
                "uncertainty_basis": "Factura conciliada",
                "notes": "Captura integrada V0.53",
            },
            files={"evidence_file": ("factura.pdf", b"%PDF-1.4\n% guided capture\n", "application/pdf")},
            follow_redirects=False,
        )
        assert response.status_code == 303
        assert response.headers["location"].startswith("/captura-guiada?source_id=")
    with SessionLocal() as session:
        record = session.scalar(select(ActivityData).where(ActivityData.notes == "Captura integrada V0.53"))
        assert record is not None
        assert record.evidence_id is not None
        assert record.quality_level == "B"
        document = session.get(EvidenceDocument, record.evidence_id)
        assert document is not None
        assert document.sha256
        assert document.source_id == record.source_id


def test_v053_invalid_inline_evidence_does_not_create_partial_record():
    with TestClient(app) as client:
        login(client)
        pending = next_pending(client)
        with SessionLocal() as session:
            before = session.scalar(select(func.count()).select_from(ActivityData))
        response = client.post(
            "/captura-guiada/registrar",
            data={
                "source_id": str(pending["id"]),
                "period_start": pending["next_start"],
                "period_end": pending["next_end"],
                "value": "99",
                "unit": pending["expected_unit"],
                "data_origin": "Registro operativo",
                "notes": "No debe persistir",
            },
            files={"evidence_file": ("soporte.pdf", b"not-a-pdf", "application/pdf")},
            follow_redirects=False,
        )
        assert response.status_code == 303
        with SessionLocal() as session:
            after = session.scalar(select(func.count()).select_from(ActivityData))
            assert after == before
            assert session.scalar(select(ActivityData).where(ActivityData.notes == "No debe persistir")) is None


def test_v053_previous_period_is_only_a_prefill_reference():
    with TestClient(app) as client:
        login(client)
        body = client.get("/api/captura-guiada").json()
        candidate_ids = [item["id"] for item in body["sources"] if item["next_start"]]
        with SessionLocal() as session:
            record = session.scalar(select(ActivityData).where(ActivityData.source_id.in_(candidate_ids)).order_by(ActivityData.id))
            assert record is not None
            record_id = record.id
            before = session.scalar(select(func.count()).select_from(ActivityData))
        page = client.get(f"/captura-guiada?copy_record_id={record_id}#registrar")
        assert page.status_code == 200
        assert "No se ha creado ningún dato" in page.text
        assert "Referencia del periodo" in page.text
        assert "requiere validación" in page.text
    with SessionLocal() as session:
        after = session.scalar(select(func.count()).select_from(ActivityData))
        assert after == before


def test_v053_sectorial_template_has_plan_data_catalogs_and_instructions():
    with TestClient(app) as client:
        login(client)
        response = client.get("/informacion/plantilla.xlsx")
        assert response.status_code == 200
        workbook = load_workbook(BytesIO(response.content), read_only=False, data_only=True)
        assert workbook.sheetnames == ["Plan de captura", "Datos", "Catálogos", "Instrucciones"]
        assert workbook["Plan de captura"]["A1"].value == "Prioridad"
        assert workbook["Plan de captura"].max_row >= 2
        assert workbook["Datos"]["A1"].value == "Fuente"
        assert workbook["Instrucciones"]["B2"].value.startswith("Revisa el Plan")


def test_v053_navigation_and_release_files_are_present():
    product_experience = (ROOT / "app/product_experience.py").read_text(encoding="utf-8")
    information = (ROOT / "app/templates/information.html").read_text(encoding="utf-8")
    assert '"Captura guiada", "/captura-guiada"' in product_experience
    assert "/captura-guiada" in information
    assert (ROOT / "app/capture_guidance.py").is_file()
    assert (ROOT / "app/capture_web.py").is_file()
    assert (ROOT / "app/templates/guided_capture.html").is_file()
