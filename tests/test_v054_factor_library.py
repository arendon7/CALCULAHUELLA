from __future__ import annotations

from io import BytesIO
import json

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.database import (
    ActivityFactorSelection,
    Base,
    EmissionCalculation,
    EmissionSource,
    ENGINE,
    SessionLocal,
    init_db,
)
from app.main import app


@pytest.fixture(autouse=True)
def fresh_database():
    # La base semilla aislada se restaura desde tests/conftest.py.
    yield


def login(client: TestClient, email: str = "consultor@calculatuhuella.local") -> None:
    response = client.post(
        "/login",
        data={"email": email, "password": "Demo2026!"},
        follow_redirects=False,
    )
    assert response.status_code == 303


def electricity_context() -> tuple[int, int]:
    with SessionLocal() as session:
        source = session.scalar(
            select(EmissionSource)
            .where(EmissionSource.name == "Electricidad")
            .options(selectinload(EmissionSource.activity_records))
        )
        assert source and source.activity_records
        return source.id, source.activity_records[0].id


def test_v054_library_page_and_api_present_professional_catalog():
    with TestClient(app) as client:
        login(client)
        page = client.get("/metodologia/biblioteca-factores")
        assert page.status_code == 200
        assert "Biblioteca profesional de factores" in page.text
        assert "Comparación controlada" in page.text
        assert "Ver pasaporte" in page.text

        api = client.get("/api/metodologia/biblioteca-factores")
        assert api.status_code == 200
        payload = api.json()
        assert payload["version"] == "1.0.0"
        assert payload["metrics"]["total"] >= 25
        assert payload["metrics"]["formal"] >= 5
        assert any(item["reporting_use"] == "Demostrativo" for item in payload["items"])
        assert any(item["decision_readiness"] == "Listo para evaluación" for item in payload["items"])


def test_v054_filters_and_passport_expose_traceable_documentation():
    with TestClient(app) as client:
        login(client)
        filtered = client.get("/api/metodologia/biblioteca-factores?reporting_use=Formal&quality=A")
        assert filtered.status_code == 200
        payload = filtered.json()
        assert payload["items"]
        assert all(item["reporting_use"] == "Formal" and item["quality_grade"] == "A" for item in payload["items"])

        passport = client.get("/metodologia/biblioteca-factores/9")
        assert passport.status_code == 200
        assert "Cadena documental" in passport.text
        assert "Electricidad SIN Colombia" in passport.text
        assert "UPME" in passport.text
        assert "DOCUMENTACIÓN" in passport.text
        assert "Restricciones" in passport.text


def test_v054_contextual_catalog_scores_factor_against_one_datum():
    source_id, record_id = electricity_context()
    url = f"/metodologia/biblioteca-factores?source_id={source_id}&activity_data_id={record_id}"
    with TestClient(app) as client:
        login(client)
        page = client.get(url)
        assert page.status_code == 200
        assert "CONTEXTO ACTIVO" in page.text
        assert "Compatibilidad" in page.text

        api = client.get(f"/api/metodologia/biblioteca-factores?source_id={source_id}&activity_data_id={record_id}")
        payload = api.json()
        assert payload["context"] == {"source_id": source_id, "activity_data_id": record_id}
        formal_electricity = next(item for item in payload["items"] if item["id"] == 9)
        assert formal_electricity["calculable"] is True
        assert formal_electricity["compatibility_score"] >= 70


def test_v054_comparison_detects_incompatibilities_and_exports_workbook():
    with TestClient(app) as client:
        login(client)
        response = client.get("/metodologia/biblioteca-factores/comparar/seleccion?ids=9&ids=10")
        assert response.status_code == 200
        assert "Compara antes de decidir" in response.text
        assert "No mezcles un factor agregado en CO₂e" in response.text

        export = client.get("/metodologia/biblioteca-factores/comparar/exportar.xlsx?ids=9&ids=10")
        assert export.status_code == 200
        workbook = load_workbook(BytesIO(export.content), data_only=True)
        assert workbook.sheetnames == ["Comparación", "Alertas", "Contexto"]
        assert workbook["Comparación"].max_row == 3
        alerts = [workbook["Alertas"].cell(row=row, column=2).value for row in range(2, workbook["Alertas"].max_row + 1)]
        assert any("agregado en CO₂e" in str(item) for item in alerts)


def test_v054_demo_factor_can_be_proposed_but_cannot_be_approved_for_formal_use():
    source_id, record_id = electricity_context()
    with TestClient(app) as client:
        login(client)
        proposed = client.post(
            f"/fuentes/{source_id}/datos/{record_id}/factores/seleccionar",
            data={
                "factor_version_id": 1,
                "rationale": "Se propone únicamente para comparar el comportamiento del entorno demostrativo.",
            },
            follow_redirects=False,
        )
        assert proposed.status_code == 303

    with SessionLocal() as session:
        selection = session.scalar(select(ActivityFactorSelection).where(
            ActivityFactorSelection.activity_data_id == record_id,
            ActivityFactorSelection.factor_version_id == 1,
        ))
        assert selection is not None
        assert selection.selection_status == "Propuesto"
        snapshot = json.loads(selection.decision_snapshot)
        assert snapshot["hard_blockers"]
        selection_id = selection.id

    with TestClient(app) as client:
        login(client, "revisor@calculatuhuella.local")
        reviewed = client.post(
            f"/fuentes/{source_id}/datos/{record_id}/factores/{selection_id}/revisar",
            data={"decision": "Aprobar", "review_notes": "La unidad coincide, pero se evalúa su aptitud formal."},
            follow_redirects=False,
        )
        assert reviewed.status_code == 303

    with SessionLocal() as session:
        selection = session.get(ActivityFactorSelection, selection_id)
        assert selection.selection_status == "Propuesto"
        assert selection.applied_at is None
        calculations = list(session.scalars(select(EmissionCalculation).where(
            EmissionCalculation.activity_data_id == record_id,
            EmissionCalculation.factor_version_id == 1,
        )))
        assert calculations == []


def test_v054_source_page_links_professional_library_for_selected_record():
    source_id, record_id = electricity_context()
    with TestClient(app) as client:
        login(client)
        page = client.get(f"/fuentes/{source_id}?activity_data_id={record_id}")
        assert page.status_code == 200
        assert "Abrir biblioteca profesional con este dato" in page.text
        assert f"activity_data_id={record_id}" in page.text
        assert "Ver pasaporte metodológico" in page.text
